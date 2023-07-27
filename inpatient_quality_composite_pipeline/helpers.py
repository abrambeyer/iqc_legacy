# library for interacting with the web browser.  Requires downloading web drivers before use.
# Drivers can be found at:  https://www.seleniumhq.org/
# For this project, we are using Google Chrome because only IE and Chrome are Vizient-approved web browsers.  Do not
# use Firefox for this project because there may be unexpected complications downloading files from Vizient's website.
from selenium import webdriver
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support.ui import Select
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.common.action_chains import ActionChains
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.by import By
from selenium.common.exceptions import TimeoutException
from selenium.common.exceptions import StaleElementReferenceException
from selenium.webdriver.support import expected_conditions
from selenium.common.exceptions import NoSuchElementException
from selenium.webdriver.support import expected_conditions
import math

#libraries for interacting with dataframes in python
import pandas as pd
import numpy as np
#library for manipulating Excel files.
from openpyxl import load_workbook
import openpyxl
import pyodbc
import os
from itertools import islice
import glob
import time
import shutil
import re
import urllib
import pathlib

################################################################################################
## AUTHOR:  R. Abram Beyer
## DESCRIPTION:  Helpers file which contains bulk of functions used in Vizient Q&A web scraper.


################################################################################################
## UPDATE LOG:

## UL001    11/19/2019  rabeyer     General maintainance.  Vizient will periodically update IDs, labels, buttons, etc.  This time they changed the AHRQ label element.
##                                  Also, the Quarters menu keeps changing causing periodic updates to the clicking logic.  updating this as well.
##                                  Added community data logic to ed-2b, ed-op-18b file naming logic.
## UL002    06/22/2020  rabeyer     Update gather_cohort_data() function to handle new Vizient cohort excel files and critical access cohort.
##                                  Update file naming and moving function to include critical access hospitals
##                                  Update update_time_period_select function to handle any pre-selected quarters menu dropdown defaults.
## UL003                            Update iqc_file_parser.py functions to handle new Critical Access calculator metrics
## UL004    10/27/2020  rabeyer     This year, Vizient changed their report templates to filter pre/post COVID.  This basically makes the templates worthless.  I needed to add
##                                  new functions to remove the covid-related custom group by lists and advanced restrictions to put the templates back.
## UL005    11/16/2020  rabeyer     update ahrq selection function to work for 2020 configuration.
## UL006    02/02/2020  rabeyer     Adding logic to exclude covid patients on demand.
## UL007    03/04/2021  rabeyer     Pandas upgrade.  XLRD no longer supports excel files other than xls.  Pandas read_excel function needs
##                                  to be updated with parameter engine='openpyxl' to fix.
## UL008    03/18/2021  rabeyer     Adding error handling on find_last_downloaded_file.  
##                                 Cause:  Sometimes the internet connection is too slow.  Some temporary file ending in .crdownload
##                                  temporarily appears and then disappears.  So, we need to ignore this type of file because it causes an error.
## UL009    05/18/2021  rabeyer     Handling focus_hosp click stale element reference error.
## UL010    11/23/2021  mbecker     PSI_09 had a vizient name change
## UL011    01/06/2021  rbeyer      Update validate_downloaded_files() function to check for standard restrictions AND advanced restrictions.
################################################################################################

def gather_hyperlink_files():
    try:
        wd = input('Enter file path for hospital hyperlink folder (template_hyperlinks).')
        # change directory to directory with file.  abspath function normalizes the directory path.
        os.chdir(os.path.abspath(wd))
    except:
        print('Something is wrong with Vizient hyperlink excel file path.')
        return
    files = os.listdir(os.curdir)
    # Filter folder files to only include '_links.xlsx' excel files.
    files = [ii for ii in files if '_links.xlsx' in ii and not ii.startswith('~')]
    return(files,wd)


################################################################################################

##UL002  BEGIN

#function to create a pandas dataframe from provide cohort excel files.
def gather_cohort_data():
    try:
        wd = input(
            'Enter file path for Vizient cohort file (File Should be like this: QA_Calculator_..._Cohorts.xlsx).')
        # change directory to directory with file.  abspath function normalizes the directory path.
        os.chdir(os.path.abspath(wd))
        # print(os.listdir())
    except:
        print('Something is wrong with cohort file path.')
        return
    try:
        # find the cohort excel file.  The file name should have QA_ and _Cohorts.xlsx in the name string.
        cohort_file_list = [file for file in os.listdir(wd) if
                            ('QA_' in file and '_Cohorts.xlsx' in file) and not (file.startswith('~'))]
        try:
            # There should only be one file in the cohort directory
            if len(cohort_file_list) > 1:
                print('More than one cohort list file.  There should only be one.')
                return
            # There should be at least one cohort file in the directory
            elif len(cohort_file_list) == 0:
                print('No cohort list files in this directory.  Check for empty directory or check file name.')
                return
            else:
                # Once you get once cohort excel file found, load it.
                try:
                    wb = openpyxl.load_workbook(filename=cohort_file_list[0], data_only=True)
                except:
                    print('Could not load workbook')
                    return

                # once you have the workbook loaded, find all the period cohort worksheets.
                # Unfortunately, the sheet names change each period and the critical access cohort is on a separate
                # sheet starting in FY2020 Period 1 so we don't have one constant worksheet name to call so we must
                # find them.
                # find the current period cohort worksheet name
                cohort_sheet_list = [sheet for sheet in wb.sheetnames if
                                     ('Q&A' in sheet and 'Cohorts' in sheet) or (sheet == 'Crit Acc. & Small Comm')]

                # if there is only one sheet, it likely lacks the critical access sheet.  Probably FY19.
                if len(cohort_sheet_list) == 1:
                    try:
                        ws = wb[cohort_sheet_list[0]]
                    except:
                        print('Something went wrong opening the worksheet.')
                        return
                    data = ws.values
                    cols = next(data)[1:]
                    data = list(data)
                    idx = [r[0] for r in data]
                    data = (islice(r, 1, None) for r in data)
                    df = pd.DataFrame(data, index=idx, columns=cols)
                    df['Hospital'] = df.index
                    df = pd.DataFrame(df, columns=['Hospital', 'Cohort'])
                    #They added a comma in LSCCMC cohort name in the 2020 calculators....
                    df['Cohort'] = df['Cohort'].str.replace(',','').str.replace('.','')
                    return (df)
                elif len(cohort_sheet_list) == 0:
                    # empty list.  failed to find any worksheets.  return and figure out the problem.
                    # if there are
                    print('Did not find any worksheets.')
                    return
                elif len(cohort_sheet_list) == 2:
                    cohort_sheet_list_index = 0
                    crit_access_list_index = 0
                    for i, item in enumerate(cohort_sheet_list):
                        if ('Q&A' in item and 'Cohorts' in item):
                            cohort_sheet_list_index = i
                        if item == 'Crit Acc. & Small Comm':
                            crit_access_list_index = i

                    # open and process cohort worksheet
                    try:
                        ws = wb[cohort_sheet_list[cohort_sheet_list_index]]
                    except:
                        print('Something went wrong opening the worksheet.')
                        return

                    data = ws.values
                    columns = list(next(data)[0:])

                    # in order to handle the 2019 Period 4 calculator, look for 'Hospital' column name as indicator
                    if 'Hospital' in columns:
                        data = ws.values
                        cols = next(data)[1:]
                        data = list(data)
                        idx = [r[0] for r in data]
                        data = (islice(r, 1, None) for r in data)
                        df = pd.DataFrame(data, index=idx, columns=cols)
                        df['Hospital'] = df.index
                        df = pd.DataFrame(df, columns=['Hospital', 'Cohort'])
                        #Vizient keeps changing the cohort names in the cohort file.
                        df['Cohort'] = np.where(df['Cohort']=='Comprehensive Academic Medical Center', 'Comprehensive Academic Medical Center', df['Cohort'])
                        df['Cohort'] = np.where(df['Cohort']=='Large, Specialized Complex Care Medical Center', 'Large Specialized Complex Care Medical Center', df['Cohort'])
                        df['Cohort'] = np.where(df['Cohort']=='Large, Specialized Medical Center', 'Large Specialized Complex Care Medical Center', df['Cohort'])
                        df['Cohort'] = np.where(df['Cohort']=='Complex Care Medical Center', 'Complex Care Medical Center', df['Cohort'])
                        df['Cohort'] = np.where(df['Cohort']=='Community Medical Center', 'Community', df['Cohort'])
                        df['Cohort'] = np.where(df['Cohort']=='Critical Access & Small Community', 'Critical Access & Small Community', df['Cohort'])
                        # They added a comma in LSCCMC cohort name in the 2020 calculators....
                        df['Cohort'] = df['Cohort'].str.replace(',','').str.replace('.','')
                        return (df)
                    else:
                        # open and process cohort worksheet
                        # Starting with FY2020, there is a Medicare ID column and Short name column instead of
                        # concatenated 'Hospital' column
                        print('must by FY2020')
                        data = ws.values
                        cols = next(data)[0:]
                        data = list(data)
                        # recreate the Hospital column from FY19.  Just concatenate the medicare ID to the Short name
                        data[:] = [list(i) for i in data]
                        [i.insert(0, str(i[0]) + ' ' + str(i[1])) for i in data]
                        data[:] = [tuple(i) for i in data]
                        idx = [r[0] for r in data]
                        data = (islice(r, 0, None) for r in data)
                        cols = list(cols)
                        cols.insert(0, 'Hospital')
                        cols = tuple(cols)
                        df = pd.DataFrame(data, index=idx, columns=cols)
                        df['Hospital'] = df.index
                        df = pd.DataFrame(df, columns=['Hospital', 'Cohort'])

                        # now do almost the same for the critical access worksheet

                        try:
                            ws_crit = wb[cohort_sheet_list[crit_access_list_index]]
                        except:
                            print('Something went wrong opening the worksheet.')
                            return

                        data_crit = ws_crit.values
                        # skip the header row
                        cols_crit = next(data_crit)
                        cols_crit = next(data_crit)
                        data_crit = list(data_crit)
                        # recreate the Hospital column from FY19.  Just concatenate the medicare ID to the Short name
                        data_crit[:] = [list(i) for i in data_crit]
                        [i.insert(0, str(i[0]) + ' ' + str(i[1])) for i in data_crit]
                        data_crit[:] = [tuple(i) for i in data_crit]
                        idx_crit = [r[0] for r in data_crit]
                        data_crit = (islice(r, 0, None) for r in data_crit)
                        cols_crit = list(cols_crit)
                        cols_crit.insert(0, 'Hospital')
                        cols_crit = tuple(cols_crit)
                        df_crit = pd.DataFrame(data_crit, index=idx_crit, columns=cols_crit)
                        df_crit['Hospital'] = df_crit.index
                        # no cohort column so we need to fill this in.
                        df_crit['Cohort'] = 'Critical Access & Small Community'
                        df_crit = pd.DataFrame(df_crit, columns=['Hospital', 'Cohort'])
                        df_list = [df, df_crit]
                        df_final = pd.concat(df_list)
                        df_final['Cohort'] = np.where(df_final['Cohort']=='Comprehensive Academic Medical Center', 'Comprehensive Academic Medical Center', df_final['Cohort'])
                        df_final['Cohort'] = np.where(df_final['Cohort']=='Large, Specialized Complex Care Medical Center', 'Large Specialized Complex Care Medical Center', df_final['Cohort'])
                        df_final['Cohort'] = np.where(df_final['Cohort']=='Large, Specialized Medical Center', 'Large Specialized Complex Care Medical Center', df_final['Cohort'])
                        df_final['Cohort'] = np.where(df_final['Cohort']=='Complex Care Medical Center', 'Complex Care Medical Center', df_final['Cohort'])
                        df_final['Cohort'] = np.where(df_final['Cohort']=='Community Medical Center', 'Community', df_final['Cohort'])
                        df_final['Cohort'] = np.where(df_final['Cohort']=='Critical Access & Small Community', 'Critical Access & Small Community', df_final['Cohort'])
                        # They added a comma in LSCCMC cohort name in the 2020 calculators....
                        df_final['Cohort'] = df_final['Cohort'].str.replace(',','').str.replace('.','')
                        return (df_final)

        except:
            print('Something went wrong finding files ending in _Cohorts.xlsx.')
            return
    except:
        print('No cohort file found.')
        return

'''
def gather_cohort_data():
    try:
        wd = input('Enter file path for Vizient cohort file (QA_Calculator_2019_Cohorts.xlsx).')
        # change directory to directory with file.  abspath function normalizes the directory path.
        os.chdir(os.path.abspath(wd))
        print(os.listdir())
    except:
        print('Something is wrong with cohort file path.')
        return
    try:
        wb = openpyxl.load_workbook(filename='QA_Calculator_2019_Cohorts.xlsx',data_only=True)
        #Navigate to the Cohor sheet
        ws = wb['2019 P3 Q&A Cohorts']
    except:
        print('No cohort file found.')
        return
    data = ws.values
    cols = next(data)[1:]
    data = list(data)
    idx = [r[0] for r in data]
    data = (islice(r, 1, None) for r in data)
    df = pd.DataFrame(data, index=idx, columns=cols)
    df['Hospital'] = df.index
    df = pd.DataFrame(df,columns=['Hospital','Cohort'])
    return(df)

'''
##UL002 END
################################################################################################

def get_report_template_links_orig():
    #find the folder with Vizient calculator template hyperlinks and put file names in a list
    try:
        file_names = gather_hyperlink_files()
    except:
        print('Problem gathering hyperlink files.')
        return
    #Get all unique report templates for all hospitals
    #initialize empty dataframe to store hyperlinks
    hyperlinks = pd.DataFrame()
    #iterate through list of hyper link files obtain from Vizient calculators and store measure name & hyperlink.
    for ii, item in enumerate(file_names[0]):
        ##UL007
        dataframe_ob = pd.DataFrame(pd.read_excel(item,sheet_name="Sheet1",engine='openpyxl'))
        dataframe_ob = pd.DataFrame(dataframe_ob,columns=['Hospital','Formal Name','Hyperlink','JobStoreID','ReportID','AdjustmentModel','AHRQ Version','Keyword/Metric','Domain'])
        hyperlinks = pd.concat([hyperlinks, dataframe_ob])
    #Remove any zero rows or null rows.  These are measures without a report template link.
    hyperlink_indices =  hyperlinks['Hyperlink'] != 0
    hyperlinks = hyperlinks[hyperlink_indices]
    hyperlink_indices2 =  hyperlinks['Hyperlink'].notnull()
    hyperlinks = hyperlinks[hyperlink_indices2]
    #Convert JobStoreID and ReportID back to integer to remove decimal point.
    hyperlinks[["JobStoreID", "ReportID"]] = hyperlinks[["JobStoreID", "ReportID"]].astype(int)
    #Remove duplicates.  Only require a unique list of report templates for all hospitals
    hyperlinks = hyperlinks.drop_duplicates()
    
    #UL003
    #Keyword/Metric for VWH only is not all caps...for some reason so I now have to convert this column to all caps.
    hyperlinks['Keyword/Metric'] = hyperlinks['Keyword/Metric'].str.upper()
    #hyperlinks['zipped_data'] = list(zip(hyperlinks.Hyperlink,hyperlinks.JobStoreID,hyperlinks.ReportID))
    #lookup_data_container = pd.Series(hyperlinks.zipped_data.values,index=hyperlinks['Formal Name'].values).to_dict()
    #hyperlinks.to_csv (r'C:/Users/NM184423/Desktop/QA_docs/template_hyperlinks', index = None, header=True)
    #hyperlinks.to_csv('here.csv', encoding='utf-8')
    return(hyperlinks,file_names[1])


def get_report_template_links():
    #find the folder with Vizient calculator template hyperlinks and put file names in a list
    try:
        file_names = gather_hyperlink_files()
    except:
        print('Problem gathering hyperlink files.')
        return
    #Get all unique report templates for all hospitals
    #initialize empty dataframe to store hyperlinks
    hyperlinks = pd.DataFrame()
    print(os.getcwd())
    print('file names:',file_names)
    #iterate through list of hyper link files obtain from Vizient calculators and store measure name & hyperlink.
    for ii, item in enumerate(file_names[0]):
        #UL007
        dataframe_ob = pd.DataFrame(pd.read_excel(item,sheet_name="Sheet1",engine='openpyxl'))
        dataframe_ob = pd.DataFrame(dataframe_ob,columns=['Hospital','Formal Name','Hyperlink','JobStoreID','ReportID','AdjustmentModel','AHRQ Version','Keyword/Metric','Domain'])
        hyperlinks = pd.concat([hyperlinks, dataframe_ob])
    #Remove any zero rows or null rows.  These are measures without a report template link.
    hyperlink_indices =  hyperlinks['Hyperlink'] != 0
    hyperlinks = hyperlinks[hyperlink_indices]
    hyperlink_indices2 =  hyperlinks['Hyperlink'].notnull()
    hyperlinks = hyperlinks[hyperlink_indices2]
    #Convert JobStoreID and ReportID back to integer to remove decimal point.
    hyperlinks[["JobStoreID", "ReportID"]] = hyperlinks[["JobStoreID", "ReportID"]].astype(int)
    hyperlinks = hyperlinks.drop_duplicates()
    
    #UL003
    #Keyword/Metric for VWH only is not all caps...for some reason so I now have to convert this column to all caps.
    hyperlinks['Keyword/Metric'] = hyperlinks['Keyword/Metric'].str.upper()
    
    return(hyperlinks)

################################################################################################

def create_hyperlink_dict(merged_df):
    merged_df = pd.DataFrame(merged_df,columns=['Cohort','Formal Name','Hyperlink','JobStoreID','ReportID','AdjustmentModel','AHRQ Version','Keyword/Metric','Domain'])
    merged_df = merged_df.drop_duplicates()
    merged_df['zipped_data'] = list(zip(merged_df.Hyperlink,merged_df.JobStoreID,merged_df.ReportID,merged_df.Cohort,merged_df['Formal Name'],merged_df['AdjustmentModel'],merged_df['AHRQ Version'],merged_df['Keyword/Metric'],merged_df['Domain']))
    merged_df['zipped_keys'] = list(zip(merged_df.Cohort,merged_df['Formal Name']))
    lookup_data_container = pd.Series(merged_df.zipped_data.values,index=merged_df.zipped_keys.values).to_dict()
    return(lookup_data_container)

################################################################################################

# function takes the hyperlink dictionary created above and generates a folder structure to store the files.
def create_folder_structure(links_dict):
    # create main folder

    new_dir_path = input('Enter path of location where you want to store the files.')
    folder_name = 'Vizient Q&A Files'
    new_dir_path = os.path.join(os.path.abspath(new_dir_path), folder_name)
    try:
        # check if folder already exists.  If it does not exist, create it.
        if os.path.isfile(new_dir_path) == False:
            os.mkdir(new_dir_path)
    except:
        pass
    # create sub folders per cohort
    cohort_names_list = [i[0] for i in links_dict.keys()]
    distinct_cohort_names_list = list(dict.fromkeys(cohort_names_list))

    for i in distinct_cohort_names_list:
        try:
            if os.path.isfile(new_dir_path) == False:
                os.mkdir(os.path.join(new_dir_path, i))
        except:
            pass
    return (new_dir_path)

################################################################################################

# Function to update the template excel files.  This will be used if the program crashes midway through
# and you need to pick up where you left off.  As reports are successfully downloaded, they will be removed
# from the template excel files.

def update_template_files(hyperlink_loc, cohort, measure):
    if cohort == 'Comprehensive Academic Medical Center':
        try:
            #UL007
            dataframe_ob = pd.DataFrame(pd.read_excel(os.path.join(os.path.abspath(hyperlink_loc),'nmh_links.xlsx'),sheet_name="Sheet1",engine='openpyxl'))
            dataframe_ob = dataframe_ob[dataframe_ob['Formal Name'] != measure]
            writer = pd.ExcelWriter('nmh_links.xlsx')
            dataframe_ob.to_excel(writer,'Sheet1')
            writer.save()
        except:
            print('Issue finding nmh_links.xlsx.')
            pass
    elif cohort == 'Large Specialized Complex Care Medical Center':
        try:
            #UL007
            dataframe_ob = pd.DataFrame(pd.read_excel(os.path.join(os.path.abspath(hyperlink_loc),'cdh_links.xlsx'),sheet_name="Sheet1",engine='openpyxl'))
            dataframe_ob = dataframe_ob[dataframe_ob['Formal Name'] != measure]
            writer = pd.ExcelWriter('cdh_links.xlsx')
            dataframe_ob.to_excel(writer,'Sheet1')
            writer.save()
        except:
            print('Issue finding cdh_links.xlsx.')
            pass
    elif cohort == 'Complex Care Medical Center':
        try:
            #UL007
            dataframe_ob3 = pd.DataFrame(pd.read_excel(os.path.join(os.path.abspath(hyperlink_loc),'lfh_links.xlsx'),sheet_name="Sheet1",engine='openpyxl'))
            dataframe_ob3 = dataframe_ob3[dataframe_ob3['Formal Name'] != measure]
            writer3 = pd.ExcelWriter('lfh_links.xlsx')
            dataframe_ob3.to_excel(writer3,'Sheet1')
            writer3.save()
            #UL007
            dataframe_ob1 = pd.DataFrame(pd.read_excel(os.path.join(os.path.abspath(hyperlink_loc),'dch_links.xlsx'),sheet_name="Sheet1",engine='openpyxl'))
            dataframe_ob1 = dataframe_ob1[dataframe_ob1['Formal Name'] != measure]
            writer1 = pd.ExcelWriter('dch_links.xlsx')
            dataframe_ob1.to_excel(writer1,'Sheet1')
            writer1.save()
            #UL007
            dataframe_ob2 = pd.DataFrame(pd.read_excel(os.path.join(os.path.abspath(hyperlink_loc),'kish_links.xlsx'),sheet_name="Sheet1",engine='openpyxl'))
            dataframe_ob2 = dataframe_ob2[dataframe_ob2['Formal Name'] != measure]
            writer2 = pd.ExcelWriter('kish_links.xlsx')
            dataframe_ob2.to_excel(writer2,'Sheet1')
            writer2.save()
            
            #UL007
            dataframe_ob4 = pd.DataFrame(pd.read_excel(os.path.join(os.path.abspath(hyperlink_loc),'hh_links.xlsx'),sheet_name="Sheet1",engine='openpyxl'))
            dataframe_ob4 = dataframe_ob4[dataframe_ob4['Formal Name'] != measure]
            writer4 = pd.ExcelWriter('hh_links.xlsx')
            dataframe_ob4.to_excel(writer4,'Sheet1')
            writer4.save()
            #UL007
            dataframe_ob5 = pd.DataFrame(pd.read_excel(os.path.join(os.path.abspath(hyperlink_loc),'mch_links.xlsx'),sheet_name="Sheet1",engine='openpyxl'))
            dataframe_ob5 = dataframe_ob5[dataframe_ob5['Formal Name'] != measure]
            writer5 = pd.ExcelWriter('mch_links.xlsx')
            dataframe_ob5.to_excel(writer5,'Sheet1')
            writer5.save()
        except:
            print('Issue finding CCMC link files for update.')
            pass
    elif cohort == 'Community Medical Center':
        try:
            #UL007
            dataframe_ob = pd.DataFrame(pd.read_excel(os.path.join(os.path.abspath(hyperlink_loc), 'comm_links.xlsx'), sheet_name="Sheet1",engine='openpyxl'))
            dataframe_ob = dataframe_ob[dataframe_ob['Formal Name'] != measure]
            writer = pd.ExcelWriter('comm_links.xlsx')
            dataframe_ob.to_excel(writer, 'Sheet1')
            writer.save()
        except:
            print('Issue finding comm_links.xlsx.')
            pass
    #UL002 adding new critical access cohort to file updater
    elif cohort == 'Critical Access & Small Community':
        try:
            #UL007
            dataframe_ob = pd.DataFrame(pd.read_excel(os.path.join(os.path.abspath(hyperlink_loc), 'vwh_links.xlsx'), sheet_name="Sheet1",engine='openpyxl'))
            dataframe_ob = dataframe_ob[dataframe_ob['Formal Name'] != measure]
            writer = pd.ExcelWriter('vwh_links.xlsx')
            dataframe_ob.to_excel(writer, 'Sheet1')
            writer.save()
        except:
            print('Issue finding vwh_links.xlsx.')
            pass


################################################################################################

# Function to query the database and get time period data
'''
def build_period_lookup_dict():
    try:
        # Establish a connection with Node A
        conn = pyodbc.connect('Driver={SQL Server};'
                              'Server=edw00pd05wva.corp.nm.org\EDWIDS1;'
                              'Database=clarity;'
                              'Trusted_Connection=yes;')
    except:
        print('Error connecting to the database.')

    time_period_choice = input(
        "Do you want the standard NM_Performance (choose 1) time periods or a custom date range (choose 2)?")

    if time_period_choice == '1':

        # input the month end date in a format the NM_Performance.period.period_lookup table will accept
        end_dts = input("Enter the Performance Close Month end datetime (format:  'xx-xx-xxxx 23:59:59'')")

        # build the query
        sql = """
        SELECT
        period_month_begin.period_type
        ,start_dd2.MONTH_NAME as [begin_month]
        ,start_dd2.YEAR as [begin_year]
        ,end_dd2.MONTH_NAME as [end_month]
        ,end_dd2.YEAR as [end_year]
        FROM
        (
        SELECT
        pl.period_type
        ,start_dd.MONTH_BEGIN_DT as start_month_begin
        ,end_dd.MONTH_BEGIN_DT as end_month_begin
        from
        NM_Performance.period.period_lookup as pl
        left join clarity.dbo.DATE_DIMENSION as start_dd
        on cast(start_dd.CALENDAR_DT as date) = cast(pl.start_dts as date)
        left join clarity.dbo.DATE_DIMENSION as end_dd
        on cast(end_dd.CALENDAR_DT as date) = cast(pl.end_dts as date)
        WHERE
        pl.period_type in ('MONTH','fscl_ytd','fscl_qtd')
        and pl.end_dts = '""" + end_dts + """'
        ) period_month_begin
        left join clarity.dbo.DATE_DIMENSION as start_dd2
        on cast(start_dd2.CALENDAR_DT as date) = cast(dateadd(mm,-1,period_month_begin.start_month_begin) as date)
        left join clarity.dbo.DATE_DIMENSION as end_dd2
        on cast(end_dd2.CALENDAR_DT as date) = cast(dateadd(mm,-1,period_month_begin.end_month_begin) as date)

        """
        # Query the database and store the results of the query to a pandas dataframe
        try:
            test_period_data = pd.DataFrame(pd.read_sql(sql, conn))
        except:
            print('Query error.')
        # replace Clarity month format with Vizient month format
        test_period_data = test_period_data.replace('January', 'Jan')
        test_period_data = test_period_data.replace('February', 'Feb')
        test_period_data = test_period_data.replace('March', 'Mar')
        test_period_data = test_period_data.replace('April', 'Apr')
        test_period_data = test_period_data.replace('May', 'May')
        test_period_data = test_period_data.replace('June', 'Jun')
        test_period_data = test_period_data.replace('July', 'Jul')
        test_period_data = test_period_data.replace('August', 'Aug')
        test_period_data = test_period_data.replace('September', 'Sep')
        test_period_data = test_period_data.replace('October', 'Oct')
        test_period_data = test_period_data.replace('November', 'Nov')
        test_period_data = test_period_data.replace('December', 'Dec')

        # format the year variables and zip everything into a dictionary
        test_period_data['begin_year'] = test_period_data['begin_year'].astype(str)
        test_period_data['end_year'] = test_period_data['end_year'].astype(str)
        test_period_data['zipped_begin'] = list(zip(test_period_data.begin_month, test_period_data.begin_year))
        test_period_data['zipped_end'] = list(zip(test_period_data.end_month, test_period_data.end_year))
        test_period_data['zipped_dates'] = list(zip(test_period_data.zipped_begin, test_period_data.zipped_end))
        period_lookup_dict = pd.Series(test_period_data.zipped_dates.values,
                                       index=test_period_data.period_type.values).to_dict()
        return (period_lookup_dict)
    if time_period_choice == '2':

        # build the query
        begin_dts = input("Enter the Custom beginning datetime (format:  'xx-01-xxxx')")
        end_dts = input("Enter the Custom Month end datetime (format:  'xx-xx-xxxx 23:59:59')")

        sql2 = """
        DECLARE @begin_dts as date; SET @begin_dts = '""" + begin_dts + """';
        DECLARE @end_dts as date; SET @end_dts = '""" + end_dts + """';


        SELECT
        bd.period_type
        ,start_dd2.MONTH_NAME as [begin_month]
        ,start_dd2.YEAR as [begin_year]
        ,end_dd2.MONTH_NAME as [end_month]
        ,end_dd2.YEAR as [end_year]
        FROM
        (
        SELECT
        'CUSTOM' as period_type
        ,start_dd.MONTH_BEGIN_DT as start_month_begin
        FROM
        clarity.dbo.DATE_DIMENSION as start_dd
        WHERE
        cast(start_dd.CALENDAR_DT as date) = cast(@begin_dts as date)
        ) bd
        LEFT JOIN
        (
        SELECT
        'CUSTOM' as period_type
        ,start_dd.MONTH_BEGIN_DT as end_month_begin
        FROM
        clarity.dbo.DATE_DIMENSION as start_dd
        WHERE
        cast(start_dd.CALENDAR_DT as date) = cast(@end_dts as date)
        ) ed
        on bd.period_type = ed.period_type
        left join clarity.dbo.DATE_DIMENSION as start_dd2
        on cast(start_dd2.CALENDAR_DT as date) = cast(bd.start_month_begin as date)
        left join clarity.dbo.DATE_DIMENSION as end_dd2
        on cast(end_dd2.CALENDAR_DT as date) = cast(ed.end_month_begin as date)
        """
        # Query the database and store the results of the query to a pandas dataframe
        try:
            test_period_data = pd.DataFrame(pd.read_sql(sql2, conn))
        except:
            print('Query error.')
        # replace Clarity month format with Vizient month format
        test_period_data = test_period_data.replace('January', 'Jan')
        test_period_data = test_period_data.replace('February', 'Feb')
        test_period_data = test_period_data.replace('March', 'Mar')
        test_period_data = test_period_data.replace('April', 'Apr')
        test_period_data = test_period_data.replace('May', 'May')
        test_period_data = test_period_data.replace('June', 'Jun')
        test_period_data = test_period_data.replace('July', 'Jul')
        test_period_data = test_period_data.replace('August', 'Aug')
        test_period_data = test_period_data.replace('September', 'Sep')
        test_period_data = test_period_data.replace('October', 'Oct')
        test_period_data = test_period_data.replace('November', 'Nov')
        test_period_data = test_period_data.replace('December', 'Dec')

        # format the year variables and zip everything into a dictionary
        test_period_data['begin_year'] = test_period_data['begin_year'].astype(str)
        test_period_data['end_year'] = test_period_data['end_year'].astype(str)
        test_period_data['zipped_begin'] = list(zip(test_period_data.begin_month, test_period_data.begin_year))
        test_period_data['zipped_end'] = list(zip(test_period_data.end_month, test_period_data.end_year))
        test_period_data['zipped_dates'] = list(zip(test_period_data.zipped_begin, test_period_data.zipped_end))
        period_lookup_dict = pd.Series(test_period_data.zipped_dates.values,
                                       index=test_period_data.period_type.values).to_dict()
        return (period_lookup_dict)
'''


def build_period_lookup_dict():
    try:
        # Establish a connection with Node A
        conn = pyodbc.connect('Driver={SQL Server};'
                              'Server=edw00pd05wva.corp.nm.org\EDWIDS1;'
                              'Database=clarity;'
                              'Trusted_Connection=yes;')
    except:
        print('Error connecting to the database.')

    time_period_choice = input(
        "Do you want the standard NM_Performance (choose 1) time periods or a custom date range (choose 2) or Vizient Q&A Quarter (choose 3)?")

    if time_period_choice == '1':

        # input the month end date in a format the NM_Performance.period.period_lookup table will accept
        end_dts = input("Enter the Performance Close Month end datetime (format:  'xx-xx-xxxx 23:59:59'')")

        # build the query
        sql = """
        SELECT
        period_month_begin.period_type
        ,start_dd2.MONTH_NAME as [begin_month]
        ,start_dd2.YEAR as [begin_year]
        ,end_dd2.MONTH_NAME as [end_month]
        ,end_dd2.YEAR as [end_year]
        FROM
        (
        SELECT
        pl.period_type
        ,start_dd.MONTH_BEGIN_DT as start_month_begin
        ,end_dd.MONTH_BEGIN_DT as end_month_begin
        from
        NM_Performance.period.period_lookup as pl
        left join clarity.dbo.DATE_DIMENSION as start_dd
        on cast(start_dd.CALENDAR_DT as date) = cast(pl.start_dts as date)
        left join clarity.dbo.DATE_DIMENSION as end_dd
        on cast(end_dd.CALENDAR_DT as date) = cast(pl.end_dts as date)
        WHERE
        pl.period_type in ('MONTH','fscl_ytd','fscl_qtd')
        and pl.end_dts = '""" + end_dts + """'
        ) period_month_begin
        left join clarity.dbo.DATE_DIMENSION as start_dd2
        on cast(start_dd2.CALENDAR_DT as date) = cast(dateadd(mm,-1,period_month_begin.start_month_begin) as date)
        left join clarity.dbo.DATE_DIMENSION as end_dd2
        on cast(end_dd2.CALENDAR_DT as date) = cast(dateadd(mm,-1,period_month_begin.end_month_begin) as date)

        """
        # Query the database and store the results of the query to a pandas dataframe
        try:
            test_period_data = pd.DataFrame(pd.read_sql(sql, conn))
        except:
            print('Query error.')
        # replace Clarity month format with Vizient month format
        test_period_data = test_period_data.replace('January', 'Jan')
        test_period_data = test_period_data.replace('February', 'Feb')
        test_period_data = test_period_data.replace('March', 'Mar')
        test_period_data = test_period_data.replace('April', 'Apr')
        test_period_data = test_period_data.replace('May', 'May')
        test_period_data = test_period_data.replace('June', 'Jun')
        test_period_data = test_period_data.replace('July', 'Jul')
        test_period_data = test_period_data.replace('August', 'Aug')
        test_period_data = test_period_data.replace('September', 'Sep')
        test_period_data = test_period_data.replace('October', 'Oct')
        test_period_data = test_period_data.replace('November', 'Nov')
        test_period_data = test_period_data.replace('December', 'Dec')

        # format the year variables and zip everything into a dictionary
        test_period_data['begin_year'] = test_period_data['begin_year'].astype(str)
        test_period_data['end_year'] = test_period_data['end_year'].astype(str)
        test_period_data['zipped_begin'] = list(zip(test_period_data.begin_month, test_period_data.begin_year))
        test_period_data['zipped_end'] = list(zip(test_period_data.end_month, test_period_data.end_year))
        test_period_data['zipped_dates'] = list(zip(test_period_data.zipped_begin, test_period_data.zipped_end))
        period_lookup_dict = pd.Series(test_period_data.zipped_dates.values,
                                       index=test_period_data.period_type.values).to_dict()
        return (period_lookup_dict)
    if time_period_choice == '2':

        # build the query
        begin_dts = input("Enter the Custom beginning datetime (format:  'xx-01-xxxx')")
        end_dts = input("Enter the Custom Month end datetime (format:  'xx-xx-xxxx 23:59:59')")

        sql2 = """
        DECLARE @begin_dts as date; SET @begin_dts = '""" + begin_dts + """';
        DECLARE @end_dts as date; SET @end_dts = '""" + end_dts + """';


        SELECT
        bd.period_type
        ,start_dd2.MONTH_NAME as [begin_month]
        ,start_dd2.YEAR as [begin_year]
        ,end_dd2.MONTH_NAME as [end_month]
        ,end_dd2.YEAR as [end_year]
        FROM
        (
        SELECT
        'CUSTOM' as period_type
        ,start_dd.MONTH_BEGIN_DT as start_month_begin
        FROM
        clarity.dbo.DATE_DIMENSION as start_dd
        WHERE
        cast(start_dd.CALENDAR_DT as date) = cast(@begin_dts as date)
        ) bd
        LEFT JOIN
        (
        SELECT
        'CUSTOM' as period_type
        ,start_dd.MONTH_BEGIN_DT as end_month_begin
        FROM
        clarity.dbo.DATE_DIMENSION as start_dd
        WHERE
        cast(start_dd.CALENDAR_DT as date) = cast(@end_dts as date)
        ) ed
        on bd.period_type = ed.period_type
        left join clarity.dbo.DATE_DIMENSION as start_dd2
        on cast(start_dd2.CALENDAR_DT as date) = cast(bd.start_month_begin as date)
        left join clarity.dbo.DATE_DIMENSION as end_dd2
        on cast(end_dd2.CALENDAR_DT as date) = cast(ed.end_month_begin as date)
        """
        # Query the database and store the results of the query to a pandas dataframe
        try:
            test_period_data = pd.DataFrame(pd.read_sql(sql2, conn))
        except:
            print('Query error.')

        # replace Clarity month format with Vizient month format
        test_period_data = test_period_data.replace('January', 'Jan')
        test_period_data = test_period_data.replace('February', 'Feb')
        test_period_data = test_period_data.replace('March', 'Mar')
        test_period_data = test_period_data.replace('April', 'Apr')
        test_period_data = test_period_data.replace('May', 'May')
        test_period_data = test_period_data.replace('June', 'Jun')
        test_period_data = test_period_data.replace('July', 'Jul')
        test_period_data = test_period_data.replace('August', 'Aug')
        test_period_data = test_period_data.replace('September', 'Sep')
        test_period_data = test_period_data.replace('October', 'Oct')
        test_period_data = test_period_data.replace('November', 'Nov')
        test_period_data = test_period_data.replace('December', 'Dec')

        # format the year variables and zip everything into a dictionary
        test_period_data['begin_year'] = test_period_data['begin_year'].astype(str)
        test_period_data['end_year'] = test_period_data['end_year'].astype(str)
        test_period_data['zipped_begin'] = list(zip(test_period_data.begin_month, test_period_data.begin_year))
        test_period_data['zipped_end'] = list(zip(test_period_data.end_month, test_period_data.end_year))
        test_period_data['zipped_dates'] = list(zip(test_period_data.zipped_begin, test_period_data.zipped_end))
        period_lookup_dict = pd.Series(test_period_data.zipped_dates.values,
                                       index=test_period_data.period_type.values).to_dict()
        return (period_lookup_dict)
    if time_period_choice == '3':
        test_period_data2 = pd.DataFrame(columns=['period_type', 'qtrs'])
        #period_type = 'QUARTER'
        answer = 'Y'
        qtr_data = ['QUARTER']
        while answer == 'Y':
            answer = input("Do you want to add another Fiscal Year Quarter?  (Y/N)").upper()
            while answer not in ['Y','N']:
                print('That is not "Y" or "N".  Please fix your answer.')
                answer = input("Do you want to add another Fiscal Year Quarter?  (Y/N)").upper()
            if answer == 'Y':
                qtr_dts = input("Enter the beginning Vizient Quarter (format:  '20XX Quarter #')")
                qtr_data.append(qtr_dts)
            elif 'N':
                pass

        #end_dts = input("Enter the Custom Month end datetime (format:  '20XX Quarter #')")
        #qtr_data = [period_type, begin_dts, end_dts]
        #test_period_data.loc[0] = qtr_data
        #test_period_data['begin_qtr'] = test_period_data['begin_qtr'].astype(str)
        #test_period_data['end_qtr'] = test_period_data['end_qtr'].astype(str)
        test_period_data = pd.DataFrame(qtr_data)
        #print(test_period_data[0].loc[1:])
        #return(np.array(test_period_data[0].loc[1:]))
        #test_period_data['zipped_begin_end'] = list(zip(np.array(test_period_data[0].loc[1:])))
        #print(test_period_data)
        value_list = [test_period_data[0].loc[0],np.array(test_period_data[0].loc[1:])]
        test_period_data2.loc[0] = value_list
        #print(test_period_data2)
        #test_period_data['zipped_begin_end'] = list(zip(test_period_data.begin_qtr, test_period_data.end_qtr))
        #period_lookup_dict = pd.Series(test_period_data.zipped_begin_end.values,index=test_period_data.period_type.values).to_dict()
        #period_lookup_dict = pd.Series(np.array(test_period_data[0].loc[1:]),index=np.array(test_period_data[0].loc[0])).to_dict()
        period_lookup_dict = pd.Series(test_period_data2.qtrs.values,index=test_period_data2.period_type.values).to_dict()
        return (period_lookup_dict)



################################################################################################

# Define functions for element/event listening for implicitly waiting for elements to load on a page.

def find_login(driver):
    login_btn = driver.find_element_by_partial_link_text('Login')
    if login_btn:
        return login_btn
    else:
        return False


def find_mobile_login(driver):
    mobile_login = driver.find_element_by_class_name('mobile-login_outer')
    if mobile_login:
        return mobile_login
    else:
        return False


def find_vizient_member_login(driver):
    vizient_member_login_btn = driver.find_element_by_partial_link_text('Log in to Vizient Member Dashboard')
    if vizient_member_login_btn:
        return vizient_member_login_btn
    else:
        return False


def find_vizient_email_input(driver):
    vizient_email_input = driver.find_element_by_name('username')
    if vizient_email_input:
        return vizient_email_input
    else:
        return False


def find_vizient_email_next(driver):
    vizient_email_next_btn = driver.find_element_by_id("idp-discovery-submit")
    if vizient_email_next_btn:
        return vizient_email_next_btn
    else:
        return False


def find_password_form(driver):
    password_form = driver.find_element_by_id("okta-signin-password")
    if password_form:
        return password_form
    else:
        return False


def find_vizient_submit_password(driver):
    vizient_submit_password = driver.find_element_by_id("okta-signin-submit")
    if vizient_submit_password:
        return vizient_submit_password
    else:
        return False


def find_adjustment_model_2019_amc(driver):
    adjustment_model_btn = driver.find_element_by_xpath('//*[@id="ctl00_ContentPlaceHolder1_radMSDRG"]')
    if adjustment_model_btn:
        return adjustment_model_btn
    else:
        return False


def find_adjustment_model_2019_comm(driver):
    adjustment_model_btn = driver.find_element_by_xpath('//*[@id="ctl00_ContentPlaceHolder1_radMSDRG_Commu"]')
    if adjustment_model_btn:
        return adjustment_model_btn
    else:
        return False



def find_adjustment_model_2018_amc(driver):
    adjustment_model_btn = driver.find_element_by_xpath(
        "//*/label[@for='ctl00_ContentPlaceHolder1_radModifiedMSDRG'][contains(text(), '2018 Risk Model (AMC)')]")
    if adjustment_model_btn:
        return adjustment_model_btn
    else:
        return False


def find_adjustment_model_2018_comm(driver):
    adjustment_model_btn = driver.find_element_by_xpath(
        "//*/label[@for='ctl00_ContentPlaceHolder1_radModifiedMSDRG_Commu'][contains(text(), '2018 Risk Model (Community)')]")
    if adjustment_model_btn:
        return adjustment_model_btn
    else:
        return False


def find_adjustment_model_2017_amc(driver):
    adjustment_model_btn = driver.find_element_by_xpath(
        "//*/label[@for='ctl00_ContentPlaceHolder1_radMSDRG'][contains(text(), '2017 Risk Model (AMC)')]")
    if adjustment_model_btn:
        return adjustment_model_btn
    else:
        return False


def find_adjustment_model_2017_comm(driver):
    adjustment_model_btn = driver.find_element_by_xpath(
        "//*/label[@for='ctl00_ContentPlaceHolder1_radMSDRG_Commu'][contains(text(), '2017 Risk Model (Community)')]")
    if adjustment_model_btn:
        return adjustment_model_btn
    else:
        return False

#UL005
# add ahrq version ability for fy21 configuration:  V2019 (Pediatric) / V2019 (Quality) / V2019 (Safety)

def find_ahrq_version_v2019(driver):
    ahrq_version_btn = driver.find_element_by_xpath("//*/label[@for='ctl00_ContentPlaceHolder1_radAHRQPrevious'][contains(text(), 'V2019 (Pediatric) / V2019 (Quality) / V2019 (Safety)')]")
        #"//*/label[@for='ctl00_ContentPlaceHolder1_radAHRQCurrent'][contains(text(), '8.0 (CMS Safety)')]")  #Vizient updates the labels...sigh...  #UL001
    if ahrq_version_btn:
        return ahrq_version_btn
    else:
        return False


def find_ahrq_version_8(driver):
    ahrq_version_btn = driver.find_element_by_xpath("//*/label[@for='ctl00_ContentPlaceHolder1_radAHRQPrevious'][contains(text(), '8.0 (CMS Safety)')]")
        #"//*/label[@for='ctl00_ContentPlaceHolder1_radAHRQCurrent'][contains(text(), '8.0 (CMS Safety)')]")  #Vizient updates the labels...sigh...  #UL001
    if ahrq_version_btn:
        return ahrq_version_btn
    else:
        return False


def find_ahrq_version_7(driver):
    ahrq_version_btn = driver.find_element_by_xpath("//*/label[@for='ctl00_ContentPlaceHolder1_radAHRQPrevious1'][contains(text(), '7.0.1 (Pediatric) / 7.0.1 (Quality) / 7.0.1 (Safety)')]")
        #"//*/label[@for='ctl00_ContentPlaceHolder1_radAHRQPrevious'][contains(text(), '7.0.1 (Pediatric) / 7.0.1 (Quality) / 7.0.1 (Safety)')]")  #UL001
    if ahrq_version_btn:
        return ahrq_version_btn
    else:
        return False


def find_ahrq_version_6(driver):
    ahrq_version_btn = driver.find_element_by_xpath(
        "//*/label[@for='ctl00_ContentPlaceHolder1_radAHRQPrevious1'][contains(text(), '6.0.2 (Pediatric) / 6.0.2 (Quality) / 6.0.2 (Safety)')]")
    if ahrq_version_btn:
        return ahrq_version_btn
    else:
        return False


def find_time_period_radio_btn1(driver):
    time_period_radio_btn1 = driver.find_element_by_id("ctl00_ContentPlaceHolder1_fromYear")
    if time_period_radio_btn1:
        return time_period_radio_btn1
    else:
        return False


def find_time_period_radio_btn2(driver):
    time_period_radio_btn2 = driver.find_element_by_id("ctl00_ContentPlaceHolder1_cmdFromYear")
    if time_period_radio_btn2:
        return time_period_radio_btn2
    else:
        return False

def find_quarters_radio_btn(driver):
    quarters_radio_btn = driver.find_element_by_id("ctl00_ContentPlaceHolder1_cmdTimePeriodQtr")
    if quarters_radio_btn:
        return quarters_radio_btn
    else:
        return False

def find_quarters_select_menu(driver):
    quarters_select_menu = driver.find_element_by_id("ctl00_ContentPlaceHolder1_cmdTimePeriodQuarters")
    if quarters_select_menu:
        return quarters_select_menu
    else:
        return False


def find_all_avail_hosp(driver):
    all_avail_hosp_btn = driver.find_element_by_id("ctl00_ContentPlaceHolder1_cmdAllAvailHosp")
    if all_avail_hosp_btn:
        return all_avail_hosp_btn
    else:
        return False


def find_excel_btn(driver):
    generate_excel_btn = driver.find_element_by_id("ctl00_ContentPlaceHolder1_imgExcel")
    if generate_excel_btn:
        return generate_excel_btn
    else:
        return False

#find delete button for PSI advanced restrictions and also for ADE and % Early Transfers.
def find_restrictions_delete_btn(driver):
    generate_restrictions_delete_btn = driver.find_element_by_id("1_imgDelete")
    if generate_restrictions_delete_btn:
        return generate_restrictions_delete_btn
    else:
        return False
    
def find_restrictions_delete_btn2(driver):
    generate_restrictions_delete_btn2 = driver.find_element_by_id("2_imgDelete")
    if generate_restrictions_delete_btn2:
        return generate_restrictions_delete_btn2
    else:
        return False
    
def find_restrictions_delete_btn3(driver):
    generate_restrictions_delete_btn3 = driver.find_element_by_id("3_imgDelete")
    if generate_restrictions_delete_btn3:
        return generate_restrictions_delete_btn3
    else:
        return False
    
def find_restrictions_delete_btn4(driver):
    generate_restrictions_delete_btn4 = driver.find_element_by_id("4_imgDelete")
    if generate_restrictions_delete_btn4:
        return generate_restrictions_delete_btn4
    else:
        return False


def find_alert_popup(driver):
    obj = driver.switch_to.alert
    if obj:
        return obj
    else:
        return False

def find_alert_popup2(driver):
    obj = driver.switch_to.alert
    if obj:
        return obj
    else:
        return False


def find_window1(driver):
    window_before = driver.window_handles[0]
    if window_before:
        return window_before
    else:
        return False


def find_window2(driver):
    window_after = driver.window_handles[1]
    if window_after:
        return window_after
    else:
        return False


def find_div1_scroll(driver):
    div_element1 = driver.find_element_by_xpath("//div[@id='divRiskAdjustment']")
    if div_element1:
        return div_element1
    else:
        return False


def find_div2_scroll(driver):
    div_element2 = driver.find_element_by_xpath("//div[@id='ctl00_ContentPlaceHolder1_PanelContent3']")
    if div_element2:
        return div_element2
    else:
        return False


def find_div3_scroll(driver):
    div_element3 = driver.find_element_by_xpath("//div[@id='ctl00_ContentPlaceHolder1_PanelContent4']")
    if div_element3:
        return div_element3
    else:
        return False


def find_div4_scroll(driver):
    div_element4 = driver.find_element_by_xpath("//div[@id='ctl00_ContentPlaceHolder1_PanelContent5']")
    if div_element4:
        return div_element4
    else:
        return False


def find_div5_scroll(driver):
    div_element5 = driver.find_element_by_xpath("//div[@id='divRiskAdjustment']")
    if div_element5:
        return div_element5
    else:
        return False

def find_adv_rest_scroll(driver):
    div_element_adv_rest = driver.find_element_by_xpath("//div[@id='tblAdvRestrictionsDiv']")
    if div_element_adv_rest:
        return div_element_adv_rest
    else:
        return False


def find_focus_hosp(driver):
    focus_hosp = driver.find_element_by_xpath("//*[@id='ctl00_ContentPlaceHolder1_cmdFocusHCO']")
    if focus_hosp:
        return focus_hosp
    else:
        return False


#UL009

def click_focus_hosp2(driver):
    my_element_xpath = "//*[@id='ctl00_ContentPlaceHolder1_cmdFocusHCO']"
    ignored_exceptions=(NoSuchElementException,StaleElementReferenceException,)
    hosp_element = WebDriverWait(driver, 120,ignored_exceptions=ignored_exceptions).until(expected_conditions.presence_of_element_located((By.XPATH, my_element_xpath)))
    try:
        hosp_element.click()
    except:
        hosp_element = WebDriverWait(driver, 120,ignored_exceptions=ignored_exceptions).until(expected_conditions.presence_of_element_located((By.XPATH, my_element_xpath)))
        hosp_element.click()
    return(driver,hosp_element)


################################################################################################

# Function to login to Vizient and open up the browser.

def vizient_login():
    # initialize Chrome
    options = webdriver.ChromeOptions()
    options.add_experimental_option("excludeSwitches", ["enable-automation"])
    options.add_experimental_option('useAutomationExtension', False)
    
    current_path = pathlib.Path(__file__).parent.absolute()
    #driver_path = r'drivers\msedgedriver.exe'
    driver_path = r'drivers\chromedriver.exe'
    
    browser = webdriver.Chrome(options=options,executable_path = os.path.abspath(os.path.join(current_path,driver_path)))
    
    
    #browser = webdriver.Edge(executable_path = os.path.abspath(os.path.join(current_path,driver_path)))
    browser.get('https://www.vizientinc.com/')
    #browser.implicitly_wait(30)
    try:
        browser.maximize_window()
    except:
        pass
    # Click the login button
    # If the screen is not maximized, it will provie the mobile buttons
    try:
        login_btn = WebDriverWait(browser, 15).until(find_login)
        # login_btn = browser.find_element_by_partial_link_text('Login')
        login_btn.click()
    except:
        try:
            mobile_login = WebDriverWait(browser, 15).until(find_mobile_login)
            # mobile_login = browser.find_element_by_class_name('mobile-login_outer')
            mobile_login.click()
        except:
            print('Was not able to find an element with that name.')

    # Click the Vizient Member Dashboard login
    try:
        vizient_member_login_btn = WebDriverWait(browser, 15).until(find_vizient_member_login)
        # vizient_member_login_btn = browser.find_element_by_partial_link_text('Log in to Vizient Member Dashboard')
        vizient_member_login_btn.click()
    except:
        print('Was not able to find an element with that name.')
    # time.sleep(0.4)

    # Enter login email
    try:
        vizient_email_input = WebDriverWait(browser, 15).until(find_vizient_email_input)
        # vizient_email_input = browser.find_element_by_name('username')
        login_email = input('Enter Vizient Login email.')
        vizient_email_input.send_keys(login_email)
    except:
        print('Was not able to find an element with that name.')
    # time.sleep(.5)

    # click submit
    vizient_email_next_btn = WebDriverWait(browser, 15).until(find_vizient_email_next)
    # vizient_email_next_btn = browser.find_element_by_id("idp-discovery-submit")
    vizient_email_next_btn.click()
    # time.sleep(1)

    # enter the login password
    password_form = WebDriverWait(browser, 15).until(find_password_form)
    # password_form = browser.find_element_by_id("okta-signin-password")
    login_password = input('Enter Vizient login password.')
    password_form.send_keys(login_password)
    # time.sleep(.7)

    # click submit
    vizient_submit_password = WebDriverWait(browser, 15).until(find_vizient_submit_password)
    # vizient_submit_password = browser.find_element_by_id("okta-signin-submit")
    vizient_submit_password.click()
    return (browser)


################################################################################################

# Open browser window of report template after login

def open_template_report(hyperlink,browser_var):
    browser_var.get(hyperlink)
    #browser_var.implicitly_wait(30)
    return(browser_var)

################################################################################################

def find_from_year(driver):
    fy = driver.find_element_by_xpath("//*[@id='ctl00_ContentPlaceHolder1_fromYear']")
    if fy:
        return fy
    else:
        return False

def find_from_month(driver):
    fm = driver.find_element_by_xpath("//*[@id='ctl00_ContentPlaceHolder1_fromMonth']")
    if fm:
        return fm
    else:
        return False

def find_to_year(driver):
    ty = driver.find_element_by_xpath("//*[@id='ctl00_ContentPlaceHolder1_toYear']")
    if ty:
        return ty
    else:
        return False

def find_to_month(driver):
    tm = driver.find_element_by_xpath("//*[@id='ctl00_ContentPlaceHolder1_toMonth']")
    if tm:
        return tm
    else:
        return False

def find_to_year_div(driver):
    tyd = driver.find_element_by_id("ToYearDiv")
    if tyd:
        return tyd
    else:
        return False

def find_to_month_div(driver):
    tmd = driver.find_element_by_id("ToMonthDiv")
    if tmd:
        return tmd
    else:
        return False

def find_from_year_div(driver):
    fyd = driver.find_element_by_id("FromYearDiv")
    if fyd:
        return fyd
    else:
        return False

def find_mult_group_by_div(driver):
    gowm = driver.find_element_by_id("ctl00_ContentPlaceHolder1_GroupByOutcomesWithMult")
    if gowm:
        return gowm
    else:
        return False

def find_from_month_div(driver):
    fmd = driver.find_element_by_id("FromMonthDiv")
    if fmd:
        return fmd
    else:
        return False

################################################################################################
def while_loop_handler_function(selenium_function):
    attempt_num = 1
    while (attempt_num != 0):
        try:
            dom_element = selenium_function
            attempt_num = 0
        except StaleElementReferenceException:
            if attempt_num == 4:
                raise
            attempt_num += 1
            time.sleep(0.5)
    return(dom_element)


################################################################################################

#UL002 BEGIN

# def update_time_period_select(browser_var2,period_dict,period_type):
#     time.sleep(1)
#     if period_type == 'QUARTER':
#         #browser_var2.implicitly_wait(30)
#         el = while_loop_handler_function(browser_var2.find_element_by_id('ctl00_ContentPlaceHolder1_cmdTimePeriodQuarters'))
#         print(period_dict['QUARTER'])
#         #attempt_el = 1
#         #while (attempt_el != 0):
#         #    try:
#         #        el = browser_var2.find_element_by_id('ctl00_ContentPlaceHolder1_cmdTimePeriodQuarters')
#         #        attempt_el = 0
#         #    except StaleElementReferenceException:
#         #        if attempt_el == 3:
#         #            raise
#         #        attempt_el += 1
#         #        time.sleep(0.5)
#         #browser_var2.implicitly_wait(30)

# # Select the element. This element is now a variable. The variable is in the drop-down?
#         option_list1 = while_loop_handler_function(el.find_elements_by_tag_name('option'))
#         for optiona in option_list1:
#             # Is the option from the drop-down at this point of iteration in our user-defined set of quarters? ALSO, is it NOT one of the following values?
#             if optiona.text in period_dict['QUARTER'] and optiona.text not in ['2018 Quarter 4', '2018 Quarter 3','2019 Quarter 1','2019 Quarter 2','2020 Quarter 2']:
#                 #print('here we are at option text part...not Q4,Q3 or 2019 Q2')
#                 #print(optiona.text)
#                 attempt_option_click = 1
#                 while (attempt_option_click != 0):
#                     try:
#                         #print('clicking on option.')
#                         optiona.click()  # select() in earlier versions of webdriver
#                         attempt_option_click = 0
#                         #print('clicked on option.')
#                     except StaleElementReferenceException:
#                         if attempt_option_click == 3:
#                             raise
#                         attempt_option_click += 1
#                         time.sleep(0.5)
#                 #browser_var2.implicitly_wait(30)
#     # At this point, we are looking for items that Vizient pre-selects, that ALSO aren't in our user-defined list of quarters to pull
#     # We are going to click those again (de-select them), whereas the ABOVE (optiona loop) is selecting items if they're from the user-defined list
#         for optionb in option_list1:
#             if '2018 Quarter 4' not in period_dict['QUARTER']:
#                 if optionb.text in ['2018 Quarter 4']:
#                     #print('got to 2018 Quarter 4 part.')
#                     attempt_option_click2 = 1
#                     while (attempt_option_click2 != 0):
#                         try:
#                             optionb.click()  # select() in earlier versions of webdriver
#                             print('clicked on 2018 quarter 4.')
#                             attempt_option_click2 = 0
#                         except StaleElementReferenceException:
#                             if attempt_option_click2 == 3:
#                                 raise
#                             attempt_option_click2 += 1
#                             time.sleep(0.5)
#                    # browser_var2.implicitly_wait(30)
#         for optionc in option_list1:
#             if '2018 Quarter 3' not in period_dict['QUARTER']:
#                 if optionc.text in ['2018 Quarter 3']:
#                     attempt_option_click3 = 1
#                     while (attempt_option_click3 != 0):
#                         try:
#                             optionc.click()  # select() in earlier versions of webdriver
#                             attempt_option_click3 = 0
#                         except StaleElementReferenceException:
#                             if attempt_option_click3 == 3:
#                                 raise
#                             attempt_option_click3 += 1
#                             time.sleep(0.5)
#                     #browser_var2.implicitly_wait(30)
#
#         for optiond in option_list1:
#             if '2019 Quarter 2' not in period_dict['QUARTER']:
#                 if optiond.text in ['2019 Quarter 2']:
#                     attempt_option_click4 = 1
#                     while (attempt_option_click4 != 0):
#                         try:
#                             optiond.click()  # select() in earlier versions of webdriver
#                             attempt_option_click4 = 0
#                         except StaleElementReferenceException:
#                             if attempt_option_click4 == 3:
#                                 raise
#                             attempt_option_click4 += 1
#                     #browser_var2.implicitly_wait(30)
#
#         for optione in option_list1:
#             if '2019 Quarter 1' not in period_dict['QUARTER']:
#                 if optione.text in ['2019 Quarter 1']:
#                     attempt_option_click5 = 1
#                     while (attempt_option_click5 != 0):
#                         try:
#                             optione.click()  # select() in earlier versions of webdriver
#                             attempt_option_click5 = 0
#                         except StaleElementReferenceException:
#                             if attempt_option_click5 == 3:
#                                 raise
#                             attempt_option_click5 += 1
#                     #browser_var2.implicitly_wait(30)
#                     #print('done with quarter selection...')
#
#         for optionf in option_list1:
#             if '2020 Quarter 2' not in period_dict['QUARTER']:
#                 if optionf.text in ['2020 Quarter 2']:
#                     attempt_option_click6 = 1
#                     while (attempt_option_click6 != 0):
#                         try:
#                             optionf.click()  # select() in earlier versions of webdriver
#                             attempt_option_click6 = 0
#                         except StaleElementReferenceException:
#                             if attempt_option_click6 == 3:
#                                 raise
#                             attempt_option_click6 += 1
#
#
#     else:
#         #update begin year
#         time.sleep(1)
#
#         #fromyeardropdown = browser_var2.find_element_by_id("FromYearDiv")
#         #fromyeardropdown.click()
#
#         fromyeardropdown = WebDriverWait(browser_var2, 120).until(find_from_year_div)
#         fromyeardropdown.click()
#
#         Select(browser_var2.find_element_by_xpath("//*[@id='ctl00_ContentPlaceHolder1_fromYear']")).select_by_visible_text(period_dict[period_type][0][1])
#         #browser_var2.implicitly_wait(30)
#         #fy = browser_var2.find_element_by_xpath("//*[@id='ctl00_ContentPlaceHolder1_fromYear']")
#         #from_year = WebDriverWait(browser_var2, 120).until(find_from_year)
#         # adjustment_model_btn = browser_var2.find_element_by_xpath("//*/label[@for='ctl00_ContentPlaceHolder1_radModifiedMSDRG'][contains(text(), '2018 Risk Model (AMC)')]")
#         # adjustment_model_btn.click()
#         #time.sleep(1)
#         #from_year_attempt = 1
#         #while True:
#             #try:
#             #    from_year = WebDriverWait(browser_var2, 120).until(find_from_year)
#             #    break
#             #except StaleElementReferenceException:
#             #    if from_year_attempt == 3:
#             #        raise
#             #    from_year_attempt += 1
#             #    time.sleep(0.5)
#
#
#         #from_year = while_loop_handler_function(WebDriverWait(browser_var2, 120).until(find_from_year))
#         #for option in from_year.find_elements_by_tag_name('option'):
#         #    if option.text == period_dict[period_type][0][1]:
#         #        option.click()
#         #        time.sleep(0.5)
#         #from_year_options = while_loop_handler_function(from_year.find_elements_by_tag_name('option'))
#         #for option in browser_var2.find_element_by_xpath("//*[@id='ctl00_ContentPlaceHolder1_fromYear']").find_elements_by_tag_name('option'):
#         #for option in from_year.find_elements_by_tag_name('option'):
#         #for option in from_year_options:
#         #    if option.text == period_dict[period_type][0][1]:
#         #        from_year_option_attempt = 1
#         #        while True:
#         #            try:
#         #                option.click()
#         #                break
#         #            except StaleElementReferenceException:
#         #                if from_year_option_attempt == 3:
#         #                    raise
#         #                from_year_option_attempt += 1
#         #                time.sleep(0.5)
#         # from_year_option_attempt = 1
#         # while (from_year_option_attempt != 0):
#         #     try:
#         #         Select(browser_var2.find_element_by_xpath("//*[@id='ctl00_ContentPlaceHolder1_fromYear']")).select_by_visible_text(period_dict[period_type][0][1])
#         #         from_year_option_attempt = 0
#         #     except StaleElementReferenceException:
#         #         if from_year_option_attempt == 4:
#         #             raise
#         #     from_year_option_attempt += 1
#         #     time.sleep(0.5)
#         #update begin month
#         time.sleep(1)
#         #time.sleep(0.5)
#         #fm = browser_var2.find_element_by_xpath("//*[@id='ctl00_ContentPlaceHolder1_fromMonth']")
#         #from_month = WebDriverWait(browser_var2, 120).until(find_from_month)
#
#         #from_month_attempt = 1
#         #while True:
#         #    try:
#         #        from_month = WebDriverWait(browser_var2, 120).until(find_from_month)
#         #        break
#         #    except StaleElementReferenceException:
#         #        if from_month_attempt == 3:
#         #            raise
#         #        from_month_attempt += 1
#         #        time.sleep(0.5)
#
#         #for option in from_month.find_elements_by_tag_name('option'):
#         #    if option.text == period_dict[period_type][0][0]:
#         #        option.click()
#         #        time.sleep(0.5)
#         #from_month = while_loop_handler_function(WebDriverWait(browser_var2, 120).until(find_from_month))
#         #time.sleep(0.5)
#         #from_month_options = while_loop_handler_function(from_month.find_elements_by_tag_name('option'))
#
#         #for option in browser_var2.find_element_by_xpath("//*[@id='ctl00_ContentPlaceHolder1_fromMonth']").find_elements_by_tag_name('option'):
#         #for option in from_month.find_elements_by_tag_name('option'):
#         #for option in from_month_options:
#         #    if option.text == period_dict[period_type][0][0]:
#         #        from_month_option_attempt = 1
#         #        while True:
#         #            try:
#         #                option.click()
#         #                break
#         #            except StaleElementReferenceException:
#         #                if from_month_option_attempt == 3:
#         #                    raise
#         #                from_month_option_attempt += 1
#         #                time.sleep(0.5)
#
#         # from_month_option_attempt = 1
#         # while (from_month_option_attempt != 0):
#         #     try:
#         #         Select(browser_var2.find_element_by_xpath("//*[@id='ctl00_ContentPlaceHolder1_fromMonth']")).select_by_visible_text(period_dict[period_type][0][0])
#         #         from_month_option_attempt = 0
#         #     except StaleElementReferenceException:
#         #         if from_month_option_attempt == 4:
#         #             raise
#         #     from_month_option_attempt += 1
#         #     time.sleep(0.5)
#
#         #frommonthdropdown = browser_var2.find_element_by_id("FromMonthDiv")
#         #frommonthdropdown.click()
#
#         frommonthdropdown = WebDriverWait(browser_var2, 120).until(find_from_month_div)
#         frommonthdropdown.click()
#
#         Select(browser_var2.find_element_by_xpath("//*[@id='ctl00_ContentPlaceHolder1_fromMonth']")).select_by_visible_text(period_dict[period_type][0][0])
#         #update end year
#         time.sleep(1)
#
#         #ty = browser_var2.find_element_by_xpath("//*[@id='ctl00_ContentPlaceHolder1_toYear']")
#         #to_year = WebDriverWait(browser_var2, 120).until(find_to_year)
#
#         #to_year_attempt = 1
#         #while True:
#         #    try:
#         #        to_year = WebDriverWait(browser_var2, 120).until(find_to_year)
#         #        break
#         #    except StaleElementReferenceException:
#         #        if to_year_attempt == 3:
#         #            raise
#         #        to_year_attempt += 1
#         #        time.sleep(0.5)
#
#         #to_year = while_loop_handler_function(WebDriverWait(browser_var2, 120).until(find_to_year))
#         #time.sleep(0.5)
#         #to_year_options = while_loop_handler_function(to_year.find_elements_by_tag_name('option'))
#
#
#         #Select To Month before To Year in order avoid Vizient's stupid popup window...
#         tomonthdropdown = WebDriverWait(browser_var2, 120).until(find_to_month_div)
#         tomonthdropdown.click()
#
#         Select(browser_var2.find_element_by_xpath("//*[@id='ctl00_ContentPlaceHolder1_toMonth']")).select_by_visible_text(period_dict[period_type][1][0])
#
#         time.sleep(1)
#
#
#         toyeardropdown = WebDriverWait(browser_var2, 120).until(find_to_year_div)
#         toyeardropdown.click()
#
#         '''
#         #Logic to handle popup window indicating the data was deprecated...
#         try:
#             WebDriverWait(browser_var2, 10).until(EC.alert_is_present(),'No data available for selected time period')
#
#             obj = WebDriverWait(browser_var2, 10).until(find_alert_popup2)
#
#             obj.accept()
#         except:
#             pass
#         '''
#         #for option in to_year.find_elements_by_tag_name('option'):
#         #    if option.text == period_dict[period_type][1][1]:
#         #        option.click()
#         #        time.sleep(0.5)
#
#         #for option in browser_var2.find_element_by_xpath("//*[@id='ctl00_ContentPlaceHolder1_toYear']").find_elements_by_tag_name('option'):
#         # for option in to_year.find_elements_by_tag_name('option'):
#         # #for option in to_year_options:
#         #     if option.text == period_dict[period_type][1][1]:
#         #         to_year_option_attempt = 1
#         #         while True:
#         #             try:
#         #                 option.click()
#         #                 break
#         #             except StaleElementReferenceException:
#         #                 if to_year_option_attempt == 3:
#         #                     raise
#         #                 to_year_option_attempt += 1
#         #                 time.sleep(0.5)
#         # time.sleep(0.5)
#         Select(browser_var2.find_element_by_xpath("//*[@id='ctl00_ContentPlaceHolder1_toYear']")).select_by_visible_text(period_dict[period_type][1][1])
#         time.sleep(1)
#         #update end month
#         #Select(browser_var2.find_element_by_xpath("//*[@id='ctl00_ContentPlaceHolder1_toMonth']")).select_by_visible_text(period_dict[period_type][1][0])
#
#         #tm = browser_var2.find_element_by_xpath("//*[@id='ctl00_ContentPlaceHolder1_toMonth']")
#
#         #to_month = WebDriverWait(browser_var2, 120).until(find_to_month)
#
#         #to_month_attempt = 1
#         #while True:
#         #    try:
#         #        to_month = WebDriverWait(browser_var2, 120).until(find_to_month)
#         #        break
#         #    except StaleElementReferenceException:
#         #        if to_month_attempt == 3:
#         #            raise
#         #        to_month_attempt += 1
#         #        time.sleep(0.5)
#
#         #to_month = while_loop_handler_function(WebDriverWait(browser_var2, 120).until(find_to_month))
#
#         #to_month_options = while_loop_handler_function(to_month.find_elements_by_tag_name('option'))
#
#         #for option in to_month.find_elements_by_tag_name('option'):
#         #    if option.text == period_dict[period_type][1][0]:
#         #        option.click()
#         #        time.sleep(0.5)
#         '''
#         tomonthdropdown = WebDriverWait(browser_var2, 120).until(find_to_month_div)
#         tomonthdropdown.click()
#
#         Select(browser_var2.find_element_by_xpath("//*[@id='ctl00_ContentPlaceHolder1_toMonth']")).select_by_visible_text(period_dict[period_type][1][0])
#         '''
#         #for option in browser_var2.find_element_by_xpath("//*[@id='ctl00_ContentPlaceHolder1_toMonth']").find_elements_by_tag_name('option'):
#         # for option in to_month.find_elements_by_tag_name('option'):
#         # #for option in to_month_options:
#         #     if option.text == period_dict[period_type][1][0]:
#         #         to_month_option_attempt = 1
#         #         while True:
#         #             try:
#         #                 option.click()
#         #                 break
#         #             except StaleElementReferenceException:
#         #                 if to_month_option_attempt == 3:
#         #                     raise
#         #                 to_month_option_attempt += 1
#         #                 time.sleep(0.5)


def update_time_period_select(browser_var2, period_dict, period_type):
    time.sleep(1)
    if period_type == 'QUARTER':
        # browser_var2.implicitly_wait(30)
        el = while_loop_handler_function(
            browser_var2.find_element_by_id('ctl00_ContentPlaceHolder1_cmdTimePeriodQuarters'))
        # print(period_dict['QUARTER'])

        # Select the element. This element is now a variable. The variable is in the drop-down?
        option_list1 = while_loop_handler_function(el.find_elements_by_tag_name('option'))

        # UL002
        # dynamically create a list of the pre-selected options in the Quarters menu because each calculator year
        # changes the defaults.  This will make the code more robust rather than hard coding quarters.
        pre_selected_options = [x.text for ind, x in enumerate(option_list1) if
                                (ind == 0) or (x.get_attribute("selected") == 'true')]
        # print(pre_selected_options)
        # First go through and click the options that are in your period dictionary but not pre-selected.
        # We do this because when you click on the menu, it automatically clicks
        # on the top option and does some funky stuff.  So, you need to first go through
        # and click on everything not pre-selected and not the first option.
        # Then go back and click again or "unclick" the options that were pre-selected but not what you wanted.
        for optiona in option_list1:
            # Is the option from the drop-down at this point of iteration in our user-defined set of quarters? ALSO, is it NOT one of the following values?
            if optiona.text in period_dict['QUARTER'] and optiona.text not in pre_selected_options:

                attempt_option_click = 1
                while (attempt_option_click != 0):
                    try:
                        # print('clicking on option.')
                        optiona.click()  # select() in earlier versions of webdriver
                        attempt_option_click = 0
                        # print('clicked on option.')
                    except StaleElementReferenceException:
                        if attempt_option_click == 3:
                            raise
                        attempt_option_click += 1
                        time.sleep(0.5)
        # now find all the menu options selected after the first round.
        selected_after_first_round = [x.text for ind, x in enumerate(option_list1) if
                                      (x.get_attribute("selected") == 'true')]

        # Unselected/Unclick by clicking again the items that are not in your period dictionary but are still
        # selected due to the preselection defaults.

        for optionb in option_list1:
            if optionb.text in selected_after_first_round and optionb.text not in period_dict['QUARTER']:
                attempt_option_click = 1
                while (attempt_option_click != 0):
                    try:
                        # print('clicking on option.')
                        optionb.click()  # select() in earlier versions of webdriver
                        attempt_option_click = 0
                        print('clicked on option.', optionb.text)
                    except StaleElementReferenceException:
                        if attempt_option_click == 3:
                            raise
                        attempt_option_click += 1
                        time.sleep(0.5)
    else:
        # update begin year
        time.sleep(1)

        # fromyeardropdown = browser_var2.find_element_by_id("FromYearDiv")
        # fromyeardropdown.click()

        fromyeardropdown = WebDriverWait(browser_var2, 120).until(find_from_year_div)
        fromyeardropdown.click()

        Select(
            browser_var2.find_element_by_xpath("//*[@id='ctl00_ContentPlaceHolder1_fromYear']")).select_by_visible_text(
            period_dict[period_type][0][1])

        time.sleep(1)

        frommonthdropdown = WebDriverWait(browser_var2, 120).until(find_from_month_div)
        frommonthdropdown.click()

        Select(browser_var2.find_element_by_xpath(
            "//*[@id='ctl00_ContentPlaceHolder1_fromMonth']")).select_by_visible_text(period_dict[period_type][0][0])
        # update end year
        time.sleep(1)

        # Select To Month before To Year in order avoid Vizient's stupid popup window...
        tomonthdropdown = WebDriverWait(browser_var2, 120).until(find_to_month_div)
        tomonthdropdown.click()

        Select(
            browser_var2.find_element_by_xpath("//*[@id='ctl00_ContentPlaceHolder1_toMonth']")).select_by_visible_text(
            period_dict[period_type][1][0])

        time.sleep(1)

        toyeardropdown = WebDriverWait(browser_var2, 120).until(find_to_year_div)
        toyeardropdown.click()

        Select(
            browser_var2.find_element_by_xpath("//*[@id='ctl00_ContentPlaceHolder1_toYear']")).select_by_visible_text(
            period_dict[period_type][1][1])
        time.sleep(1)




#UL002 END
#################################################################################################

# Function to select multiple group by
# EDAC and Readmission measure templates can only be filter by quarter so we need to group by discharge month
# in order to get monthly values.

def update_group_by_select(browser_var2,period_dict,period_type):
    time.sleep(0.5)
    if period_type == 'QUARTER':

        multgroupbydiv = WebDriverWait(browser_var2, 120).until(find_mult_group_by_div)
        multgroupbydiv.click()

        Select(browser_var2.find_element_by_xpath("//*[@id='ctl00_ContentPlaceHolder1_cmdMeasuresByWithMult']")).select_by_visible_text('Discharge Month 1st Admit')
        time.sleep(0.5)

    else:
        #update begin year
        time.sleep(0.5)
        pass






################################################################################################

def choose_adjustment_model(browser_var2, link_dict, key):

    if link_dict[key][5].replace(' ','').upper() == '2019RISKMODEL(AMC)':
        print(link_dict[key][5].replace(' ','').upper())
        #print(link_dict[key][5])
        # Click the Risk Adjustment Model radio button
        #adjustment_model_btn = WebDriverWait(browser_var2, 45).until(find_adjustment_model_2018_amc)
        # adjustment_model_btn = browser_var2.find_element_by_xpath("//*/label[@for='ctl00_ContentPlaceHolder1_radModifiedMSDRG'][contains(text(), '2018 Risk Model (AMC)')]")
        # adjustment_model_btn.click()
        adjustment_model_btn_attempt = 1
        while True:
            try:
                adjustment_model_btn = WebDriverWait(browser_var2, 45).until(find_adjustment_model_2019_amc)
                break
            except StaleElementReferenceException:
                if adjustment_model_btn_attempt == 3:
                    raise
                adjustment_model_btn_attempt += 1
        #ActionChains(browser_var2).move_to_element(adjustment_model_btn).click().perform()

        adjustment_model_btn_click_attempt = 1
        print('Click Attempt:',adjustment_model_btn_click_attempt)
        while True:
            try:
                adjustment_model_btn.click()
                #ActionChains(browser_var2).move_to_element(adjustment_model_btn).click().perform()
                break
            except StaleElementReferenceException:
                if adjustment_model_btn_click_attempt == 3:
                    raise
                adjustment_model_btn_click_attempt += 1
                
                
                
    elif link_dict[key][5].replace(' ','').upper() == '2019RISKMODEL(COMMUNITY)':

        
        #adjustment_model_btn = WebDriverWait(browser_var2, 45).until(find_adjustment_model_2018_comm)
        # adjustment_model_btn = browser_var2.find_element_by_xpath("//*/label[@for='ctl00_ContentPlaceHolder1_radModifiedMSDRG_Commu'][contains(text(), '2018 Risk Model (Community)')]")
        # adjustment_model_btn.click()
        #ActionChains(browser_var2).move_to_element(adjustment_model_btn).click().perform()
        # print('yep')
        adjustment_model_btn_attempt = 1
        while True:
            try:
                adjustment_model_btn = WebDriverWait(browser_var2, 45).until(find_adjustment_model_2019_comm)
                break
            except StaleElementReferenceException:
                if adjustment_model_btn_attempt == 3:
                    raise
                adjustment_model_btn_attempt += 1
        # ActionChains(browser_var2).move_to_element(adjustment_model_btn).click().perform()

        adjustment_model_btn_click_attempt = 1
        while True:
            try:
                adjustment_model_btn.click()
                #ActionChains(browser_var2).move_to_element(adjustment_model_btn).click().perform()
                break
            except StaleElementReferenceException:
                if adjustment_model_btn_click_attempt == 3:
                    raise
                adjustment_model_btn_click_attempt += 1
    # print(link_dict[key][5])
    elif link_dict[key][5] == '2018 Risk Model (AMC)':
        # Click the Risk Adjustment Model radio button
        #adjustment_model_btn = WebDriverWait(browser_var2, 45).until(find_adjustment_model_2018_amc)
        # adjustment_model_btn = browser_var2.find_element_by_xpath("//*/label[@for='ctl00_ContentPlaceHolder1_radModifiedMSDRG'][contains(text(), '2018 Risk Model (AMC)')]")
        # adjustment_model_btn.click()
        adjustment_model_btn_attempt = 1
        while True:
            try:
                adjustment_model_btn = WebDriverWait(browser_var2, 45).until(find_adjustment_model_2018_amc)
                break
            except StaleElementReferenceException:
                if adjustment_model_btn_attempt == 3:
                    raise
                adjustment_model_btn_attempt += 1
        #ActionChains(browser_var2).move_to_element(adjustment_model_btn).click().perform()

        adjustment_model_btn_click_attempt = 1
        while True:
            try:
                ActionChains(browser_var2).move_to_element(adjustment_model_btn).click().perform()
                break
            except StaleElementReferenceException:
                if adjustment_model_btn_click_attempt == 3:
                    raise
                adjustment_model_btn_click_attempt += 1
    elif link_dict[key][5] == '2018 Risk Model (Community)':

        #adjustment_model_btn = WebDriverWait(browser_var2, 45).until(find_adjustment_model_2018_comm)
        # adjustment_model_btn = browser_var2.find_element_by_xpath("//*/label[@for='ctl00_ContentPlaceHolder1_radModifiedMSDRG_Commu'][contains(text(), '2018 Risk Model (Community)')]")
        # adjustment_model_btn.click()
        #ActionChains(browser_var2).move_to_element(adjustment_model_btn).click().perform()
        # print('yep')
        adjustment_model_btn_attempt = 1
        while True:
            try:
                adjustment_model_btn = WebDriverWait(browser_var2, 45).until(find_adjustment_model_2018_comm)
                break
            except StaleElementReferenceException:
                if adjustment_model_btn_attempt == 3:
                    raise
                adjustment_model_btn_attempt += 1
        # ActionChains(browser_var2).move_to_element(adjustment_model_btn).click().perform()

        adjustment_model_btn_click_attempt = 1
        while True:
            try:
                ActionChains(browser_var2).move_to_element(adjustment_model_btn).click().perform()
                break
            except StaleElementReferenceException:
                if adjustment_model_btn_click_attempt == 3:
                    raise
                adjustment_model_btn_click_attempt += 1
    elif link_dict[key][5] == '2017 Risk Model (AMC)':
        #adjustment_model_btn = WebDriverWait(browser_var2, 45).until(find_adjustment_model_2017_amc)
        # adjustment_model_btn = browser_var2.find_element_by_xpath("//*/label[@for='ctl00_ContentPlaceHolder1_radMSDRG'][contains(text(), '2017 Risk Model (AMC)')]")
        # adjustment_model_btn.click()
        #ActionChains(browser_var2).move_to_element(adjustment_model_btn).click().perform()
        adjustment_model_btn_attempt = 1
        while True:
            try:
                adjustment_model_btn = WebDriverWait(browser_var2, 45).until(find_adjustment_model_2017_amc)
                break
            except StaleElementReferenceException:
                if adjustment_model_btn_attempt == 3:
                    raise
                adjustment_model_btn_attempt += 1
        # ActionChains(browser_var2).move_to_element(adjustment_model_btn).click().perform()

        adjustment_model_btn_click_attempt = 1
        while True:
            try:
                ActionChains(browser_var2).move_to_element(adjustment_model_btn).click().perform()
                break
            except StaleElementReferenceException:
                if adjustment_model_btn_click_attempt == 3:
                    raise
                adjustment_model_btn_click_attempt += 1
    elif link_dict[key][5] == '2017 Risk Model (Community)':
        #adjustment_model_btn = WebDriverWait(browser_var2, 45).until(find_adjustment_model_2017_comm)
        # adjustment_model_btn = browser_var2.find_element_by_xpath("//*/label[@for='ctl00_ContentPlaceHolder1_radMSDRG_Commu'][contains(text(), '2017 Risk Model (Community)')]")
        # adjustment_model_btn.click()
        #ActionChains(browser_var2).move_to_element(adjustment_model_btn).click().perform()
        adjustment_model_btn_attempt = 1
        while True:
            try:
                adjustment_model_btn = WebDriverWait(browser_var2, 45).until(find_adjustment_model_2017_comm)
                break
            except StaleElementReferenceException:
                if adjustment_model_btn_attempt == 3:
                    raise
                adjustment_model_btn_attempt += 1
        # ActionChains(browser_var2).move_to_element(adjustment_model_btn).click().perform()

        adjustment_model_btn_click_attempt = 1
        while True:
            try:
                ActionChains(browser_var2).move_to_element(adjustment_model_btn).click().perform()
                break
            except StaleElementReferenceException:
                if adjustment_model_btn_click_attempt == 3:
                    raise
                adjustment_model_btn_click_attempt += 1
    # except:
    #    print('Something is wrong with the Risk Adjustment Model button.')
    return (browser_var2)


################################################################################################

def choose_ahrq_version(browser_var2, link_dict, key):
    # print(link_dict[key][6])
    try:
        #UL005
        if link_dict[key][6] == 'V2019 (Pediatric) / V2019 (Quality) / V2019 (Safety)':

            # Click the Risk Adjustment Model radio button
            ahrq_version_btn = WebDriverWait(browser_var2, 30).until(find_ahrq_version_v2019)
            ActionChains(browser_var2).move_to_element(ahrq_version_btn).click().perform()
        
        elif link_dict[key][6] == '8.0 (CMS Safety)':

            # Click the Risk Adjustment Model radio button
            ahrq_version_btn = WebDriverWait(browser_var2, 30).until(find_ahrq_version_8)
            # ahrq_version_btn = browser_var2.find_element_by_xpath("//*/label[@for='ctl00_ContentPlaceHolder1_radAHRQCurrent'][contains(text(), '8.0 (CMS Safety)')]")
            # ahrq_version_btn.click()
            ActionChains(browser_var2).move_to_element(ahrq_version_btn).click().perform()
            # print('yep_ahrq')
        elif link_dict[key][6] == '7.0.1 (Pediatric) / 7.0.1 (Quality) / 7.0.1 (Safety)':
            ahrq_version_btn = WebDriverWait(browser_var2, 30).until(find_ahrq_version_7)
            # ahrq_version_btn = browser_var2.find_element_by_xpath("//*/label[@for='ctl00_ContentPlaceHolder1_radAHRQPrevious'][contains(text(), '7.0.1 (Pediatric) / 7.0.1 (Quality) / 7.0.1 (Safety)')]")
            # ahrq_version_btn.click()
            ActionChains(browser_var2).move_to_element(ahrq_version_btn).click().perform()
        elif link_dict[key][5] == '6.0.2 (Pediatric) / 6.0.2 (Quality) / 6.0.2 (Safety)':
            ahrq_version_btn = WebDriverWait(browser_var2, 30).until(find_ahrq_version_6)
            # ahrq_version_btn = browser_var2.find_element_by_xpath("//*/label[@for='ctl00_ContentPlaceHolder1_radAHRQPrevious1'][contains(text(), '6.0.2 (Pediatric) / 6.0.2 (Quality) / 6.0.2 (Safety)')]")
            # ahrq_version_btn.click()
            ActionChains(browser_var2).move_to_element(ahrq_version_btn).click().perform()
    except:
        print('Something is wrong with the AHRQ Version button.')
    return (browser_var2)

################################################################################################

def find_last_downloaded_file(dir):
    list_of_files = glob.glob(dir + '/*')
    latest_file = max(list_of_files, key=os.path.getctime)
    ##UL008
    while latest_file.endswith('.crdownload') == True:
        print('accidentally grabbed crdownload temp file.  Trying again...')
        time.sleep(1)
        list_of_files = glob.glob(dir + '/*')
        latest_file = max(list_of_files, key=os.path.getctime)
    print(latest_file)
    return(latest_file)

################################################################################################

def rename_and_move_file(file, hospital_type, measure_name, period_type, new_file_dir):
    # os.chdir('C:/Data/Downloads')
    if hospital_type == 'Complex Care Medical Center':
        hospital_type2 = 'CCMC'
    elif hospital_type == 'Comprehensive Academic Medical Center':
        hospital_type2 = 'AMC'
    elif (hospital_type == 'Large Specialized Complex Care Medical Center') or (hospital_type == 'Large, Specialized Complex Care Medical Center'):
        hospital_type2 = 'LSCCMC'
    #UL002 add critical access to file naming options
    elif hospital_type == 'Critical Access & Small Community':
        hospital_type2 = 'CASC'
    else:
        hospital_type2 = 'COMM'

    new_hospital_name = (str(hospital_type2).replace(" ", "_")).replace("-", "_").upper()
    new_measure_name = (str(measure_name).replace(" ", "_")).replace("-", "_").upper()
    new_period_type = (str(period_type).replace(" ", "_")).replace("-", "_").upper()
    new_file_name = new_hospital_name + '_' + new_measure_name + '_' + new_period_type + '.xlsx'

    new_path = os.path.join(new_file_dir, str(hospital_type), str(new_measure_name), new_file_name)
    main_cohort_path = os.path.join(new_file_dir, str(hospital_type), str(new_measure_name))

    if not os.path.exists(main_cohort_path):
        os.makedirs(main_cohort_path)
    if os.path.exists(new_path):
        print('File already exists!')
        print(new_path)
        overwrite_decision = input("Do you really want to overwrite this file? Choose: 'YES' or 'NO'")

        if overwrite_decision.upper() == 'YES':
            shutil.move(file, new_path)
        else:
            print('Please handle the existing file or change the directory location.')
            exit()
    # print('check check...')
    # print(file)
    # print(new_path)
    shutil.move(file, new_path)
    # os.remove(file)

################################################################################################

# List of measures to exclude for testing.  These are excluded because Vizient currently has
# the time period radio button disabled on these templates.  We need to wait for them to re-activate them.
#list_of_total_revisits_measures = ['Readmission - Cardiology','Readmission - CT Surgery','Readmission - CT Surgery','Readmission - Gastroenterology','Readmission - Medicine General','Readmission - Neurology','Readmission - Neurosurgery','Readmission - Oncology','Readmission - Ortho/Spine','Readmission - Pulmonary/Critical Care','Readmission - Solid Organ Transplant','Readmission - Surgery General','Readmission - Trauma','Readmission - Vascular Surgery','Excess Days  - Cardiology','Excess Days - CT','Excess Days - Gastroenterology','Excess Days - Medicine General','Excess Days - Neurology','Excess Days - Neurosurgery','Excess Days - Oncology','Excess Days - Ortho/Spine','Excess Days - Pulmonary/Critical Care','Excess Days - Solid Organ Transplant','Excess Days - Surgery General','Excess Days - Trauma','Excess Days - Vascular Surgery']

list_of_total_revisits_measures = ['Excess Days  - Cardiology', 'Excess Days - CT', 'Excess Days - Gastroenterology', 'Excess Days - Medicine General', 'Excess Days - Neurology', 'Excess Days - Neurosurgery', 'Excess Days - Oncology', 'Excess Days - Ortho/Spine', 'Excess Days - Pulmonary/Critical Care', 'Excess Days - Solid Organ Transplant', 'Excess Days - Surgery General', 'Excess Days - Trauma', 'Excess Days - Vascular Surgery', 'Readmission - CT Surgery', 'Readmission - Cardiology', 'Readmission - Gastroenterology', 'Readmission - Medicine General', 'Readmission - Neurology', 'Readmission - Neurosurgery', 'Readmission - Oncology', 'Readmission - Ortho/Spine', 'Readmission - Pulmonary/Critical Care', 'Readmission - Solid Organ Transplant', 'Readmission - Surgery General', 'Readmission - Trauma', 'Readmission - Vascular Surgery', 'Readmissions - Medical', 'Readmissions - Surgical']

list_of_measures_to_run_twice_num_denom = ['Adverse Drug Events Rate','% Early Transfers Out']


################################################################################################

# Query the database and get core measure values for ED-2B and ED-OP-18B

def get_core_measure_metric_vals(new_file_path,begin_time,end_time):
    new_file_path = os.path.abspath(new_file_path)
    # connect to SQL Server
    core_conn = pyodbc.connect('Driver={SQL Server};'
                          'Server=edw00pd05wva.corp.nm.org\EDWIDS1;'
                          'Database=clarity;'
                          'Trusted_Connection=yes;')

    # set the begin and end dates for the queries based on case_date.
    begin_dts = begin_time
    end_dts = end_time
    #begin_dts = input("Core Measures:  Enter the beginning datetime (format:  'xx-01-xxxx')")
    #end_dts = input("Core Measures:  Enter the end datetime (format:  'xx-xx-xxxx 23:59:59')")

    # Query for ED-2B
    sql_ed2b = """
    DECLARE @begin_dts as datetime; SET @begin_dts = '%s';
    DECLARE @end_dts as datetime; SET @end_dts = '%s';

    SELECT
    medians.*
    ,num_size.[Num Size]
    ,denom_size.[Denom Size]
    FROM
    (
    select
    CASE
                WHEN mvm.attrib_id = 1001 THEN '140281 NORTHWESTERN_MEMORIAL'
                WHEN mvm.attrib_id = 2678 THEN '140242 NORTHWESTERN_CDH'
                WHEN mvm.attrib_id = 1113 THEN '140130 NORTHWESTERN_LAKEFOREST'
                WHEN mvm.attrib_id = 4904 THEN '140286 NORTHWESTERN_KISH'
                WHEN mvm.attrib_id = 2679 THEN '140211 NORTHWESTERN_DELNOR'
                WHEN mvm.attrib_id = 4905 THEN '141340 NORTHWESTERN_VALLEYW'
				--UL003
				WHEN mvm.attrib_id = 7628 THEN '140116 NORTHWESTERN_MCHENRY'
                WHEN mvm.attrib_id = 7627 THEN '149916 NORTHWESTERN_HUNTLEY'
            END as 'Hospital'
            ,'Effectiveness' as Domain
            ,'ED-2B' as Measure
			, mvm.value as [Metric Value]
    from
    NM_Performance.metric.metric_values_merged mvm
    join NM_Performance.period.period_lookup as pl 
    on pl.period_id = mvm.period_id
    and pl.end_dts =  @end_dts
    and pl.period_type = 'fscl_ytd'
    where
    metric_id in 
    (
    11990  -- ED-2b: ED-2b: Median Admit Decision Time to ED Departure Time for Admitted Patients
    )
    and
    mvm.attrib_id
    in
    (
    1001	--Northwestern Memorial Hospital
    ,1113	--Northwestern Medicine Lake Forest Hospital
    ,2678	--Central DuPage Hospital
    ,2679	--Delnor Hospital
    ,4904	--Kishwaukee Hospital
    ,4905	--Valley West Community Hospital
	--UL003
	,7628	--Northwestern Medicine McHenry Hospital
    ,7627	--Northwestern Medicine Huntley Hospital

    )
    ) medians

    LEFT JOIN

    (
    select
    CASE
                WHEN mvm.attrib_id = 1001 THEN '140281 NORTHWESTERN_MEMORIAL'
                WHEN mvm.attrib_id = 2678 THEN '140242 NORTHWESTERN_CDH'
                WHEN mvm.attrib_id = 1113 THEN '140130 NORTHWESTERN_LAKEFOREST'
                WHEN mvm.attrib_id = 4904 THEN '140286 NORTHWESTERN_KISH'
                WHEN mvm.attrib_id = 2679 THEN '140211 NORTHWESTERN_DELNOR'
                WHEN mvm.attrib_id = 4905 THEN '141340 NORTHWESTERN_VALLEYW'
				--UL003
				WHEN mvm.attrib_id = 7628 THEN '140116 NORTHWESTERN_MCHENRY'
                WHEN mvm.attrib_id = 7627 THEN '149916 NORTHWESTERN_HUNTLEY'
            END as 'Hospital'
            ,'Effectiveness' as Domain
            ,'ED-2B' as Measure
			, mvm.value as [Num Size]
    from
    NM_Performance.metric.metric_values_merged mvm
    join NM_Performance.period.period_lookup as pl 
    on pl.period_id = mvm.period_id
    and pl.end_dts =  @end_dts
    and pl.period_type = 'fscl_ytd'
    where
    metric_id in 
    (
    11988	--ED-2b: Admit Decision Time to ED Departure Time for Admitted Patients  SUM
    )
    and
    mvm.attrib_id
    in
    (
    1001	--Northwestern Memorial Hospital
    ,1113	--Northwestern Medicine Lake Forest Hospital
    ,2678	--Central DuPage Hospital
    ,2679	--Delnor Hospital
    ,4904	--Kishwaukee Hospital
    ,4905	--Valley West Community Hospital
	--UL003
	,7628	--Northwestern Medicine McHenry Hospital
    ,7627	--Northwestern Medicine Huntley Hospital

    )
    ) num_size
    on 
    num_size.Hospital = medians.hospital

    LEFT JOIN

    (
    select
    CASE
                WHEN mvm.attrib_id = 1001 THEN '140281 NORTHWESTERN_MEMORIAL'
                WHEN mvm.attrib_id = 2678 THEN '140242 NORTHWESTERN_CDH'
                WHEN mvm.attrib_id = 1113 THEN '140130 NORTHWESTERN_LAKEFOREST'
                WHEN mvm.attrib_id = 4904 THEN '140286 NORTHWESTERN_KISH'
                WHEN mvm.attrib_id = 2679 THEN '140211 NORTHWESTERN_DELNOR'
                WHEN mvm.attrib_id = 4905 THEN '141340 NORTHWESTERN_VALLEYW'
				--UL003
				WHEN mvm.attrib_id = 7628 THEN '140116 NORTHWESTERN_MCHENRY'
                WHEN mvm.attrib_id = 7627 THEN '149916 NORTHWESTERN_HUNTLEY'
            END as 'Hospital'
            ,'Effectiveness' as Domain
            ,'ED-2B' as Measure
			, mvm.value as [Denom Size]
    from
    NM_Performance.metric.metric_values_merged mvm
    join NM_Performance.period.period_lookup as pl 
    on pl.period_id = mvm.period_id
    and pl.end_dts =  @end_dts
    and pl.period_type = 'fscl_ytd'
    where
    metric_id in 
    (
    11989	--ED-2b: Admit Decision Time to ED Departure Time for Admitted Patients  COUNT
    )
    and
    mvm.attrib_id
    in
    (
    1001	--Northwestern Memorial Hospital
    ,1113	--Northwestern Medicine Lake Forest Hospital
    ,2678	--Central DuPage Hospital
    ,2679	--Delnor Hospital
    ,4904	--Kishwaukee Hospital
    ,4905	--Valley West Community Hospital
	--UL003
	,7628	--Northwestern Medicine McHenry Hospital
    ,7627	--Northwestern Medicine Huntley Hospital
    )
    ) denom_size
    on denom_size.Hospital = medians.Hospital
    order BY
    medians.Hospital
    """ % (begin_dts,end_dts)

    # Query for ED-OP-18B
    sql_op18b = """
    DECLARE @begin_dts as datetime; SET @begin_dts = '%s';
    DECLARE @end_dts as datetime; SET @end_dts = '%s';


    SELECT
    medians.*
    ,num_size.[Num Size]
    ,denom_size.[Denom Size]
    FROM
    (
    select
    CASE
                WHEN mvm.attrib_id = 1001 THEN '140281 NORTHWESTERN_MEMORIAL'
                WHEN mvm.attrib_id = 2678 THEN '140242 NORTHWESTERN_CDH'
                WHEN mvm.attrib_id = 1113 THEN '140130 NORTHWESTERN_LAKEFOREST'
                WHEN mvm.attrib_id = 4904 THEN '140286 NORTHWESTERN_KISH'
                WHEN mvm.attrib_id = 2679 THEN '140211 NORTHWESTERN_DELNOR'
                WHEN mvm.attrib_id = 4905 THEN '141340 NORTHWESTERN_VALLEYW'
				--UL003
				WHEN mvm.attrib_id = 7628 THEN '140116 NORTHWESTERN_MCHENRY'
                WHEN mvm.attrib_id = 7627 THEN '149916 NORTHWESTERN_HUNTLEY'
            END as 'Hospital'
            ,'Effectiveness' as Domain
            ,'ED-OP-18B' as Measure
            , mvm.value as [Metric Value]
    from
    NM_Performance.metric.metric_values_merged mvm
    join NM_Performance.period.period_lookup as pl 
    on pl.period_id = mvm.period_id
    and pl.end_dts =  @end_dts
    and pl.period_type = 'fscl_ytd'
    where
    metric_id in 
    (
    12066	--OP-18b: Median Time from ED Arrival to ED Departure for Discharged ED Patients
    )
    and
    mvm.attrib_id
    in  
    (
    1001	--Northwestern Memorial Hospital
    ,1113	--Northwestern Medicine Lake Forest Hospital
    ,2678	--Central DuPage Hospital
    ,2679	--Delnor Hospital
    ,4904	--Kishwaukee Hospital
    ,4905	--Valley West Community Hospital
	--UL003
	,7628	--Northwestern Medicine McHenry Hospital
    ,7627	--Northwestern Medicine Huntley Hospital
    )
    ) medians

    LEFT JOIN

    (
    select
    CASE
                WHEN mvm.attrib_id = 1001 THEN '140281 NORTHWESTERN_MEMORIAL'
                WHEN mvm.attrib_id = 2678 THEN '140242 NORTHWESTERN_CDH'
                WHEN mvm.attrib_id = 1113 THEN '140130 NORTHWESTERN_LAKEFOREST'
                WHEN mvm.attrib_id = 4904 THEN '140286 NORTHWESTERN_KISH'
                WHEN mvm.attrib_id = 2679 THEN '140211 NORTHWESTERN_DELNOR'
                WHEN mvm.attrib_id = 4905 THEN '141340 NORTHWESTERN_VALLEYW'
				--UL003
				WHEN mvm.attrib_id = 7628 THEN '140116 NORTHWESTERN_MCHENRY'
                WHEN mvm.attrib_id = 7627 THEN '149916 NORTHWESTERN_HUNTLEY'
            END as 'Hospital'
            ,'Effectiveness' as Domain
            ,'ED-2B' as Measure
            , mvm.value as [Num Size]
    from
    NM_Performance.metric.metric_values_merged mvm
    join NM_Performance.period.period_lookup as pl 
    on pl.period_id = mvm.period_id
    and pl.end_dts =  @end_dts
    and pl.period_type = 'fscl_ytd'
    where
    metric_id in 
    (
    12064	--OP-18b: Median Time from ED Arrival to ED Departure for Discharged ED Patients SUM
    )
    and
    mvm.attrib_id
    in
    (
    1001	--Northwestern Memorial Hospital
    ,1113	--Northwestern Medicine Lake Forest Hospital
    ,2678	--Central DuPage Hospital
    ,2679	--Delnor Hospital
    ,4904	--Kishwaukee Hospital
    ,4905	--Valley West Community Hospital
	--UL003
	,7628	--Northwestern Medicine McHenry Hospital
    ,7627	--Northwestern Medicine Huntley Hospital
    )
    ) num_size
    on 
    num_size.Hospital = medians.hospital

    LEFT JOIN

    (
    select
    CASE
                WHEN mvm.attrib_id = 1001 THEN '140281 NORTHWESTERN_MEMORIAL'
                WHEN mvm.attrib_id = 2678 THEN '140242 NORTHWESTERN_CDH'
                WHEN mvm.attrib_id = 1113 THEN '140130 NORTHWESTERN_LAKEFOREST'
                WHEN mvm.attrib_id = 4904 THEN '140286 NORTHWESTERN_KISH'
                WHEN mvm.attrib_id = 2679 THEN '140211 NORTHWESTERN_DELNOR'
                WHEN mvm.attrib_id = 4905 THEN '141340 NORTHWESTERN_VALLEYW'
				--UL003
				WHEN mvm.attrib_id = 7628 THEN '140116 NORTHWESTERN_MCHENRY'
                WHEN mvm.attrib_id = 7627 THEN '149916 NORTHWESTERN_HUNTLEY'
            END as 'Hospital'
            ,'Effectiveness' as Domain
            ,'ED-2B' as Measure
            , mvm.value as [Denom Size]
    from
    NM_Performance.metric.metric_values_merged mvm
    join NM_Performance.period.period_lookup as pl 
    on pl.period_id = mvm.period_id
    and pl.end_dts =  @end_dts
    and pl.period_type = 'fscl_ytd'
    where
    metric_id in 
    (
    12065	--OP-18b: Median Time from ED Arrival to ED Departure for Discharged ED Patients COUNT
    )
    and
    mvm.attrib_id
    in
    (
    1001	--Northwestern Memorial Hospital
    ,1113	--Northwestern Medicine Lake Forest Hospital
    ,2678	--Central DuPage Hospital
    ,2679	--Delnor Hospital
    ,4904	--Kishwaukee Hospital
    ,4905	--Valley West Community Hospital
	--UL003
	,7628	--Northwestern Medicine McHenry Hospital
    ,7627	--Northwestern Medicine Huntley Hospital
    )
    ) denom_size
    on denom_size.Hospital = medians.Hospital
    order BY
    medians.Hospital
    """ % (begin_dts, end_dts)

    # run the queries and store the results in a pandas dataframe
    ed_2b = pd.DataFrame(pd.read_sql(sql_ed2b, core_conn))
    ed_op18b = pd.DataFrame(pd.read_sql(sql_op18b, core_conn))
    # close the database connection
    core_conn.close()

    list_of_cohort_dirs = os.listdir(new_file_path)

    for i in list_of_cohort_dirs:
        sub_dir = os.path.join(new_file_path, i)

        sub_dir_ed2b = os.path.join(sub_dir, 'ED_2B')
        sub_dir_op18b = os.path.join(sub_dir, 'ED_OP_18B')

        try:
            # check if folder already exists.  If it does not exist, create it.
            if os.path.isfile(sub_dir_ed2b) == False:
                os.mkdir(sub_dir_ed2b)
        except:
            pass

        try:
            # check if folder already exists.  If it does not exist, create it.
            if os.path.isfile(sub_dir_op18b) == False:
                os.mkdir(sub_dir_op18b)
        except:
            pass

        if i == 'Complex Care Medical Center':
            filename_ed2b = 'CCMC_ED_2B_CUSTOM.xlsx'
        elif i == 'Comprehensive Academic Medical Center':
            filename_ed2b = 'AMC_ED_2B_CUSTOM.xlsx'
        elif i == 'Large Specialized Complex Care Medical Center':
            filename_ed2b = 'LSCCMC_ED_2B_CUSTOM.xlsx'
        elif i == 'Community Medical Center':   #UL001
            filename_ed2b = 'COMM_ED_2B_CUSTOM.xlsx'
        elif i == 'Critical Access & Small Community':   #UL003
            filename_ed2b = 'CASC_ED_2B_CUSTOM.xlsx'

        if i == 'Complex Care Medical Center':
            filename_op18b = 'CCMC_ED_OP_18B_CUSTOM.xlsx'
        elif i == 'Comprehensive Academic Medical Center':
            filename_op18b = 'AMC_ED_OP_18B_CUSTOM.xlsx'
        elif i == 'Large Specialized Complex Care Medical Center':
            filename_op18b = 'LSCCMC_ED_OP_18B_CUSTOM.xlsx'
        elif i == 'Community Medical Center': #UL001
            filename_op18b = 'COMM_ED_OP_18B_CUSTOM.xlsx'
        elif i == 'Critical Access & Small Community': #UL003
            filename_op18b = 'CASC_ED_OP_18B_CUSTOM.xlsx'

        new_file_name_ed2b = os.path.join(sub_dir_ed2b, filename_ed2b)
        new_file_name_op18b = os.path.join(sub_dir_op18b, filename_op18b)

        ed_2b.to_excel(new_file_name_ed2b, sheet_name='CDPRM Report')
        ed_op18b.to_excel(new_file_name_op18b, sheet_name='CDPRM Report')
        
################################################################################################
        
##UL004   New functions to handle 2020 Covid custom group by lists and advanced restrictions.  
#         NOTE:  We likely won't need these functions in future years.  This is a special case.
        
##UL004  New functions to remove the custom group by list at the top of the template.

#UL004  Function to find the first instance of the text 'Edit Custom List.' 
def find_edit_custom_list_link(driver):
    #get a list of all instances of the text 'Edit Custom List.'  
    #If that phrase is present in the HTML, then there has been a custom list created.  
    #There could be several, but we care about the first one at the top of the page.
    edit_custom_list_link = driver.find_elements_by_xpath("//*[contains(text(),'" + 'Edit Custom List' + "')]")
    if edit_custom_list_link:
        return(edit_custom_list_link)
    else:
        return(edit_custom_list_link)   
        
#UL004
#After clicking the 'Edit Custom List' link, you need to click 'Remove Custom Group By'
#There are actually many of these buttons on the DOM, so we need to do a conditional click.
#If PSI measure, then click the second button id.  If not, then click the first.
def find_remove_custom_group_by_bttn(driver):
    #Click the button with the value 'Remove Custom Group By' to remove the custom list.
    #remove_custom_group_by_bttn = driver.find_element_by_css_selector(".sbttn[value='Remove Custom Group By']")
    remove_custom_group_by_bttn = driver.find_elements_by_xpath("//*[@id = 'ctl00_ContentPlaceHolder1_ucctl11_Custom2_cmdRemoveCustomGroupby' or @id = 'ctl00_ContentPlaceHolder1_ucctl11_Custom_cmdRemoveCustomGroupby']")
    if remove_custom_group_by_bttn:
        return(remove_custom_group_by_bttn)
    else:
        return(False)  

#UL004
#Put it all together
def remove_custom_covid_list_from_template(browser_var,measure_name_var):
    
    #give it a couple seconds to load
    time.sleep(2)
    #find the first appearance of text 'Edit Custom List.'  If it doesn't exist,
    #then there isn't this restriction.  READM, EDAC, THK don't have this restriction.
    edit_custom_list_list = find_edit_custom_list_link(browser_var)
    
    #some templates do not use a custom list.  Check to see if there is a custom list.  If not, pass.  If yes, then
    #click the link to open up the editor.
    if len(edit_custom_list_list) == 1:
        edit_custom_list_list[0].click()
    else:
        return(browser_var)
    
    #click Remove Custom Group By
    remove_custom_group_by_bttn = WebDriverWait(browser_var, 3).until(find_remove_custom_group_by_bttn)
    
    #if it's a PSI, then choose the second button id element.  Else, 
    if measure_name_var in ['PSI-03 O/E','PSI-11 O/E','PSI-09 O/E','PSI-06 O/E','PSI-13 O/E']:
        remove_custom_group_by_bttn[1].click()
    else:
        remove_custom_group_by_bttn[0].click()
        
    obj = WebDriverWait(browser_var, 3).until(find_alert_popup)

    obj.accept()
    
    return(browser_var)     
    

#UL004
#Vizient also added some multiple group by to the covid/pre-covid report templates 
#we need to remove these to get back to the default template.
#By default, all the templates should be grouped by Hospital and not 
#by discharge month.  The first step is to remove the multiple-group-by
#then choose Hospital/Hospital System from the Group By dropdown menu.
#It seems only PSI have the multiple group by applied but, to be safe, 
#I'm going to do this to all templates.

def find_multiple_group_by_bttn(driver):
    multiple_group_by_bttn = driver.find_element_by_id('ctl00_ContentPlaceHolder1_lbtnGroupBy')
    if multiple_group_by_bttn:
        return(multiple_group_by_bttn)
    else:
        return(False)
#UL004   
def find_multiple_group_by_selected_list(driver):
    multiple_selected_list = driver.find_element_by_id('ctl00_ContentPlaceHolder1_ucctl11_Multi_lstSpecific_GroupBy')
    if multiple_selected_list:
        return(multiple_selected_list)
    else:
        return(False)
#UL004
def find_multiple_group_by_remove_bttn(driver):
    multiple_group_by_remove_bttn = driver.find_element_by_id('btnRemove')
    if multiple_group_by_remove_bttn:
        return(multiple_group_by_remove_bttn)
    else:
        return(False)
#UL004
def find_multiple_group_by_save_bttn(driver):
    multiple_group_by_save_bttn = driver.find_element_by_css_selector(".sbttn[name='cmdSaveGroupby']")
    if multiple_group_by_save_bttn:
        return(multiple_group_by_save_bttn)
    else:
        return(False)

#UL004
#Vizient also added some multiple group by to the covid/pre-covid report templates 
#we need to remove these to get back to the default template.
def remove_multiple_group_by_from_template(browser_var):
    
    time.sleep(0.2)
    #find and click 'Multiple Group By.'  This will always be on the template.  I 
    #will do this to every template even if there are no selections.
    multiple_group_by_bttn = WebDriverWait(browser_var, 10).until(find_multiple_group_by_bttn)
    #click the link
    multiple_group_by_bttn.click()
    time.sleep(0.2)
    #loop over all the selected items in the multiple-group-by menu.  Select them and remove them.  Then save.
    
    multiple_group_by_selected_list = WebDriverWait(browser_var, 3).until(find_multiple_group_by_selected_list)
    
    time.sleep(0.2)
    mult_list = multiple_group_by_selected_list.find_elements_by_tag_name('option')
    
    if mult_list:
        if len(mult_list) > 0:
            time.sleep(0.2)
            #click all selected multiple group-by options
            for i, item in enumerate(while_loop_handler_function(multiple_group_by_selected_list.find_elements_by_tag_name('option'))):
                item.click()
                time.sleep(0.2)   
    
    #click all selected multiple group-by options
    #for i, item in enumerate(while_loop_handler_function(multiple_group_by_selected_list.find_elements_by_tag_name('option'))):
    #    item.click()
    #time.sleep(0.2)   
    #click the remove button to remove all the above selected items.
    
    multiple_group_by_remove_bttn = WebDriverWait(browser_var, 3).until(find_multiple_group_by_remove_bttn)
    
    multiple_group_by_remove_bttn.click()
    time.sleep(0.2)
    #click the save button
    
    multiple_group_by_save_bttn = WebDriverWait(browser_var, 3).until(find_multiple_group_by_save_bttn)
    
    multiple_group_by_save_bttn.click()
    
    return(browser_var)



#UL004
def update_group_by_select_to_default(browser_var,measure_name_var):
    time.sleep(0.2)

    multgroupbydiv = WebDriverWait(browser_var, 3).until(find_mult_group_by_div)
    multgroupbydiv.click()
    
    #if psi, then select AHRQ, else Hospital/Hospital System
    if measure_name_var in ['PSI-03 O/E','PSI-11 O/E','PSI-09 O/E','PSI-06 O/E','PSI-13 O/E']:
        Select(browser_var.find_element_by_xpath("//*[@id='ctl00_ContentPlaceHolder1_cmdMeasuresByWithMult']")).select_by_visible_text('AHRQ Safety')
        time.sleep(0.2)
        multgroupbydiv.click()
        time.sleep(0.2)
    else:
        Select(browser_var.find_element_by_xpath("//*[@id='ctl00_ContentPlaceHolder1_cmdMeasuresByWithMult']")).select_by_visible_text('Hospital / Hospital System')
        time.sleep(0.2)
        multgroupbydiv.click()
        time.sleep(0.2)
    return(browser_var)



#UL004
    
#Only PSI measures:  Vizient added a new advanced restriction to remove covid-19 sub service lines (gen med & pulm)
#To get back to default setting, first remove this restriction.
def find_first_delete_bttn_adv_restriction(driver):
    first_del_bttn = driver.find_element_by_id('1_imgDelete')
    if first_del_bttn:
        return(first_del_bttn)
    else:
        return False

#UL004
#This function checks whether the measure is a PSI.
#If the measure is a PSI, then it will remove the first advanced restriction
#on the report template.  This restriction is always the Covid-19 sub service line filter.
#if the measure is not a psi, then this will do nothing.
def remove_psi_first_covid_adv_restriction(browser_var,measure_name_var):
    
    time.sleep(0.3)
    
    if measure_name_var in ['PSI-03 O/E','PSI-11 O/E','PSI-09 O/E','PSI-06 O/E','PSI-13 O/E']:
        first_psi_covid_adv_restriction = WebDriverWait(browser_var, 3).until(find_first_delete_bttn_adv_restriction)
        first_psi_covid_adv_restriction.click()
    else:
        pass
    
    return(browser_var)

#UL004
#This function checks whether the measure is DCOST, MORT or LOS Gen Med.
#If it is, we need to remove the 'General Medicine-COVID-19' sub service line restriction.
#For some reason, this only exists on teh mort, dcost, and los templates
#and not readmission or edac.

#find and select the edit image that is associated with subservice line
def find_edit_subservice_line_img(driver):
    #find and click the edit button assigned to SubServiceLine
    edit_subservice_line_img = driver.find_element_by_xpath("//img[@title='Edit' and contains(@onclick, '12|UHCSubServiceLine')]")
    if edit_subservice_line_img:
        return(edit_subservice_line_img)
    else:
        return False

#UL004
#after clicking the edit button, a popup menu will appear to select
#subservice line advanced restrictions.
def find_adv_rest_master_list(driver):
    #Select the master list from the pop up Advanced Restrictions menu
    adv_rest_master_list = driver.find_element_by_id("lstMasterList")
    if adv_rest_master_list:
        return(adv_rest_master_list)
    else:
        return False

#after modifying the advanced restrictions, then click the 'ok' button.
#find and select the 'ok' button
def find_adv_rest_save_bttn(driver):
    #find and click the 'ok' button
    adv_rest_ok_bttn = driver.find_element_by_id("ctl00_ContentPlaceHolder1_ucctl11_btnSave")
    if adv_rest_ok_bttn:
        return(adv_rest_ok_bttn)
    else:
        return False

#UL004
#this function is used to remove the stupid covid 19 subservice line restriction on
#only the general medicine templates (but not readmission or edac for some reason...Just mort,los,dcost).
#It uses the 3 functions defined above.

#Critical access uses a different filter and does not have gen med or pulmonary subservice line breakdown
#so we are safe to exclude it from the logic of this function.  We will have to handle critical access in a different way.

def remove_covid_19_gen_med_subservice_rest(browser_var,measure_name_var):
    
    #only do this if the measure is Gen Med related.
    
    if measure_name_var in ['DCOST O/E - Medicine General','LOS O/E - Medicine General','Mortality O/E - Medicine General']:
        
        #step one.  click the edit image that is linked to subservice line restrictions.
        subservice_line_edit = WebDriverWait(browser_var, 3).until(find_edit_subservice_line_img)
        subservice_line_edit.click()
        time.sleep(0.3)
        #step two.  select the master list of all subservice lines and scroll through it.  Unclick covid-19 gen med
        # and click the gen med - gastro.

        master_list = WebDriverWait(browser_var, 3).until(find_adv_rest_master_list)
        
        option_list1 = while_loop_handler_function(master_list.find_elements_by_tag_name('option'))
        
        pre_selected_options = [x.text for ind, x in enumerate(option_list1) if (ind==0) or (x.get_attribute("selected") == 'true')]
        
        #First go through and click the options that are in your period dictionary but not pre-selected.
        #We do this because when you click on the menu, it automatically clicks 
        #on the top option and does some funky stuff.  So, you need to first go through
        #and click on everything not pre-selected and not the first option. 
        #Then go back and click again or "unclick" the options that were pre-selected but not what you wanted.
        
        #In this case, we only want 'General Medicine- Gastroenterology'.  So, we are going to check it against the pre-selected
        #items and make sure it is selected.
        
        for optiona in option_list1:
            # Is the option from the drop-down at this point of iteration in our user-defined set of quarters? ALSO, is it NOT one of the following values?
            if optiona.text in ['General Medicine- Gastroenterology'] and optiona.text not in pre_selected_options:

                attempt_option_click = 1
                while (attempt_option_click != 0):
                    try:
                        optiona.click()  # select() in earlier versions of webdriver
                        attempt_option_click = 0
                    except StaleElementReferenceException:
                        if attempt_option_click == 3:
                            raise
                        attempt_option_click += 1
                        time.sleep(0.5)
                        
        #now find all the menu options selected after the first round.
        selected_after_first_round = [x.text for ind, x in enumerate(option_list1) if (x.get_attribute("selected") == 'true')]
        
        #Unselected/Unclick by clicking again the items that are not in your period dictionary but are still
        #selected due to the preselection defaults.
        
        for optionb in option_list1:
            if optionb.text in selected_after_first_round and optionb.text not in ['General Medicine- Gastroenterology']:
                attempt_option_click = 1
                while (attempt_option_click != 0):
                    try:
                        #print('clicking on option.')
                        optionb.click()  # select() in earlier versions of webdriver
                        attempt_option_click = 0
                        #print('clicked on option.',optionb.text)
                    except StaleElementReferenceException:
                        if attempt_option_click == 3:
                            raise
                        attempt_option_click += 1
                        time.sleep(0.5)
        #click OK.
        adv_rest_ok_bttn = WebDriverWait(browser_var, 3).until(find_adv_rest_save_bttn)
        adv_rest_ok_bttn.click()
        time.sleep(0.3)
    return(browser_var)


#UL004
    
def find_delete_subservice_line_img(driver):
    #find and click the edit button assigned to SubServiceLine
    delete_subservice_line_img = driver.find_element_by_xpath("//img[@title='Delete' and contains(@onclick, '12|UHCSubServiceLine')]")
    if delete_subservice_line_img:
        return(delete_subservice_line_img)
    else:
        return False
    
#this function is used to remove the stupid covid 19 subservice line restriction on
#only the pulmonary/critical care templates (but not readmission or edac for some reason...Just mort,los,dcost).

#Critical access uses a different filter and does not have gen med or pulmonary subservice line breakdown
#so we are safe to exclude it from the logic of this function.  We will have to handle critical access in a different way.

def remove_covid_19_pulmonary_subservice_rest(browser_var,measure_name_var):
    
    #only do this if the measure is Pulmonary/Critical Care related.
    
    if measure_name_var in ['DCOST O/E - Pulmonary/Critical Care','LOS O/E - Pulmonary/Critical Care','Mortality O/E - Pulmonary/Critical Care']:
        
        #step one.  click the delete image that is linked to subservice line restrictions.
        subservice_rest_delete = WebDriverWait(browser_var, 3).until(find_delete_subservice_line_img)
        subservice_rest_delete.click()
        time.sleep(0.3)
    return(browser_var)


#UL004
    
#The next two functions are used to remove the discharge month restrictions on all the 
#EDAC and READM report templates.  Vizient added all these discharge month restrictions
#to force a certain time period on the report.  Obviously, if you want to pull numbers
#outside of that time period, the results will be blank.  So...we have to remove them from
#every edac and readm report template.  For some reason, this was only done to edac and readm.

#find and select all 'Delete' images that contain the javascript "DischargeMonth'
def find_delete_discharge_month(driver):
    #find and click the delete image for discharge month
    delete_discharge_month = driver.find_elements_by_xpath("//img[@title='Delete' and contains(@onclick, 'DischargeMonth')]")
    if delete_discharge_month:
        return(delete_discharge_month)
    else:
        return([])

#UL004
#if the measure name is edac or readm, then check whether there are any 
#restrictions for 'DischargeMonth.'  If you get any results, click the first element.
#Keep doing this until there aren't anymore on the DOM.
def remove_covid_19_readm_edac_discharge_month(browser_var,measure_name_var):
    
    #only do this if the measure is readm or edac
    if measure_name_var in list_of_total_revisits_measures:
        time.sleep(0.3)
        #find any advanced restrictions associated with discharge month
        try:
            discharge_month = find_delete_discharge_month(browser_var)
        except:
            discharge_month = []
            return(browser_var)
        #Since the html ids and css change when things are added/removed to the DOM, 
        #we have to use a while loop and continuously check/delete while there are still
        #discharge month advanced restrictions.
        while len(discharge_month) > 0:
            try:
                discharge_month = find_delete_discharge_month(browser_var)
                discharge_month[0].click()
                time.sleep(0.3)
            except:
                discharge_month = []
        time.sleep(0.3)
        return(browser_var)
    else:
        time.sleep(0.3)
        return(browser_var)


#UL004
#All the revisits measures also have a discharge month restriction for pre-covid. 
#we need to remove that restriction.
def remove_covid_19_revisits_discharge_month(browser_var,measure_name_var):
    
    #only do this if the measure is readm or edac
    if measure_name_var in ['OP Procedure Revisits - Urological',\
                           'OP Procedure Revisits - Colonoscopy',\
                            'OP Procedure Revisits - Biliary',\
                            'OP Procedure Revisits - Arthroscopy',\
                            'Urinary Procedures Revisits within 7-days',\
                            'Colonscopy Revisits within 7-days',\
                           'Arthroscopy Revisits within 7-days']:
        time.sleep(0.3)
        #find any advanced restrictions associated with discharge month
        try:
            discharge_month = find_delete_discharge_month(browser_var)
        except:
            discharge_month = []
            return(browser_var)
        #Since the html ids and css change when things are added/removed to the DOM, 
        #we have to use a while loop and continuously check/delete while there are still
        #discharge month advanced restrictions.
        while len(discharge_month) > 0:
            try:
                discharge_month = find_delete_discharge_month(browser_var)
                discharge_month[0].click()
                time.sleep(0.3)
            except:
                discharge_month = []
        time.sleep(0.3)
        return(browser_var)
    else:
        time.sleep(0.3)
        return(browser_var)


#UL004
#Critical Access templates use a combination of discharge month restriction
#or subservice line or both.
        
def remove_covid_19_crit_access_discharge_month(browser_var,cohort_nm):
    
    #only do this if the measure is readm or edac
    if cohort_nm in ['Critical Access & Small Community']:
        time.sleep(0.3)
        #find any advanced restrictions associated with discharge month
        try:
            discharge_month = find_delete_discharge_month(browser_var)
        except:
            discharge_month = []
            return(browser_var)
        #Since the html ids and css change when things are added/removed to the DOM, 
        #we have to use a while loop and continuously check/delete while there are still
        #discharge month advanced restrictions.
        while len(discharge_month) > 0:
            try:
                discharge_month = find_delete_discharge_month(browser_var)
                discharge_month[0].click()
                time.sleep(0.3)
            except:
                discharge_month = []
        time.sleep(0.3)
        return(browser_var)
    else:
        time.sleep(0.3)
        return(browser_var)


#UL004
#Same thing as above for critical access to handle the subservice line restrictions

#find and select any delete image for subservice line
def find_delete_subservice_line_img2(driver):
    #find and click the edit button assigned to SubServiceLine
    delete_subservice_line_img = driver.find_elements_by_xpath("//img[@title='Delete' and contains(@onclick, 'UHCSubServiceLine')]")
    if delete_subservice_line_img:
        return(delete_subservice_line_img)
    else:
        return([])    
    
    
    
def remove_covid_19_crit_access_subservice_line(browser_var,cohort_nm):
    
    #only do this if cohort is critical access
    if cohort_nm in ['Critical Access & Small Community']:
        time.sleep(0.3)
        #find any advanced restrictions associated with discharge month
        try:
            subservice_line = find_delete_subservice_line_img2(browser_var)
        except:
            subservice_line = []
            return(browser_var)
        #Since the html ids and css change when things are added/removed to the DOM, 
        #we have to use a while loop and continuously check/delete while there are still
        #subservice line advanced restrictions.
        while len(subservice_line) > 0:
            try:
                subservice_line = find_delete_subservice_line_img2(browser_var)
                subservice_line[0].click()
                time.sleep(0.3)
            except:
                subservice_line = []
        time.sleep(0.3)
        return(browser_var)
    else:
        time.sleep(0.3)
        return(browser_var)


################################################################################################
################################################################################################
#UL006   Add new functions to remove Covid patients from each report
        
    
def find_adv_restrictions_add(driver):
    ar1 = driver.find_elements_by_xpath('//*[@id="advMenu"]/span')[0]
    if ar1:
        return ar1
    else:
        return False
    
def find_adv_restrictions_menu_id(driver):
    armi = driver.find_element_by_id("ctl00_ContentPlaceHolder1_RadMenu1")
    if armi:
        return armi
    else:
        return False
    
def find_rad_span_list(driver,ar_menu):
    r_span_list = ar_menu.find_elements_by_xpath('.//span[@class = "rmText rmExpandRight"]')
    if r_span_list:
        return r_span_list
    else:
        return False
    
def find_diag_proc_option(driver):
    dpo = driver.find_element_by_xpath('//*[@id="ctl00_ContentPlaceHolder1_RadMenu1"]/ul/li/div/ul/li[3]/a/span')
    if dpo:
        return dpo
    else:
        return False
    
def find_any_diag_option(driver):
    ado = driver.find_elements_by_xpath('//*[@id="ctl00_ContentPlaceHolder1_RadMenu1"]/ul/li/div/ul/li[3]/div/ul/li[1]/a/span')[0]
    if ado:
        return ado
    else:
        return False
    
def find_exclude_only(driver):
    ex_o = driver.find_element_by_xpath("//select[@id='ddlOperator']/option[text()='Exclude Only']")
    if ex_o:
        return ex_o
    else:
        return False
    
    
def find_search_bar(driver):
    sbar = driver.find_element_by_id("SearchArea")
    if sbar:
        return sbar
    else:
        return False
    
    
def find_search_bttn_adv_rest(driver):
    s_bttn = driver.find_element_by_xpath('//*[@id="SearchButton"]')
    if s_bttn:
        return s_bttn
    else:
        return False
    
def find_all_box_bttn(driver):
    ab_bttn = driver.find_element_by_xpath('//*[@id="gvMaster_ctl02_chkSelect"]')
    if ab_bttn:
        return ab_bttn
    else:
        return False
    
def find_add_selections_bttn(driver):
    a_select_bttn = driver.find_element_by_xpath('//*[@id="Button1"]')
    if a_select_bttn:
        return a_select_bttn
    else:
        return False
    
def find_send_to_cdb_bttn(driver):
    cdb_bttn = driver.find_element_by_xpath('//*[@id="Button3"]')
    if cdb_bttn:
        return cdb_bttn
    else:
        return False
    
    
#UL006
        
    
def find_adv_rest_index_apply(driver):
    any_diag_index_apply = driver.find_elements_by_xpath("//*[contains(@onchange, 'AnyDiagnosis')]")
    if any_diag_index_apply:
        return any_diag_index_apply
    else:
        return False
    
    
    
    
'''
def add_covid_advanced_restriction(driver_var):
    #click the advanced restrictions menu
    time.sleep(0.5)
    adv_rest = WebDriverWait(driver_var, 10).until(find_adv_restrictions_add)
    adv_rest.click()
    time.sleep(3)
    
    
    # click the 'Diagnosis/Procedure option'
    #adv_rest_menu_id = WebDriverWait(driver_var, 10).until(find_adv_restrictions_menu_id)
    #print(adv_rest_menu_id)
    time.sleep(0.5)
    diag_proc_option = WebDriverWait(driver_var, 10).until(find_diag_proc_option)
    time.sleep(1)
    diag_proc_option.click()
    
    
    time.sleep(0.5)
    any_diag_option = WebDriverWait(driver_var, 10).until(find_any_diag_option) 
    time.sleep(1)
    any_diag_option.click()
    time.sleep(1)
    #a popup window will appear.  Switch control to the new window.
    
    browser_driver1 = driver_var.window_handles[0]
    browser_driver2 = driver_var.window_handles[1]
    
    driver_var.switch_to.window(browser_driver2)
    
    #click the 'Exclude Only' dropdown menu option
    exclude_only_option = WebDriverWait(driver_var, 10).until(find_exclude_only)
    time.sleep(0.5)
    exclude_only_option.click()
    
    #Type the U071 ICD-10 code into the search bar to find the Covid ICD-10 code
    search_bar = WebDriverWait(driver_var, 10).until(find_search_bar)
    time.sleep(0.5)
    #Type Covid ICD10 Code
    search_bar.send_keys("U071")
    time.sleep(1)
    #click 'Search'
    search_bttn = WebDriverWait(driver_var, 10).until(find_search_bttn_adv_rest)
    time.sleep(0.5)
    search_bttn.click()
    time.sleep(0.5)
    #select the U071 option
    all_box_bttn = WebDriverWait(driver_var, 10).until(find_all_box_bttn)
    all_box_bttn.click()
    time.sleep(1)
    #click 'Add Selections'
    add_selections_bttn = WebDriverWait(driver_var, 10).until(find_add_selections_bttn)
    time.sleep(0.5)
    add_selections_bttn.click()
    time.sleep(0.5)
    #click 'Send to CDB'
    send_to_cdb_bttn = WebDriverWait(driver_var, 10).until(find_send_to_cdb_bttn)
    time.sleep(0.5)
    send_to_cdb_bttn.click()
    
    #return driver control back to first window
    driver_var.switch_to.window(browser_driver1)
    
    return(driver_var)
'''

def add_covid_advanced_restriction(driver_var,meas_type = 'OTHER'):
    
    if meas_type == 'THK':
        
        return(driver_var)
    else:
        print('other...')
        #click the advanced restrictions menu
        time.sleep(0.5)
        adv_rest = WebDriverWait(driver_var, 10).until(find_adv_restrictions_add)
        adv_rest.click()
        time.sleep(3)


        # click the 'Diagnosis/Procedure option'
        #adv_rest_menu_id = WebDriverWait(driver_var, 10).until(find_adv_restrictions_menu_id)
        #print(adv_rest_menu_id)
        time.sleep(0.5)
        diag_proc_option = WebDriverWait(driver_var, 10).until(find_diag_proc_option)
        time.sleep(1)
        diag_proc_option.click()

        '''
        #time.sleep(0.5)
        rad_span_list = adv_rest_menu_id.find_elements_by_xpath('.//span[@class = "rmText rmExpandRight"]')
        print(rad_span_list)
        #rad_span_list = find_rad_span_list(driver_var,adv_rest_menu_id)
        #print(rad_span_list)
        #loop over spans under the advanced restriction menu and only click the one that has text 'Diagnosis/Procedure'
        for i in rad_span_list:
            if i.text == 'Diagnosis / Procedure':
                i.click()
        #time.sleep(0.5)
        #click the 'Any Diagnosis' option
        '''
        time.sleep(0.5)
        any_diag_option = WebDriverWait(driver_var, 10).until(find_any_diag_option) 
        time.sleep(1)
        any_diag_option.click()
        time.sleep(1)
        #a popup window will appear.  Switch control to the new window.

        browser_driver1 = driver_var.window_handles[0]
        browser_driver2 = driver_var.window_handles[1]

        driver_var.switch_to.window(browser_driver2)

        #click the 'Exclude Only' dropdown menu option
        exclude_only_option = WebDriverWait(driver_var, 10).until(find_exclude_only)
        time.sleep(0.5)
        exclude_only_option.click()

        #Type the U071 ICD-10 code into the search bar to find the Covid ICD-10 code
        search_bar = WebDriverWait(driver_var, 10).until(find_search_bar)
        time.sleep(0.5)
        #Type Covid ICD10 Code
        search_bar.send_keys("U071")
        time.sleep(1)
        #click 'Search'
        search_bttn = WebDriverWait(driver_var, 10).until(find_search_bttn_adv_rest)
        time.sleep(0.5)
        search_bttn.click()
        time.sleep(0.5)
        #select the U071 option
        all_box_bttn = WebDriverWait(driver_var, 10).until(find_all_box_bttn)
        all_box_bttn.click()
        time.sleep(1)
        #click 'Add Selections'
        add_selections_bttn = WebDriverWait(driver_var, 10).until(find_add_selections_bttn)
        time.sleep(0.5)
        add_selections_bttn.click()
        time.sleep(0.5)
        #click 'Send to CDB'
        send_to_cdb_bttn = WebDriverWait(driver_var, 10).until(find_send_to_cdb_bttn)
        time.sleep(0.5)
        send_to_cdb_bttn.click()

        #return driver control back to first window
        driver_var.switch_to.window(browser_driver1)
        
        if meas_type == 'READM':
            time.sleep(0.5)
            
            adv_rest_apply_index_menu = WebDriverWait(driver_var, 10).until(find_adv_rest_index_apply)
            
            
            for i, item in enumerate(adv_rest_apply_index_menu[0].find_elements_by_tag_name('option')):
                
                if item.text == 'Both':
                    item.click()
            return(driver_var)
        else:
            return(driver_var) 
################################################################################################
################################################################################################






################################################################################################
# Main function to loop over report templates and perform downloads.

def loop_template_download(link_dict, period_dict1,period_dict2,period_dict3, driver_var, exclusion_list,file_dir,hyperlink_loc,remove_covid_pats = False):
    report_counter = 0

    #assign correct path to Downloads folder.  This will change depending on which machine you run the script on.
    try:
        len(os.listdir('C:/Users/NM184423/Downloads'))
        download_folder_dir = 'C:/Users/NM184423/Downloads'
    except:
        len(os.listdir(r'H:\Downloads'))
        download_folder_dir = r'H:\Downloads'

    for i in link_dict.keys():
        if link_dict[i][4] == 'THK Complication':
            for p2 in period_dict2.keys():
                #print('got to THK')
                # get number of files in Downloads folder
                num_already_downloaded_files = len(os.listdir(download_folder_dir))
                print(download_folder_dir)
                print('num downloaded:',num_already_downloaded_files)
                # open the template
                time.sleep(1)
                driver_var = open_template_report(link_dict[i][0], driver_var)
                driver_var.implicitly_wait(30)
                time.sleep(1)
                try:
                    driver_var.maximize_window()
                except:
                    pass

                # scroll down
                div_element1 = WebDriverWait(driver_var, 120).until(find_div1_scroll)
                # div_element1 = driver_var.find_element_by_xpath("//div[@id='divRiskAdjustment']")
                driver_var.execute_script(
                    "return arguments[0].scrollIntoView(0, document.documentElement.scrollHeight-10);", div_element1)
                print('right before choose adjustment.')
                
                #BEGIN FY21 COVID FILTER FUNCTIONS#
                ###################################
                
                #UL004
                '''
                driver_var = remove_custom_covid_list_from_template(driver_var,link_dict[i][4])
                
                time.sleep(0.5)
                #UL004
                #driver_var = remove_multiple_group_by_from_template(driver_var)
                #time.sleep(0.5)
                #UL004
                driver_var = update_group_by_select_to_default(driver_var,link_dict[i][4])
                time.sleep(0.5)
                #UL004
                driver_var = remove_psi_first_covid_adv_restriction(driver_var,link_dict[i][4])
                time.sleep(0.5)
                #UL004
                driver_var = remove_covid_19_gen_med_subservice_rest(driver_var,link_dict[i][4])
                
                #UL004
                driver_var = remove_covid_19_pulmonary_subservice_rest(driver_var,link_dict[i][4])
                
                #UL004
                driver_var = remove_covid_19_readm_edac_discharge_month(driver_var,link_dict[i][4])
                
                #UL004
                driver_var = remove_covid_19_revisits_discharge_month(driver_var,link_dict[i][4])
                
                #UL004
                driver_var = remove_covid_19_crit_access_discharge_month(driver_var,link_dict[i][3])
                
                #UL004
                driver_var = remove_covid_19_crit_access_subservice_line(driver_var,link_dict[i][3])
                '''
                
                
                
                
                #END FY21 COVID FILTER FUNCTIONS#
                ##################################
                
                # Click Risk Adjustment Model button
                driver_var = choose_adjustment_model(driver_var, link_dict, i)
                time.sleep(1)
                # Click AHRQ Version button
                driver_var = choose_ahrq_version(driver_var, link_dict, i)

                # Update Multiple Group By Drop down
                update_group_by_select(driver_var, period_dict2, p2)

                # scroll down
                div_element2 = WebDriverWait(driver_var, 60).until(find_div2_scroll)
                # div_element2 = driver_var.find_element_by_xpath("//div[@id='ctl00_ContentPlaceHolder1_PanelContent3']")
                driver_var.execute_script(
                    "return arguments[0].scrollIntoView(0, document.documentElement.scrollHeight-10);", div_element2)

                # click the From/To Time Period Radio Button
                try:
                    #time_period_radio_btn = WebDriverWait(driver_var, 45).until(find_time_period_radio_btn1)
                    # time_period_radio_btn = driver_var.find_element_by_id("ctl00_ContentPlaceHolder1_fromYear")
                    # time_period_radio_btn.click()
                    #ActionChains(driver_var).move_to_element(time_period_radio_btn).click().perform()
                    time_period_btn_attempt = 1
                    while True:
                        try:
                            time_period_radio_btn = WebDriverWait(driver_var, 45).until(find_time_period_radio_btn1)
                            break
                        except StaleElementReferenceException:
                            if time_period_btn_attempt == 3:
                                raise
                            time_period_btn_attempt += 1

                    time_period_btn_click_attempt = 1
                    while True:
                        try:
                            ActionChains(driver_var).move_to_element(time_period_radio_btn).click().perform()
                            break
                        except StaleElementReferenceException:
                            if time_period_btn_click_attempt == 3:
                                raise
                            time_period_btn_click_attempt += 1

                except:
                    #time_period_radio_btn = WebDriverWait(driver_var, 45).until(find_time_period_radio_btn2)
                    # time_period_radio_btn = driver_var.find_element_by_id("ctl00_ContentPlaceHolder1_cmdFromYear")
                    #ActionChains(driver_var).move_to_element(time_period_radio_btn).click().perform()
                    time_period_btn_attempt = 1
                    while True:
                        try:
                            time_period_radio_btn = WebDriverWait(driver_var, 45).until(find_time_period_radio_btn2)
                            break
                        except StaleElementReferenceException:
                            if time_period_btn_attempt == 3:
                                raise
                            time_period_btn_attempt += 1

                    time_period_btn_click_attempt = 1
                    while True:
                        try:
                            ActionChains(driver_var).move_to_element(time_period_radio_btn).click().perform()
                            break
                        except StaleElementReferenceException:
                            if time_period_btn_click_attempt == 3:
                                raise
                            time_period_btn_click_attempt += 1
                driver_var.implicitly_wait(10)
                # update time period dropdown menus
                update_time_period_select(driver_var, period_dict2, p2)

                #time.sleep(0.5)

                # scroll down
                div_element3 = WebDriverWait(driver_var, 30).until(find_div3_scroll)
                # div_element3 = driver_var.find_element_by_xpath("//div[@id='ctl00_ContentPlaceHolder1_PanelContent4']")
                driver_var.execute_script(
                    "return arguments[0].scrollIntoView(0, document.documentElement.scrollHeight-10);", div_element3)

                time.sleep(2)

                # Set focus hospital to NMH
                try:
                    element = WebDriverWait(driver_var, 120).until(
                        EC.presence_of_element_located((By.CSS_SELECTOR, "option[value='140281']"))
                    )
                    #print("Option loaded")
                except TimeoutException:
                    print("Time exceeded!")

                time.sleep(1.5)
                '''

                focus_hosp = WebDriverWait(driver_var, 150).until(find_focus_hosp)
                # Select(driver_var.find_element_by_xpath("//*[@id='ctl00_ContentPlaceHolder1_cmdFocusHCO']")).select_by_visible_text('NORTHWESTERN_MEMORIAL 140281')
                time.sleep(1)
                Select(focus_hosp).select_by_visible_text('NORTHWESTERN_MEMORIAL 140281')
                '''
                #focus_hosp = WebDriverWait(driver_var, 150).until(find_focus_hosp)
                #fh = driver_var.find_element_by_id('ctl00_ContentPlaceHolder1_cmdFocusHCO')
                #time.sleep(1.5)

                #focus_hosp_attempt = 1
                #while (focus_hosp_attempt != 0):
                #    try:
                #        focus_hosp = WebDriverWait(driver_var, 150).until(find_focus_hosp)
                #        focus_hosp_attempt = 0
                #    except StaleElementReferenceException:
                #        if focus_hosp_attempt == 4:
                #            raise
                #        focus_hosp_attempt += 1
                #        time.sleep(0.5)

                #UL009 begin
                
                
                '''
                focus_hosp = while_loop_handler_function(WebDriverWait(driver_var, 150).until(find_focus_hosp))
                time.sleep(0.5)
                try:
                    focus_hosp.click()
                except:
                    focus_hosp = while_loop_handler_function(WebDriverWait(driver_var, 150).until(find_focus_hosp))
                    time.sleep(0.5)
                    focus_hosp.click()
                time.sleep(0.5)
                '''
                
                driver_var, focus_hosp = click_focus_hosp2(driver_var)
                
                    
                
                try:
                    Select(focus_hosp).select_by_visible_text('NORTHWESTERN_MEMORIAL 140281')
                except:
                    
                    driver_var, focus_hosp = click_focus_hosp2(driver_var)
                    Select(focus_hosp).select_by_visible_text('NORTHWESTERN_MEMORIAL 140281')
                #UL009 end
                
                #time.sleep(1)
                #focus_hosp_options = while_loop_handler_function(focus_hosp.find_elements_by_tag_name('option'))
                #time.sleep(0.5)
                #print('first group hosp select')
                #for option in focus_hosp.find_elements_by_tag_name('option'):
                # for option1 in focus_hosp_options:
                #     if option1.text == 'NORTHWESTERN_MEMORIAL 140281':
                #
                #         #print(option1.get_attribute("value"))
                #         #option.click()
                #         #time.sleep(0.5)
                #         focus_hosp_click_attempt = 1
                #         while (focus_hosp_click_attempt != 0):
                #             try:
                #                 option1.click()
                #                 focus_hosp_click_attempt = 0
                #             except StaleElementReferenceException:
                #                 if focus_hosp_click_attempt == 4:
                #                     raise
                #                 focus_hosp_click_attempt += 1
                #                 time.sleep(0.5)

                # scroll down
                div_element4 = WebDriverWait(driver_var, 30).until(find_div4_scroll)
                # div_element4 = driver_var.find_element_by_xpath("//div[@id='ctl00_ContentPlaceHolder1_PanelContent5']")
                driver_var.execute_script(
                    "return arguments[0].scrollIntoView(0, document.documentElement.scrollHeight-10);", div_element4)

                time.sleep(0.5)

                # Click 'All available hospitals in the database
                #all_avail_hosp_btn = WebDriverWait(driver_var, 30).until(find_all_avail_hosp)
                # all_avail_hosp_btn = driver_var.find_element_by_id("ctl00_ContentPlaceHolder1_cmdAllAvailHosp")
                #all_avail_hosp_btn.click()

                all_avail_hosp_attempt = 1
                while True:
                    try:
                        all_avail_hosp_btn = WebDriverWait(driver_var, 30).until(find_all_avail_hosp)
                        break
                    except StaleElementReferenceException:
                        if all_avail_hosp_attempt == 3:
                            raise
                        all_avail_hosp_attempt += 1

                all_avail_hosp_click_attempt = 1
                while True:
                    try:
                        all_avail_hosp_btn.click()
                        break
                    except StaleElementReferenceException:
                        if all_avail_hosp_click_attempt == 3:
                            raise
                        all_avail_hosp_click_attempt += 1

                # scroll down
                div_element5 = WebDriverWait(driver_var, 30).until(find_div5_scroll)  # type: object
                # div_element5 = driver_var.find_element_by_xpath("//div[@id='divRiskAdjustment']")
                driver_var.execute_script("return arguments[0].scrollIntoView(0, document.documentElement.scrollHeight-10);", div_element5)

                # ActionChains(driver_var).move_to_element(all_avail_hosp_btn).click().perform()
                time.sleep(2)

                # Click Download button
                # generate_excel_btn = driver_var.find_element_by_id("ctl00_ContentPlaceHolder1_imgExcel")
                """try:
                    element = WebDriverWait(driver_var, 120).until(
                        EC.presence_of_element_located((By.ID, "ctl00_ContentPlaceHolder1_imgExcel"))
                    )
                    print("Excel image clickable.")
                except TimeoutException:
                    print("Time exceeded!")
                """
                
                #UL006
                
                if remove_covid_pats == True:
                    driver_var = add_covid_advanced_restriction(driver_var,'THK')
                
                time.sleep(0.5)
                #generate_excel_btn = WebDriverWait(driver_var, 10).until(find_excel_btn)
                #generate_excel_btn.click()

                excel_btn_attempt = 1
                while True:
                    try:
                        generate_excel_btn = WebDriverWait(driver_var, 10).until(find_excel_btn)
                        break
                    except StaleElementReferenceException:
                        if excel_btn_attempt == 3:
                            raise
                        excel_btn_attempt += 1

                excel_btn_click_attempt = 1
                while True:
                    try:
                        generate_excel_btn.click()
                        break
                    except StaleElementReferenceException:
                        if excel_btn_click_attempt == 3:
                            raise
                        excel_btn_click_attempt += 1
                #webdriver.ActionChains(driver_var).move_to_element(generate_excel_btn).click(generate_excel_btn).perform()
                #ActionChains(driver_var).move_to_element(generate_excel_btn).click().perform()
                time.sleep(1)
                # Accept Pop-up window

                try:
                    WebDriverWait(driver_var, 10).until(EC.alert_is_present(),
                                                       'Timed out waiting for PA creation ' +
                                                       'confirmation popup to appear.')

                    obj = WebDriverWait(driver_var, 10).until(find_alert_popup)

                    obj.accept()
                except:
                    pass
                print('checking whether downloaded...')
                #print('before download check while loop')
                while len(os.listdir(download_folder_dir)) <= num_already_downloaded_files:
                    time.sleep(1)
                print('downloaded...')
                #print('after download check while loop')
                #time.sleep(1.5)
                #print('before window check while loop')
                # wait for the pop up browser window to display so you can close it.
                while len(driver_var.window_handles) < 2:
                    time.sleep(0.5)
                #print('after window check while loop')
                window_before = driver_var.window_handles[0]
                window_after = driver_var.window_handles[1]
                # switch control to popup window and close it
                driver_var.switch_to.window(window_after)
                driver_var.close()
                # Switch control back to the original window.
                driver_var.switch_to.window(window_before)
                time.sleep(2)
                # find most recently-added file to Downloads folder and rename it.
                try:
                    latest_file = find_last_downloaded_file('C:/Users/NM184423/Downloads')
                except:
                    latest_file = find_last_downloaded_file('H:/Downloads')
                latest_file = os.path.abspath(latest_file)

                time.sleep(0.500)

                rename_and_move_file(latest_file, link_dict[i][3], link_dict[i][7], p2, file_dir)
                time.sleep(0.500)
                report_counter += 1
                time.sleep(2)
                update_template_files(hyperlink_loc, link_dict[i][3], link_dict[i][4])
                time.sleep(2)

        elif link_dict[i][4] in list_of_total_revisits_measures:
            for p3 in period_dict3.keys():
                #print('got to effectiveness')
                #print(p3)
                # get number of files in Downloads folder

                num_already_downloaded_files = len(os.listdir(download_folder_dir))

                # open the template
                time.sleep(1)
                driver_var = open_template_report(link_dict[i][0], driver_var)
                time.sleep(1)
                driver_var.implicitly_wait(30)
                try:
                    driver_var.maximize_window()
                except:
                    pass

                # scroll down
                div_element1 = WebDriverWait(driver_var, 120).until(find_div1_scroll)
                # div_element1 = driver_var.find_element_by_xpath("//div[@id='divRiskAdjustment']")
                driver_var.execute_script("return arguments[0].scrollIntoView(0, document.documentElement.scrollHeight-10);",div_element1)
                
                #BEGIN FY21 COVID FILTER FUNCTIONS#
                ###################################
                
                #UL004
                '''
                driver_var = remove_custom_covid_list_from_template(driver_var,link_dict[i][4])
                time.sleep(0.5)
                #UL004
                #driver_var = remove_multiple_group_by_from_template(driver_var)
                #time.sleep(0.5)
                #UL004
                driver_var = update_group_by_select_to_default(driver_var,link_dict[i][4])
                time.sleep(0.5)
                #UL004
                driver_var = remove_psi_first_covid_adv_restriction(driver_var,link_dict[i][4])
                time.sleep(0.5)
                #UL004
                driver_var = remove_covid_19_gen_med_subservice_rest(driver_var,link_dict[i][4])
                time.sleep(0.5)
                #UL004
                driver_var = remove_covid_19_pulmonary_subservice_rest(driver_var,link_dict[i][4])
                time.sleep(0.5)
                #UL004
                driver_var = remove_covid_19_readm_edac_discharge_month(driver_var,link_dict[i][4])
                time.sleep(0.5)
                #UL004
                driver_var = remove_covid_19_revisits_discharge_month(driver_var,link_dict[i][4])
                time.sleep(0.5)
                #UL004
                driver_var = remove_covid_19_crit_access_discharge_month(driver_var,link_dict[i][3])
                time.sleep(0.5)
                #UL004
                driver_var = remove_covid_19_crit_access_subservice_line(driver_var,link_dict[i][3])
                '''
                
                
                #END FY21 COVID FILTER FUNCTIONS#
                ##################################
                
                # Click Risk Adjustment Model button
                driver_var = choose_adjustment_model(driver_var, link_dict, i)

                # Click AHRQ Version button
                driver_var = choose_ahrq_version(driver_var, link_dict, i)

                # Update Multiple Group By Drop down
                update_group_by_select(driver_var, period_dict3, p3)

                # scroll down
                div_element2 = WebDriverWait(driver_var, 60).until(find_div2_scroll)
                # div_element2 = driver_var.find_element_by_xpath("//div[@id='ctl00_ContentPlaceHolder1_PanelContent3']")
                driver_var.execute_script("return arguments[0].scrollIntoView(0, document.documentElement.scrollHeight-10);",div_element2)

                # click the From/To Time Period Radio Button
                try:
                    #quarters_period_radio_btn = WebDriverWait(driver_var, 45).until(find_quarters_radio_btn)
                    # time_period_radio_btn = driver_var.find_element_by_id("ctl00_ContentPlaceHolder1_fromYear")
                    # time_period_radio_btn.click()
                    #ActionChains(driver_var).move_to_element(quarters_period_radio_btn).click().perform()

                    quarters_period_radio_btn_attempt = 1
                    while True:
                        try:
                            quarters_period_radio_btn = WebDriverWait(driver_var, 45).until(find_quarters_radio_btn)
                            break
                        except StaleElementReferenceException:
                            if quarters_period_radio_btn_attempt == 3:
                                raise
                            quarters_period_radio_btn_attempt += 1

                    quarters_period_radio_btn_click_attempt = 1
                    while True:
                        try:
                            ActionChains(driver_var).move_to_element(quarters_period_radio_btn).click().perform()
                            break
                        except StaleElementReferenceException:
                            if quarters_period_radio_btn_click_attempt == 3:
                                raise
                            quarters_period_radio_btn_click_attempt += 1

                except:
                    print('Cannot click on Quarters radio button.')
                driver_var.implicitly_wait(10)
                # update time period dropdown menus
                #print('HERE I AM!!!!!!!')
                update_time_period_select(driver_var, period_dict3, p3)

                #time.sleep(0.5)

                # scroll down
                div_element3 = WebDriverWait(driver_var, 30).until(find_div3_scroll)
                # div_element3 = driver_var.find_element_by_xpath("//div[@id='ctl00_ContentPlaceHolder1_PanelContent4']")
                driver_var.execute_script("return arguments[0].scrollIntoView(0, document.documentElement.scrollHeight-10);",div_element3)

                #time.sleep(2)

                # Set focus hospital to NMH
                try:
                    element = WebDriverWait(driver_var, 120).until(
                        EC.presence_of_element_located((By.CSS_SELECTOR, "option[value='140281']"))
                    )
                    #print("Option loaded")
                except TimeoutException:
                    print("Time exceeded!")

                time.sleep(1.5)
                '''
                                focus_hosp = WebDriverWait(driver_var, 150).until(find_focus_hosp)
                                # Select(driver_var.find_element_by_xpath("//*[@id='ctl00_ContentPlaceHolder1_cmdFocusHCO']")).select_by_visible_text('NORTHWESTERN_MEMORIAL 140281')
                                time.sleep(1)
                                Select(focus_hosp).select_by_visible_text('NORTHWESTERN_MEMORIAL 140281')
                                '''
                #focus_hosp = WebDriverWait(driver_var, 150).until(find_focus_hosp)

                #focus_hosp_attempt = 1
                #while (focus_hosp_attempt != 0):
                #    try:
                #        focus_hosp = WebDriverWait(driver_var, 150).until(find_focus_hosp)
                #        focus_hosp_attempt = 0
                #    except StaleElementReferenceException:
                #        if focus_hosp_attempt == 4:
                #            raise
                #        focus_hosp_attempt += 1
                #        time.sleep(0.5)
                
                #UL009 begin
                '''
                focus_hosp = while_loop_handler_function(WebDriverWait(driver_var, 150).until(find_focus_hosp))
                time.sleep(0.5)
                #focus_hosp.click()
                try:
                    focus_hosp.click()
                except:
                    focus_hosp = while_loop_handler_function(WebDriverWait(driver_var, 150).until(find_focus_hosp))
                    time.sleep(0.5)
                    focus_hosp.click()
                time.sleep(0.5)
                Select(focus_hosp).select_by_visible_text('NORTHWESTERN_MEMORIAL 140281')
                '''
                
                driver_var, focus_hosp = click_focus_hosp2(driver_var)
                
                    
                
                try:
                    Select(focus_hosp).select_by_visible_text('NORTHWESTERN_MEMORIAL 140281')
                except:
                    
                    driver_var, focus_hosp = click_focus_hosp2(driver_var)
                    Select(focus_hosp).select_by_visible_text('NORTHWESTERN_MEMORIAL 140281')
                
                #UL009 end
                
                #time.sleep(1)
                #focus_hosp_options = while_loop_handler_function(focus_hosp.find_elements_by_tag_name('option'))
                #time.sleep(0.5)
                #print('second group hosp select')
                #for option in focus_hosp.find_elements_by_tag_name('option'):
                # for option2 in focus_hosp_options:
                #     if option2.text == 'NORTHWESTERN_MEMORIAL 140281':
                #         #print(option2.get_attribute("value"))
                #         # option.click()
                #         # time.sleep(0.5)
                #         focus_hosp_click_attempt = 1
                #         while (focus_hosp_click_attempt != 0):
                #             try:
                #                 option2.click()
                #                 focus_hosp_click_attempt = 0
                #             except StaleElementReferenceException:
                #                 if focus_hosp_click_attempt == 4:
                #                     raise
                #                 focus_hosp_click_attempt += 1
                #                 time.sleep(0.5)

                # scroll down
                div_element4 = WebDriverWait(driver_var, 30).until(find_div4_scroll)
                # div_element4 = driver_var.find_element_by_xpath("//div[@id='ctl00_ContentPlaceHolder1_PanelContent5']")
                driver_var.execute_script(
                    "return arguments[0].scrollIntoView(0, document.documentElement.scrollHeight-10);", div_element4)

                time.sleep(0.5)

                # Click 'All available hospitals in the database
                #all_avail_hosp_btn = WebDriverWait(driver_var, 30).until(find_all_avail_hosp)
                # all_avail_hosp_btn = driver_var.find_element_by_id("ctl00_ContentPlaceHolder1_cmdAllAvailHosp")
                #all_avail_hosp_btn.click()

                all_avail_hosp_attempt = 1
                while True:
                    try:
                        all_avail_hosp_btn = WebDriverWait(driver_var, 30).until(find_all_avail_hosp)
                        break
                    except StaleElementReferenceException:
                        if all_avail_hosp_attempt == 3:
                            raise
                        all_avail_hosp_attempt += 1

                all_avail_hosp_click_attempt = 1
                while True:
                    try:
                        all_avail_hosp_btn.click()
                        break
                    except StaleElementReferenceException:
                        if all_avail_hosp_click_attempt == 3:
                            raise
                        all_avail_hosp_click_attempt += 1

                # scroll down
                div_element5 = WebDriverWait(driver_var, 30).until(find_div5_scroll)  # type: object
                # div_element5 = driver_var.find_element_by_xpath("//div[@id='divRiskAdjustment']")
                driver_var.execute_script(
                    "return arguments[0].scrollIntoView(0, document.documentElement.scrollHeight-10);", div_element5)

                # ActionChains(driver_var).move_to_element(all_avail_hosp_btn).click().perform()
                time.sleep(2)

                # Click Download button
                # generate_excel_btn = driver_var.find_element_by_id("ctl00_ContentPlaceHolder1_imgExcel")
                """try:
                    element = WebDriverWait(driver_var, 120).until(
                        EC.presence_of_element_located((By.ID, "ctl00_ContentPlaceHolder1_imgExcel"))
                    )
                    print("Excel image clickable.")
                except TimeoutException:
                    print("Time exceeded!")
                """
                
                #UL006
                
                if remove_covid_pats == True:
                    driver_var = add_covid_advanced_restriction(driver_var,'READM')
                    
                time.sleep(0.5)
                
                #generate_excel_btn = WebDriverWait(driver_var, 10).until(find_excel_btn)
                #generate_excel_btn.click()

                excel_btn_attempt = 1
                while True:
                    try:
                        generate_excel_btn = WebDriverWait(driver_var, 10).until(find_excel_btn)
                        break
                    except StaleElementReferenceException:
                        if excel_btn_attempt == 3:
                            raise
                        excel_btn_attempt += 1

                excel_btn_click_attempt = 1
                while True:
                    try:
                        generate_excel_btn.click()
                        break
                    except StaleElementReferenceException:
                        if excel_btn_click_attempt == 3:
                            raise
                        excel_btn_click_attempt += 1
                # webdriver.ActionChains(driver_var).move_to_element(generate_excel_btn).click(generate_excel_btn).perform()
                # ActionChains(driver_var).move_to_element(generate_excel_btn).click().perform()
                time.sleep(1)
                # Accept Pop-up window

                try:
                    WebDriverWait(driver_var, 10).until(EC.alert_is_present(),
                                                        'Timed out waiting for PA creation ' +
                                                        'confirmation popup to appear.')

                    obj = WebDriverWait(driver_var, 10).until(find_alert_popup)

                    obj.accept()
                except:
                    pass
                # print('before download check while loop')
                while len(os.listdir(download_folder_dir)) <= num_already_downloaded_files:
                    time.sleep(1)
                # print('after download check while loop')
                # time.sleep(1.5)
                # print('before window check while loop')
                # wait for the pop up browser window to display so you can close it.
                while len(driver_var.window_handles) < 2:
                    time.sleep(0.5)
                # print('after window check while loop')
                window_before = driver_var.window_handles[0]
                window_after = driver_var.window_handles[1]
                # switch control to popup window and close it
                driver_var.switch_to.window(window_after)
                driver_var.close()
                # Switch control back to the original window.
                driver_var.switch_to.window(window_before)
                time.sleep(2)
                # find most recently-added file to Downloads folder and rename it.
                try:
                    latest_file = find_last_downloaded_file('C:/Users/NM184423/Downloads')
                except:
                    latest_file = find_last_downloaded_file('H:/Downloads')
                latest_file = os.path.abspath(latest_file)

                time.sleep(0.500)

                rename_and_move_file(latest_file, link_dict[i][3], link_dict[i][7], p3, file_dir)
                time.sleep(0.500)
                report_counter += 1
                time.sleep(2)
                update_template_files(hyperlink_loc, link_dict[i][3], link_dict[i][4])
                time.sleep(2)
        # leave a filter in the code for now until Vizient resolves report radio button issue...
        #print(link_dict[i][4])
        elif link_dict[i][4] not in list_of_total_revisits_measures and link_dict[i][4] in ['PSI-03 O/E','PSI-11 O/E','PSI-09 O/E','PSI-06 O/E','PSI-13 O/E']:
            for p1 in period_dict1.keys():
                print('PSI!')
                print('Measure: ',link_dict[i][4])
                # get number of files in Downloads folder

                num_already_downloaded_files = len(os.listdir(download_folder_dir))

                # open the template
                time.sleep(1)
                driver_var = open_template_report(link_dict[i][0], driver_var)
                time.sleep(1)
                driver_var.implicitly_wait(30)
                try:
                    driver_var.maximize_window()
                except:
                    pass

                # scroll down
                div_element1 = WebDriverWait(driver_var, 120).until(find_div1_scroll)
                # div_element1 = driver_var.find_element_by_xpath("//div[@id='divRiskAdjustment']")
                driver_var.execute_script(
                    "return arguments[0].scrollIntoView(0, document.documentElement.scrollHeight-10);", div_element1)
                
                
                #BEGIN FY21 COVID FILTER FUNCTIONS#
                ###################################
                
                #UL004
                '''
                driver_var = remove_custom_covid_list_from_template(driver_var,link_dict[i][4])
                
                #UL004
                driver_var = remove_multiple_group_by_from_template(driver_var)
                
                #UL004
                driver_var = update_group_by_select_to_default(driver_var,link_dict[i][4])
                
                #UL004
                driver_var = remove_psi_first_covid_adv_restriction(driver_var,link_dict[i][4])
                
                #UL004
                driver_var = remove_covid_19_gen_med_subservice_rest(driver_var,link_dict[i][4])
                
                #UL004
                driver_var = remove_covid_19_pulmonary_subservice_rest(driver_var,link_dict[i][4])
                
                #UL004
                driver_var = remove_covid_19_readm_edac_discharge_month(driver_var,link_dict[i][4])
                
                #UL004
                driver_var = remove_covid_19_revisits_discharge_month(driver_var,link_dict[i][4])
                
                #UL004
                driver_var = remove_covid_19_crit_access_discharge_month(driver_var,link_dict[i][3])
                
                #UL004
                driver_var = remove_covid_19_crit_access_subservice_line(driver_var,link_dict[i][3])
                
                
                '''
                #END FY21 COVID FILTER FUNCTIONS#
                ##################################
                
                
                # Click Risk Adjustment Model button
                driver_var = choose_adjustment_model(driver_var, link_dict, i)

                # Click AHRQ Version button
                driver_var = choose_ahrq_version(driver_var, link_dict, i)

                # Update Multiple Group By Drop down
                update_group_by_select(driver_var, period_dict1, p1)

                # scroll down
                div_element2 = WebDriverWait(driver_var, 60).until(find_div2_scroll)
                # div_element2 = driver_var.find_element_by_xpath("//div[@id='ctl00_ContentPlaceHolder1_PanelContent3']")
                driver_var.execute_script(
                    "return arguments[0].scrollIntoView(0, document.documentElement.scrollHeight-10);", div_element2)

                # click the From/To Time Period Radio Button

                try:
                    #time_period_radio_btn = WebDriverWait(driver_var, 45).until(find_time_period_radio_btn1)
                    # time_period_radio_btn = driver_var.find_element_by_id("ctl00_ContentPlaceHolder1_fromYear")
                    # time_period_radio_btn.click()
                    #ActionChains(driver_var).move_to_element(time_period_radio_btn).click().perform()
                    time_period_btn_attempt = 1
                    while True:
                        try:
                            time_period_radio_btn = WebDriverWait(driver_var, 45).until(find_time_period_radio_btn1)
                            break
                        except StaleElementReferenceException:
                            if time_period_btn_attempt == 3:
                                raise
                            time_period_btn_attempt += 1

                    time_period_btn_click_attempt = 1
                    while True:
                        try:
                            ActionChains(driver_var).move_to_element(time_period_radio_btn).click().perform()
                            break
                        except StaleElementReferenceException:
                            if time_period_btn_click_attempt == 3:
                                raise
                            time_period_btn_click_attempt += 1

                except:
                    #time_period_radio_btn = WebDriverWait(driver_var, 45).until(find_time_period_radio_btn2)
                    # time_period_radio_btn = driver_var.find_element_by_id("ctl00_ContentPlaceHolder1_cmdFromYear")
                    #ActionChains(driver_var).move_to_element(time_period_radio_btn).click().perform()
                    time_period_btn_attempt = 1
                    while True:
                        try:
                            time_period_radio_btn = WebDriverWait(driver_var, 45).until(find_time_period_radio_btn2)
                            break
                        except StaleElementReferenceException:
                            if time_period_btn_attempt == 3:
                                raise
                            time_period_btn_attempt += 1

                    time_period_btn_click_attempt = 1
                    while True:
                        try:
                            ActionChains(driver_var).move_to_element(time_period_radio_btn).click().perform()
                            break
                        except StaleElementReferenceException:
                            if time_period_btn_click_attempt == 3:
                                raise
                            time_period_btn_click_attempt += 1
                driver_var.implicitly_wait(10)


                # update time period downdown menus
                update_time_period_select(driver_var, period_dict1, p1)

                #time.sleep(0.5)

                # scroll down
                div_element3 = WebDriverWait(driver_var, 30).until(find_div3_scroll)
                # div_element3 = driver_var.find_element_by_xpath("//div[@id='ctl00_ContentPlaceHolder1_PanelContent4']")
                driver_var.execute_script(
                    "return arguments[0].scrollIntoView(0, document.documentElement.scrollHeight-10);", div_element3)

                #time.sleep(2)

                # Set focus hospital to NMH

                try:
                    element = WebDriverWait(driver_var, 120).until(
                        EC.presence_of_element_located((By.CSS_SELECTOR, "option[value='140281']"))
                    )
                    #print("Option loaded")
                except TimeoutException:
                    print("Time exceeded!")

                time.sleep(1.5)
                '''
                                focus_hosp = WebDriverWait(driver_var, 150).until(find_focus_hosp)
                                # Select(driver_var.find_element_by_xpath("//*[@id='ctl00_ContentPlaceHolder1_cmdFocusHCO']")).select_by_visible_text('NORTHWESTERN_MEMORIAL 140281')
                                time.sleep(1)
                                Select(focus_hosp).select_by_visible_text('NORTHWESTERN_MEMORIAL 140281')
                                '''
                #focus_hosp = WebDriverWait(driver_var, 150).until(find_focus_hosp)
                #focus_hosp_attempt = 1
                #while (focus_hosp_attempt != 0):
                #    try:
                #        focus_hosp = WebDriverWait(driver_var, 150).until(find_focus_hosp)
                #        focus_hosp_attempt = 0
                #    except StaleElementReferenceException:
                #        if focus_hosp_attempt == 4:
                #            raise
                #        focus_hosp_attempt += 1
                #        time.sleep(0.5)
                
                #UL009 begin
                '''
                focus_hosp = while_loop_handler_function(WebDriverWait(driver_var, 150).until(find_focus_hosp))
                time.sleep(0.5)
                #focus_hosp.click()
                try:
                    focus_hosp.click()
                except:
                    focus_hosp = while_loop_handler_function(WebDriverWait(driver_var, 150).until(find_focus_hosp))
                    time.sleep(0.5)
                    focus_hosp.click()
                time.sleep(0.5)
                Select(focus_hosp).select_by_visible_text('NORTHWESTERN_MEMORIAL 140281')
                '''
                driver_var, focus_hosp = click_focus_hosp2(driver_var)
                
                    
                
                try:
                    Select(focus_hosp).select_by_visible_text('NORTHWESTERN_MEMORIAL 140281')
                except:
                    
                    driver_var, focus_hosp = click_focus_hosp2(driver_var)
                    Select(focus_hosp).select_by_visible_text('NORTHWESTERN_MEMORIAL 140281')
                
                #UL009 end
                #time.sleep(1)
                #focus_hosp_options = while_loop_handler_function(focus_hosp.find_elements_by_tag_name('option'))
                #time.sleep(0.5)
                #print('third group hosp select')
                #for option in focus_hosp.find_elements_by_tag_name('option'):
                # for option3 in focus_hosp_options:
                #     if option3.text == 'NORTHWESTERN_MEMORIAL 140281':
                #         #print(option.get_attribute("value"))
                #         # option.click()
                #         # time.sleep(0.5)
                #         focus_hosp_click_attempt = 1
                #         while (focus_hosp_click_attempt != 0):
                #             try:
                #                 option3.click()
                #                 focus_hosp_click_attempt = 0
                #             except StaleElementReferenceException:
                #                 if focus_hosp_click_attempt == 4:
                #                     raise
                #                 focus_hosp_click_attempt += 1
                #                 time.sleep(0.5)


                # scroll down
                div_element4 = WebDriverWait(driver_var, 30).until(find_div4_scroll)
                # div_element4 = driver_var.find_element_by_xpath("//div[@id='ctl00_ContentPlaceHolder1_PanelContent5']")
                driver_var.execute_script(
                    "return arguments[0].scrollIntoView(0, document.documentElement.scrollHeight-10);", div_element4)

                time.sleep(0.5)

                # Click 'All available hospitals in the database
                #all_avail_hosp_btn = WebDriverWait(driver_var, 30).until(find_all_avail_hosp)
                # all_avail_hosp_btn = driver_var.find_element_by_id("ctl00_ContentPlaceHolder1_cmdAllAvailHosp")
                #all_avail_hosp_btn.click()

                all_avail_hosp_attempt = 1
                while True:
                    try:
                        all_avail_hosp_btn = WebDriverWait(driver_var, 30).until(find_all_avail_hosp)
                        break
                    except StaleElementReferenceException:
                        if all_avail_hosp_attempt == 3:
                            raise
                        all_avail_hosp_attempt += 1

                all_avail_hosp_click_attempt = 1
                while True:
                    try:
                        all_avail_hosp_btn.click()
                        break
                    except StaleElementReferenceException:
                        if all_avail_hosp_click_attempt == 3:
                            raise
                        all_avail_hosp_click_attempt += 1


                # Because Vizient's stupid group by filter in the CBD report
                # will eliminate rows unnecessarily, we must remove PSI safety Indicator from Advanced Restrictions
                # in order to see the correct results and get correct zero rows.

                # scroll down
                adv_rest_div_element = WebDriverWait(driver_var, 30).until(find_adv_rest_scroll)
                # div_element4 = driver_var.find_element_by_xpath("//div[@id='ctl00_ContentPlaceHolder1_PanelContent5']")
                driver_var.execute_script("return arguments[0].scrollIntoView(0, document.documentElement.scrollHeight-10);",adv_rest_div_element)

                time.sleep(1)

                delete_btn_attempt = 1
                while True:
                    try:
                        generate_delete_btn = WebDriverWait(driver_var, 10).until(find_restrictions_delete_btn)
                        break
                    except StaleElementReferenceException:
                        if delete_btn_attempt == 3:
                            raise
                        delete_btn_attempt += 1

                delete_btn_click_attempt = 1
                while True:
                    try:
                        generate_delete_btn.click()
                        break
                    except StaleElementReferenceException:
                        if delete_btn_click_attempt == 3:
                            raise
                        delete_btn_click_attempt += 1

                time.sleep(1)
                
                #UL006
                
                if remove_covid_pats == True:
                    driver_var = add_covid_advanced_restriction(driver_var)
                    
                time.sleep(0.5)
                
                # scroll down
                div_element5 = WebDriverWait(driver_var, 30).until(find_div5_scroll)  # type: object
                # div_element5 = driver_var.find_element_by_xpath("//div[@id='divRiskAdjustment']")
                driver_var.execute_script("return arguments[0].scrollIntoView(0, document.documentElement.scrollHeight-10);", div_element5)

                # ActionChains(driver_var).move_to_element(all_avail_hosp_btn).click().perform()
                time.sleep(2)

                # Click Download button
                # generate_excel_btn = driver_var.find_element_by_id("ctl00_ContentPlaceHolder1_imgExcel")
                """try:
                    element = WebDriverWait(driver_var, 120).until(
                        EC.presence_of_element_located((By.ID, "ctl00_ContentPlaceHolder1_imgExcel"))
                    )
                    print("Excel image clickable.")
                except TimeoutException:
                    print("Time exceeded!")
                """
                #generate_excel_btn = WebDriverWait(driver_var, 10).until(find_excel_btn)
                #generate_excel_btn.click()

                excel_btn_attempt = 1
                while True:
                    try:
                        generate_excel_btn = WebDriverWait(driver_var, 10).until(find_excel_btn)
                        break
                    except StaleElementReferenceException:
                        if excel_btn_attempt == 3:
                            raise
                        excel_btn_attempt += 1

                excel_btn_click_attempt = 1
                while True:
                    try:
                        generate_excel_btn.click()
                        break
                    except StaleElementReferenceException:
                        if excel_btn_click_attempt == 3:
                            raise
                        excel_btn_click_attempt += 1
                #webdriver.ActionChains(driver_var).move_to_element(generate_excel_btn).click(generate_excel_btn).perform()
                #ActionChains(driver_var).move_to_element(generate_excel_btn).click().perform()
                time.sleep(1)
                # Accept Pop-up window

                try:
                    WebDriverWait(driver_var, 10).until(EC.alert_is_present(),
                                                       'Timed out waiting for PA creation ' +
                                                       'confirmation popup to appear.')

                    obj = WebDriverWait(driver_var, 10).until(find_alert_popup)

                    obj.accept()
                except:
                    pass
                #print('before download check while loop')
                while len(os.listdir(download_folder_dir)) <= num_already_downloaded_files:
                    time.sleep(1)
                #print('after download check while loop')
                #time.sleep(1.5)
                #print('before window check while loop')
                # wait for the pop up browser window to display so you can close it.
                while len(driver_var.window_handles) < 2:
                    time.sleep(0.5)
                #print('after window check while loop')
                window_before = driver_var.window_handles[0]
                window_after = driver_var.window_handles[1]
                # switch control to popup window and close it
                driver_var.switch_to.window(window_after)
                driver_var.close()
                # Switch control back to the original window.
                driver_var.switch_to.window(window_before)
                time.sleep(2)
                # find most recently-added file to Downloads folder and rename it.
                try:
                    latest_file = find_last_downloaded_file('C:/Users/NM184423/Downloads')
                except:
                    latest_file = find_last_downloaded_file('H:/Downloads')
                latest_file = os.path.abspath(latest_file)

                time.sleep(0.500)

                rename_and_move_file(latest_file, link_dict[i][3], link_dict[i][7], p1, file_dir)
                time.sleep(0.500)
                report_counter += 1
                time.sleep(2)
                update_template_files(hyperlink_loc, link_dict[i][3], link_dict[i][4])
                time.sleep(2)
                
        #UL003
        # Critical Access Measures:  Adverse Drug Events Rate & % Early Transfers Out
        # Require their reports run twice...
        # Run once with default settings to get numerator (Cases)
        # Run second time removing all Advanced Restrictions to get denominator (Cases)
        # Once reports are downloaded, then will need to take numerator/denominator to get true result.
        elif link_dict[i][4] in list_of_measures_to_run_twice_num_denom:
            for p1 in period_dict1.keys():
                #run twice.
                for item in ['NUM','DENOM']:
                    #first time run in default setting to get numerator
                    if item == 'NUM':
                        #if item is 'NUM' then run the ADE or % Early Transfer report in default mode.
                        # The 'Cases' column will give the numerator of the ratio.
                        print('Measure: ',link_dict[i][4])
                        # get number of files in Downloads folder
        
                        num_already_downloaded_files = len(os.listdir(download_folder_dir))
                        # open the template
                        time.sleep(1)
                        driver_var = open_template_report(link_dict[i][0], driver_var)
                        time.sleep(1)
                        driver_var.implicitly_wait(30)
                        try:
                            driver_var.maximize_window()
                        except:
                            pass
        
                        # scroll down
                        div_element1 = WebDriverWait(driver_var, 120).until(find_div1_scroll)
                        # div_element1 = driver_var.find_element_by_xpath("//div[@id='divRiskAdjustment']")
                        driver_var.execute_script(
                            "return arguments[0].scrollIntoView(0, document.documentElement.scrollHeight-10);", div_element1)
                        
                        #BEGIN FY21 COVID FILTER FUNCTIONS#
                        ###################################
                        
                        #UL004
                        '''
                        driver_var = remove_custom_covid_list_from_template(driver_var,link_dict[i][4])
                        
                        #UL004
                        driver_var = remove_multiple_group_by_from_template(driver_var)
                        
                        #UL004
                        driver_var = update_group_by_select_to_default(driver_var,link_dict[i][4])
                        
                        #UL004
                        driver_var = remove_psi_first_covid_adv_restriction(driver_var,link_dict[i][4])
                        
                        #UL004
                        driver_var = remove_covid_19_gen_med_subservice_rest(driver_var,link_dict[i][4])
                        
                        #UL004
                        driver_var = remove_covid_19_pulmonary_subservice_rest(driver_var,link_dict[i][4])
                        
                        #UL004
                        driver_var = remove_covid_19_readm_edac_discharge_month(driver_var,link_dict[i][4])
                        
                        #UL004
                        driver_var = remove_covid_19_revisits_discharge_month(driver_var,link_dict[i][4])
                        
                        #UL004
                        driver_var = remove_covid_19_crit_access_discharge_month(driver_var,link_dict[i][3])
                        
                        #UL004
                        driver_var = remove_covid_19_crit_access_subservice_line(driver_var,link_dict[i][3])
                        
                        '''
                        
                        
                        #END FY21 COVID FILTER FUNCTIONS#
                        ##################################
                        
                        
                        # Click Risk Adjustment Model button
                        driver_var = choose_adjustment_model(driver_var, link_dict, i)
        
                        # Click AHRQ Version button
                        driver_var = choose_ahrq_version(driver_var, link_dict, i)
        
                        # Update Multiple Group By Drop down
                        update_group_by_select(driver_var, period_dict1, p1)
        
                        # scroll down
                        div_element2 = WebDriverWait(driver_var, 60).until(find_div2_scroll)
                        # div_element2 = driver_var.find_element_by_xpath("//div[@id='ctl00_ContentPlaceHolder1_PanelContent3']")
                        driver_var.execute_script(
                            "return arguments[0].scrollIntoView(0, document.documentElement.scrollHeight-10);", div_element2)
        
                        # click the From/To Time Period Radio Button
        
                        try:
                            
                            time_period_btn_attempt = 1
                            while True:
                                try:
                                    time_period_radio_btn = WebDriverWait(driver_var, 45).until(find_time_period_radio_btn1)
                                    break
                                except StaleElementReferenceException:
                                    if time_period_btn_attempt == 3:
                                        raise
                                    time_period_btn_attempt += 1
        
                            time_period_btn_click_attempt = 1
                            while True:
                                try:
                                    ActionChains(driver_var).move_to_element(time_period_radio_btn).click().perform()
                                    break
                                except StaleElementReferenceException:
                                    if time_period_btn_click_attempt == 3:
                                        raise
                                    time_period_btn_click_attempt += 1
        
                        except:
                            
                            time_period_btn_attempt = 1
                            while True:
                                try:
                                    time_period_radio_btn = WebDriverWait(driver_var, 45).until(find_time_period_radio_btn2)
                                    break
                                except StaleElementReferenceException:
                                    if time_period_btn_attempt == 3:
                                        raise
                                    time_period_btn_attempt += 1
        
                            time_period_btn_click_attempt = 1
                            while True:
                                try:
                                    ActionChains(driver_var).move_to_element(time_period_radio_btn).click().perform()
                                    break
                                except StaleElementReferenceException:
                                    if time_period_btn_click_attempt == 3:
                                        raise
                                    time_period_btn_click_attempt += 1
                        driver_var.implicitly_wait(10)
        
        
                        # update time period downdown menus
                        update_time_period_select(driver_var, period_dict1, p1)
        
                        #time.sleep(0.5)
        
                        # scroll down
                        div_element3 = WebDriverWait(driver_var, 30).until(find_div3_scroll)
                        # div_element3 = driver_var.find_element_by_xpath("//div[@id='ctl00_ContentPlaceHolder1_PanelContent4']")
                        driver_var.execute_script(
                            "return arguments[0].scrollIntoView(0, document.documentElement.scrollHeight-10);", div_element3)
        
                        #time.sleep(2)
        
                        # Set focus hospital to NMH
        
                        try:
                            element = WebDriverWait(driver_var, 120).until(
                                EC.presence_of_element_located((By.CSS_SELECTOR, "option[value='140281']"))
                            )
                            #print("Option loaded")
                        except TimeoutException:
                            print("Time exceeded!")
        
                        time.sleep(1.5)
                        '''
                                        focus_hosp = WebDriverWait(driver_var, 150).until(find_focus_hosp)
                                        # Select(driver_var.find_element_by_xpath("//*[@id='ctl00_ContentPlaceHolder1_cmdFocusHCO']")).select_by_visible_text('NORTHWESTERN_MEMORIAL 140281')
                                        time.sleep(1)
                                        Select(focus_hosp).select_by_visible_text('NORTHWESTERN_MEMORIAL 140281')
                                        '''
                        #UL009 begin
                        '''
                        
                        focus_hosp = while_loop_handler_function(WebDriverWait(driver_var, 150).until(find_focus_hosp))
                        time.sleep(0.5)
                        #focus_hosp.click()
                        try:
                            focus_hosp.click()
                        except:
                            focus_hosp = while_loop_handler_function(WebDriverWait(driver_var, 150).until(find_focus_hosp))
                            time.sleep(0.5)
                            focus_hosp.click()
                        time.sleep(0.5)
                        Select(focus_hosp).select_by_visible_text('NORTHWESTERN_MEMORIAL 140281')
                        '''
                        
                        driver_var, focus_hosp = click_focus_hosp2(driver_var)
                
                    
                
                        try:
                            Select(focus_hosp).select_by_visible_text('NORTHWESTERN_MEMORIAL 140281')
                        except:
                            
                            driver_var, focus_hosp = click_focus_hosp2(driver_var)
                            Select(focus_hosp).select_by_visible_text('NORTHWESTERN_MEMORIAL 140281')
                        
                        #UL009 end
        
                        # scroll down
                        div_element4 = WebDriverWait(driver_var, 30).until(find_div4_scroll)
                        
                        driver_var.execute_script(
                            "return arguments[0].scrollIntoView(0, document.documentElement.scrollHeight-10);", div_element4)
        
                        time.sleep(0.5)
        
        
                        all_avail_hosp_attempt = 1
                        while True:
                            try:
                                all_avail_hosp_btn = WebDriverWait(driver_var, 30).until(find_all_avail_hosp)
                                break
                            except StaleElementReferenceException:
                                if all_avail_hosp_attempt == 3:
                                    raise
                                all_avail_hosp_attempt += 1
        
                        all_avail_hosp_click_attempt = 1
                        while True:
                            try:
                                all_avail_hosp_btn.click()
                                break
                            except StaleElementReferenceException:
                                if all_avail_hosp_click_attempt == 3:
                                    raise
                                all_avail_hosp_click_attempt += 1
                                
                        # scroll down
                        div_element5 = WebDriverWait(driver_var, 30).until(find_div5_scroll)  # type: object
                        # div_element5 = driver_var.find_element_by_xpath("//div[@id='divRiskAdjustment']")
                        driver_var.execute_script("return arguments[0].scrollIntoView(0, document.documentElement.scrollHeight-10);", div_element5)
        
                        # ActionChains(driver_var).move_to_element(all_avail_hosp_btn).click().perform()
                        time.sleep(2)
        
                        # Click Download button
                        # generate_excel_btn = driver_var.find_element_by_id("ctl00_ContentPlaceHolder1_imgExcel")
                        """try:
                            element = WebDriverWait(driver_var, 120).until(
                                EC.presence_of_element_located((By.ID, "ctl00_ContentPlaceHolder1_imgExcel"))
                            )
                            print("Excel image clickable.")
                        except TimeoutException:
                            print("Time exceeded!")
                        """
                        
                        #UL006
                
                        if remove_covid_pats == True:
                            driver_var = add_covid_advanced_restriction(driver_var)
                            
                        time.sleep(0.5)
                        
                        excel_btn_attempt = 1
                        while True:
                            try:
                                generate_excel_btn = WebDriverWait(driver_var, 10).until(find_excel_btn)
                                break
                            except StaleElementReferenceException:
                                if excel_btn_attempt == 3:
                                    raise
                                excel_btn_attempt += 1
        
                        excel_btn_click_attempt = 1
                        while True:
                            try:
                                generate_excel_btn.click()
                                break
                            except StaleElementReferenceException:
                                if excel_btn_click_attempt == 3:
                                    raise
                                excel_btn_click_attempt += 1
                        time.sleep(1)
                        # Accept Pop-up window
        
                        try:
                            WebDriverWait(driver_var, 10).until(EC.alert_is_present(),
                                                               'Timed out waiting for PA creation ' +
                                                               'confirmation popup to appear.')
        
                            obj = WebDriverWait(driver_var, 10).until(find_alert_popup)
        
                            obj.accept()
                        except:
                            pass
                        #print('before download check while loop')
                        while len(os.listdir(download_folder_dir)) <= num_already_downloaded_files:
                            time.sleep(1)
                            
                        while len(driver_var.window_handles) < 2:
                            time.sleep(0.5)
                        #print('after window check while loop')
                        window_before = driver_var.window_handles[0]
                        window_after = driver_var.window_handles[1]
                        # switch control to popup window and close it
                        driver_var.switch_to.window(window_after)
                        driver_var.close()
                        # Switch control back to the original window.
                        driver_var.switch_to.window(window_before)
                        time.sleep(2)
                        # find most recently-added file to Downloads folder and rename it.
                        try:
                            latest_file = find_last_downloaded_file('C:/Users/NM184423/Downloads')
                        except:
                            latest_file = find_last_downloaded_file('H:/Downloads')
                        latest_file = os.path.abspath(latest_file)
        
                        time.sleep(0.500)
                        
                        #UL003
                        #if 'NUM' then rename the measure_name to ..._NUM.  For example, ADE_NUM.
                        #UL003
                        updated_measure_name = link_dict[i][7] + '_NUM'
                        #UL003
                        #rename_and_move_file(latest_file, link_dict[i][3], link_dict[i][7], p1, file_dir)
                        rename_and_move_file(latest_file, link_dict[i][3], updated_measure_name, p1, file_dir)
                        time.sleep(0.500)
                        report_counter += 1
                        time.sleep(2)
                        update_template_files(hyperlink_loc, link_dict[i][3], link_dict[i][4])
                        time.sleep(2)
                        
                if item == 'DENOM':
                        #Adverse Drug Events only has one advanced restriction so only need
                        #delete one advanced restriction.
                        if link_dict[i][4] == 'Adverse Drug Events Rate':
                            
                            #if item is 'DENOM' then DO NOT run the ADE or % Early Transfer report in default mode.
                            # Need to remove all advanced restrictions by clicking the delete button
                            # THEN run the report to get the denominator.
                            # The 'Cases' column will give the denominator of the ratio.
                            print('Measure: ',link_dict[i][4])
                            # get number of files in Downloads folder
            
                            num_already_downloaded_files = len(os.listdir(download_folder_dir))
                            # open the template
                            time.sleep(0.5)
                            driver_var = open_template_report(link_dict[i][0], driver_var)
                            time.sleep(1)
                            driver_var.implicitly_wait(30)
                            try:
                                driver_var.maximize_window()
                            except:
                                pass
            
                            # scroll down
                            div_element1 = WebDriverWait(driver_var, 120).until(find_div1_scroll)
                            # div_element1 = driver_var.find_element_by_xpath("//div[@id='divRiskAdjustment']")
                            driver_var.execute_script(
                                "return arguments[0].scrollIntoView(0, document.documentElement.scrollHeight-10);", div_element1)
                            
                            #BEGIN FY21 COVID FILTER FUNCTIONS#
                            ###################################
                            
                            #UL004
                            '''
                            driver_var = remove_custom_covid_list_from_template(driver_var,link_dict[i][4])
                            
                            #UL004
                            driver_var = remove_multiple_group_by_from_template(driver_var)
                            
                            #UL004
                            driver_var = update_group_by_select_to_default(driver_var,link_dict[i][4])
                            
                            #UL004
                            driver_var = remove_psi_first_covid_adv_restriction(driver_var,link_dict[i][4])
                            
                            #UL004
                            driver_var = remove_covid_19_gen_med_subservice_rest(driver_var,link_dict[i][4])
                            
                            #UL004
                            driver_var = remove_covid_19_pulmonary_subservice_rest(driver_var,link_dict[i][4])
                            
                            #UL004
                            driver_var = remove_covid_19_readm_edac_discharge_month(driver_var,link_dict[i][4])
                            
                            #UL004
                            driver_var = remove_covid_19_revisits_discharge_month(driver_var,link_dict[i][4])
                            
                            #UL004
                            driver_var = remove_covid_19_crit_access_discharge_month(driver_var,link_dict[i][3])
                            
                            #UL004
                            driver_var = remove_covid_19_crit_access_subservice_line(driver_var,link_dict[i][3])
                            
                            '''
                            
                            #END FY21 COVID FILTER FUNCTIONS#
                            ##################################
                            
                            
                            # Click Risk Adjustment Model button
                            driver_var = choose_adjustment_model(driver_var, link_dict, i)
            
                            # Click AHRQ Version button
                            driver_var = choose_ahrq_version(driver_var, link_dict, i)
            
                            # Update Multiple Group By Drop down
                            update_group_by_select(driver_var, period_dict1, p1)
            
                            # scroll down
                            div_element2 = WebDriverWait(driver_var, 60).until(find_div2_scroll)
                            # div_element2 = driver_var.find_element_by_xpath("//div[@id='ctl00_ContentPlaceHolder1_PanelContent3']")
                            driver_var.execute_script(
                                "return arguments[0].scrollIntoView(0, document.documentElement.scrollHeight-10);", div_element2)
            
                            # click the From/To Time Period Radio Button
            
                            try:
                                
                                time_period_btn_attempt = 1
                                while True:
                                    try:
                                        time_period_radio_btn = WebDriverWait(driver_var, 45).until(find_time_period_radio_btn1)
                                        break
                                    except StaleElementReferenceException:
                                        if time_period_btn_attempt == 3:
                                            raise
                                        time_period_btn_attempt += 1
            
                                time_period_btn_click_attempt = 1
                                while True:
                                    try:
                                        ActionChains(driver_var).move_to_element(time_period_radio_btn).click().perform()
                                        break
                                    except StaleElementReferenceException:
                                        if time_period_btn_click_attempt == 3:
                                            raise
                                        time_period_btn_click_attempt += 1
            
                            except:
                                
                                time_period_btn_attempt = 1
                                while True:
                                    try:
                                        time_period_radio_btn = WebDriverWait(driver_var, 45).until(find_time_period_radio_btn2)
                                        break
                                    except StaleElementReferenceException:
                                        if time_period_btn_attempt == 3:
                                            raise
                                        time_period_btn_attempt += 1
            
                                time_period_btn_click_attempt = 1
                                while True:
                                    try:
                                        ActionChains(driver_var).move_to_element(time_period_radio_btn).click().perform()
                                        break
                                    except StaleElementReferenceException:
                                        if time_period_btn_click_attempt == 3:
                                            raise
                                        time_period_btn_click_attempt += 1
                            driver_var.implicitly_wait(10)
            
            
                            # update time period downdown menus
                            update_time_period_select(driver_var, period_dict1, p1)
            
                            #time.sleep(0.5)
            
                            # scroll down
                            div_element3 = WebDriverWait(driver_var, 30).until(find_div3_scroll)
                            # div_element3 = driver_var.find_element_by_xpath("//div[@id='ctl00_ContentPlaceHolder1_PanelContent4']")
                            driver_var.execute_script(
                                "return arguments[0].scrollIntoView(0, document.documentElement.scrollHeight-10);", div_element3)
            
                            #time.sleep(2)
            
                            # Set focus hospital to NMH
            
                            try:
                                element = WebDriverWait(driver_var, 120).until(
                                    EC.presence_of_element_located((By.CSS_SELECTOR, "option[value='140281']"))
                                )
                                #print("Option loaded")
                            except TimeoutException:
                                print("Time exceeded!")
            
                            time.sleep(1.5)
                            '''
                                            focus_hosp = WebDriverWait(driver_var, 150).until(find_focus_hosp)
                                            # Select(driver_var.find_element_by_xpath("//*[@id='ctl00_ContentPlaceHolder1_cmdFocusHCO']")).select_by_visible_text('NORTHWESTERN_MEMORIAL 140281')
                                            time.sleep(1)
                                            Select(focus_hosp).select_by_visible_text('NORTHWESTERN_MEMORIAL 140281')
                                            '''
                            #UL009 begin
                            '''
                            focus_hosp = while_loop_handler_function(WebDriverWait(driver_var, 150).until(find_focus_hosp))
                            time.sleep(0.5)
                            #focus_hosp.click()
                            try:
                                focus_hosp.click()
                            except:
                                focus_hosp = while_loop_handler_function(WebDriverWait(driver_var, 150).until(find_focus_hosp))
                                time.sleep(0.5)
                                focus_hosp.click()
                            time.sleep(0.5)
                            Select(focus_hosp).select_by_visible_text('NORTHWESTERN_MEMORIAL 140281')
                            '''
                            
                            driver_var, focus_hosp = click_focus_hosp2(driver_var)
                
                    
                
                            try:
                                Select(focus_hosp).select_by_visible_text('NORTHWESTERN_MEMORIAL 140281')
                            except:
                                
                                driver_var, focus_hosp = click_focus_hosp2(driver_var)
                                Select(focus_hosp).select_by_visible_text('NORTHWESTERN_MEMORIAL 140281')
                            
                            #UL009 end
                            
                            # scroll down
                            div_element4 = WebDriverWait(driver_var, 30).until(find_div4_scroll)
                            
                            driver_var.execute_script(
                                "return arguments[0].scrollIntoView(0, document.documentElement.scrollHeight-10);", div_element4)
            
                            time.sleep(0.5)
            
            
                            all_avail_hosp_attempt = 1
                            while True:
                                try:
                                    all_avail_hosp_btn = WebDriverWait(driver_var, 30).until(find_all_avail_hosp)
                                    break
                                except StaleElementReferenceException:
                                    if all_avail_hosp_attempt == 3:
                                        raise
                                    all_avail_hosp_attempt += 1
            
                            all_avail_hosp_click_attempt = 1
                            while True:
                                try:
                                    all_avail_hosp_btn.click()
                                    break
                                except StaleElementReferenceException:
                                    if all_avail_hosp_click_attempt == 3:
                                        raise
                                    all_avail_hosp_click_attempt += 1
                                    
                            # UL003
                            # Similar to PSI's, we need to remove the advanced restrictions
                            # For some reason, on the Critical Access calculator, 2 measures (ADE, % Early Transfers)
                            # the report template does not give the final correct calculations.
                            # According to Vizient, you actually have to run the report twice. 
                            # Adverse Drug Events Rate only has 1 advanced restriction so we only need to 
                            # Click on 1 delete button in the advanced restrictions section similar to the PSI measures.
            
                            # scroll down
                            adv_rest_div_element = WebDriverWait(driver_var, 30).until(find_adv_rest_scroll)
                            # div_element4 = driver_var.find_element_by_xpath("//div[@id='ctl00_ContentPlaceHolder1_PanelContent5']")
                            driver_var.execute_script("return arguments[0].scrollIntoView(0, document.documentElement.scrollHeight-10);",adv_rest_div_element)
            
                            time.sleep(1)
                            #UL003
                            delete_btn_attempt = 1
                            while True:
                                try:
                                    generate_delete_btn = WebDriverWait(driver_var, 10).until(find_restrictions_delete_btn)
                                    break
                                except StaleElementReferenceException:
                                    if delete_btn_attempt == 3:
                                        raise
                                    delete_btn_attempt += 1
                            #UL003
                            delete_btn_click_attempt = 1
                            while True:
                                try:
                                    generate_delete_btn.click()
                                    break
                                except StaleElementReferenceException:
                                    if delete_btn_click_attempt == 3:
                                        raise
                                    delete_btn_click_attempt += 1
            
                            time.sleep(1)
            
                            # scroll down
                            div_element5 = WebDriverWait(driver_var, 30).until(find_div5_scroll)  # type: object
                            # div_element5 = driver_var.find_element_by_xpath("//div[@id='divRiskAdjustment']")
                            driver_var.execute_script("return arguments[0].scrollIntoView(0, document.documentElement.scrollHeight-10);", div_element5)
            
                            # ActionChains(driver_var).move_to_element(all_avail_hosp_btn).click().perform()
                            time.sleep(2)
            
                            # Click Download button
                            # generate_excel_btn = driver_var.find_element_by_id("ctl00_ContentPlaceHolder1_imgExcel")
                            """try:
                                element = WebDriverWait(driver_var, 120).until(
                                    EC.presence_of_element_located((By.ID, "ctl00_ContentPlaceHolder1_imgExcel"))
                                )
                                print("Excel image clickable.")
                            except TimeoutException:
                                print("Time exceeded!")
                            """
                            
                            #UL006
                
                            if remove_covid_pats == True:
                                driver_var = add_covid_advanced_restriction(driver_var)
                                
                            time.sleep(0.5)
                            
                            excel_btn_attempt = 1
                            while True:
                                try:
                                    generate_excel_btn = WebDriverWait(driver_var, 10).until(find_excel_btn)
                                    break
                                except StaleElementReferenceException:
                                    if excel_btn_attempt == 3:
                                        raise
                                    excel_btn_attempt += 1
            
                            excel_btn_click_attempt = 1
                            while True:
                                try:
                                    generate_excel_btn.click()
                                    break
                                except StaleElementReferenceException:
                                    if excel_btn_click_attempt == 3:
                                        raise
                                    excel_btn_click_attempt += 1
                            time.sleep(1)
                            # Accept Pop-up window
            
                            try:
                                WebDriverWait(driver_var, 10).until(EC.alert_is_present(),
                                                                   'Timed out waiting for PA creation ' +
                                                                   'confirmation popup to appear.')
            
                                obj = WebDriverWait(driver_var, 10).until(find_alert_popup)
            
                                obj.accept()
                            except:
                                pass
                            #print('before download check while loop')
                            while len(os.listdir(download_folder_dir)) <= num_already_downloaded_files:
                                time.sleep(1)
                                
                            while len(driver_var.window_handles) < 2:
                                time.sleep(0.5)
                            #print('after window check while loop')
                            window_before = driver_var.window_handles[0]
                            window_after = driver_var.window_handles[1]
                            # switch control to popup window and close it
                            driver_var.switch_to.window(window_after)
                            driver_var.close()
                            # Switch control back to the original window.
                            driver_var.switch_to.window(window_before)
                            time.sleep(2)
                            # find most recently-added file to Downloads folder and rename it.
                            try:
                                latest_file = find_last_downloaded_file('C:/Users/NM184423/Downloads')
                            except:
                                latest_file = find_last_downloaded_file('H:/Downloads')
                            latest_file = os.path.abspath(latest_file)
            
                            time.sleep(0.500)
                            
                            #UL003
                            #if 'DENOM' then rename the measure_name to ..._NUM.  For example, ADE_NUM.
                            #UL003
                            updated_measure_name = link_dict[i][7] + '_DENOM'
                            #UL003
                            #rename_and_move_file(latest_file, link_dict[i][3], link_dict[i][7], p1, file_dir)
                            rename_and_move_file(latest_file, link_dict[i][3], updated_measure_name, p1, file_dir)
                            time.sleep(0.500)
                            report_counter += 1
                            time.sleep(2)
                            update_template_files(hyperlink_loc, link_dict[i][3], link_dict[i][4])
                            time.sleep(2)
                            
                        #% Early Transfer Out has 4 advanced restrictions so we need to remove
                        #4 different advanced restrictions by clicking the delete button.
                        elif link_dict[i][4] == '% Early Transfers Out':
                            
                            #if item is 'DENOM' then DO NOT run the ADE or % Early Transfer report in default mode.
                            #first remove all advanced restrictions THEN run the report.
                            # The 'Cases' column will give the denominator of the ratio.
                            print('Measure: ',link_dict[i][4])
                            # get number of files in Downloads folder
            
                            num_already_downloaded_files = len(os.listdir(download_folder_dir))
                            # open the template
                            time.sleep(0.5)
                            driver_var = open_template_report(link_dict[i][0], driver_var)
                            time.sleep(1)
                            driver_var.implicitly_wait(30)
                            try:
                                driver_var.maximize_window()
                            except:
                                pass
            
                            # scroll down
                            div_element1 = WebDriverWait(driver_var, 120).until(find_div1_scroll)
                            # div_element1 = driver_var.find_element_by_xpath("//div[@id='divRiskAdjustment']")
                            driver_var.execute_script(
                                "return arguments[0].scrollIntoView(0, document.documentElement.scrollHeight-10);", div_element1)
                            
                            #BEGIN FY21 COVID FILTER FUNCTIONS#
                            ###################################
                            
                            #UL004
                            '''
                            driver_var = remove_custom_covid_list_from_template(driver_var,link_dict[i][4])
                            
                            #UL004
                            driver_var = remove_multiple_group_by_from_template(driver_var)
                            
                            #UL004
                            driver_var = update_group_by_select_to_default(driver_var,link_dict[i][4])
                            
                            #UL004
                            driver_var = remove_psi_first_covid_adv_restriction(driver_var,link_dict[i][4])
                            
                            #UL004
                            driver_var = remove_covid_19_gen_med_subservice_rest(driver_var,link_dict[i][4])
                            
                            #UL004
                            driver_var = remove_covid_19_pulmonary_subservice_rest(driver_var,link_dict[i][4])
                            
                            #UL004
                            driver_var = remove_covid_19_readm_edac_discharge_month(driver_var,link_dict[i][4])
                            
                            #UL004
                            driver_var = remove_covid_19_revisits_discharge_month(driver_var,link_dict[i][4])
                            
                            #UL004
                            driver_var = remove_covid_19_crit_access_discharge_month(driver_var,link_dict[i][3])
                            
                            #UL004
                            driver_var = remove_covid_19_crit_access_subservice_line(driver_var,link_dict[i][3])
                            
                            '''
                            
                            #END FY21 COVID FILTER FUNCTIONS#
                            ##################################
                            
                            
                            # Click Risk Adjustment Model button
                            driver_var = choose_adjustment_model(driver_var, link_dict, i)
            
                            # Click AHRQ Version button
                            driver_var = choose_ahrq_version(driver_var, link_dict, i)
            
                            # Update Multiple Group By Drop down
                            update_group_by_select(driver_var, period_dict1, p1)
            
                            # scroll down
                            div_element2 = WebDriverWait(driver_var, 60).until(find_div2_scroll)
                            # div_element2 = driver_var.find_element_by_xpath("//div[@id='ctl00_ContentPlaceHolder1_PanelContent3']")
                            driver_var.execute_script(
                                "return arguments[0].scrollIntoView(0, document.documentElement.scrollHeight-10);", div_element2)
            
                            # click the From/To Time Period Radio Button
            
                            try:
                                
                                time_period_btn_attempt = 1
                                while True:
                                    try:
                                        time_period_radio_btn = WebDriverWait(driver_var, 45).until(find_time_period_radio_btn1)
                                        break
                                    except StaleElementReferenceException:
                                        if time_period_btn_attempt == 3:
                                            raise
                                        time_period_btn_attempt += 1
            
                                time_period_btn_click_attempt = 1
                                while True:
                                    try:
                                        ActionChains(driver_var).move_to_element(time_period_radio_btn).click().perform()
                                        break
                                    except StaleElementReferenceException:
                                        if time_period_btn_click_attempt == 3:
                                            raise
                                        time_period_btn_click_attempt += 1
            
                            except:
                                
                                time_period_btn_attempt = 1
                                while True:
                                    try:
                                        time_period_radio_btn = WebDriverWait(driver_var, 45).until(find_time_period_radio_btn2)
                                        break
                                    except StaleElementReferenceException:
                                        if time_period_btn_attempt == 3:
                                            raise
                                        time_period_btn_attempt += 1
            
                                time_period_btn_click_attempt = 1
                                while True:
                                    try:
                                        ActionChains(driver_var).move_to_element(time_period_radio_btn).click().perform()
                                        break
                                    except StaleElementReferenceException:
                                        if time_period_btn_click_attempt == 3:
                                            raise
                                        time_period_btn_click_attempt += 1
                            driver_var.implicitly_wait(10)
            
            
                            # update time period downdown menus
                            update_time_period_select(driver_var, period_dict1, p1)
            
                            #time.sleep(0.5)
            
                            # scroll down
                            div_element3 = WebDriverWait(driver_var, 30).until(find_div3_scroll)
                            # div_element3 = driver_var.find_element_by_xpath("//div[@id='ctl00_ContentPlaceHolder1_PanelContent4']")
                            driver_var.execute_script(
                                "return arguments[0].scrollIntoView(0, document.documentElement.scrollHeight-10);", div_element3)
            
                            #time.sleep(2)
            
                            # Set focus hospital to NMH
            
                            try:
                                element = WebDriverWait(driver_var, 120).until(
                                    EC.presence_of_element_located((By.CSS_SELECTOR, "option[value='140281']"))
                                )
                                #print("Option loaded")
                            except TimeoutException:
                                print("Time exceeded!")
            
                            time.sleep(1.5)
                            '''
                                            focus_hosp = WebDriverWait(driver_var, 150).until(find_focus_hosp)
                                            # Select(driver_var.find_element_by_xpath("//*[@id='ctl00_ContentPlaceHolder1_cmdFocusHCO']")).select_by_visible_text('NORTHWESTERN_MEMORIAL 140281')
                                            time.sleep(1)
                                            Select(focus_hosp).select_by_visible_text('NORTHWESTERN_MEMORIAL 140281')
                                            '''
                            
                            #UL009 begin
                            '''
                            focus_hosp = while_loop_handler_function(WebDriverWait(driver_var, 150).until(find_focus_hosp))
                            time.sleep(0.5)
                            #focus_hosp.click()
                            try:
                                focus_hosp.click()
                            except:
                                focus_hosp = while_loop_handler_function(WebDriverWait(driver_var, 150).until(find_focus_hosp))
                                time.sleep(0.5)
                                focus_hosp.click()
                            time.sleep(0.5)
                            Select(focus_hosp).select_by_visible_text('NORTHWESTERN_MEMORIAL 140281')
                            '''
                            
                            driver_var, focus_hosp = click_focus_hosp2(driver_var)
                
                    
                
                            try:
                                Select(focus_hosp).select_by_visible_text('NORTHWESTERN_MEMORIAL 140281')
                            except:
                                
                                driver_var, focus_hosp = click_focus_hosp2(driver_var)
                                Select(focus_hosp).select_by_visible_text('NORTHWESTERN_MEMORIAL 140281')
                            
                            #UL009 end
            
                            # scroll down
                            div_element4 = WebDriverWait(driver_var, 30).until(find_div4_scroll)
                            
                            driver_var.execute_script(
                                "return arguments[0].scrollIntoView(0, document.documentElement.scrollHeight-10);", div_element4)
            
                            time.sleep(0.5)
            
            
                            all_avail_hosp_attempt = 1
                            while True:
                                try:
                                    all_avail_hosp_btn = WebDriverWait(driver_var, 30).until(find_all_avail_hosp)
                                    break
                                except StaleElementReferenceException:
                                    if all_avail_hosp_attempt == 3:
                                        raise
                                    all_avail_hosp_attempt += 1
            
                            all_avail_hosp_click_attempt = 1
                            while True:
                                try:
                                    all_avail_hosp_btn.click()
                                    break
                                except StaleElementReferenceException:
                                    if all_avail_hosp_click_attempt == 3:
                                        raise
                                    all_avail_hosp_click_attempt += 1
                                    
                            # UL003
                            # Similar to PSI's, we need to remove the advanced restrictions
                            # For some reason, on the Critical Access calculator, 2 measures (ADE, % Early Transfers)
                            # the report template does not give the final correct calculations.
                            # According to Vizient, you actually have to run the report twice. 
                            #% Early Transfer out, there are 4 advanced restrictions.  We need 
                            # to remove all advanced restrictions by clicking 4 different delete buttons.
            
                            # scroll down
                            adv_rest_div_element = WebDriverWait(driver_var, 30).until(find_adv_rest_scroll)
                            # div_element4 = driver_var.find_element_by_xpath("//div[@id='ctl00_ContentPlaceHolder1_PanelContent5']")
                            driver_var.execute_script("return arguments[0].scrollIntoView(0, document.documentElement.scrollHeight-10);",adv_rest_div_element)
            
                            #start deleting advanced restrictions starting from bottom. 
                            
                            
                            time.sleep(1)
                            
                            #click delete button 4.
                            
                            delete_btn_attempt = 1
                            while True:
                                try:
                                    generate_delete_btn4 = WebDriverWait(driver_var, 10).until(find_restrictions_delete_btn4)
                                    break
                                except StaleElementReferenceException:
                                    if delete_btn_attempt == 3:
                                        raise
                                    delete_btn_attempt += 1
            
                            delete_btn_click_attempt = 1
                            while True:
                                try:
                                    generate_delete_btn4.click()
                                    break
                                except StaleElementReferenceException:
                                    if delete_btn_click_attempt == 3:
                                        raise
                                    delete_btn_click_attempt += 1
                                    
                                    
                            time.sleep(1)
                            
                            #click delete button 3
                            delete_btn_attempt = 1
                            while True:
                                try:
                                    generate_delete_btn3 = WebDriverWait(driver_var, 10).until(find_restrictions_delete_btn3)
                                    break
                                except StaleElementReferenceException:
                                    if delete_btn_attempt == 3:
                                        raise
                                    delete_btn_attempt += 1
            
                            delete_btn_click_attempt = 1
                            while True:
                                try:
                                    generate_delete_btn3.click()
                                    break
                                except StaleElementReferenceException:
                                    if delete_btn_click_attempt == 3:
                                        raise
                                    delete_btn_click_attempt += 1
                                        
                            time.sleep(1)
                            
                            
                            #click delete button 2
                            delete_btn_attempt = 1
                            while True:
                                try:
                                    generate_delete_btn2 = WebDriverWait(driver_var, 10).until(find_restrictions_delete_btn2)
                                    break
                                except StaleElementReferenceException:
                                    if delete_btn_attempt == 3:
                                        raise
                                    delete_btn_attempt += 1
            
                            delete_btn_click_attempt = 1
                            while True:
                                try:
                                    generate_delete_btn2.click()
                                    break
                                except StaleElementReferenceException:
                                    if delete_btn_click_attempt == 3:
                                        raise
                                    delete_btn_click_attempt += 1
                                        
                            time.sleep(1)
                            
            
                            delete_btn_attempt = 1
                            while True:
                                try:
                                    generate_delete_btn = WebDriverWait(driver_var, 10).until(find_restrictions_delete_btn)
                                    break
                                except StaleElementReferenceException:
                                    if delete_btn_attempt == 3:
                                        raise
                                    delete_btn_attempt += 1
            
                            delete_btn_click_attempt = 1
                            while True:
                                try:
                                    generate_delete_btn.click()
                                    break
                                except StaleElementReferenceException:
                                    if delete_btn_click_attempt == 3:
                                        raise
                                    delete_btn_click_attempt += 1
            
                            time.sleep(1)
            
                            # scroll down
                            div_element5 = WebDriverWait(driver_var, 30).until(find_div5_scroll)  # type: object
                            # div_element5 = driver_var.find_element_by_xpath("//div[@id='divRiskAdjustment']")
                            driver_var.execute_script("return arguments[0].scrollIntoView(0, document.documentElement.scrollHeight-10);", div_element5)
            
                            # ActionChains(driver_var).move_to_element(all_avail_hosp_btn).click().perform()
                            time.sleep(2)
            
                            # Click Download button
                            # generate_excel_btn = driver_var.find_element_by_id("ctl00_ContentPlaceHolder1_imgExcel")
                            """try:
                                element = WebDriverWait(driver_var, 120).until(
                                    EC.presence_of_element_located((By.ID, "ctl00_ContentPlaceHolder1_imgExcel"))
                                )
                                print("Excel image clickable.")
                            except TimeoutException:
                                print("Time exceeded!")
                            """
                            
                            #UL006
                
                            if remove_covid_pats == True:
                                driver_var = add_covid_advanced_restriction(driver_var)
                                
                            time.sleep(0.5)
            
                            excel_btn_attempt = 1
                            while True:
                                try:
                                    generate_excel_btn = WebDriverWait(driver_var, 10).until(find_excel_btn)
                                    break
                                except StaleElementReferenceException:
                                    if excel_btn_attempt == 3:
                                        raise
                                    excel_btn_attempt += 1
            
                            excel_btn_click_attempt = 1
                            while True:
                                try:
                                    generate_excel_btn.click()
                                    break
                                except StaleElementReferenceException:
                                    if excel_btn_click_attempt == 3:
                                        raise
                                    excel_btn_click_attempt += 1
                            time.sleep(1)
                            # Accept Pop-up window
            
                            try:
                                WebDriverWait(driver_var, 10).until(EC.alert_is_present(),
                                                                   'Timed out waiting for PA creation ' +
                                                                   'confirmation popup to appear.')
            
                                obj = WebDriverWait(driver_var, 10).until(find_alert_popup)
            
                                obj.accept()
                            except:
                                pass
                            #print('before download check while loop')
                            while len(os.listdir(download_folder_dir)) <= num_already_downloaded_files:
                                time.sleep(1)
                                
                            while len(driver_var.window_handles) < 2:
                                time.sleep(0.5)
                            #print('after window check while loop')
                            window_before = driver_var.window_handles[0]
                            window_after = driver_var.window_handles[1]
                            # switch control to popup window and close it
                            driver_var.switch_to.window(window_after)
                            driver_var.close()
                            # Switch control back to the original window.
                            driver_var.switch_to.window(window_before)
                            time.sleep(2)
                            # find most recently-added file to Downloads folder and rename it.
                            try:
                                latest_file = find_last_downloaded_file('C:/Users/NM184423/Downloads')
                            except:
                                latest_file = find_last_downloaded_file('H:/Downloads')
                            latest_file = os.path.abspath(latest_file)
            
                            time.sleep(0.500)
                            
                            #UL003
                            #if 'DENOM' then rename the measure_name to ..._DENOM.  For example, ADE_DENOM.
                            #UL003
                            updated_measure_name = link_dict[i][7] + '_DENOM'
                            #UL003
                            #rename_and_move_file(latest_file, link_dict[i][3], link_dict[i][7], p1, file_dir)
                            rename_and_move_file(latest_file, link_dict[i][3], updated_measure_name, p1, file_dir)
                            time.sleep(0.500)
                            report_counter += 1
                            time.sleep(2)
                            update_template_files(hyperlink_loc, link_dict[i][3], link_dict[i][4])
                            time.sleep(2)
        elif link_dict[i][4] not in list_of_total_revisits_measures:
            for p1 in period_dict1.keys():
                print('Measure: ',link_dict[i][4])
                # get number of files in Downloads folder

                num_already_downloaded_files = len(os.listdir(download_folder_dir))
                # open the template
                time.sleep(0.5)
                driver_var = open_template_report(link_dict[i][0], driver_var)
                time.sleep(1)
                driver_var.implicitly_wait(30)
                try:
                    driver_var.maximize_window()
                except:
                    pass

                # scroll down
                div_element1 = WebDriverWait(driver_var, 120).until(find_div1_scroll)
                # div_element1 = driver_var.find_element_by_xpath("//div[@id='divRiskAdjustment']")
                driver_var.execute_script(
                    "return arguments[0].scrollIntoView(0, document.documentElement.scrollHeight-10);", div_element1)
                
                #BEGIN FY21 COVID FILTER FUNCTIONS#
                ###################################
                
                #UL004
                '''
                driver_var = remove_custom_covid_list_from_template(driver_var,link_dict[i][4])
                
                #UL004
                driver_var = remove_multiple_group_by_from_template(driver_var)
                
                #UL004
                driver_var = update_group_by_select_to_default(driver_var,link_dict[i][4])
                
                #UL004
                driver_var = remove_psi_first_covid_adv_restriction(driver_var,link_dict[i][4])
                
                #UL004
                driver_var = remove_covid_19_gen_med_subservice_rest(driver_var,link_dict[i][4])
                
                #UL004
                driver_var = remove_covid_19_pulmonary_subservice_rest(driver_var,link_dict[i][4])
                
                #UL004
                driver_var = remove_covid_19_readm_edac_discharge_month(driver_var,link_dict[i][4])
                
                #UL004
                driver_var = remove_covid_19_revisits_discharge_month(driver_var,link_dict[i][4])
                
                #UL004
                driver_var = remove_covid_19_crit_access_discharge_month(driver_var,link_dict[i][3])
                
                #UL004
                driver_var = remove_covid_19_crit_access_subservice_line(driver_var,link_dict[i][3])
                
                '''
                
                
                #END FY21 COVID FILTER FUNCTIONS#
                ##################################
                
                
                # Click Risk Adjustment Model button
                driver_var = choose_adjustment_model(driver_var, link_dict, i)

                # Click AHRQ Version button
                driver_var = choose_ahrq_version(driver_var, link_dict, i)

                # Update Multiple Group By Drop down
                update_group_by_select(driver_var, period_dict1, p1)

                # scroll down
                div_element2 = WebDriverWait(driver_var, 60).until(find_div2_scroll)
                # div_element2 = driver_var.find_element_by_xpath("//div[@id='ctl00_ContentPlaceHolder1_PanelContent3']")
                driver_var.execute_script(
                    "return arguments[0].scrollIntoView(0, document.documentElement.scrollHeight-10);", div_element2)

                # click the From/To Time Period Radio Button

                try:
                    #time_period_radio_btn = WebDriverWait(driver_var, 45).until(find_time_period_radio_btn1)
                    # time_period_radio_btn = driver_var.find_element_by_id("ctl00_ContentPlaceHolder1_fromYear")
                    # time_period_radio_btn.click()
                    #ActionChains(driver_var).move_to_element(time_period_radio_btn).click().perform()
                    time_period_btn_attempt = 1
                    while True:
                        try:
                            time_period_radio_btn = WebDriverWait(driver_var, 45).until(find_time_period_radio_btn1)
                            break
                        except StaleElementReferenceException:
                            if time_period_btn_attempt == 3:
                                raise
                            time_period_btn_attempt += 1

                    time_period_btn_click_attempt = 1
                    while True:
                        try:
                            ActionChains(driver_var).move_to_element(time_period_radio_btn).click().perform()
                            break
                        except StaleElementReferenceException:
                            if time_period_btn_click_attempt == 3:
                                raise
                            time_period_btn_click_attempt += 1

                except:
                    #time_period_radio_btn = WebDriverWait(driver_var, 45).until(find_time_period_radio_btn2)
                    # time_period_radio_btn = driver_var.find_element_by_id("ctl00_ContentPlaceHolder1_cmdFromYear")
                    #ActionChains(driver_var).move_to_element(time_period_radio_btn).click().perform()
                    time_period_btn_attempt = 1
                    while True:
                        try:
                            time_period_radio_btn = WebDriverWait(driver_var, 45).until(find_time_period_radio_btn2)
                            break
                        except StaleElementReferenceException:
                            if time_period_btn_attempt == 3:
                                raise
                            time_period_btn_attempt += 1

                    time_period_btn_click_attempt = 1
                    while True:
                        try:
                            ActionChains(driver_var).move_to_element(time_period_radio_btn).click().perform()
                            break
                        except StaleElementReferenceException:
                            if time_period_btn_click_attempt == 3:
                                raise
                            time_period_btn_click_attempt += 1
                driver_var.implicitly_wait(10)


                # update time period downdown menus
                update_time_period_select(driver_var, period_dict1, p1)

                #time.sleep(0.5)

                # scroll down
                div_element3 = WebDriverWait(driver_var, 30).until(find_div3_scroll)
                # div_element3 = driver_var.find_element_by_xpath("//div[@id='ctl00_ContentPlaceHolder1_PanelContent4']")
                driver_var.execute_script(
                    "return arguments[0].scrollIntoView(0, document.documentElement.scrollHeight-10);", div_element3)

                #time.sleep(2)

                # Set focus hospital to NMH

                try:
                    element = WebDriverWait(driver_var, 120).until(
                        EC.presence_of_element_located((By.CSS_SELECTOR, "option[value='140281']"))
                    )
                    #print("Option loaded")
                except TimeoutException:
                    print("Time exceeded!")

                time.sleep(1.5)
                '''
                                focus_hosp = WebDriverWait(driver_var, 150).until(find_focus_hosp)
                                # Select(driver_var.find_element_by_xpath("//*[@id='ctl00_ContentPlaceHolder1_cmdFocusHCO']")).select_by_visible_text('NORTHWESTERN_MEMORIAL 140281')
                                time.sleep(1)
                                Select(focus_hosp).select_by_visible_text('NORTHWESTERN_MEMORIAL 140281')
                                '''
                #focus_hosp = WebDriverWait(driver_var, 150).until(find_focus_hosp)
                #focus_hosp_attempt = 1
                #while (focus_hosp_attempt != 0):
                #    try:
                #        focus_hosp = WebDriverWait(driver_var, 150).until(find_focus_hosp)
                #        focus_hosp_attempt = 0
                #    except StaleElementReferenceException:
                #        if focus_hosp_attempt == 4:
                #            raise
                #        focus_hosp_attempt += 1
                #        time.sleep(0.5)
                
                #UL009 begin
                '''
                focus_hosp = while_loop_handler_function(WebDriverWait(driver_var, 150).until(find_focus_hosp))
                time.sleep(0.5)
                #focus_hosp.click()
                try:
                    focus_hosp.click()
                except:
                    focus_hosp = while_loop_handler_function(WebDriverWait(driver_var, 150).until(find_focus_hosp))
                    time.sleep(0.5)
                    focus_hosp.click()
                time.sleep(0.5)
                Select(focus_hosp).select_by_visible_text('NORTHWESTERN_MEMORIAL 140281')
                '''
                
                driver_var, focus_hosp = click_focus_hosp2(driver_var)
                
                    
                
                try:
                    Select(focus_hosp).select_by_visible_text('NORTHWESTERN_MEMORIAL 140281')
                except:
                    
                    driver_var, focus_hosp = click_focus_hosp2(driver_var)
                    Select(focus_hosp).select_by_visible_text('NORTHWESTERN_MEMORIAL 140281')
                #UL009 end
                
                
                #time.sleep(1)
                #focus_hosp_options = while_loop_handler_function(focus_hosp.find_elements_by_tag_name('option'))
                #time.sleep(0.5)
                #print('third group hosp select')
                #for option in focus_hosp.find_elements_by_tag_name('option'):
                # for option3 in focus_hosp_options:
                #     if option3.text == 'NORTHWESTERN_MEMORIAL 140281':
                #         #print(option.get_attribute("value"))
                #         # option.click()
                #         # time.sleep(0.5)
                #         focus_hosp_click_attempt = 1
                #         while (focus_hosp_click_attempt != 0):
                #             try:
                #                 option3.click()
                #                 focus_hosp_click_attempt = 0
                #             except StaleElementReferenceException:
                #                 if focus_hosp_click_attempt == 4:
                #                     raise
                #                 focus_hosp_click_attempt += 1
                #                 time.sleep(0.5)


                # scroll down
                div_element4 = WebDriverWait(driver_var, 30).until(find_div4_scroll)
                # div_element4 = driver_var.find_element_by_xpath("//div[@id='ctl00_ContentPlaceHolder1_PanelContent5']")
                driver_var.execute_script(
                    "return arguments[0].scrollIntoView(0, document.documentElement.scrollHeight-10);", div_element4)

                time.sleep(0.5)

                # Click 'All available hospitals in the database
                #all_avail_hosp_btn = WebDriverWait(driver_var, 30).until(find_all_avail_hosp)
                # all_avail_hosp_btn = driver_var.find_element_by_id("ctl00_ContentPlaceHolder1_cmdAllAvailHosp")
                #all_avail_hosp_btn.click()

                all_avail_hosp_attempt = 1
                while True:
                    try:
                        all_avail_hosp_btn = WebDriverWait(driver_var, 30).until(find_all_avail_hosp)
                        break
                    except StaleElementReferenceException:
                        if all_avail_hosp_attempt == 3:
                            raise
                        all_avail_hosp_attempt += 1

                all_avail_hosp_click_attempt = 1
                while True:
                    try:
                        all_avail_hosp_btn.click()
                        break
                    except StaleElementReferenceException:
                        if all_avail_hosp_click_attempt == 3:
                            raise
                        all_avail_hosp_click_attempt += 1

                # scroll down
                div_element5 = WebDriverWait(driver_var, 30).until(find_div5_scroll)  # type: object
                # div_element5 = driver_var.find_element_by_xpath("//div[@id='divRiskAdjustment']")
                driver_var.execute_script("return arguments[0].scrollIntoView(0, document.documentElement.scrollHeight-10);", div_element5)

                # ActionChains(driver_var).move_to_element(all_avail_hosp_btn).click().perform()
                time.sleep(2)

                # Click Download button
                # generate_excel_btn = driver_var.find_element_by_id("ctl00_ContentPlaceHolder1_imgExcel")
                """try:
                    element = WebDriverWait(driver_var, 120).until(
                        EC.presence_of_element_located((By.ID, "ctl00_ContentPlaceHolder1_imgExcel"))
                    )
                    print("Excel image clickable.")
                except TimeoutException:
                    print("Time exceeded!")
                """
                #generate_excel_btn = WebDriverWait(driver_var, 10).until(find_excel_btn)
                #generate_excel_btn.click()
                
                #UL006
                
                if remove_covid_pats == True:
                    driver_var = add_covid_advanced_restriction(driver_var)
                    
                time.sleep(0.5)

                excel_btn_attempt = 1
                while True:
                    try:
                        generate_excel_btn = WebDriverWait(driver_var, 10).until(find_excel_btn)
                        break
                    except StaleElementReferenceException:
                        if excel_btn_attempt == 3:
                            raise
                        excel_btn_attempt += 1

                excel_btn_click_attempt = 1
                while True:
                    try:
                        generate_excel_btn.click()
                        break
                    except StaleElementReferenceException:
                        if excel_btn_click_attempt == 3:
                            raise
                        excel_btn_click_attempt += 1
                #webdriver.ActionChains(driver_var).move_to_element(generate_excel_btn).click(generate_excel_btn).perform()
                #ActionChains(driver_var).move_to_element(generate_excel_btn).click().perform()
                time.sleep(1)
                # Accept Pop-up window

                try:
                    WebDriverWait(driver_var, 10).until(EC.alert_is_present(),
                                                       'Timed out waiting for PA creation ' +
                                                       'confirmation popup to appear.')

                    obj = WebDriverWait(driver_var, 10).until(find_alert_popup)

                    obj.accept()
                except:
                    pass
                #print('before download check while loop')
                while len(os.listdir(download_folder_dir)) <= num_already_downloaded_files:
                    time.sleep(1)
                #print('after download check while loop')
                #time.sleep(1.5)
                #print('before window check while loop')
                # wait for the pop up browser window to display so you can close it.
                while len(driver_var.window_handles) < 2:
                    time.sleep(0.5)
                #print('after window check while loop')
                window_before = driver_var.window_handles[0]
                window_after = driver_var.window_handles[1]
                # switch control to popup window and close it
                driver_var.switch_to.window(window_after)
                driver_var.close()
                # Switch control back to the original window.
                driver_var.switch_to.window(window_before)
                time.sleep(2)
                # find most recently-added file to Downloads folder and rename it.
                try:
                    latest_file = find_last_downloaded_file('C:/Users/NM184423/Downloads')
                except:
                    latest_file = find_last_downloaded_file('H:/Downloads')
                latest_file = os.path.abspath(latest_file)

                time.sleep(0.500)

                rename_and_move_file(latest_file, link_dict[i][3], link_dict[i][7], p1, file_dir)
                time.sleep(0.500)
                report_counter += 1
                time.sleep(2)
                update_template_files(hyperlink_loc, link_dict[i][3], link_dict[i][4])
                time.sleep(2)
    return (report_counter)


################################################################################################

# define one function to put it all together.
#UL006  adding new parameter to control whether or not we remove covid patients.
def core_scraper_function(remove_covid = False):
    # step 1: What period is this for?  Enter in the end date of for the Performance Close month.  Import
    #         the time period data and generate a helper dictionary from it.
    print('Choose time period for most measures.')
    period_helper_dict1 = build_period_lookup_dict()

    print('Choose time period for THK measure.')
    period_helper_dict2 = build_period_lookup_dict()

    print('Choose time period for Readmission/Excess Days.')
    period_helper_dict3 = build_period_lookup_dict()

    begin_dts_cm = input("Core Measures:  Enter the beginning datetime (format:  'xx-01-xxxx')")
    end_dts_cm = input("Core Measures:  Enter the end datetime (format:  'xx-xx-xxxx 23:59:59')")

    # step 2:  Import cohort data from the Vizient documentation file.
    cohort_helper_df = gather_cohort_data()

    # step 3:  Import the Vizient template hyperlink file and generate a helper dictionary from it.
    print('before getting report template links')
    hyperlinks_helper_df = get_report_template_links_orig()
    print('after getting report template links')
    print(hyperlinks_helper_df)
    print('after that...')
    print('Bug occurs here somewhere.  Root Cause:  originally, hyperlinks_helper_df as a list containing a df and list of link file names.  for some reason, I changed it.  I do not remember why')
    merged_hyperlink_helper_df = pd.merge(cohort_helper_df, hyperlinks_helper_df[0], on='Hospital')
    hyperlink_helper_dict_final = create_hyperlink_dict(merged_hyperlink_helper_df)

    # step 4:  Create main folder structure to store the files in.
    file_directory_name_helper = create_folder_structure(hyperlink_helper_dict_final)
    # step 5:  Login to Vizient main page
    browser_helper_obj = vizient_login()
    #browser_helper_obj.implicitly_wait(30)
    # step 6:  Loop over hyperlink helper dictionary and time period helper dictionary for every
    #          Vizient cohort and measure hyperlink, download the data, rename the file and store the excel file
    start = time.time()
    #print('final dictionary:')
    #print(hyperlink_helper_dict_final)
    num_reports = loop_template_download(hyperlink_helper_dict_final, period_helper_dict1,period_helper_dict2,period_helper_dict3, browser_helper_obj,list_of_total_revisits_measures,file_directory_name_helper,hyperlinks_helper_df[1],remove_covid)

    get_core_measure_metric_vals(file_directory_name_helper,begin_dts_cm,end_dts_cm)

    end = time.time()
    elapsed_time = end - start

    return ([num_reports, elapsed_time])


################################################################################################
################################################################################################
################################################################################################
################################################################################################

## BELOW:  Functions for wrangling the scraped data into individual excel files sorted by domain and measure ##

################################################################################################
################################################################################################
################################################################################################
################################################################################################


################################################################################################


# takes an openpyxl excel worksheet object and returns the index of the first column with data populated
def find_first_ws_col(worksheet_var):
    for i in range(1,worksheet_var.max_column+1):
        if worksheet_var.cell(row=1,column=i).value is None:
            pass
        else:
            return(i)

################################################################################################


# takes an openpyxl excel worksheet object and returns the index of the first row with data populated
# Vizient Q&A reports typically have 'Hospital' or 'AHRQ Safety' as the first column header
def find_ws_header_row(worksheet_var, first_val_col):
    for i in range(1,worksheet_var.max_row+1):
        if worksheet_var.cell(row=i,column=first_val_col).value == 'Hospital' or worksheet_var.cell(row=i,column=first_val_col).value == 'AHRQ Safety' or worksheet_var.cell(row=i,column=first_val_col).value == 'Encounter Month':
            return(i)

################################################################################################



# calculate EDAC (Excess Days Per 100 Index Encounters) from multiple groupby discharge month rows
def calculate_edac(df):
    #print('calculating edac!')
    #if first column is 'Excess Days', then calculate n events instead.
    if df.columns[0] == 'Excess Days':
        #sum each column
        summed_df = pd.DataFrame({'value sum': df.sum()})
        total_excess_days = summed_df.loc['Excess Days']['value sum']
        data_val = total_excess_days
    elif df.columns[0] == 'Total Index Encounters' and len(df.columns) == 2:
        #sum each column
        summed_df = pd.DataFrame({'value sum': df.sum()})
        total_excess_days = summed_df.loc['Total Index Encounters']['value sum']
        data_val = total_excess_days
    #else calculate the measure value for excess days.
    else:
        #sum each column
        summed_df = pd.DataFrame({'value sum': df.sum()})
        #isolate summed values
        total_index_enc = summed_df.loc['Total Index Encounters']['value sum']
        #print(total_index_enc)
        total_revisit_days = summed_df.loc['Total Revisit Days']['value sum']
        #print(total_revisit_days)
        total_exp_inpat_days = summed_df.loc['Expected Inpatient Days']['value sum']
        #print(total_exp_inpat_days)
        #calculate EDAC.
        #print('before calc.')
        data_val = (total_revisit_days - total_exp_inpat_days)/(total_index_enc/100)
        #print('after calc.')
    return(data_val)

################################################################################################

# calculate READM (PCT HWR Inpatient) from multiple groupby discharge month rows
def calculate_readm(df):
    #print('calculating readm')
    #if first column is 'Total Revisits', then calculate n events instead.
    if df.columns[0] == 'Revisit Inpatient Cases':
        #sum each column
        summed_df = pd.DataFrame({'value sum': df.sum()})
        total_revisits = summed_df.loc['Revisit Inpatient Cases']['value sum']
        data_val = total_revisits
    elif df.columns[0] == 'Total Index Encounters' and len(df.columns) == 2:
        #sum each column
        summed_df = pd.DataFrame({'value sum': df.sum()})
        total_revisits = summed_df.loc['Total Index Encounters']['value sum']
        data_val = total_revisits
    else:
        #sum each column
        summed_df = pd.DataFrame({'value sum': df.sum()})
        #print('EDAC summed df: ')
        #isolate summed values
        revisit_inpatient_cases = summed_df.loc['Revisit Inpatient Cases']['value sum']
        total_index_encounters = summed_df.loc['Total Index Encounters']['value sum']
        #calculate PCT HWR Inpatient.
        data_val = (revisit_inpatient_cases/total_index_encounters) * 100.0
    return(data_val)


################################################################################################


# takes the header row variable from find_ws_header_row and returns a list of column indexes which have data populated.
def find_ws_populated_cols(worksheet_var,header_row):
    cols_with_values = []
    for i in range(1,worksheet_var.max_column+1):
        if worksheet_var.cell(row=header_row,column=i).value is not None:
            cols_with_values.append(i-1)
    return(cols_with_values)

################################################################################################

# Create a lookup dictionary used to wrangle excel files
'''
def create_hyperlink_dict_wrangle(merged_df):
    merged_df = pd.DataFrame(merged_df,columns=['Cohort','Formal Name','Keyword/Metric','Domain'])
    merged_df = merged_df.drop_duplicates()
    merged_df['zipped_data'] = list(zip(merged_df.Cohort,merged_df['Domain'],merged_df['Formal Name'],merged_df['Keyword/Metric']))
    merged_df['zipped_keys'] = list(zip(merged_df.Cohort,merged_df['Keyword/Metric']))
    lookup_data_container = pd.Series(merged_df.zipped_data.values,index=merged_df.zipped_keys.values).to_dict()
    return(lookup_data_container)
'''
 #Create a lookup dictionary used to wrangle excel files
def create_hyperlink_dict_wrangle(merged_df):
    merged_df = pd.DataFrame(merged_df,columns=['Cohort','Formal Name','Keyword/Metric','Domain'])
    merged_df = merged_df.drop_duplicates()
    merged_df['zipped_data'] = list(zip(merged_df.Cohort,merged_df['Domain'],merged_df['Formal Name'],merged_df['Keyword/Metric']))
    merged_df['zipped_keys'] = list(zip(merged_df.Cohort,merged_df['Keyword/Metric']))
    lookup_data_container = pd.Series(merged_df.zipped_data.values,index=merged_df.zipped_keys.values).to_dict()
    return(lookup_data_container)
################################################################################################

# create folder structure to store final wrangled excel files in
def create_final_excel_folder_structure():
    # create main folder

    new_dir_path = input('Enter path of location where you want to store the files.')
    folder_name = 'Final Excel Files'
    new_dir_path = os.path.join(os.path.abspath(new_dir_path), folder_name)
    try:
        # check if folder already exists.  If it does not exist, create it.
        if os.path.isfile(new_dir_path) == False:
            os.mkdir(new_dir_path)
    except:
        pass
    return (new_dir_path)

################################################################################################

# function puts together the above defined functions.

# Step 1:  Takes a file path object for folder full of Vizient Q&A reports and file name list object.
# Step 2:  Join the path and the file name
# Step 3:  Open the excel file using openpyxl
# Step 4:  Open the first worksheet of the workbook
# Step 5:  Find first column with data, find first row with data
# Step 6:  Read in the data into a pandas dataframe

def open_excel_file(path_obj, file_list_obj):
    #Create filename path in order to open the excel file
    file_loc = os.path.join(os.path.abspath(path_obj),file_list_obj[0])
    dirname = os.path.dirname(file_loc)
    #save the excel workbook object in a variable
    wb = openpyxl.load_workbook(file_loc)
    #save the excel worksheet object in a variable
    wb_sheetnames = wb.sheetnames
    #take the first sheet in the workbook
    ws = wb[wb_sheetnames[0]]
    #find the first column number with a value in it
    first_col = find_first_ws_col(ws)
    #find the first row number with 'Hospital' or 'AHRQ Safety' indicating the header row of the Vizient files
    header_row = find_ws_header_row(ws,first_col)
    #create a list of all columns which have values
    populated_columns = find_ws_populated_cols(ws,header_row)
    #parse the Excel spreadsheet to create a pandas dataframe
    #UL007
    xlsx_file = pd.DataFrame(pd.read_excel(file_loc,sheet_name=wb_sheetnames[0],skiprows=header_row-1,usecols=populated_columns,engine='openpyxl'))
    return(xlsx_file)


################################################################################################

# filter down to NM hospitals and extract data values.  If a value is missing, fill in the row with a 0 or 'LV'.
def filter_func(excel, dictionary, keys, row_filter_criteria, colname, measure_name, edac_readm_choice_num,
                edac_readm_month_list='No List'):
    if measure_name in ['PSI_03', 'PSI_06', 'PSI_09', 'PSI_11', 'PSI_13']:
        try:
            hosp_index = excel.index[row_filter_criteria].tolist()
            data_val = excel.iloc[hosp_index[0]][colname]
        except:
            data_val = 0
        row_data = [dictionary[keys][1], dictionary[keys][2], data_val]
        return (row_data)

    # if measure is EDAC or READMISSION, there will be multiple columns due to multiple group by discharge month.
    elif measure_name.split('_')[0] in ['EDAC', 'READM']:
        try:
            # Filter rows to the correct hospital and/or metric
            hosp_index = excel.index[row_filter_criteria].tolist()
            # filter columns to correct metric columns
            data_val = excel.iloc[hosp_index][colname]
            # filter custom months if requested
            if edac_readm_choice_num == 2:
                data_val['Encounter Month'] = data_val['Encounter Month'].str.strip()
                data_val = data_val[data_val['Encounter Month'].isin(edac_readm_month_list)]
            else:
                pass
            # calculate EDAC
            if measure_name.split('_')[0] in ['EDAC']:
                data_val = calculate_edac(data_val)
            # calculate READM
            elif measure_name.split('_')[0] in ['READM']:
                data_val = calculate_readm(data_val)
        except:
            data_val = 'LV'
        row_data = [dictionary[keys][1], dictionary[keys][2], data_val]
        return (row_data)

    else:
        try:
            hosp_index = excel.index[row_filter_criteria].tolist()
            data_val = excel.iloc[hosp_index[0]][colname]
        except:
            data_val = 'LV'
        row_data = [dictionary[keys][1], dictionary[keys][2], data_val]
        return (row_data)

################################################################################################

# For each CCMC hospital, filter down to NM hospitals and extract data values.  If a value is missing, fill in the row with a 0 or 'LV'.
def filter_func_ccmc(excel, dictionary, keys, criteria_list, colname, measure_name, edac_readm_choice_num,
                     edac_readm_month_list='No List'):
    result_list = []
    for i in criteria_list:
        if measure_name in ['PSI_03', 'PSI_06', 'PSI_09', 'PSI_11', 'PSI_13']:

            if measure_name == 'PSI_03':
                try:
                    hosp_index = excel.index[
                        (excel['Hospital'] == i) & (excel['AHRQ Safety'] == 'PSI03 Pressure Ulcer Rate')].tolist()
                    data_val = excel.iloc[hosp_index[0]][colname]
                except:
                    data_val = 0


            elif measure_name == 'PSI_06':
                try:
                    hosp_index = excel.index[(excel['Hospital'] == i) & (
                            excel['AHRQ Safety'] == 'PSI06 Iatrogenic Pneumothorax Rate')].tolist()
                    data_val = excel.iloc[hosp_index[0]][colname]
                except:
                    data_val = 0

            elif measure_name == 'PSI_09':
                try:
                    hosp_index = excel.index[(excel['Hospital'] == i) & (
                            #excel['AHRQ Safety'] == 'PSI09 Perioperative Hemorrhage or Hematoma Rate')].tolist()   #UL010
                            excel['AHRQ Safety'] == 'PSI09 Postoperative Hemorrhage or Hematoma Rate')].tolist()    #UL010
                    data_val = excel.iloc[hosp_index[0]][colname]
                except:
                    data_val = 0

            elif measure_name == 'PSI_11':
                try:
                    hosp_index = excel.index[(excel['Hospital'] == i) & (
                            excel['AHRQ Safety'] == 'PSI11 Postoperative Respiratory Failure Rate')].tolist()
                    data_val = excel.iloc[hosp_index[0]][colname]
                except:
                    data_val = 0

            elif measure_name == 'PSI_13':
                try:
                    hosp_index = excel.index[
                        (excel['Hospital'] == i) & (excel['AHRQ Safety'] == 'PSI13 Postoperative Sepsis Rate')].tolist()
                    data_val = excel.iloc[hosp_index[0]][colname]
                except:
                    data_val = 0

            row_data = [i, dictionary[keys][1], dictionary[keys][2], data_val]
            result_list.append(row_data)

        # if measure is EDAC or READMISSION, there will be multiple columns due to multiple group by discharge month.
        elif measure_name.split('_')[0] in ['EDAC', 'READM']:
            try:
                # Filter rows to the correct hospital and/or metric
                hosp_index = excel.index[excel['Hospital'] == i].tolist()
                # filter columns to correct metric columns
                data_val = excel.iloc[hosp_index][colname]
                print(data_val)
                # filter custom months if requested
                if edac_readm_choice_num == 2:
                    data_val['Encounter Month'] = data_val['Encounter Month'].str.strip()
                    data_val = data_val[data_val['Encounter Month'].isin(edac_readm_month_list)]
                else:
                    pass
                # calculate EDAC
                if measure_name.split('_')[0] in ['EDAC']:
                    data_val = calculate_edac(data_val)
                # calculate READM
                elif measure_name.split('_')[0] in ['READM']:
                    data_val = calculate_readm(data_val)
            except:
                data_val = 'LV'
            row_data = [i, dictionary[keys][1], dictionary[keys][2], data_val]
            result_list.append(row_data)

        else:
            try:
                hosp_index = excel.index[excel['Hospital'] == i].tolist()
                data_val = excel.iloc[hosp_index[0]][colname]
            except:
                data_val = 'LV'
            row_data = [i, dictionary[keys][1], dictionary[keys][2], data_val]
            result_list.append(row_data)
    return (result_list)

################################################################################################
'''
# For each CCMC hospital, filter down to NM hospitals and extract data values.  If a value is missing, fill in the row with a 0 or 'LV'.
def filter_func_all_cohort_hosps(excel, dictionary, keys, criteria_list, colname, measure_name, edac_readm_choice_num,
                                 edac_readm_month_list='No List'):
    result_list = []
    excel['Hospital'] = excel['Hospital'].str.upper()
    excel['Hospital'] = excel['Hospital']
    for i in criteria_list:
        i = i.upper()
        # split the medicare ID from the hospital name because Vizient will periodically update hospital names
        # filter using the medicare ID instead, which should be more reliable.
        i2 = i.split(" ", 1)
        #print('Hospital: ', i)
        if measure_name in ['PSI_03', 'PSI_06', 'PSI_09', 'PSI_11', 'PSI_13']:

            if measure_name == 'PSI_03':
                try:

                    hosp_index = excel.index[
                        (excel['Hospital'].str.contains(i2[0])) & (
                                    excel['AHRQ Safety'] == 'PSI03 Pressure Ulcer Rate')].tolist()
                    # hosp_index = excel.index[
                    #    (excel['Hospital'] == i) & (excel['AHRQ Safety'] == 'PSI03 Pressure Ulcer Rate')].tolist()
                    data_val = excel.iloc[hosp_index[0]][colname]

                    # Sometimes there are NaN.  We do not want blank/NaN.  Instead impute 'Missing'
                    if math.isnan(data_val) == True or data_val == '':
                        data_val = 'Missing'
                except:
                    data_val = 'Missing'


            elif measure_name == 'PSI_06':
                try:
                    hosp_index = excel.index[(excel['Hospital'].str.contains(i2[0])) & (
                            excel['AHRQ Safety'] == 'PSI06 Iatrogenic Pneumothorax Rate')].tolist()
                    # hosp_index = excel.index[(excel['Hospital'] == i) & (
                    #            excel['AHRQ Safety'] == 'PSI06 Iatrogenic Pneumothorax Rate')].tolist()
                    data_val = excel.iloc[hosp_index[0]][colname]

                    # Sometimes there are NaN.  We do not want blank/NaN.  Instead impute 'Missing'
                    if math.isnan(data_val) == True or data_val == '':
                        data_val = 'Missing'
                except:
                    data_val = 'Missing'

            elif measure_name == 'PSI_09':
                try:
                    hosp_index = excel.index[(excel['Hospital'].str.contains(i2[0])) & (
                            #excel['AHRQ Safety'] == 'PSI09 Perioperative Hemorrhage or Hematoma Rate')].tolist()   #UL010
                            excel['AHRQ Safety'] == 'PSI09 Postoperative Hemorrhage or Hematoma Rate')].tolist()    #UL010
                    # hosp_index = excel.index[(excel['Hospital'] == i) & (
                    #            excel['AHRQ Safety'] == 'PSI09 Perioperative Hemorrhage or Hematoma Rate')].tolist()
                    data_val = excel.iloc[hosp_index[0]][colname]

                    # Sometimes there are NaN.  We do not want blank/NaN.  Instead impute 'Missing'
                    if math.isnan(data_val) == True or data_val == '':
                        data_val = 'Missing'
                except:
                    data_val = 'Missing'

            elif measure_name == 'PSI_11':
                try:
                    hosp_index = excel.index[(excel['Hospital'].str.contains(i2[0])) & (
                            excel['AHRQ Safety'] == 'PSI11 Postoperative Respiratory Failure Rate')].tolist()
                    # hosp_index = excel.index[(excel['Hospital'] == i) & (
                    #            excel['AHRQ Safety'] == 'PSI11 Postoperative Respiratory Failure Rate')].tolist()
                    data_val = excel.iloc[hosp_index[0]][colname]

                    # Sometimes there are NaN.  We do not want blank/NaN.  Instead impute 'Missing'
                    if math.isnan(data_val) == True or data_val == '':
                        data_val = 'Missing'
                except:
                    data_val = 'Missing'

            elif measure_name == 'PSI_13':
                try:
                    hosp_index = excel.index[
                        (excel['Hospital'].str.contains(i2[0])) & (
                                    excel['AHRQ Safety'] == 'PSI13 Postoperative Sepsis Rate')].tolist()
                    # hosp_index = excel.index[
                    #    (excel['Hospital'] == i) & (excel['AHRQ Safety'] == 'PSI13 Postoperative Sepsis Rate')].tolist()
                    data_val = excel.iloc[hosp_index[0]][colname]

                    # Sometimes there are NaN.  We do not want blank/NaN.  Instead impute 'Missing'
                    if math.isnan(data_val) == True or data_val == '':
                        data_val = 'Missing'

                except:
                    data_val = 'Missing'

            row_data = [i, dictionary[keys][1], dictionary[keys][2], data_val]
            result_list.append(row_data)

        # if measure is EDAC or READMISSION, there will be multiple columns due to multiple group by discharge month.
        elif measure_name.split('_')[0] in ['EDAC',
                                            'READM'] and colname != 'Pct Total Readmit' and colname != 'Readmit Rate Num Cases (Readmit Cases)' and colname != 'Readmit Rate Denom Cases':

            try:
                # Filter rows to the correct hospital and/or metric
                hosp_index = excel.index[excel['Hospital'].str.contains(i2[0])].tolist()

                data_val = excel.iloc[hosp_index][colname]

                if data_val.empty == False:

                    # filter custom months if requested
                    if edac_readm_choice_num == 2:

                        data_val['Encounter Month'] = data_val['Encounter Month'].str.strip()
                        data_val = data_val[data_val['Encounter Month'].isin(edac_readm_month_list)]
                    else:
                        pass

                    # Only calculate EDAC/READM if there are values.  Sometimes hospitals have
                    # rows but not the specific month we are looking for.  If there are no rows
                    # then just set to 'Missing.'
                    if data_val.empty == False:
                        # calculate EDAC
                        if measure_name.split('_')[0] in ['EDAC']:
                            data_val = calculate_edac(data_val)
                        # calculate READM
                        elif measure_name.split('_')[0] in ['READM']:
                            data_val = calculate_readm(data_val)
                    else:
                        data_val = 'Missing'

                    # Sometimes there are NaN.  We do not want blank/NaN.  Instead impute 'Missing'
                    if math.isnan(data_val) == True or data_val == '':
                        data_val = 'Missing'
                else:
                    data_val = 'Missing'
            except:
                data_val = 'Missing'
            row_data = [i, dictionary[keys][1], dictionary[keys][2], data_val]
            result_list.append(row_data)

        else:
            try:
                hosp_index = excel.index[excel['Hospital'].str.contains(i2[0])].tolist()
                # hosp_index = excel.index[excel['Hospital'] == i].tolist()
                data_val = excel.iloc[hosp_index[0]][colname]
                # Sometimes there are NaN.  We do not want blank/NaN.  Instead impute 'Missing'
                if math.isnan(data_val) == True or data_val == '':
                    data_val = 'Missing'

            except:
                data_val = 'Missing'
            row_data = [i, dictionary[keys][1], dictionary[keys][2], data_val]
            result_list.append(row_data)
    return (result_list)
'''
#UL003

def filter_func_all_cohort_hosps(excel, dictionary, keys, criteria_list, colname, measure_name, edac_readm_choice_num,
                                 edac_readm_month_list='No List'):
    result_list = []
    excel['Hospital'] = excel['Hospital'].str.upper()
    excel['Hospital'] = excel['Hospital']
    for i in criteria_list:
        i = i.upper()
        # split the medicare ID from the hospital name because Vizient will periodically update hospital names
        # filter using the medicare ID instead, which should be more reliable.
        i2 = i.split(" ", 1)
        #print('Hospital: ', i)
        if measure_name in ['PSI_03', 'PSI_06', 'PSI_09', 'PSI_11', 'PSI_13']:

            if measure_name == 'PSI_03':
                try:

                    hosp_index = excel.index[
                        (excel['Hospital'].str.contains(i2[0])) & (
                                    excel['AHRQ Safety'] == 'PSI03 Pressure Ulcer Rate')].tolist()
                    # hosp_index = excel.index[
                    #    (excel['Hospital'] == i) & (excel['AHRQ Safety'] == 'PSI03 Pressure Ulcer Rate')].tolist()
                    data_val = excel.iloc[hosp_index[0]][colname]

                    # Sometimes there are NaN.  We do not want blank/NaN.  Instead impute 'Missing'
                    if math.isnan(data_val) == True or data_val == '':
                        data_val = 'Missing'
                except:
                    data_val = 'Missing'


            elif measure_name == 'PSI_06':
                try:
                    hosp_index = excel.index[(excel['Hospital'].str.contains(i2[0])) & (
                            excel['AHRQ Safety'] == 'PSI06 Iatrogenic Pneumothorax Rate')].tolist()
                    # hosp_index = excel.index[(excel['Hospital'] == i) & (
                    #            excel['AHRQ Safety'] == 'PSI06 Iatrogenic Pneumothorax Rate')].tolist()
                    data_val = excel.iloc[hosp_index[0]][colname]

                    # Sometimes there are NaN.  We do not want blank/NaN.  Instead impute 'Missing'
                    if math.isnan(data_val) == True or data_val == '':
                        data_val = 'Missing'
                except:
                    data_val = 'Missing'

            elif measure_name == 'PSI_09':
                try:
                    hosp_index = excel.index[(excel['Hospital'].str.contains(i2[0])) & (
                            #excel['AHRQ Safety'] == 'PSI09 Perioperative Hemorrhage or Hematoma Rate')].tolist()   #UL010
                            excel['AHRQ Safety'] == 'PSI09 Postoperative Hemorrhage or Hematoma Rate')].tolist()    
                    # hosp_index = excel.index[(excel['Hospital'] == i) & (
                    #            excel['AHRQ Safety'] == 'PSI09 Perioperative Hemorrhage or Hematoma Rate')].tolist()
                    data_val = excel.iloc[hosp_index[0]][colname]

                    # Sometimes there are NaN.  We do not want blank/NaN.  Instead impute 'Missing'
                    if math.isnan(data_val) == True or data_val == '':
                        data_val = 'Missing'
                except:
                    data_val = 'Missing'

            elif measure_name == 'PSI_11':
                try:
                    hosp_index = excel.index[(excel['Hospital'].str.contains(i2[0])) & (
                            excel['AHRQ Safety'] == 'PSI11 Postoperative Respiratory Failure Rate')].tolist()
                    # hosp_index = excel.index[(excel['Hospital'] == i) & (
                    #            excel['AHRQ Safety'] == 'PSI11 Postoperative Respiratory Failure Rate')].tolist()
                    data_val = excel.iloc[hosp_index[0]][colname]

                    # Sometimes there are NaN.  We do not want blank/NaN.  Instead impute 'Missing'
                    if math.isnan(data_val) == True or data_val == '':
                        data_val = 'Missing'
                except:
                    data_val = 'Missing'

            elif measure_name == 'PSI_13':
                try:
                    hosp_index = excel.index[
                        (excel['Hospital'].str.contains(i2[0])) & (
                                    excel['AHRQ Safety'] == 'PSI13 Postoperative Sepsis Rate')].tolist()
                    # hosp_index = excel.index[
                    #    (excel['Hospital'] == i) & (excel['AHRQ Safety'] == 'PSI13 Postoperative Sepsis Rate')].tolist()
                    data_val = excel.iloc[hosp_index[0]][colname]

                    # Sometimes there are NaN.  We do not want blank/NaN.  Instead impute 'Missing'
                    if math.isnan(data_val) == True or data_val == '':
                        data_val = 'Missing'

                except:
                    data_val = 'Missing'

            row_data = [i, dictionary[keys][1], dictionary[keys][2], data_val]
            result_list.append(row_data)

        # if measure is EDAC or READMISSION, there will be multiple columns due to multiple group by discharge month.
        elif measure_name.split('_')[0] in ['EDAC',
                                            'READM'] and colname != 'Pct Total Readmit' and colname != 'Readmit Rate Num Cases (Readmit Cases)' and colname != 'Readmit Rate Denom Cases':

            try:
                # Filter rows to the correct hospital and/or metric
                hosp_index = excel.index[excel['Hospital'].str.contains(i2[0])].tolist()

                data_val = excel.iloc[hosp_index][colname]

                if data_val.empty == False:

                    # filter custom months if requested
                    if edac_readm_choice_num == 2:

                        data_val['Encounter Month'] = data_val['Encounter Month'].str.strip()
                        data_val = data_val[data_val['Encounter Month'].isin(edac_readm_month_list)]
                    else:
                        pass

                    # Only calculate EDAC/READM if there are values.  Sometimes hospitals have
                    # rows but not the specific month we are looking for.  If there are no rows
                    # then just set to 'Missing.'
                    if data_val.empty == False:
                        # calculate EDAC
                        if measure_name.split('_')[0] in ['EDAC']:
                            data_val = calculate_edac(data_val)
                        # calculate READM
                        elif measure_name.split('_')[0] in ['READM']:
                            data_val = calculate_readm(data_val)
                    else:
                        data_val = 'Missing'

                    # Sometimes there are NaN.  We do not want blank/NaN.  Instead impute 'Missing'
                    if math.isnan(data_val) == True or data_val == '':
                        data_val = 'Missing'
                else:
                    data_val = 'Missing'
            except:
                data_val = 'Missing'
            row_data = [i, dictionary[keys][1], dictionary[keys][2], data_val]
            result_list.append(row_data)

        else:
            try:
                hosp_index = excel.index[excel['Hospital'].str.contains(i2[0])].tolist()
                # hosp_index = excel.index[excel['Hospital'] == i].tolist()
                data_val = excel.iloc[hosp_index[0]][colname]
                # Sometimes there are NaN.  We do not want blank/NaN.  Instead impute 'Missing'
                if math.isnan(data_val) == True or data_val == '':
                    data_val = 'Missing'

            except:
                data_val = 'Missing'
            #print(data_val)    
            #UL003
            if dictionary[keys][2] in ['Adverse Drug Events Rate','% Early Transfers Out']:
                row_data = [i, dictionary[keys][1], measure_name, data_val]
            else:
                row_data = [i, dictionary[keys][1], dictionary[keys][2], data_val]

            result_list.append(row_data)
            #print(row_data)

    return (result_list)



################################################################################################

# Conditionally implement filter functions depending on Hospital cohort and domain.
# Require conditions in order to select correct data column and filter rows correctly.

def file_filter_switcher(cohort_nm, domain_nm, measure_nm, excel_file_name, lookup_dict, keyname, edac_readm_choice_num,
                         edac_readm_month_list='No List'):
    ccmc_hosp_list = ['140286 NORTHWESTERN_KISH', '140211 NORTHWESTERN_DELNOR', '140130 NORTHWESTERN_LAKEFOREST']

    if cohort_nm == 'Comprehensive Academic Medical Center' and domain_nm == 'Mortality':

        results = filter_func(excel_file_name, lookup_dict, keyname,
                              excel_file_name['Hospital'] == '140281 NORTHWESTERN_MEMORIAL', 'Mortality Index',
                              measure_nm, edac_readm_choice_num)
        return (results)

    elif cohort_nm == 'Comprehensive Academic Medical Center' and domain_nm == 'Effectiveness':

        if measure_nm.split('_')[0] == 'EDAC':

            if edac_readm_choice_num == 1:
                results = filter_func(excel_file_name, lookup_dict, keyname,
                                      excel_file_name['Hospital'] == '140281 NORTHWESTERN_MEMORIAL',
                                      ['Total Index Encounters', 'Total Revisit Days', 'Expected Inpatient Days',
                                       'Encounter Month'], measure_nm, edac_readm_choice_num)
                # 'Excess Days Per 100 Index Encounters', measure_nm)
            else:
                results = filter_func(excel_file_name, lookup_dict, keyname,
                                      excel_file_name['Hospital'] == '140281 NORTHWESTERN_MEMORIAL',
                                      ['Total Index Encounters', 'Total Revisit Days', 'Expected Inpatient Days',
                                       'Encounter Month'], measure_nm, edac_readm_choice_num, edac_readm_month_list)
                # 'Excess Days Per 100 Index Encounters', measure_nm)
            return (results)

        elif measure_nm.split('_')[0] == 'READM':

            if edac_readm_choice_num == 1:
                results = filter_func(excel_file_name, lookup_dict, keyname,
                                      excel_file_name['Hospital'] == '140281 NORTHWESTERN_MEMORIAL',
                                      ['Encounter Month', 'Revisit Inpatient Cases', 'Total Index Encounters'],
                                      # excel_file_name['Hospital'] == '140281 NORTHWESTERN_MEMORIAL', 'PCT HWR Inpatient',
                                      measure_nm, edac_readm_choice_num)
            else:
                results = filter_func(excel_file_name, lookup_dict, keyname,
                                      excel_file_name['Hospital'] == '140281 NORTHWESTERN_MEMORIAL',
                                      ['Encounter Month', 'Revisit Inpatient Cases', 'Total Index Encounters'],
                                      # excel_file_name['Hospital'] == '140281 NORTHWESTERN_MEMORIAL', 'PCT HWR Inpatient',
                                      measure_nm, edac_readm_choice_num, edac_readm_month_list)
            return (results)

    elif cohort_nm == 'Comprehensive Academic Medical Center' and domain_nm == 'Efficiency':
        if measure_nm.split('_')[0] == 'DCOST':

            results = filter_func(excel_file_name, lookup_dict, keyname,
                                  excel_file_name['Hospital'] == '140281 NORTHWESTERN_MEMORIAL', 'Direct Cost Index',
                                  measure_nm, edac_readm_choice_num)
            return (results)

        elif measure_nm.split('_')[0] == 'LOS':

            results = filter_func(excel_file_name, lookup_dict, keyname,
                                  excel_file_name['Hospital'] == '140281 NORTHWESTERN_MEMORIAL', 'LOS Index',
                                  measure_nm, edac_readm_choice_num)
            return (results)

    elif cohort_nm == 'Comprehensive Academic Medical Center' and domain_nm == 'Safety':
        if measure_nm.split('_')[0] == 'PSI':
            if measure_nm == 'PSI_03':

                results = filter_func(excel_file_name, lookup_dict, keyname,
                                      (excel_file_name['Hospital'] == '140281 NORTHWESTERN_MEMORIAL') & (
                                              excel_file_name['AHRQ Safety'] == 'PSI03 Pressure Ulcer Rate'),
                                      'O/E Ratio', measure_nm, edac_readm_choice_num)
                return (results)

            elif measure_nm == 'PSI_06':

                results = filter_func(excel_file_name, lookup_dict, keyname,
                                      (excel_file_name['Hospital'] == '140281 NORTHWESTERN_MEMORIAL') & (
                                              excel_file_name[
                                                  'AHRQ Safety'] == 'PSI06 Iatrogenic Pneumothorax Rate'),
                                      'O/E Ratio', measure_nm, edac_readm_choice_num)
                return (results)

            elif measure_nm == 'PSI_09':

                results = filter_func(excel_file_name, lookup_dict, keyname,
                                      (excel_file_name['Hospital'] == '140281 NORTHWESTERN_MEMORIAL') & (
                                              excel_file_name[
                                                  #'AHRQ Safety'] == 'PSI09 Perioperative Hemorrhage or Hematoma Rate'),   #UL010
                                                  'AHRQ Safety'] == 'PSI09 Postoperative Hemorrhage or Hematoma Rate'),    #UL010
                                      'O/E Ratio', measure_nm, edac_readm_choice_num)
                return (results)

            elif measure_nm == 'PSI_11':

                results = filter_func(excel_file_name, lookup_dict, keyname,
                                      (excel_file_name['Hospital'] == '140281 NORTHWESTERN_MEMORIAL') & (
                                              excel_file_name[
                                                  'AHRQ Safety'] == 'PSI11 Postoperative Respiratory Failure Rate'),
                                      'O/E Ratio', measure_nm, edac_readm_choice_num)
                return (results)

            elif measure_nm == 'PSI_13':

                results = filter_func(excel_file_name, lookup_dict, keyname,
                                      (excel_file_name['Hospital'] == '140281 NORTHWESTERN_MEMORIAL') & (
                                              excel_file_name['AHRQ Safety'] == 'PSI13 Postoperative Sepsis Rate'),
                                      'O/E Ratio', measure_nm, edac_readm_choice_num)
                return (results)

        elif measure_nm == 'THK':

            results = filter_func(excel_file_name, lookup_dict, keyname,
                                  (excel_file_name['Hospital'] == '140281 NORTHWESTERN_MEMORIAL'), 'Rate', measure_nm,
                                  edac_readm_choice_num)
            return (results)

    elif cohort_nm == 'Comprehensive Academic Medical Center' and domain_nm == 'Outpatient':

        results = filter_func(excel_file_name, lookup_dict, keyname,
                              (excel_file_name['Hospital'] == '140281 NORTHWESTERN_MEMORIAL'), 'Pct Total Readmit',
                              measure_nm, edac_readm_choice_num)
        return (results)

    # LSCCMC logic:
    elif cohort_nm == 'Large Specialized Complex Care Medical Center' and domain_nm == 'Mortality':

        results = filter_func(excel_file_name, lookup_dict, keyname,
                              excel_file_name['Hospital'] == '140242 NORTHWESTERN_CDH', 'Mortality Index', measure_nm,
                              edac_readm_choice_num)
        return (results)

    elif cohort_nm == 'Large Specialized Complex Care Medical Center' and domain_nm == 'Effectiveness':

        if measure_nm.split('_')[0] == 'EDAC':

            if edac_readm_choice_num == 1:
                results = filter_func(excel_file_name, lookup_dict, keyname,
                                      excel_file_name['Hospital'] == '140242 NORTHWESTERN_CDH',
                                      ['Total Index Encounters', 'Total Revisit Days', 'Expected Inpatient Days',
                                       'Encounter Month'], measure_nm, edac_readm_choice_num)
                # 'Excess Days Per 100 Index Encounters', measure_nm)
            else:
                results = filter_func(excel_file_name, lookup_dict, keyname,
                                      excel_file_name['Hospital'] == '140242 NORTHWESTERN_CDH',
                                      ['Total Index Encounters', 'Total Revisit Days', 'Expected Inpatient Days',
                                       'Encounter Month'], measure_nm, edac_readm_choice_num, edac_readm_month_list)
                # 'Excess Days Per 100 Index Encounters', measure_nm)
            return (results)

        elif measure_nm.split('_')[0] == 'READM':

            if edac_readm_choice_num == 1:
                results = filter_func(excel_file_name, lookup_dict, keyname,
                                      excel_file_name['Hospital'] == '140242 NORTHWESTERN_CDH',
                                      ['Encounter Month', 'Revisit Inpatient Cases', 'Total Index Encounters'],
                                      # excel_file_name['Hospital'] == '140281 NORTHWESTERN_MEMORIAL', 'PCT HWR Inpatient',
                                      measure_nm, edac_readm_choice_num)
            else:
                results = filter_func(excel_file_name, lookup_dict, keyname,
                                      excel_file_name['Hospital'] == '140242 NORTHWESTERN_CDH',
                                      ['Encounter Month', 'Revisit Inpatient Cases', 'Total Index Encounters'],
                                      # excel_file_name['Hospital'] == '140281 NORTHWESTERN_MEMORIAL', 'PCT HWR Inpatient',
                                      measure_nm, edac_readm_choice_num, edac_readm_month_list)
            return (results)

    elif cohort_nm == 'Large Specialized Complex Care Medical Center' and domain_nm == 'Efficiency':
        if measure_nm.split('_')[0] == 'DCOST':

            results = filter_func(excel_file_name, lookup_dict, keyname,
                                  excel_file_name['Hospital'] == '140242 NORTHWESTERN_CDH', 'Direct Cost Index',
                                  measure_nm, edac_readm_choice_num)
            return (results)

        elif measure_nm.split('_')[0] == 'LOS':

            results = filter_func(excel_file_name, lookup_dict, keyname,
                                  excel_file_name['Hospital'] == '140242 NORTHWESTERN_CDH', 'LOS Index', measure_nm,
                                  edac_readm_choice_num)
            return (results)

    elif cohort_nm == 'Large Specialized Complex Care Medical Center' and domain_nm == 'Safety':
        if measure_nm.split('_')[0] == 'PSI':
            if measure_nm == 'PSI_03':

                results = filter_func(excel_file_name, lookup_dict, keyname,
                                      (excel_file_name['Hospital'] == '140242 NORTHWESTERN_CDH') & (
                                              excel_file_name['AHRQ Safety'] == 'PSI03 Pressure Ulcer Rate'),
                                      'O/E Ratio', measure_nm, edac_readm_choice_num)
                return (results)

            elif measure_nm == 'PSI_06':

                results = filter_func(excel_file_name, lookup_dict, keyname,
                                      (excel_file_name['Hospital'] == '140242 NORTHWESTERN_CDH') & (excel_file_name[
                                                                                                        'AHRQ Safety'] == 'PSI06 Iatrogenic Pneumothorax Rate'),
                                      'O/E Ratio', measure_nm, edac_readm_choice_num)
                return (results)

            elif measure_nm == 'PSI_09':

                results = filter_func(excel_file_name, lookup_dict, keyname,
                                      (excel_file_name['Hospital'] == '140242 NORTHWESTERN_CDH') & (excel_file_name[
                                                                                                        #'AHRQ Safety'] == 'PSI09 Perioperative Hemorrhage or Hematoma Rate'),   #UL010
                                                                                                        'AHRQ Safety'] == 'PSI09 Postoperative Hemorrhage or Hematoma Rate'),    #UL010
                                      'O/E Ratio', measure_nm, edac_readm_choice_num)
                return (results)

            elif measure_nm == 'PSI_11':

                results = filter_func(excel_file_name, lookup_dict, keyname,
                                      (excel_file_name['Hospital'] == '140242 NORTHWESTERN_CDH') & (excel_file_name[
                                                                                                        'AHRQ Safety'] == 'PSI11 Postoperative Respiratory Failure Rate'),
                                      'O/E Ratio', measure_nm, edac_readm_choice_num)
                return (results)

            elif measure_nm == 'PSI_13':

                results = filter_func(excel_file_name, lookup_dict, keyname,
                                      (excel_file_name['Hospital'] == '140242 NORTHWESTERN_CDH') & (
                                              excel_file_name['AHRQ Safety'] == 'PSI13 Postoperative Sepsis Rate'),
                                      'O/E Ratio', measure_nm, edac_readm_choice_num)
                return (results)

        elif measure_nm == 'THK':

            results = filter_func(excel_file_name, lookup_dict, keyname,
                                  (excel_file_name['Hospital'] == '140242 NORTHWESTERN_CDH'), 'Rate', measure_nm,
                                  edac_readm_choice_num)
            return (results)


    elif cohort_nm == 'Large Specialized Complex Care Medical Center' and domain_nm == 'Outpatient':

        results = filter_func(excel_file_name, lookup_dict, keyname,
                              (excel_file_name['Hospital'] == '140242 NORTHWESTERN_CDH'), 'Pct Total Readmit',
                              measure_nm, edac_readm_choice_num)
        return (results)
    # CCMC Logic:
    elif cohort_nm == 'Complex Care Medical Center' and domain_nm == 'Mortality':
        # print('got to ccmc')

        results = filter_func_ccmc(excel_file_name, lookup_dict, keyname, ccmc_hosp_list, 'Mortality Index', measure_nm,
                                   edac_readm_choice_num)
        return (results)

    elif cohort_nm == 'Complex Care Medical Center' and domain_nm == 'Effectiveness':

        if measure_nm.split('_')[0] == 'EDAC':

            if edac_readm_choice_num == 1:
                results = filter_func_ccmc(excel_file_name, lookup_dict, keyname, ccmc_hosp_list,
                                           ['Total Index Encounters', 'Total Revisit Days', 'Expected Inpatient Days',
                                            'Encounter Month'], measure_nm, edac_readm_choice_num)
                # 'Excess Days Per 100 Index Encounters', measure_nm)
            else:
                results = filter_func_ccmc(excel_file_name, lookup_dict, keyname, ccmc_hosp_list,
                                           ['Total Index Encounters', 'Total Revisit Days', 'Expected Inpatient Days',
                                            'Encounter Month'], measure_nm, edac_readm_choice_num,
                                           edac_readm_month_list)
                # 'Excess Days Per 100 Index Encounters', measure_nm)

            return (results)

        elif measure_nm.split('_')[0] == 'READM':

            if edac_readm_choice_num == 1:
                results = filter_func_ccmc(excel_file_name, lookup_dict, keyname, ccmc_hosp_list,
                                           ['Encounter Month', 'Revisit Inpatient Cases', 'Total Index Encounters'],
                                           measure_nm, edac_readm_choice_num)
            else:
                results = filter_func_ccmc(excel_file_name, lookup_dict, keyname, ccmc_hosp_list,
                                           ['Encounter Month', 'Revisit Inpatient Cases', 'Total Index Encounters'],
                                           measure_nm, edac_readm_choice_num, edac_readm_month_list)
            return (results)

    elif cohort_nm == 'Complex Care Medical Center' and domain_nm == 'Efficiency':
        if measure_nm.split('_')[0] == 'DCOST':

            results = filter_func_ccmc(excel_file_name, lookup_dict, keyname, ccmc_hosp_list, 'Direct Cost Index',
                                       measure_nm, edac_readm_choice_num)
            return (results)

        elif measure_nm.split('_')[0] == 'LOS':

            results = filter_func_ccmc(excel_file_name, lookup_dict, keyname, ccmc_hosp_list, 'LOS Index', measure_nm,
                                       edac_readm_choice_num)
            return (results)

    elif cohort_nm == 'Complex Care Medical Center' and domain_nm == 'Safety':
        if measure_nm.split('_')[0] == 'PSI':
            if measure_nm == 'PSI_03':

                results = filter_func_ccmc(excel_file_name, lookup_dict, keyname, ccmc_hosp_list, 'O/E Ratio',
                                           measure_nm, edac_readm_choice_num)
                return (results)

            elif measure_nm == 'PSI_06':

                results = filter_func_ccmc(excel_file_name, lookup_dict, keyname, ccmc_hosp_list, 'O/E Ratio',
                                           measure_nm, edac_readm_choice_num)
                return (results)

            elif measure_nm == 'PSI_09':

                results = filter_func_ccmc(excel_file_name, lookup_dict, keyname, ccmc_hosp_list, 'O/E Ratio',
                                           measure_nm, edac_readm_choice_num)
                return (results)

            elif measure_nm == 'PSI_11':

                results = filter_func_ccmc(excel_file_name, lookup_dict, keyname, ccmc_hosp_list, 'O/E Ratio',
                                           measure_nm, edac_readm_choice_num)
                return (results)

            elif measure_nm == 'PSI_13':

                results = filter_func_ccmc(excel_file_name, lookup_dict, keyname, ccmc_hosp_list, 'O/E Ratio',
                                           measure_nm, edac_readm_choice_num)
                return (results)

        elif measure_nm == 'THK':

            results = filter_func_ccmc(excel_file_name, lookup_dict, keyname, ccmc_hosp_list, 'Rate', measure_nm,
                                       edac_readm_choice_num)
            return (results)

    elif cohort_nm == 'Complex Care Medical Center' and domain_nm == 'Outpatient':

        results = filter_func_ccmc(excel_file_name, lookup_dict, keyname, ccmc_hosp_list, 'Pct Total Readmit',
                                   measure_nm, edac_readm_choice_num)
        return (results)

###############################################################################################

# Conditionally implement filter functions depending on Hospital cohort and domain.
# Require conditions in order to select correct data column and filter rows correctly.
'''
def file_filter_switcher_all_cohort_hosps(cohort_nm, domain_nm, measure_nm, excel_file_name, lookup_dict, keyname,
                                          edac_readm_choice_num, camc_cohort_list, lsccmc_cohort_list, ccmc_cohort_list,
                                          comm_cohort_list, edac_readm_month_list='No List'):
    #print(excel_file_name)
    if cohort_nm == 'Comprehensive Academic Medical Center' and domain_nm == 'Mortality':

        results = filter_func_all_cohort_hosps(excel_file_name, lookup_dict, keyname, camc_cohort_list,
                                               'Mortality Index', measure_nm, edac_readm_choice_num)
        return (results)

    elif cohort_nm == 'Comprehensive Academic Medical Center' and domain_nm == 'Effectiveness':

        if measure_nm.split('_')[0] == 'EDAC':

            if edac_readm_choice_num == 1:
                results = filter_func_all_cohort_hosps(excel_file_name, lookup_dict, keyname, camc_cohort_list,
                                                       ['Total Index Encounters', 'Total Revisit Days',
                                                        'Expected Inpatient Days', 'Encounter Month'], measure_nm,
                                                       edac_readm_choice_num)
                # 'Excess Days Per 100 Index Encounters', measure_nm)
            else:
                results = filter_func_all_cohort_hosps(excel_file_name, lookup_dict, keyname, camc_cohort_list,
                                                       ['Total Index Encounters', 'Total Revisit Days',
                                                        'Expected Inpatient Days', 'Encounter Month'], measure_nm,
                                                       edac_readm_choice_num, edac_readm_month_list)
                # 'Excess Days Per 100 Index Encounters', measure_nm)

            return (results)

        elif measure_nm.split('_')[0] == 'READM':

            if edac_readm_choice_num == 1:
                results = filter_func_all_cohort_hosps(excel_file_name, lookup_dict, keyname, camc_cohort_list,
                                                       ['Encounter Month', 'Revisit Inpatient Cases',
                                                        'Total Index Encounters'],
                                                       measure_nm, edac_readm_choice_num)
            else:
                results = filter_func_all_cohort_hosps(excel_file_name, lookup_dict, keyname, camc_cohort_list,
                                                       ['Encounter Month', 'Revisit Inpatient Cases',
                                                        'Total Index Encounters'],
                                                       measure_nm, edac_readm_choice_num, edac_readm_month_list)
            return (results)

    elif cohort_nm == 'Comprehensive Academic Medical Center' and domain_nm == 'Efficiency':
        if measure_nm.split('_')[0] == 'DCOST':

            results = filter_func_all_cohort_hosps(excel_file_name, lookup_dict, keyname, camc_cohort_list,
                                                   'Direct Cost Index',
                                                   measure_nm, edac_readm_choice_num)
            return (results)

        elif measure_nm.split('_')[0] == 'LOS':

            results = filter_func_all_cohort_hosps(excel_file_name, lookup_dict, keyname, camc_cohort_list, 'LOS Index',
                                                   measure_nm, edac_readm_choice_num)
            return (results)

    elif cohort_nm == 'Comprehensive Academic Medical Center' and domain_nm == 'Safety':
        if measure_nm.split('_')[0] == 'PSI':
            if measure_nm == 'PSI_03':

                results = filter_func_all_cohort_hosps(excel_file_name, lookup_dict, keyname, camc_cohort_list,
                                                       'O/E Ratio',
                                                       measure_nm, edac_readm_choice_num)
                return (results)

            elif measure_nm == 'PSI_06':

                results = filter_func_all_cohort_hosps(excel_file_name, lookup_dict, keyname, camc_cohort_list,
                                                       'O/E Ratio',
                                                       measure_nm, edac_readm_choice_num)
                return (results)

            elif measure_nm == 'PSI_09':

                results = filter_func_all_cohort_hosps(excel_file_name, lookup_dict, keyname, camc_cohort_list,
                                                       'O/E Ratio',
                                                       measure_nm, edac_readm_choice_num)
                return (results)

            elif measure_nm == 'PSI_11':

                results = filter_func_all_cohort_hosps(excel_file_name, lookup_dict, keyname, camc_cohort_list,
                                                       'O/E Ratio',
                                                       measure_nm, edac_readm_choice_num)
                return (results)

            elif measure_nm == 'PSI_13':

                results = filter_func_all_cohort_hosps(excel_file_name, lookup_dict, keyname, camc_cohort_list,
                                                       'O/E Ratio',
                                                       measure_nm, edac_readm_choice_num)
                return (results)

        elif measure_nm == 'THK':

            results = filter_func_all_cohort_hosps(excel_file_name, lookup_dict, keyname, camc_cohort_list, 'Rate',
                                                   measure_nm, edac_readm_choice_num)
            return (results)

    elif cohort_nm == 'Comprehensive Academic Medical Center' and domain_nm == 'Outpatient':

        results = filter_func_all_cohort_hosps(excel_file_name, lookup_dict, keyname, camc_cohort_list,
                                               'Pct Total Readmit',
                                               measure_nm, edac_readm_choice_num)
        return (results)

    # LSCCMC logic:
    elif cohort_nm == 'Large Specialized Complex Care Medical Center' and domain_nm == 'Mortality':

        results = filter_func_all_cohort_hosps(excel_file_name, lookup_dict, keyname, lsccmc_cohort_list,
                                               'Mortality Index', measure_nm, edac_readm_choice_num)
        return (results)

    elif cohort_nm == 'Large Specialized Complex Care Medical Center' and domain_nm == 'Effectiveness':
        #print('got into LSCCMC effectiveness')
        if measure_nm.split('_')[0] == 'EDAC':
            #print('got to EDAC')
            if edac_readm_choice_num == 1:
                results = filter_func_all_cohort_hosps(excel_file_name, lookup_dict, keyname, lsccmc_cohort_list,
                                                       ['Total Index Encounters', 'Total Revisit Days',
                                                        'Expected Inpatient Days', 'Encounter Month'], measure_nm,
                                                       edac_readm_choice_num)
                # 'Excess Days Per 100 Index Encounters', measure_nm)
            else:
                results = filter_func_all_cohort_hosps(excel_file_name, lookup_dict, keyname, lsccmc_cohort_list,
                                                       ['Total Index Encounters', 'Total Revisit Days',
                                                        'Expected Inpatient Days', 'Encounter Month'], measure_nm,
                                                       edac_readm_choice_num, edac_readm_month_list)
                # 'Excess Days Per 100 Index Encounters', measure_nm)

            return (results)

        elif measure_nm.split('_')[0] == 'READM':

            if edac_readm_choice_num == 1:
                results = filter_func_all_cohort_hosps(excel_file_name, lookup_dict, keyname, lsccmc_cohort_list,
                                                       ['Encounter Month', 'Revisit Inpatient Cases',
                                                        'Total Index Encounters'],
                                                       measure_nm, edac_readm_choice_num)
            else:
                results = filter_func_all_cohort_hosps(excel_file_name, lookup_dict, keyname, lsccmc_cohort_list,
                                                       ['Encounter Month', 'Revisit Inpatient Cases',
                                                        'Total Index Encounters'],
                                                       measure_nm, edac_readm_choice_num, edac_readm_month_list)
            return (results)

    elif cohort_nm == 'Large Specialized Complex Care Medical Center' and domain_nm == 'Efficiency':
        if measure_nm.split('_')[0] == 'DCOST':

            results = filter_func_all_cohort_hosps(excel_file_name, lookup_dict, keyname, lsccmc_cohort_list,
                                                   'Direct Cost Index',
                                                   measure_nm, edac_readm_choice_num)
            return (results)

        elif measure_nm.split('_')[0] == 'LOS':

            results = filter_func_all_cohort_hosps(excel_file_name, lookup_dict, keyname, lsccmc_cohort_list,
                                                   'LOS Index', measure_nm, edac_readm_choice_num)
            return (results)

    elif cohort_nm == 'Large Specialized Complex Care Medical Center' and domain_nm == 'Safety':
        if measure_nm.split('_')[0] == 'PSI':
            if measure_nm == 'PSI_03':

                results = filter_func_all_cohort_hosps(excel_file_name, lookup_dict, keyname, lsccmc_cohort_list,
                                                       'O/E Ratio',
                                                       measure_nm, edac_readm_choice_num)
                return (results)

            elif measure_nm == 'PSI_06':

                results = filter_func_all_cohort_hosps(excel_file_name, lookup_dict, keyname, lsccmc_cohort_list,
                                                       'O/E Ratio',
                                                       measure_nm, edac_readm_choice_num)
                return (results)

            elif measure_nm == 'PSI_09':

                results = filter_func_all_cohort_hosps(excel_file_name, lookup_dict, keyname, lsccmc_cohort_list,
                                                       'O/E Ratio',
                                                       measure_nm, edac_readm_choice_num)
                return (results)

            elif measure_nm == 'PSI_11':

                results = filter_func_all_cohort_hosps(excel_file_name, lookup_dict, keyname, lsccmc_cohort_list,
                                                       'O/E Ratio',
                                                       measure_nm, edac_readm_choice_num)
                return (results)

            elif measure_nm == 'PSI_13':

                results = filter_func_all_cohort_hosps(excel_file_name, lookup_dict, keyname, lsccmc_cohort_list,
                                                       'O/E Ratio',
                                                       measure_nm, edac_readm_choice_num)
                return (results)

        elif measure_nm == 'THK':

            results = filter_func_all_cohort_hosps(excel_file_name, lookup_dict, keyname, lsccmc_cohort_list, 'Rate',
                                                   measure_nm, edac_readm_choice_num)
            return (results)



    elif cohort_nm == 'Large Specialized Complex Care Medical Center' and domain_nm == 'Outpatient':

        results = filter_func_all_cohort_hosps(excel_file_name, lookup_dict, keyname, lsccmc_cohort_list,
                                               'Pct Total Readmit',
                                               measure_nm, edac_readm_choice_num)
        return (results)

    # CCMC Logic:
    elif cohort_nm == 'Complex Care Medical Center' and domain_nm == 'Mortality':
        # print('got to ccmc')

        results = filter_func_all_cohort_hosps(excel_file_name, lookup_dict, keyname, ccmc_cohort_list,
                                               'Mortality Index', measure_nm, edac_readm_choice_num)
        return (results)

    elif cohort_nm == 'Complex Care Medical Center' and domain_nm == 'Effectiveness':

        if measure_nm.split('_')[0] == 'EDAC':

            if edac_readm_choice_num == 1:
                results = filter_func_all_cohort_hosps(excel_file_name, lookup_dict, keyname, ccmc_cohort_list,
                                                       ['Total Index Encounters', 'Total Revisit Days',
                                                        'Expected Inpatient Days', 'Encounter Month'], measure_nm,
                                                       edac_readm_choice_num)
                # 'Excess Days Per 100 Index Encounters', measure_nm)
            else:
                results = filter_func_all_cohort_hosps(excel_file_name, lookup_dict, keyname, ccmc_cohort_list,
                                                       ['Total Index Encounters', 'Total Revisit Days',
                                                        'Expected Inpatient Days', 'Encounter Month'], measure_nm,
                                                       edac_readm_choice_num, edac_readm_month_list)
                # 'Excess Days Per 100 Index Encounters', measure_nm)

            return (results)

        elif measure_nm.split('_')[0] == 'READM':

            if edac_readm_choice_num == 1:
                results = filter_func_all_cohort_hosps(excel_file_name, lookup_dict, keyname, ccmc_cohort_list,
                                                       ['Encounter Month', 'Revisit Inpatient Cases',
                                                        'Total Index Encounters'],
                                                       measure_nm, edac_readm_choice_num)
            else:
                results = filter_func_all_cohort_hosps(excel_file_name, lookup_dict, keyname, ccmc_cohort_list,
                                                       ['Encounter Month', 'Revisit Inpatient Cases',
                                                        'Total Index Encounters'],
                                                       measure_nm, edac_readm_choice_num, edac_readm_month_list)
            return (results)

    elif cohort_nm == 'Complex Care Medical Center' and domain_nm == 'Efficiency':
        if measure_nm.split('_')[0] == 'DCOST':

            results = filter_func_all_cohort_hosps(excel_file_name, lookup_dict, keyname, ccmc_cohort_list,
                                                   'Direct Cost Index',
                                                   measure_nm, edac_readm_choice_num)
            return (results)

        elif measure_nm.split('_')[0] == 'LOS':

            results = filter_func_all_cohort_hosps(excel_file_name, lookup_dict, keyname, ccmc_cohort_list, 'LOS Index',
                                                   measure_nm, edac_readm_choice_num)
            return (results)

    elif cohort_nm == 'Complex Care Medical Center' and domain_nm == 'Safety':
        if measure_nm.split('_')[0] == 'PSI':
            if measure_nm == 'PSI_03':

                results = filter_func_all_cohort_hosps(excel_file_name, lookup_dict, keyname, ccmc_cohort_list,
                                                       'O/E Ratio',
                                                       measure_nm, edac_readm_choice_num)
                return (results)

            elif measure_nm == 'PSI_06':

                results = filter_func_all_cohort_hosps(excel_file_name, lookup_dict, keyname, ccmc_cohort_list,
                                                       'O/E Ratio',
                                                       measure_nm, edac_readm_choice_num)
                return (results)

            elif measure_nm == 'PSI_09':

                results = filter_func_all_cohort_hosps(excel_file_name, lookup_dict, keyname, ccmc_cohort_list,
                                                       'O/E Ratio',
                                                       measure_nm, edac_readm_choice_num)
                return (results)

            elif measure_nm == 'PSI_11':

                results = filter_func_all_cohort_hosps(excel_file_name, lookup_dict, keyname, ccmc_cohort_list,
                                                       'O/E Ratio',
                                                       measure_nm, edac_readm_choice_num)
                return (results)

            elif measure_nm == 'PSI_13':

                results = filter_func_all_cohort_hosps(excel_file_name, lookup_dict, keyname, ccmc_cohort_list,
                                                       'O/E Ratio',
                                                       measure_nm, edac_readm_choice_num)
                return (results)

        elif measure_nm == 'THK':

            results = filter_func_all_cohort_hosps(excel_file_name, lookup_dict, keyname, ccmc_cohort_list, 'Rate',
                                                   measure_nm, edac_readm_choice_num)
            return (results)

    elif cohort_nm == 'Complex Care Medical Center' and domain_nm == 'Outpatient':

        results = filter_func_all_cohort_hosps(excel_file_name, lookup_dict, keyname, ccmc_cohort_list,
                                               'Pct Total Readmit',
                                               measure_nm, edac_readm_choice_num)
        return (results)

    # COMM Logic:
    elif cohort_nm == 'Community Medical Center' and domain_nm == 'Mortality':
        # print('got to ccmc')

        results = filter_func_all_cohort_hosps(excel_file_name, lookup_dict, keyname, comm_cohort_list,
                                               'Mortality Index', measure_nm, edac_readm_choice_num)
        return (results)

    elif cohort_nm == 'Community Medical Center' and domain_nm == 'Effectiveness':

        if measure_nm.split('_')[0] == 'EDAC':

            if edac_readm_choice_num == 1:
                results = filter_func_all_cohort_hosps(excel_file_name, lookup_dict, keyname, comm_cohort_list,
                                                       ['Total Index Encounters', 'Total Revisit Days',
                                                        'Expected Inpatient Days', 'Encounter Month'], measure_nm,
                                                       edac_readm_choice_num)
                # 'Excess Days Per 100 Index Encounters', measure_nm)
            else:
                results = filter_func_all_cohort_hosps(excel_file_name, lookup_dict, keyname, comm_cohort_list,
                                                       ['Total Index Encounters', 'Total Revisit Days',
                                                        'Expected Inpatient Days', 'Encounter Month'], measure_nm,
                                                       edac_readm_choice_num, edac_readm_month_list)
                # 'Excess Days Per 100 Index Encounters', measure_nm)

            return (results)

        elif measure_nm.split('_')[0] == 'READM':

            if edac_readm_choice_num == 1:
                results = filter_func_all_cohort_hosps(excel_file_name, lookup_dict, keyname, comm_cohort_list,
                                                       ['Encounter Month', 'Revisit Inpatient Cases',
                                                        'Total Index Encounters'],
                                                       measure_nm, edac_readm_choice_num)
            else:
                results = filter_func_all_cohort_hosps(excel_file_name, lookup_dict, keyname, comm_cohort_list,
                                                       ['Encounter Month', 'Revisit Inpatient Cases',
                                                        'Total Index Encounters'],
                                                       measure_nm, edac_readm_choice_num, edac_readm_month_list)
            return (results)

    elif cohort_nm == 'Community Medical Center' and domain_nm == 'Efficiency':
        if measure_nm.split('_')[0] == 'DCOST':

            results = filter_func_all_cohort_hosps(excel_file_name, lookup_dict, keyname, comm_cohort_list,
                                                   'Direct Cost Index',
                                                   measure_nm, edac_readm_choice_num)
            return (results)

        elif measure_nm.split('_')[0] == 'LOS':

            results = filter_func_all_cohort_hosps(excel_file_name, lookup_dict, keyname, comm_cohort_list, 'LOS Index',
                                                   measure_nm, edac_readm_choice_num)
            return (results)

    elif cohort_nm == 'Community Medical Center' and domain_nm == 'Safety':
        if measure_nm.split('_')[0] == 'PSI':
            if measure_nm == 'PSI_03':

                results = filter_func_all_cohort_hosps(excel_file_name, lookup_dict, keyname, comm_cohort_list,
                                                       'O/E Ratio',
                                                       measure_nm, edac_readm_choice_num)
                return (results)

            elif measure_nm == 'PSI_06':

                results = filter_func_all_cohort_hosps(excel_file_name, lookup_dict, keyname, comm_cohort_list,
                                                       'O/E Ratio',
                                                       measure_nm, edac_readm_choice_num)
                return (results)

            elif measure_nm == 'PSI_09':

                results = filter_func_all_cohort_hosps(excel_file_name, lookup_dict, keyname, comm_cohort_list,
                                                       'O/E Ratio',
                                                       measure_nm, edac_readm_choice_num)
                return (results)

            elif measure_nm == 'PSI_11':

                results = filter_func_all_cohort_hosps(excel_file_name, lookup_dict, keyname, comm_cohort_list,
                                                       'O/E Ratio',
                                                       measure_nm, edac_readm_choice_num)
                return (results)

            elif measure_nm == 'PSI_13':

                results = filter_func_all_cohort_hosps(excel_file_name, lookup_dict, keyname, comm_cohort_list,
                                                       'O/E Ratio',
                                                       measure_nm, edac_readm_choice_num)
                return (results)

        elif measure_nm == 'THK':

            results = filter_func_all_cohort_hosps(excel_file_name, lookup_dict, keyname, comm_cohort_list, 'Rate',
                                                   measure_nm, edac_readm_choice_num)
            return (results)

    elif cohort_nm == 'Community Medical Center' and domain_nm == 'Outpatient':

        results = filter_func_all_cohort_hosps(excel_file_name, lookup_dict, keyname, comm_cohort_list,
                                               'Pct Total Readmit',
                                               measure_nm, edac_readm_choice_num)
        return (results)
'''
#UL003

def file_filter_switcher_all_cohort_hosps(cohort_nm, domain_nm, measure_nm, excel_file_name, lookup_dict, keyname,
                                          edac_readm_choice_num, camc_cohort_list, lsccmc_cohort_list, ccmc_cohort_list,
                                          comm_cohort_list,casc_cohort_list,edac_readm_month_list='No List'):
    
    #print(excel_file_name)
    if cohort_nm == 'Comprehensive Academic Medical Center' and domain_nm == 'Mortality':

        results = filter_func_all_cohort_hosps(excel_file_name, lookup_dict, keyname, camc_cohort_list,
                                               'Mortality Index', measure_nm, edac_readm_choice_num)
        return (results)

    elif cohort_nm == 'Comprehensive Academic Medical Center' and domain_nm == 'Effectiveness':

        if measure_nm.split('_')[0] == 'EDAC':

            if edac_readm_choice_num == 1:
                results = filter_func_all_cohort_hosps(excel_file_name, lookup_dict, keyname, camc_cohort_list,
                                                       ['Total Index Encounters', 'Total Revisit Days',
                                                        'Expected Inpatient Days', 'Encounter Month'], measure_nm,
                                                       edac_readm_choice_num)
                # 'Excess Days Per 100 Index Encounters', measure_nm)
            else:
                results = filter_func_all_cohort_hosps(excel_file_name, lookup_dict, keyname, camc_cohort_list,
                                                       ['Total Index Encounters', 'Total Revisit Days',
                                                        'Expected Inpatient Days', 'Encounter Month'], measure_nm,
                                                       edac_readm_choice_num, edac_readm_month_list)
                # 'Excess Days Per 100 Index Encounters', measure_nm)

            return (results)
        #UL003
        elif measure_nm.split('_')[0] == 'READM' and measure_nm.split('_')[1] not in ['ARTHRO','CHOL','COLON','URI']:

            if edac_readm_choice_num == 1:
                results = filter_func_all_cohort_hosps(excel_file_name, lookup_dict, keyname, camc_cohort_list,
                                                       ['Encounter Month', 'Revisit Inpatient Cases',
                                                        'Total Index Encounters'],
                                                       measure_nm, edac_readm_choice_num)
            else:
                results = filter_func_all_cohort_hosps(excel_file_name, lookup_dict, keyname, camc_cohort_list,
                                                       ['Encounter Month', 'Revisit Inpatient Cases',
                                                        'Total Index Encounters'],
                                                       measure_nm, edac_readm_choice_num, edac_readm_month_list)
            return (results)
        
        elif measure_nm.split('_')[0] == 'READM' and domain_nm == 'Effectiveness' and measure_nm.split('_')[1] in ['ARTHRO','CHOL','COLON','URI']:

            results = filter_func_all_cohort_hosps(excel_file_name, lookup_dict, keyname, camc_cohort_list,
                                               'Pct Total Readmit',
                                               measure_nm, edac_readm_choice_num)
            return (results)

    elif cohort_nm == 'Comprehensive Academic Medical Center' and domain_nm == 'Efficiency':
        if measure_nm.split('_')[0] == 'DCOST':

            results = filter_func_all_cohort_hosps(excel_file_name, lookup_dict, keyname, camc_cohort_list,
                                                   'Direct Cost Index',
                                                   measure_nm, edac_readm_choice_num)
            return (results)

        elif measure_nm.split('_')[0] == 'LOS':

            results = filter_func_all_cohort_hosps(excel_file_name, lookup_dict, keyname, camc_cohort_list, 'LOS Index',
                                                   measure_nm, edac_readm_choice_num)
            return (results)

    elif cohort_nm == 'Comprehensive Academic Medical Center' and domain_nm == 'Safety':
        if measure_nm.split('_')[0] == 'PSI':
            if measure_nm == 'PSI_03':

                results = filter_func_all_cohort_hosps(excel_file_name, lookup_dict, keyname, camc_cohort_list,
                                                       'O/E Ratio',
                                                       measure_nm, edac_readm_choice_num)
                return (results)

            elif measure_nm == 'PSI_06':

                results = filter_func_all_cohort_hosps(excel_file_name, lookup_dict, keyname, camc_cohort_list,
                                                       'O/E Ratio',
                                                       measure_nm, edac_readm_choice_num)
                return (results)

            elif measure_nm == 'PSI_09':

                results = filter_func_all_cohort_hosps(excel_file_name, lookup_dict, keyname, camc_cohort_list,
                                                       'O/E Ratio',
                                                       measure_nm, edac_readm_choice_num)
                return (results)

            elif measure_nm == 'PSI_11':

                results = filter_func_all_cohort_hosps(excel_file_name, lookup_dict, keyname, camc_cohort_list,
                                                       'O/E Ratio',
                                                       measure_nm, edac_readm_choice_num)
                return (results)

            elif measure_nm == 'PSI_13':

                results = filter_func_all_cohort_hosps(excel_file_name, lookup_dict, keyname, camc_cohort_list,
                                                       'O/E Ratio',
                                                       measure_nm, edac_readm_choice_num)
                return (results)

        elif measure_nm == 'THK':

            results = filter_func_all_cohort_hosps(excel_file_name, lookup_dict, keyname, camc_cohort_list, 'Rate',
                                                   measure_nm, edac_readm_choice_num)
            return (results)

    elif cohort_nm == 'Comprehensive Academic Medical Center' and domain_nm == 'Outpatient':
        
        results = filter_func_all_cohort_hosps(excel_file_name, lookup_dict, keyname, camc_cohort_list,
                                               'Pct Total Readmit',
                                               measure_nm, edac_readm_choice_num)
        return (results)

    # LSCCMC logic:
    elif cohort_nm == 'Large Specialized Complex Care Medical Center' and domain_nm == 'Mortality':

        results = filter_func_all_cohort_hosps(excel_file_name, lookup_dict, keyname, lsccmc_cohort_list,
                                               'Mortality Index', measure_nm, edac_readm_choice_num)
        return (results)

    elif cohort_nm == 'Large Specialized Complex Care Medical Center' and domain_nm == 'Effectiveness':
        #print('got into LSCCMC effectiveness')
        if measure_nm.split('_')[0] == 'EDAC':
            #print('got to EDAC')
            if edac_readm_choice_num == 1:
                results = filter_func_all_cohort_hosps(excel_file_name, lookup_dict, keyname, lsccmc_cohort_list,
                                                       ['Total Index Encounters', 'Total Revisit Days',
                                                        'Expected Inpatient Days', 'Encounter Month'], measure_nm,
                                                       edac_readm_choice_num)
                # 'Excess Days Per 100 Index Encounters', measure_nm)
            else:
                results = filter_func_all_cohort_hosps(excel_file_name, lookup_dict, keyname, lsccmc_cohort_list,
                                                       ['Total Index Encounters', 'Total Revisit Days',
                                                        'Expected Inpatient Days', 'Encounter Month'], measure_nm,
                                                       edac_readm_choice_num, edac_readm_month_list)
                # 'Excess Days Per 100 Index Encounters', measure_nm)

            return (results)

        #UL003
        elif measure_nm.split('_')[0] == 'READM' and measure_nm.split('_')[1] not in ['ARTHRO','CHOL','COLON','URI']:

            if edac_readm_choice_num == 1:
                results = filter_func_all_cohort_hosps(excel_file_name, lookup_dict, keyname, lsccmc_cohort_list,
                                                       ['Encounter Month', 'Revisit Inpatient Cases',
                                                        'Total Index Encounters'],
                                                       measure_nm, edac_readm_choice_num)
            else:
                results = filter_func_all_cohort_hosps(excel_file_name, lookup_dict, keyname, lsccmc_cohort_list,
                                                       ['Encounter Month', 'Revisit Inpatient Cases',
                                                        'Total Index Encounters'],
                                                       measure_nm, edac_readm_choice_num, edac_readm_month_list)
            return (results)
        
        elif measure_nm.split('_')[0] == 'READM' and domain_nm == 'Effectiveness' and measure_nm.split('_')[1] in ['ARTHRO','CHOL','COLON','URI']:

            results = filter_func_all_cohort_hosps(excel_file_name, lookup_dict, keyname, lsccmc_cohort_list,
                                               'Pct Total Readmit',
                                               measure_nm, edac_readm_choice_num)
            return (results)

    elif cohort_nm == 'Large Specialized Complex Care Medical Center' and domain_nm == 'Efficiency':
        if measure_nm.split('_')[0] == 'DCOST':

            results = filter_func_all_cohort_hosps(excel_file_name, lookup_dict, keyname, lsccmc_cohort_list,
                                                   'Direct Cost Index',
                                                   measure_nm, edac_readm_choice_num)
            return (results)

        elif measure_nm.split('_')[0] == 'LOS':

            results = filter_func_all_cohort_hosps(excel_file_name, lookup_dict, keyname, lsccmc_cohort_list,
                                                   'LOS Index', measure_nm, edac_readm_choice_num)
            return (results)

    elif cohort_nm == 'Large Specialized Complex Care Medical Center' and domain_nm == 'Safety':
        if measure_nm.split('_')[0] == 'PSI':
            if measure_nm == 'PSI_03':

                results = filter_func_all_cohort_hosps(excel_file_name, lookup_dict, keyname, lsccmc_cohort_list,
                                                       'O/E Ratio',
                                                       measure_nm, edac_readm_choice_num)
                return (results)

            elif measure_nm == 'PSI_06':

                results = filter_func_all_cohort_hosps(excel_file_name, lookup_dict, keyname, lsccmc_cohort_list,
                                                       'O/E Ratio',
                                                       measure_nm, edac_readm_choice_num)
                return (results)

            elif measure_nm == 'PSI_09':

                results = filter_func_all_cohort_hosps(excel_file_name, lookup_dict, keyname, lsccmc_cohort_list,
                                                       'O/E Ratio',
                                                       measure_nm, edac_readm_choice_num)
                return (results)

            elif measure_nm == 'PSI_11':

                results = filter_func_all_cohort_hosps(excel_file_name, lookup_dict, keyname, lsccmc_cohort_list,
                                                       'O/E Ratio',
                                                       measure_nm, edac_readm_choice_num)
                return (results)

            elif measure_nm == 'PSI_13':

                results = filter_func_all_cohort_hosps(excel_file_name, lookup_dict, keyname, lsccmc_cohort_list,
                                                       'O/E Ratio',
                                                       measure_nm, edac_readm_choice_num)
                return (results)

        elif measure_nm == 'THK':

            results = filter_func_all_cohort_hosps(excel_file_name, lookup_dict, keyname, lsccmc_cohort_list, 'Rate',
                                                   measure_nm, edac_readm_choice_num)
            return (results)



    elif cohort_nm == 'Large Specialized Complex Care Medical Center' and domain_nm == 'Outpatient':

        results = filter_func_all_cohort_hosps(excel_file_name, lookup_dict, keyname, lsccmc_cohort_list,
                                               'Pct Total Readmit',
                                               measure_nm, edac_readm_choice_num)
        return (results)

    # CCMC Logic:
    elif cohort_nm == 'Complex Care Medical Center' and domain_nm == 'Mortality':
        # print('got to ccmc')

        results = filter_func_all_cohort_hosps(excel_file_name, lookup_dict, keyname, ccmc_cohort_list,
                                               'Mortality Index', measure_nm, edac_readm_choice_num)
        return (results)

    elif cohort_nm == 'Complex Care Medical Center' and domain_nm == 'Effectiveness':

        if measure_nm.split('_')[0] == 'EDAC':

            if edac_readm_choice_num == 1:
                results = filter_func_all_cohort_hosps(excel_file_name, lookup_dict, keyname, ccmc_cohort_list,
                                                       ['Total Index Encounters', 'Total Revisit Days',
                                                        'Expected Inpatient Days', 'Encounter Month'], measure_nm,
                                                       edac_readm_choice_num)
                # 'Excess Days Per 100 Index Encounters', measure_nm)
            else:
                results = filter_func_all_cohort_hosps(excel_file_name, lookup_dict, keyname, ccmc_cohort_list,
                                                       ['Total Index Encounters', 'Total Revisit Days',
                                                        'Expected Inpatient Days', 'Encounter Month'], measure_nm,
                                                       edac_readm_choice_num, edac_readm_month_list)
                # 'Excess Days Per 100 Index Encounters', measure_nm)

            return (results)

        #UL003
        elif measure_nm.split('_')[0] == 'READM' and measure_nm.split('_')[1] not in ['ARTHRO','CHOL','COLON','URI']:

            if edac_readm_choice_num == 1:
                results = filter_func_all_cohort_hosps(excel_file_name, lookup_dict, keyname, ccmc_cohort_list,
                                                       ['Encounter Month', 'Revisit Inpatient Cases',
                                                        'Total Index Encounters'],
                                                       measure_nm, edac_readm_choice_num)
            else:
                results = filter_func_all_cohort_hosps(excel_file_name, lookup_dict, keyname, ccmc_cohort_list,
                                                       ['Encounter Month', 'Revisit Inpatient Cases',
                                                        'Total Index Encounters'],
                                                       measure_nm, edac_readm_choice_num, edac_readm_month_list)
            return (results)
        
        elif measure_nm.split('_')[0] == 'READM' and domain_nm == 'Effectiveness' and measure_nm.split('_')[1] in ['ARTHRO','CHOL','COLON','URI']:

            results = filter_func_all_cohort_hosps(excel_file_name, lookup_dict, keyname, ccmc_cohort_list,
                                               'Pct Total Readmit',
                                               measure_nm, edac_readm_choice_num)
            return (results)

    elif cohort_nm == 'Complex Care Medical Center' and domain_nm == 'Efficiency':
        if measure_nm.split('_')[0] == 'DCOST':

            results = filter_func_all_cohort_hosps(excel_file_name, lookup_dict, keyname, ccmc_cohort_list,
                                                   'Direct Cost Index',
                                                   measure_nm, edac_readm_choice_num)
            return (results)

        elif measure_nm.split('_')[0] == 'LOS':

            results = filter_func_all_cohort_hosps(excel_file_name, lookup_dict, keyname, ccmc_cohort_list, 'LOS Index',
                                                   measure_nm, edac_readm_choice_num)
            return (results)

    elif cohort_nm == 'Complex Care Medical Center' and domain_nm == 'Safety':
        if measure_nm.split('_')[0] == 'PSI':
            if measure_nm == 'PSI_03':

                results = filter_func_all_cohort_hosps(excel_file_name, lookup_dict, keyname, ccmc_cohort_list,
                                                       'O/E Ratio',
                                                       measure_nm, edac_readm_choice_num)
                return (results)

            elif measure_nm == 'PSI_06':

                results = filter_func_all_cohort_hosps(excel_file_name, lookup_dict, keyname, ccmc_cohort_list,
                                                       'O/E Ratio',
                                                       measure_nm, edac_readm_choice_num)
                return (results)

            elif measure_nm == 'PSI_09':

                results = filter_func_all_cohort_hosps(excel_file_name, lookup_dict, keyname, ccmc_cohort_list,
                                                       'O/E Ratio',
                                                       measure_nm, edac_readm_choice_num)
                return (results)

            elif measure_nm == 'PSI_11':

                results = filter_func_all_cohort_hosps(excel_file_name, lookup_dict, keyname, ccmc_cohort_list,
                                                       'O/E Ratio',
                                                       measure_nm, edac_readm_choice_num)
                return (results)

            elif measure_nm == 'PSI_13':

                results = filter_func_all_cohort_hosps(excel_file_name, lookup_dict, keyname, ccmc_cohort_list,
                                                       'O/E Ratio',
                                                       measure_nm, edac_readm_choice_num)
                return (results)

        elif measure_nm == 'THK':

            results = filter_func_all_cohort_hosps(excel_file_name, lookup_dict, keyname, ccmc_cohort_list, 'Rate',
                                                   measure_nm, edac_readm_choice_num)
            return (results)

    elif cohort_nm == 'Complex Care Medical Center' and domain_nm == 'Outpatient':

        results = filter_func_all_cohort_hosps(excel_file_name, lookup_dict, keyname, ccmc_cohort_list,
                                               'Pct Total Readmit',
                                               measure_nm, edac_readm_choice_num)
        return (results)
    

    # COMM Logic:
    elif cohort_nm == 'Community Medical Center' and domain_nm == 'Mortality':
        # print('got to ccmc')

        results = filter_func_all_cohort_hosps(excel_file_name, lookup_dict, keyname, comm_cohort_list,
                                               'Mortality Index', measure_nm, edac_readm_choice_num)
        return (results)

    elif cohort_nm == 'Community Medical Center' and domain_nm == 'Effectiveness':
        
        
        if measure_nm.split('_')[0] == 'EDAC':

            if edac_readm_choice_num == 1:
                results = filter_func_all_cohort_hosps(excel_file_name, lookup_dict, keyname, comm_cohort_list,
                                                       ['Total Index Encounters', 'Total Revisit Days',
                                                        'Expected Inpatient Days', 'Encounter Month'], measure_nm,
                                                       edac_readm_choice_num)
                # 'Excess Days Per 100 Index Encounters', measure_nm)
            else:
                results = filter_func_all_cohort_hosps(excel_file_name, lookup_dict, keyname, comm_cohort_list,
                                                       ['Total Index Encounters', 'Total Revisit Days',
                                                        'Expected Inpatient Days', 'Encounter Month'], measure_nm,
                                                       edac_readm_choice_num, edac_readm_month_list)
                # 'Excess Days Per 100 Index Encounters', measure_nm)

            return (results)

        #UL003
        elif measure_nm.split('_')[0] == 'READM' and measure_nm.split('_')[1] not in ['ARTHRO','CHOL','COLON','URI']:
            
            if edac_readm_choice_num == 1:
                results = filter_func_all_cohort_hosps(excel_file_name, lookup_dict, keyname, comm_cohort_list,
                                                       ['Encounter Month', 'Revisit Inpatient Cases',
                                                        'Total Index Encounters'],
                                                       measure_nm, edac_readm_choice_num)
            else:
                results = filter_func_all_cohort_hosps(excel_file_name, lookup_dict, keyname, comm_cohort_list,
                                                       ['Encounter Month', 'Revisit Inpatient Cases',
                                                        'Total Index Encounters'],
                                                       measure_nm, edac_readm_choice_num, edac_readm_month_list)
            return (results)
        
        elif measure_nm.split('_')[0] == 'READM' and domain_nm == 'Effectiveness' and measure_nm.split('_')[1] in ['ARTHRO','CHOL','COLON','URI']:

            results = filter_func_all_cohort_hosps(excel_file_name, lookup_dict, keyname, comm_cohort_list,
                                               'Pct Total Readmit',
                                               measure_nm, edac_readm_choice_num)
            return (results)

    elif cohort_nm == 'Community Medical Center' and domain_nm == 'Efficiency':
        if measure_nm.split('_')[0] == 'DCOST':

            results = filter_func_all_cohort_hosps(excel_file_name, lookup_dict, keyname, comm_cohort_list,
                                                   'Direct Cost Index',
                                                   measure_nm, edac_readm_choice_num)
            return (results)

        elif measure_nm.split('_')[0] == 'LOS':

            results = filter_func_all_cohort_hosps(excel_file_name, lookup_dict, keyname, comm_cohort_list, 'LOS Index',
                                                   measure_nm, edac_readm_choice_num)
            return (results)

    elif cohort_nm == 'Community Medical Center' and domain_nm == 'Safety':
        if measure_nm.split('_')[0] == 'PSI':
            if measure_nm == 'PSI_03':

                results = filter_func_all_cohort_hosps(excel_file_name, lookup_dict, keyname, comm_cohort_list,
                                                       'O/E Ratio',
                                                       measure_nm, edac_readm_choice_num)
                return (results)

            elif measure_nm == 'PSI_06':

                results = filter_func_all_cohort_hosps(excel_file_name, lookup_dict, keyname, comm_cohort_list,
                                                       'O/E Ratio',
                                                       measure_nm, edac_readm_choice_num)
                return (results)

            elif measure_nm == 'PSI_09':

                results = filter_func_all_cohort_hosps(excel_file_name, lookup_dict, keyname, comm_cohort_list,
                                                       'O/E Ratio',
                                                       measure_nm, edac_readm_choice_num)
                return (results)

            elif measure_nm == 'PSI_11':

                results = filter_func_all_cohort_hosps(excel_file_name, lookup_dict, keyname, comm_cohort_list,
                                                       'O/E Ratio',
                                                       measure_nm, edac_readm_choice_num)
                return (results)

            elif measure_nm == 'PSI_13':

                results = filter_func_all_cohort_hosps(excel_file_name, lookup_dict, keyname, comm_cohort_list,
                                                       'O/E Ratio',
                                                       measure_nm, edac_readm_choice_num)
                return (results)

        elif measure_nm == 'THK':

            results = filter_func_all_cohort_hosps(excel_file_name, lookup_dict, keyname, comm_cohort_list, 'Rate',
                                                   measure_nm, edac_readm_choice_num)
            return (results)

    elif cohort_nm == 'Community Medical Center' and domain_nm == 'Outpatient':

        results = filter_func_all_cohort_hosps(excel_file_name, lookup_dict, keyname, comm_cohort_list,
                                               'Pct Total Readmit',
                                               measure_nm, edac_readm_choice_num)
        return (results)
    
    # UL003
    # CASC Logic:
    elif cohort_nm == 'Critical Access & Small Community' and domain_nm == 'Mortality':
        # print('got to ccmc')

        results = filter_func_all_cohort_hosps(excel_file_name, lookup_dict, keyname, casc_cohort_list,
                                               'Mortality Index', measure_nm, edac_readm_choice_num)
        return (results)

    elif cohort_nm == 'Critical Access & Small Community' and domain_nm == 'Efficiency':
        if measure_nm.split('_')[0] == 'DCOST':

            results = filter_func_all_cohort_hosps(excel_file_name, lookup_dict, keyname, casc_cohort_list,
                                                   'Direct Cost Index',
                                                   measure_nm, edac_readm_choice_num)
            return (results)

        elif measure_nm.split('_')[0] == 'LOS':

            results = filter_func_all_cohort_hosps(excel_file_name, lookup_dict, keyname, casc_cohort_list, 'LOS Index',
                                                   measure_nm, edac_readm_choice_num)
            return (results)
        
    elif cohort_nm == 'Critical Access & Small Community' and domain_nm == 'Effectiveness' and measure_nm.split('_')[0] not in ['ADE','EARLY']:

        if measure_nm.split('_')[0] == 'EDAC':

            if edac_readm_choice_num == 1:
                results = filter_func_all_cohort_hosps(excel_file_name, lookup_dict, keyname, casc_cohort_list,
                                                       ['Total Index Encounters', 'Total Revisit Days',
                                                        'Expected Inpatient Days', 'Encounter Month'], measure_nm,
                                                       edac_readm_choice_num)
                # 'Excess Days Per 100 Index Encounters', measure_nm)
            else:
                results = filter_func_all_cohort_hosps(excel_file_name, lookup_dict, keyname, casc_cohort_list,
                                                       ['Total Index Encounters', 'Total Revisit Days',
                                                        'Expected Inpatient Days', 'Encounter Month'], measure_nm,
                                                       edac_readm_choice_num, edac_readm_month_list)
                # 'Excess Days Per 100 Index Encounters', measure_nm)

            return (results)
        #UL003
        elif measure_nm.split('_')[0] == 'READM' and measure_nm.split('_')[1] not in ['ARTHRO','CHOL','COLON','URI']:

            if edac_readm_choice_num == 1:
                results = filter_func_all_cohort_hosps(excel_file_name, lookup_dict, keyname, casc_cohort_list,
                                                       ['Encounter Month', 'Revisit Inpatient Cases',
                                                        'Total Index Encounters'],
                                                       measure_nm, edac_readm_choice_num)
            else:
                results = filter_func_all_cohort_hosps(excel_file_name, lookup_dict, keyname, casc_cohort_list,
                                                       ['Encounter Month', 'Revisit Inpatient Cases',
                                                        'Total Index Encounters'],
                                                       measure_nm, edac_readm_choice_num, edac_readm_month_list)
            return (results)
        
        elif measure_nm.split('_')[0] == 'READM' and domain_nm == 'Effectiveness' and measure_nm.split('_')[1] in ['ARTHRO','CHOL','COLON','URI']:

            results = filter_func_all_cohort_hosps(excel_file_name, lookup_dict, keyname, casc_cohort_list,
                                               'Pct Total Readmit',
                                               measure_nm, edac_readm_choice_num)
            return (results)
        elif domain_nm == 'Effectiveness' and measure_nm.split('_')[1] in ['REVISIT','RETURN']:

            results = filter_func_all_cohort_hosps(excel_file_name, lookup_dict, keyname, casc_cohort_list,
                                               'Pct Total Readmit',
                                               measure_nm, edac_readm_choice_num)
            return (results)
        
    elif measure_nm.split('_')[0] in ['ADE','EARLY']:
        #print('HERE I AM!')
        results = filter_func_all_cohort_hosps(excel_file_name, lookup_dict, keyname, casc_cohort_list,
                                               'Cases',
                                               measure_nm, edac_readm_choice_num)
        return (results)
        

################################################################################################

# Conditionally implement filter functions depending on Hospital cohort and domain.
# Require conditions in order to select correct data column and filter rows correctly.

'''
def file_filter_switcher_all_cohort_hosps_n_events(cohort_nm, domain_nm, measure_nm, excel_file_name, lookup_dict,
                                                   keyname, edac_readm_choice_num, camc_cohort_list, lsccmc_cohort_list,
                                                   ccmc_cohort_list, comm_cohort_list, edac_readm_month_list='No List'):
    #print('got to n size switcher')

    if cohort_nm == 'Comprehensive Academic Medical Center' and domain_nm == 'Mortality':

        results = filter_func_all_cohort_hosps(excel_file_name, lookup_dict, keyname, camc_cohort_list, 'Deaths (Obs)',
                                               measure_nm, edac_readm_choice_num)
        return (results)

    elif cohort_nm == 'Comprehensive Academic Medical Center' and domain_nm == 'Effectiveness':

        if measure_nm.split('_')[0] == 'EDAC':

            if edac_readm_choice_num == 1:
                results = filter_func_all_cohort_hosps(excel_file_name, lookup_dict, keyname, camc_cohort_list,
                                                       ['Excess Days', 'Encounter Month'], measure_nm,
                                                       edac_readm_choice_num)

            else:
                results = filter_func_all_cohort_hosps(excel_file_name, lookup_dict, keyname, camc_cohort_list,
                                                       ['Excess Days', 'Encounter Month'], measure_nm,
                                                       edac_readm_choice_num, edac_readm_month_list)

            return (results)

        elif measure_nm.split('_')[0] == 'READM':

            if edac_readm_choice_num == 1:
                results = filter_func_all_cohort_hosps(excel_file_name, lookup_dict, keyname, camc_cohort_list,
                                                       ['Revisit Inpatient Cases', 'Encounter Month'],
                                                       measure_nm, edac_readm_choice_num)
            else:
                results = filter_func_all_cohort_hosps(excel_file_name, lookup_dict, keyname, camc_cohort_list,
                                                       ['Revisit Inpatient Cases', 'Encounter Month'],
                                                       measure_nm, edac_readm_choice_num, edac_readm_month_list)
            return (results)

    elif cohort_nm == 'Comprehensive Academic Medical Center' and domain_nm == 'Efficiency':
        if measure_nm.split('_')[0] == 'DCOST':

            results = filter_func_all_cohort_hosps(excel_file_name, lookup_dict, keyname, camc_cohort_list,
                                                   'Mean Direct Cost (Obs)',
                                                   measure_nm, edac_readm_choice_num)
            return (results)

        elif measure_nm.split('_')[0] == 'LOS':

            results = filter_func_all_cohort_hosps(excel_file_name, lookup_dict, keyname, camc_cohort_list,
                                                   'Mean LOS (Obs)', measure_nm, edac_readm_choice_num)
            return (results)

    elif cohort_nm == 'Comprehensive Academic Medical Center' and domain_nm == 'Safety':
        if measure_nm.split('_')[0] == 'PSI':
            if measure_nm == 'PSI_03':

                results = filter_func_all_cohort_hosps(excel_file_name, lookup_dict, keyname, camc_cohort_list,
                                                       'AHRQ Safety Numerator',
                                                       measure_nm, edac_readm_choice_num)
                return (results)

            elif measure_nm == 'PSI_06':

                results = filter_func_all_cohort_hosps(excel_file_name, lookup_dict, keyname, camc_cohort_list,
                                                       'AHRQ Safety Numerator',
                                                       measure_nm, edac_readm_choice_num)
                return (results)

            elif measure_nm == 'PSI_09':

                results = filter_func_all_cohort_hosps(excel_file_name, lookup_dict, keyname, camc_cohort_list,
                                                       'AHRQ Safety Numerator',
                                                       measure_nm, edac_readm_choice_num)
                return (results)

            elif measure_nm == 'PSI_11':

                results = filter_func_all_cohort_hosps(excel_file_name, lookup_dict, keyname, camc_cohort_list,
                                                       'AHRQ Safety Numerator',
                                                       measure_nm, edac_readm_choice_num)
                return (results)

            elif measure_nm == 'PSI_13':

                results = filter_func_all_cohort_hosps(excel_file_name, lookup_dict, keyname, camc_cohort_list,
                                                       'AHRQ Safety Numerator',
                                                       measure_nm, edac_readm_choice_num)
                return (results)

        elif measure_nm == 'THK':

            results = filter_func_all_cohort_hosps(excel_file_name, lookup_dict, keyname, camc_cohort_list,
                                                   'Total THK cases with Complications', measure_nm,
                                                   edac_readm_choice_num)
            return (results)

    elif cohort_nm == 'Comprehensive Academic Medical Center' and domain_nm == 'Outpatient':

        results = filter_func_all_cohort_hosps(excel_file_name, lookup_dict, keyname, camc_cohort_list,
                                               'Readmit Rate Num Cases (Readmit Cases)',
                                               measure_nm, edac_readm_choice_num)
        return (results)

    # LSCCMC logic:
    elif cohort_nm == 'Large Specialized Complex Care Medical Center' and domain_nm == 'Mortality':

        results = filter_func_all_cohort_hosps(excel_file_name, lookup_dict, keyname, lsccmc_cohort_list,
                                               'Deaths (Obs)', measure_nm, edac_readm_choice_num)
        return (results)

    elif cohort_nm == 'Large Specialized Complex Care Medical Center' and domain_nm == 'Effectiveness':
        print('got into LSCCMC effectiveness')
        if measure_nm.split('_')[0] == 'EDAC':
            #print('got to EDAC')
            if edac_readm_choice_num == 1:
                results = filter_func_all_cohort_hosps(excel_file_name, lookup_dict, keyname, lsccmc_cohort_list,
                                                       ['Excess Days', 'Encounter Month'], measure_nm,
                                                       edac_readm_choice_num)
                # 'Excess Days Per 100 Index Encounters', measure_nm)
            else:
                results = filter_func_all_cohort_hosps(excel_file_name, lookup_dict, keyname, lsccmc_cohort_list,
                                                       ['Excess Days', 'Encounter Month'], measure_nm,
                                                       edac_readm_choice_num, edac_readm_month_list)
                # 'Excess Days Per 100 Index Encounters', measure_nm)

            return (results)

        elif measure_nm.split('_')[0] == 'READM':

            if edac_readm_choice_num == 1:
                results = filter_func_all_cohort_hosps(excel_file_name, lookup_dict, keyname, lsccmc_cohort_list,
                                                       ['Revisit Inpatient Cases', 'Encounter Month'],
                                                       measure_nm, edac_readm_choice_num)
            else:
                results = filter_func_all_cohort_hosps(excel_file_name, lookup_dict, keyname, lsccmc_cohort_list,
                                                       ['Revisit Inpatient Cases', 'Encounter Month'],
                                                       measure_nm, edac_readm_choice_num, edac_readm_month_list)
            return (results)

    elif cohort_nm == 'Large Specialized Complex Care Medical Center' and domain_nm == 'Efficiency':
        if measure_nm.split('_')[0] == 'DCOST':

            results = filter_func_all_cohort_hosps(excel_file_name, lookup_dict, keyname, lsccmc_cohort_list,
                                                   'Mean Direct Cost (Obs)',
                                                   measure_nm, edac_readm_choice_num)
            return (results)

        elif measure_nm.split('_')[0] == 'LOS':

            results = filter_func_all_cohort_hosps(excel_file_name, lookup_dict, keyname, lsccmc_cohort_list,
                                                   'Mean LOS (Obs)', measure_nm, edac_readm_choice_num)
            return (results)

    elif cohort_nm == 'Large Specialized Complex Care Medical Center' and domain_nm == 'Safety':
        if measure_nm.split('_')[0] == 'PSI':
            if measure_nm == 'PSI_03':

                results = filter_func_all_cohort_hosps(excel_file_name, lookup_dict, keyname, lsccmc_cohort_list,
                                                       'AHRQ Safety Numerator',
                                                       measure_nm, edac_readm_choice_num)
                return (results)

            elif measure_nm == 'PSI_06':

                results = filter_func_all_cohort_hosps(excel_file_name, lookup_dict, keyname, lsccmc_cohort_list,
                                                       'AHRQ Safety Numerator',
                                                       measure_nm, edac_readm_choice_num)
                return (results)

            elif measure_nm == 'PSI_09':

                results = filter_func_all_cohort_hosps(excel_file_name, lookup_dict, keyname, lsccmc_cohort_list,
                                                       'AHRQ Safety Numerator',
                                                       measure_nm, edac_readm_choice_num)
                return (results)

            elif measure_nm == 'PSI_11':

                results = filter_func_all_cohort_hosps(excel_file_name, lookup_dict, keyname, lsccmc_cohort_list,
                                                       'AHRQ Safety Numerator',
                                                       measure_nm, edac_readm_choice_num)
                return (results)

            elif measure_nm == 'PSI_13':

                results = filter_func_all_cohort_hosps(excel_file_name, lookup_dict, keyname, lsccmc_cohort_list,
                                                       'AHRQ Safety Numerator',
                                                       measure_nm, edac_readm_choice_num)
                return (results)

        elif measure_nm == 'THK':

            results = filter_func_all_cohort_hosps(excel_file_name, lookup_dict, keyname, lsccmc_cohort_list,
                                                   'Total THK cases with Complications', measure_nm,
                                                   edac_readm_choice_num)
            return (results)



    elif cohort_nm == 'Large Specialized Complex Care Medical Center' and domain_nm == 'Outpatient':

        results = filter_func_all_cohort_hosps(excel_file_name, lookup_dict, keyname, lsccmc_cohort_list,
                                               'Readmit Rate Num Cases (Readmit Cases)',
                                               measure_nm, edac_readm_choice_num)
        return (results)

    # CCMC Logic:
    elif cohort_nm == 'Complex Care Medical Center' and domain_nm == 'Mortality':
        # print('got to ccmc')

        results = filter_func_all_cohort_hosps(excel_file_name, lookup_dict, keyname, ccmc_cohort_list, 'Deaths (Obs)',
                                               measure_nm, edac_readm_choice_num)
        return (results)

    elif cohort_nm == 'Complex Care Medical Center' and domain_nm == 'Effectiveness':

        if measure_nm.split('_')[0] == 'EDAC':

            if edac_readm_choice_num == 1:
                results = filter_func_all_cohort_hosps(excel_file_name, lookup_dict, keyname, ccmc_cohort_list,
                                                       ['Excess Days', 'Encounter Month'], measure_nm,
                                                       edac_readm_choice_num)
                # 'Excess Days Per 100 Index Encounters', measure_nm)
            else:
                results = filter_func_all_cohort_hosps(excel_file_name, lookup_dict, keyname, ccmc_cohort_list,
                                                       ['Excess Days', 'Encounter Month'], measure_nm,
                                                       edac_readm_choice_num, edac_readm_month_list)
                # 'Excess Days Per 100 Index Encounters', measure_nm)

            return (results)

        elif measure_nm.split('_')[0] == 'READM':

            if edac_readm_choice_num == 1:
                results = filter_func_all_cohort_hosps(excel_file_name, lookup_dict, keyname, ccmc_cohort_list,
                                                       ['Revisit Inpatient Cases', 'Encounter Month'],
                                                       measure_nm, edac_readm_choice_num)
            else:
                results = filter_func_all_cohort_hosps(excel_file_name, lookup_dict, keyname, ccmc_cohort_list,
                                                       ['Revisit Inpatient Cases', 'Encounter Month'],
                                                       measure_nm, edac_readm_choice_num, edac_readm_month_list)
            return (results)

    elif cohort_nm == 'Complex Care Medical Center' and domain_nm == 'Efficiency':
        if measure_nm.split('_')[0] == 'DCOST':

            results = filter_func_all_cohort_hosps(excel_file_name, lookup_dict, keyname, ccmc_cohort_list,
                                                   'Mean Direct Cost (Obs)',
                                                   measure_nm, edac_readm_choice_num)
            return (results)

        elif measure_nm.split('_')[0] == 'LOS':

            results = filter_func_all_cohort_hosps(excel_file_name, lookup_dict, keyname, ccmc_cohort_list,
                                                   'Mean LOS (Obs)', measure_nm, edac_readm_choice_num)
            return (results)

    elif cohort_nm == 'Complex Care Medical Center' and domain_nm == 'Safety':
        if measure_nm.split('_')[0] == 'PSI':
            if measure_nm == 'PSI_03':

                results = filter_func_all_cohort_hosps(excel_file_name, lookup_dict, keyname, ccmc_cohort_list,
                                                       'AHRQ Safety Numerator',
                                                       measure_nm, edac_readm_choice_num)
                return (results)

            elif measure_nm == 'PSI_06':

                results = filter_func_all_cohort_hosps(excel_file_name, lookup_dict, keyname, ccmc_cohort_list,
                                                       'AHRQ Safety Numerator',
                                                       measure_nm, edac_readm_choice_num)
                return (results)

            elif measure_nm == 'PSI_09':

                results = filter_func_all_cohort_hosps(excel_file_name, lookup_dict, keyname, ccmc_cohort_list,
                                                       'AHRQ Safety Numerator',
                                                       measure_nm, edac_readm_choice_num)
                return (results)

            elif measure_nm == 'PSI_11':

                results = filter_func_all_cohort_hosps(excel_file_name, lookup_dict, keyname, ccmc_cohort_list,
                                                       'AHRQ Safety Numerator',
                                                       measure_nm, edac_readm_choice_num)
                return (results)

            elif measure_nm == 'PSI_13':

                results = filter_func_all_cohort_hosps(excel_file_name, lookup_dict, keyname, ccmc_cohort_list,
                                                       'AHRQ Safety Numerator',
                                                       measure_nm, edac_readm_choice_num)
                return (results)

        elif measure_nm == 'THK':

            results = filter_func_all_cohort_hosps(excel_file_name, lookup_dict, keyname, ccmc_cohort_list,
                                                   'Total THK cases with Complications', measure_nm,
                                                   edac_readm_choice_num)
            return (results)

    elif cohort_nm == 'Complex Care Medical Center' and domain_nm == 'Outpatient':

        results = filter_func_all_cohort_hosps(excel_file_name, lookup_dict, keyname, ccmc_cohort_list,
                                               'Readmit Rate Num Cases (Readmit Cases)',
                                               measure_nm, edac_readm_choice_num)
        return (results)

    # COMM Logic:
    elif cohort_nm == 'Community Medical Center' and domain_nm == 'Mortality':
        # print('got to ccmc')

        results = filter_func_all_cohort_hosps(excel_file_name, lookup_dict, keyname, comm_cohort_list, 'Deaths (Obs)',
                                               measure_nm, edac_readm_choice_num)
        return (results)

    elif cohort_nm == 'Community Medical Center' and domain_nm == 'Effectiveness':

        if measure_nm.split('_')[0] == 'EDAC':

            if edac_readm_choice_num == 1:
                results = filter_func_all_cohort_hosps(excel_file_name, lookup_dict, keyname, comm_cohort_list,
                                                       ['Excess Days', 'Encounter Month'], measure_nm,
                                                       edac_readm_choice_num)
                # 'Excess Days Per 100 Index Encounters', measure_nm)
            else:
                results = filter_func_all_cohort_hosps(excel_file_name, lookup_dict, keyname, comm_cohort_list,
                                                       ['Excess Days', 'Encounter Month'], measure_nm,
                                                       edac_readm_choice_num, edac_readm_month_list)
                # 'Excess Days Per 100 Index Encounters', measure_nm)

            return (results)

        elif measure_nm.split('_')[0] == 'READM':

            if edac_readm_choice_num == 1:
                results = filter_func_all_cohort_hosps(excel_file_name, lookup_dict, keyname, comm_cohort_list,
                                                       ['Revisit Inpatient Cases', 'Encounter Month'],
                                                       measure_nm, edac_readm_choice_num)
            else:
                results = filter_func_all_cohort_hosps(excel_file_name, lookup_dict, keyname, comm_cohort_list,
                                                       ['Revisit Inpatient Cases', 'Encounter Month'],
                                                       measure_nm, edac_readm_choice_num, edac_readm_month_list)
            return (results)

    elif cohort_nm == 'Community Medical Center' and domain_nm == 'Efficiency':
        if measure_nm.split('_')[0] == 'DCOST':

            results = filter_func_all_cohort_hosps(excel_file_name, lookup_dict, keyname, comm_cohort_list,
                                                   'Mean Direct Cost (Obs)',
                                                   measure_nm, edac_readm_choice_num)
            return (results)

        elif measure_nm.split('_')[0] == 'LOS':

            results = filter_func_all_cohort_hosps(excel_file_name, lookup_dict, keyname, comm_cohort_list,
                                                   'Mean LOS (Obs)', measure_nm, edac_readm_choice_num)
            return (results)

    elif cohort_nm == 'Community Medical Center' and domain_nm == 'Safety':
        if measure_nm.split('_')[0] == 'PSI':
            if measure_nm == 'PSI_03':

                results = filter_func_all_cohort_hosps(excel_file_name, lookup_dict, keyname, comm_cohort_list,
                                                       'AHRQ Safety Numerator',
                                                       measure_nm, edac_readm_choice_num)
                return (results)

            elif measure_nm == 'PSI_06':

                results = filter_func_all_cohort_hosps(excel_file_name, lookup_dict, keyname, comm_cohort_list,
                                                       'AHRQ Safety Numerator',
                                                       measure_nm, edac_readm_choice_num)
                return (results)

            elif measure_nm == 'PSI_09':

                results = filter_func_all_cohort_hosps(excel_file_name, lookup_dict, keyname, comm_cohort_list,
                                                       'AHRQ Safety Numerator',
                                                       measure_nm, edac_readm_choice_num)
                return (results)

            elif measure_nm == 'PSI_11':

                results = filter_func_all_cohort_hosps(excel_file_name, lookup_dict, keyname, comm_cohort_list,
                                                       'AHRQ Safety Numerator',
                                                       measure_nm, edac_readm_choice_num)
                return (results)

            elif measure_nm == 'PSI_13':

                results = filter_func_all_cohort_hosps(excel_file_name, lookup_dict, keyname, comm_cohort_list,
                                                       'AHRQ Safety Numerator',
                                                       measure_nm, edac_readm_choice_num)
                return (results)

        elif measure_nm == 'THK':

            results = filter_func_all_cohort_hosps(excel_file_name, lookup_dict, keyname, comm_cohort_list,
                                                   'Total THK cases with Complications', measure_nm,
                                                   edac_readm_choice_num)
            return (results)

    elif cohort_nm == 'Community Medical Center' and domain_nm == 'Outpatient':

        results = filter_func_all_cohort_hosps(excel_file_name, lookup_dict, keyname, comm_cohort_list,
                                               'Readmit Rate Num Cases (Readmit Cases)',
                                               measure_nm, edac_readm_choice_num)
        return (results)
'''
#UL003

def file_filter_switcher_all_cohort_hosps_n_events(cohort_nm, domain_nm, measure_nm, excel_file_name, lookup_dict,
                                                   keyname, edac_readm_choice_num, camc_cohort_list, lsccmc_cohort_list,
                                                   ccmc_cohort_list, comm_cohort_list,casc_cohort_list, edac_readm_month_list='No List'):
    #print('got to n size switcher')

    if cohort_nm == 'Comprehensive Academic Medical Center' and domain_nm == 'Mortality':

        results = filter_func_all_cohort_hosps(excel_file_name, lookup_dict, keyname, camc_cohort_list, 'Deaths (Obs)',
                                               measure_nm, edac_readm_choice_num)
        return (results)

    elif cohort_nm == 'Comprehensive Academic Medical Center' and domain_nm == 'Effectiveness':

        if measure_nm.split('_')[0] == 'EDAC':

            if edac_readm_choice_num == 1:
                results = filter_func_all_cohort_hosps(excel_file_name, lookup_dict, keyname, camc_cohort_list,
                                                       ['Excess Days', 'Encounter Month'], measure_nm,
                                                       edac_readm_choice_num)

            else:
                results = filter_func_all_cohort_hosps(excel_file_name, lookup_dict, keyname, camc_cohort_list,
                                                       ['Excess Days', 'Encounter Month'], measure_nm,
                                                       edac_readm_choice_num, edac_readm_month_list)

            return (results)

        elif measure_nm.split('_')[0] == 'READM' and measure_nm.split('_')[1] not in ['ARTHRO','CHOL','COLON','URI']:

            if edac_readm_choice_num == 1:
                results = filter_func_all_cohort_hosps(excel_file_name, lookup_dict, keyname, camc_cohort_list,
                                                       ['Revisit Inpatient Cases', 'Encounter Month'],
                                                       measure_nm, edac_readm_choice_num)
            else:
                results = filter_func_all_cohort_hosps(excel_file_name, lookup_dict, keyname, camc_cohort_list,
                                                       ['Revisit Inpatient Cases', 'Encounter Month'],
                                                       measure_nm, edac_readm_choice_num, edac_readm_month_list)
            return (results)
        
        #UL003
        elif measure_nm.split('_')[0] == 'READM' and domain_nm == 'Effectiveness' and measure_nm.split('_')[1] in ['ARTHRO','CHOL','COLON','URI']:

            results = filter_func_all_cohort_hosps(excel_file_name, lookup_dict, keyname, camc_cohort_list,
                                               'Readmit Rate Num Cases (Readmit Cases)',
                                               measure_nm, edac_readm_choice_num)
            return (results)

    elif cohort_nm == 'Comprehensive Academic Medical Center' and domain_nm == 'Efficiency':
        if measure_nm.split('_')[0] == 'DCOST':

            results = filter_func_all_cohort_hosps(excel_file_name, lookup_dict, keyname, camc_cohort_list,
                                                   'Mean Direct Cost (Obs)',
                                                   measure_nm, edac_readm_choice_num)
            return (results)

        elif measure_nm.split('_')[0] == 'LOS':

            results = filter_func_all_cohort_hosps(excel_file_name, lookup_dict, keyname, camc_cohort_list,
                                                   'Mean LOS (Obs)', measure_nm, edac_readm_choice_num)
            return (results)

    elif cohort_nm == 'Comprehensive Academic Medical Center' and domain_nm == 'Safety':
        if measure_nm.split('_')[0] == 'PSI':
            if measure_nm == 'PSI_03':

                results = filter_func_all_cohort_hosps(excel_file_name, lookup_dict, keyname, camc_cohort_list,
                                                       'AHRQ Safety Numerator',
                                                       measure_nm, edac_readm_choice_num)
                return (results)

            elif measure_nm == 'PSI_06':

                results = filter_func_all_cohort_hosps(excel_file_name, lookup_dict, keyname, camc_cohort_list,
                                                       'AHRQ Safety Numerator',
                                                       measure_nm, edac_readm_choice_num)
                return (results)

            elif measure_nm == 'PSI_09':

                results = filter_func_all_cohort_hosps(excel_file_name, lookup_dict, keyname, camc_cohort_list,
                                                       'AHRQ Safety Numerator',
                                                       measure_nm, edac_readm_choice_num)
                return (results)

            elif measure_nm == 'PSI_11':

                results = filter_func_all_cohort_hosps(excel_file_name, lookup_dict, keyname, camc_cohort_list,
                                                       'AHRQ Safety Numerator',
                                                       measure_nm, edac_readm_choice_num)
                return (results)

            elif measure_nm == 'PSI_13':

                results = filter_func_all_cohort_hosps(excel_file_name, lookup_dict, keyname, camc_cohort_list,
                                                       'AHRQ Safety Numerator',
                                                       measure_nm, edac_readm_choice_num)
                return (results)

        elif measure_nm == 'THK':

            results = filter_func_all_cohort_hosps(excel_file_name, lookup_dict, keyname, camc_cohort_list,
                                                   'Total THK cases with Complications', measure_nm,
                                                   edac_readm_choice_num)
            return (results)

    elif cohort_nm == 'Comprehensive Academic Medical Center' and domain_nm == 'Outpatient':

        results = filter_func_all_cohort_hosps(excel_file_name, lookup_dict, keyname, camc_cohort_list,
                                               'Readmit Rate Num Cases (Readmit Cases)',
                                               measure_nm, edac_readm_choice_num)
        return (results)
    

    # LSCCMC logic:
    elif cohort_nm == 'Large Specialized Complex Care Medical Center' and domain_nm == 'Mortality':

        results = filter_func_all_cohort_hosps(excel_file_name, lookup_dict, keyname, lsccmc_cohort_list,
                                               'Deaths (Obs)', measure_nm, edac_readm_choice_num)
        return (results)

    elif cohort_nm == 'Large Specialized Complex Care Medical Center' and domain_nm == 'Effectiveness':
        print('got into LSCCMC effectiveness')
        if measure_nm.split('_')[0] == 'EDAC':
            #print('got to EDAC')
            if edac_readm_choice_num == 1:
                results = filter_func_all_cohort_hosps(excel_file_name, lookup_dict, keyname, lsccmc_cohort_list,
                                                       ['Excess Days', 'Encounter Month'], measure_nm,
                                                       edac_readm_choice_num)
                # 'Excess Days Per 100 Index Encounters', measure_nm)
            else:
                results = filter_func_all_cohort_hosps(excel_file_name, lookup_dict, keyname, lsccmc_cohort_list,
                                                       ['Excess Days', 'Encounter Month'], measure_nm,
                                                       edac_readm_choice_num, edac_readm_month_list)
                # 'Excess Days Per 100 Index Encounters', measure_nm)

            return (results)

        elif measure_nm.split('_')[0] == 'READM' and measure_nm.split('_')[1] not in ['ARTHRO','CHOL','COLON','URI']:

            if edac_readm_choice_num == 1:
                results = filter_func_all_cohort_hosps(excel_file_name, lookup_dict, keyname, lsccmc_cohort_list,
                                                       ['Revisit Inpatient Cases', 'Encounter Month'],
                                                       measure_nm, edac_readm_choice_num)
            else:
                results = filter_func_all_cohort_hosps(excel_file_name, lookup_dict, keyname, lsccmc_cohort_list,
                                                       ['Revisit Inpatient Cases', 'Encounter Month'],
                                                       measure_nm, edac_readm_choice_num, edac_readm_month_list)
            return (results)
        
        #UL003
        elif measure_nm.split('_')[0] == 'READM' and domain_nm == 'Effectiveness' and measure_nm.split('_')[1] in ['ARTHRO','CHOL','COLON','URI']:

            results = filter_func_all_cohort_hosps(excel_file_name, lookup_dict, keyname, lsccmc_cohort_list,
                                               'Readmit Rate Num Cases (Readmit Cases)',
                                               measure_nm, edac_readm_choice_num)
            return (results)

    elif cohort_nm == 'Large Specialized Complex Care Medical Center' and domain_nm == 'Efficiency':
        if measure_nm.split('_')[0] == 'DCOST':

            results = filter_func_all_cohort_hosps(excel_file_name, lookup_dict, keyname, lsccmc_cohort_list,
                                                   'Mean Direct Cost (Obs)',
                                                   measure_nm, edac_readm_choice_num)
            return (results)

        elif measure_nm.split('_')[0] == 'LOS':

            results = filter_func_all_cohort_hosps(excel_file_name, lookup_dict, keyname, lsccmc_cohort_list,
                                                   'Mean LOS (Obs)', measure_nm, edac_readm_choice_num)
            return (results)

    elif cohort_nm == 'Large Specialized Complex Care Medical Center' and domain_nm == 'Safety':
        if measure_nm.split('_')[0] == 'PSI':
            if measure_nm == 'PSI_03':

                results = filter_func_all_cohort_hosps(excel_file_name, lookup_dict, keyname, lsccmc_cohort_list,
                                                       'AHRQ Safety Numerator',
                                                       measure_nm, edac_readm_choice_num)
                return (results)

            elif measure_nm == 'PSI_06':

                results = filter_func_all_cohort_hosps(excel_file_name, lookup_dict, keyname, lsccmc_cohort_list,
                                                       'AHRQ Safety Numerator',
                                                       measure_nm, edac_readm_choice_num)
                return (results)

            elif measure_nm == 'PSI_09':

                results = filter_func_all_cohort_hosps(excel_file_name, lookup_dict, keyname, lsccmc_cohort_list,
                                                       'AHRQ Safety Numerator',
                                                       measure_nm, edac_readm_choice_num)
                return (results)

            elif measure_nm == 'PSI_11':

                results = filter_func_all_cohort_hosps(excel_file_name, lookup_dict, keyname, lsccmc_cohort_list,
                                                       'AHRQ Safety Numerator',
                                                       measure_nm, edac_readm_choice_num)
                return (results)

            elif measure_nm == 'PSI_13':

                results = filter_func_all_cohort_hosps(excel_file_name, lookup_dict, keyname, lsccmc_cohort_list,
                                                       'AHRQ Safety Numerator',
                                                       measure_nm, edac_readm_choice_num)
                return (results)

        elif measure_nm == 'THK':

            results = filter_func_all_cohort_hosps(excel_file_name, lookup_dict, keyname, lsccmc_cohort_list,
                                                   'Total THK cases with Complications', measure_nm,
                                                   edac_readm_choice_num)
            return (results)



    elif cohort_nm == 'Large Specialized Complex Care Medical Center' and domain_nm == 'Outpatient':

        results = filter_func_all_cohort_hosps(excel_file_name, lookup_dict, keyname, lsccmc_cohort_list,
                                               'Readmit Rate Num Cases (Readmit Cases)',
                                               measure_nm, edac_readm_choice_num)
        return (results)
    
    #UL003
    elif cohort_nm == 'Large Specialized Complex Care Medical Center' and domain_nm == 'Effectiveness' and measure_nm.split('_')[1] in ['ARTHRO','CHOL','COLON','URI']:

        results = filter_func_all_cohort_hosps(excel_file_name, lookup_dict, keyname, lsccmc_cohort_list,
                                               'Readmit Rate Num Cases (Readmit Cases)',
                                               measure_nm, edac_readm_choice_num)
        return (results)

    # CCMC Logic:
    elif cohort_nm == 'Complex Care Medical Center' and domain_nm == 'Mortality':
        # print('got to ccmc')

        results = filter_func_all_cohort_hosps(excel_file_name, lookup_dict, keyname, ccmc_cohort_list, 'Deaths (Obs)',
                                               measure_nm, edac_readm_choice_num)
        return (results)
    

    elif cohort_nm == 'Complex Care Medical Center' and domain_nm == 'Effectiveness':

        if measure_nm.split('_')[0] == 'EDAC':

            if edac_readm_choice_num == 1:
                results = filter_func_all_cohort_hosps(excel_file_name, lookup_dict, keyname, ccmc_cohort_list,
                                                       ['Excess Days', 'Encounter Month'], measure_nm,
                                                       edac_readm_choice_num)
                # 'Excess Days Per 100 Index Encounters', measure_nm)
            else:
                results = filter_func_all_cohort_hosps(excel_file_name, lookup_dict, keyname, ccmc_cohort_list,
                                                       ['Excess Days', 'Encounter Month'], measure_nm,
                                                       edac_readm_choice_num, edac_readm_month_list)
                # 'Excess Days Per 100 Index Encounters', measure_nm)

            return (results)

        elif measure_nm.split('_')[0] == 'READM' and measure_nm.split('_')[1] not in ['ARTHRO','CHOL','COLON','URI']:

            if edac_readm_choice_num == 1:
                results = filter_func_all_cohort_hosps(excel_file_name, lookup_dict, keyname, ccmc_cohort_list,
                                                       ['Revisit Inpatient Cases', 'Encounter Month'],
                                                       measure_nm, edac_readm_choice_num)
            else:
                results = filter_func_all_cohort_hosps(excel_file_name, lookup_dict, keyname, ccmc_cohort_list,
                                                       ['Revisit Inpatient Cases', 'Encounter Month'],
                                                       measure_nm, edac_readm_choice_num, edac_readm_month_list)
            return (results)
        
        #UL003
        elif measure_nm.split('_')[0] == 'READM' and domain_nm == 'Effectiveness' and measure_nm.split('_')[1] in ['ARTHRO','CHOL','COLON','URI']:

            results = filter_func_all_cohort_hosps(excel_file_name, lookup_dict, keyname, ccmc_cohort_list,
                                               'Readmit Rate Num Cases (Readmit Cases)',
                                               measure_nm, edac_readm_choice_num)
            return (results)

    elif cohort_nm == 'Complex Care Medical Center' and domain_nm == 'Efficiency':
        if measure_nm.split('_')[0] == 'DCOST':

            results = filter_func_all_cohort_hosps(excel_file_name, lookup_dict, keyname, ccmc_cohort_list,
                                                   'Mean Direct Cost (Obs)',
                                                   measure_nm, edac_readm_choice_num)
            return (results)

        elif measure_nm.split('_')[0] == 'LOS':

            results = filter_func_all_cohort_hosps(excel_file_name, lookup_dict, keyname, ccmc_cohort_list,
                                                   'Mean LOS (Obs)', measure_nm, edac_readm_choice_num)
            return (results)

    elif cohort_nm == 'Complex Care Medical Center' and domain_nm == 'Safety':
        if measure_nm.split('_')[0] == 'PSI':
            if measure_nm == 'PSI_03':

                results = filter_func_all_cohort_hosps(excel_file_name, lookup_dict, keyname, ccmc_cohort_list,
                                                       'AHRQ Safety Numerator',
                                                       measure_nm, edac_readm_choice_num)
                return (results)

            elif measure_nm == 'PSI_06':

                results = filter_func_all_cohort_hosps(excel_file_name, lookup_dict, keyname, ccmc_cohort_list,
                                                       'AHRQ Safety Numerator',
                                                       measure_nm, edac_readm_choice_num)
                return (results)

            elif measure_nm == 'PSI_09':

                results = filter_func_all_cohort_hosps(excel_file_name, lookup_dict, keyname, ccmc_cohort_list,
                                                       'AHRQ Safety Numerator',
                                                       measure_nm, edac_readm_choice_num)
                return (results)

            elif measure_nm == 'PSI_11':

                results = filter_func_all_cohort_hosps(excel_file_name, lookup_dict, keyname, ccmc_cohort_list,
                                                       'AHRQ Safety Numerator',
                                                       measure_nm, edac_readm_choice_num)
                return (results)

            elif measure_nm == 'PSI_13':

                results = filter_func_all_cohort_hosps(excel_file_name, lookup_dict, keyname, ccmc_cohort_list,
                                                       'AHRQ Safety Numerator',
                                                       measure_nm, edac_readm_choice_num)
                return (results)

        elif measure_nm == 'THK':

            results = filter_func_all_cohort_hosps(excel_file_name, lookup_dict, keyname, ccmc_cohort_list,
                                                   'Total THK cases with Complications', measure_nm,
                                                   edac_readm_choice_num)
            return (results)

    elif cohort_nm == 'Complex Care Medical Center' and domain_nm == 'Outpatient':

        results = filter_func_all_cohort_hosps(excel_file_name, lookup_dict, keyname, ccmc_cohort_list,
                                               'Readmit Rate Num Cases (Readmit Cases)',
                                               measure_nm, edac_readm_choice_num)
        return (results)
    

    # COMM Logic:
    elif cohort_nm == 'Community Medical Center' and domain_nm == 'Mortality':
        # print('got to ccmc')

        results = filter_func_all_cohort_hosps(excel_file_name, lookup_dict, keyname, comm_cohort_list, 'Deaths (Obs)',
                                               measure_nm, edac_readm_choice_num)
        return (results)

    elif cohort_nm == 'Community Medical Center' and domain_nm == 'Effectiveness':

        if measure_nm.split('_')[0] == 'EDAC':

            if edac_readm_choice_num == 1:
                results = filter_func_all_cohort_hosps(excel_file_name, lookup_dict, keyname, comm_cohort_list,
                                                       ['Excess Days', 'Encounter Month'], measure_nm,
                                                       edac_readm_choice_num)
                # 'Excess Days Per 100 Index Encounters', measure_nm)
            else:
                results = filter_func_all_cohort_hosps(excel_file_name, lookup_dict, keyname, comm_cohort_list,
                                                       ['Excess Days', 'Encounter Month'], measure_nm,
                                                       edac_readm_choice_num, edac_readm_month_list)
                # 'Excess Days Per 100 Index Encounters', measure_nm)

            return (results)

        elif measure_nm.split('_')[0] == 'READM' and measure_nm.split('_')[1] not in ['ARTHRO','CHOL','COLON','URI']:

            if edac_readm_choice_num == 1:
                results = filter_func_all_cohort_hosps(excel_file_name, lookup_dict, keyname, comm_cohort_list,
                                                       ['Revisit Inpatient Cases', 'Encounter Month'],
                                                       measure_nm, edac_readm_choice_num)
            else:
                results = filter_func_all_cohort_hosps(excel_file_name, lookup_dict, keyname, comm_cohort_list,
                                                       ['Revisit Inpatient Cases', 'Encounter Month'],
                                                       measure_nm, edac_readm_choice_num, edac_readm_month_list)
            return (results)
        
        #UL003
        elif measure_nm.split('_')[0] == 'READM' and domain_nm == 'Effectiveness' and measure_nm.split('_')[1] in ['ARTHRO','CHOL','COLON','URI']:

            results = filter_func_all_cohort_hosps(excel_file_name, lookup_dict, keyname, comm_cohort_list,
                                               'Readmit Rate Num Cases (Readmit Cases)',
                                               measure_nm, edac_readm_choice_num)
            return (results)

    elif cohort_nm == 'Community Medical Center' and domain_nm == 'Efficiency':
        if measure_nm.split('_')[0] == 'DCOST':

            results = filter_func_all_cohort_hosps(excel_file_name, lookup_dict, keyname, comm_cohort_list,
                                                   'Mean Direct Cost (Obs)',
                                                   measure_nm, edac_readm_choice_num)
            return (results)

        elif measure_nm.split('_')[0] == 'LOS':

            results = filter_func_all_cohort_hosps(excel_file_name, lookup_dict, keyname, comm_cohort_list,
                                                   'Mean LOS (Obs)', measure_nm, edac_readm_choice_num)
            return (results)

    elif cohort_nm == 'Community Medical Center' and domain_nm == 'Safety':
        if measure_nm.split('_')[0] == 'PSI':
            if measure_nm == 'PSI_03':

                results = filter_func_all_cohort_hosps(excel_file_name, lookup_dict, keyname, comm_cohort_list,
                                                       'AHRQ Safety Numerator',
                                                       measure_nm, edac_readm_choice_num)
                return (results)

            elif measure_nm == 'PSI_06':

                results = filter_func_all_cohort_hosps(excel_file_name, lookup_dict, keyname, comm_cohort_list,
                                                       'AHRQ Safety Numerator',
                                                       measure_nm, edac_readm_choice_num)
                return (results)

            elif measure_nm == 'PSI_09':

                results = filter_func_all_cohort_hosps(excel_file_name, lookup_dict, keyname, comm_cohort_list,
                                                       'AHRQ Safety Numerator',
                                                       measure_nm, edac_readm_choice_num)
                return (results)

            elif measure_nm == 'PSI_11':

                results = filter_func_all_cohort_hosps(excel_file_name, lookup_dict, keyname, comm_cohort_list,
                                                       'AHRQ Safety Numerator',
                                                       measure_nm, edac_readm_choice_num)
                return (results)

            elif measure_nm == 'PSI_13':

                results = filter_func_all_cohort_hosps(excel_file_name, lookup_dict, keyname, comm_cohort_list,
                                                       'AHRQ Safety Numerator',
                                                       measure_nm, edac_readm_choice_num)
                return (results)

        elif measure_nm == 'THK':

            results = filter_func_all_cohort_hosps(excel_file_name, lookup_dict, keyname, comm_cohort_list,
                                                   'Total THK cases with Complications', measure_nm,
                                                   edac_readm_choice_num)
            return (results)

    elif cohort_nm == 'Community Medical Center' and domain_nm == 'Outpatient':

        results = filter_func_all_cohort_hosps(excel_file_name, lookup_dict, keyname, comm_cohort_list,
                                               'Readmit Rate Num Cases (Readmit Cases)',
                                               measure_nm, edac_readm_choice_num)
        return (results)
    
    #UL003
    # CASC Logic:
    elif cohort_nm == 'Critical Access & Small Community' and domain_nm == 'Mortality':

        results = filter_func_all_cohort_hosps(excel_file_name, lookup_dict, keyname, casc_cohort_list, 'Deaths (Obs)',
                                               measure_nm, edac_readm_choice_num)
        return (results)
    
    elif cohort_nm == 'Critical Access & Small Community' and domain_nm == 'Efficiency':
        if measure_nm.split('_')[0] == 'DCOST':

            results = filter_func_all_cohort_hosps(excel_file_name, lookup_dict, keyname, casc_cohort_list,
                                                   'Mean Direct Cost (Obs)',
                                                   measure_nm, edac_readm_choice_num)
            return (results)

        elif measure_nm.split('_')[0] == 'LOS':

            results = filter_func_all_cohort_hosps(excel_file_name, lookup_dict, keyname, casc_cohort_list,
                                                   'Mean LOS (Obs)', measure_nm, edac_readm_choice_num)
            return (results)
        
    elif cohort_nm == 'Critical Access & Small Community' and domain_nm == 'Effectiveness' and measure_nm.split('_')[0] not in ['ADE','EARLY']:

        if measure_nm.split('_')[0] == 'EDAC':

            if edac_readm_choice_num == 1:
                results = filter_func_all_cohort_hosps(excel_file_name, lookup_dict, keyname, casc_cohort_list,
                                                       ['Excess Days', 'Encounter Month'], measure_nm,
                                                       edac_readm_choice_num)

            else:
                results = filter_func_all_cohort_hosps(excel_file_name, lookup_dict, keyname, casc_cohort_list,
                                                       ['Excess Days', 'Encounter Month'], measure_nm,
                                                       edac_readm_choice_num, edac_readm_month_list)

            return (results)

        elif measure_nm.split('_')[0] == 'READM' and measure_nm.split('_')[1] not in ['ARTHRO','CHOL','COLON','URI']:

            if edac_readm_choice_num == 1:
                results = filter_func_all_cohort_hosps(excel_file_name, lookup_dict, keyname, casc_cohort_list,
                                                       ['Revisit Inpatient Cases', 'Encounter Month'],
                                                       measure_nm, edac_readm_choice_num)
            else:
                results = filter_func_all_cohort_hosps(excel_file_name, lookup_dict, keyname, casc_cohort_list,
                                                       ['Revisit Inpatient Cases', 'Encounter Month'],
                                                       measure_nm, edac_readm_choice_num, edac_readm_month_list)
            return (results)
        
        #UL003
        elif measure_nm.split('_')[0] == 'READM' and domain_nm == 'Effectiveness' and measure_nm.split('_')[1] in ['ARTHRO','CHOL','COLON','URI']:

            results = filter_func_all_cohort_hosps(excel_file_name, lookup_dict, keyname, casc_cohort_list,
                                               'Readmit Rate Num Cases (Readmit Cases)',
                                               measure_nm, edac_readm_choice_num)
            return (results)
        
        elif domain_nm == 'Effectiveness' and measure_nm.split('_')[1] in ['REVISIT','RETURN']:

            results = filter_func_all_cohort_hosps(excel_file_name, lookup_dict, keyname, casc_cohort_list,
                                               'Readmit Rate Num Cases (Readmit Cases)',
                                               measure_nm, edac_readm_choice_num)
            return (results)
        
    elif measure_nm.split('_')[0] in ['ADE','EARLY']:

        results = filter_func_all_cohort_hosps(excel_file_name, lookup_dict, keyname, casc_cohort_list,
                                               'Cases',
                                               measure_nm, edac_readm_choice_num)
        return (results)

###############################################################################################

# same function as above but for denominator columns.
'''
def file_filter_switcher_all_cohort_hosps_d_events(cohort_nm, domain_nm, measure_nm, excel_file_name, lookup_dict,
                                                   keyname, edac_readm_choice_num, camc_cohort_list, lsccmc_cohort_list,
                                                   ccmc_cohort_list, comm_cohort_list, edac_readm_month_list='No List'):
    #print('got to n size switcher')

    if cohort_nm == 'Comprehensive Academic Medical Center' and domain_nm == 'Mortality':

        results = filter_func_all_cohort_hosps(excel_file_name, lookup_dict, keyname, camc_cohort_list, 'Cases',
                                               measure_nm, edac_readm_choice_num)
        return (results)

    elif cohort_nm == 'Comprehensive Academic Medical Center' and domain_nm == 'Effectiveness':

        if measure_nm.split('_')[0] == 'EDAC':

            if edac_readm_choice_num == 1:
                results = filter_func_all_cohort_hosps(excel_file_name, lookup_dict, keyname, camc_cohort_list,
                                                       ['Total Index Encounters', 'Encounter Month'], measure_nm,
                                                       edac_readm_choice_num)

            else:
                results = filter_func_all_cohort_hosps(excel_file_name, lookup_dict, keyname, camc_cohort_list,
                                                       ['Total Index Encounters', 'Encounter Month'], measure_nm,
                                                       edac_readm_choice_num, edac_readm_month_list)

            return (results)

        elif measure_nm.split('_')[0] == 'READM':

            if edac_readm_choice_num == 1:
                results = filter_func_all_cohort_hosps(excel_file_name, lookup_dict, keyname, camc_cohort_list,
                                                       ['Total Index Encounters', 'Encounter Month'],
                                                       measure_nm, edac_readm_choice_num)
            else:
                results = filter_func_all_cohort_hosps(excel_file_name, lookup_dict, keyname, camc_cohort_list,
                                                       ['Total Index Encounters', 'Encounter Month'],
                                                       measure_nm, edac_readm_choice_num, edac_readm_month_list)
            return (results)

    elif cohort_nm == 'Comprehensive Academic Medical Center' and domain_nm == 'Efficiency':
        if measure_nm.split('_')[0] == 'DCOST':

            results = filter_func_all_cohort_hosps(excel_file_name, lookup_dict, keyname, camc_cohort_list,
                                                   'Cases',
                                                   measure_nm, edac_readm_choice_num)
            return (results)

        elif measure_nm.split('_')[0] == 'LOS':

            results = filter_func_all_cohort_hosps(excel_file_name, lookup_dict, keyname, camc_cohort_list,
                                                   'Cases', measure_nm, edac_readm_choice_num)
            return (results)

    elif cohort_nm == 'Comprehensive Academic Medical Center' and domain_nm == 'Safety':
        if measure_nm.split('_')[0] == 'PSI':
            if measure_nm == 'PSI_03':

                results = filter_func_all_cohort_hosps(excel_file_name, lookup_dict, keyname, camc_cohort_list,
                                                       'AHRQ Safety Denominator',
                                                       measure_nm, edac_readm_choice_num)
                return (results)

            elif measure_nm == 'PSI_06':

                results = filter_func_all_cohort_hosps(excel_file_name, lookup_dict, keyname, camc_cohort_list,
                                                       'AHRQ Safety Denominator',
                                                       measure_nm, edac_readm_choice_num)
                return (results)

            elif measure_nm == 'PSI_09':

                results = filter_func_all_cohort_hosps(excel_file_name, lookup_dict, keyname, camc_cohort_list,
                                                       'AHRQ Safety Denominator',
                                                       measure_nm, edac_readm_choice_num)
                return (results)

            elif measure_nm == 'PSI_11':

                results = filter_func_all_cohort_hosps(excel_file_name, lookup_dict, keyname, camc_cohort_list,
                                                       'AHRQ Safety Denominator',
                                                       measure_nm, edac_readm_choice_num)
                return (results)

            elif measure_nm == 'PSI_13':

                results = filter_func_all_cohort_hosps(excel_file_name, lookup_dict, keyname, camc_cohort_list,
                                                       'AHRQ Safety Denominator',
                                                       measure_nm, edac_readm_choice_num)
                return (results)

        elif measure_nm == 'THK':

            results = filter_func_all_cohort_hosps(excel_file_name, lookup_dict, keyname, camc_cohort_list,
                                                   'Total THK cases', measure_nm,
                                                   edac_readm_choice_num)
            return (results)

    elif cohort_nm == 'Comprehensive Academic Medical Center' and domain_nm == 'Outpatient':

        results = filter_func_all_cohort_hosps(excel_file_name, lookup_dict, keyname, camc_cohort_list,
                                               'Readmit Rate Denom Cases',
                                               measure_nm, edac_readm_choice_num)
        return (results)

    # LSCCMC logic:
    elif cohort_nm == 'Large Specialized Complex Care Medical Center' and domain_nm == 'Mortality':

        results = filter_func_all_cohort_hosps(excel_file_name, lookup_dict, keyname, lsccmc_cohort_list,
                                               'Cases', measure_nm, edac_readm_choice_num)
        return (results)

    elif cohort_nm == 'Large Specialized Complex Care Medical Center' and domain_nm == 'Effectiveness':
        #print('got into LSCCMC effectiveness')
        if measure_nm.split('_')[0] == 'EDAC':
            print('got to EDAC')
            if edac_readm_choice_num == 1:
                results = filter_func_all_cohort_hosps(excel_file_name, lookup_dict, keyname, lsccmc_cohort_list,
                                                       ['Total Index Encounters', 'Encounter Month'], measure_nm,
                                                       edac_readm_choice_num)
                # 'Excess Days Per 100 Index Encounters', measure_nm)
            else:
                results = filter_func_all_cohort_hosps(excel_file_name, lookup_dict, keyname, lsccmc_cohort_list,
                                                       ['Total Index Encounters', 'Encounter Month'], measure_nm,
                                                       edac_readm_choice_num, edac_readm_month_list)
                # 'Excess Days Per 100 Index Encounters', measure_nm)

            return (results)

        elif measure_nm.split('_')[0] == 'READM':

            if edac_readm_choice_num == 1:
                results = filter_func_all_cohort_hosps(excel_file_name, lookup_dict, keyname, lsccmc_cohort_list,
                                                       ['Total Index Encounters', 'Encounter Month'],
                                                       measure_nm, edac_readm_choice_num)
            else:
                results = filter_func_all_cohort_hosps(excel_file_name, lookup_dict, keyname, lsccmc_cohort_list,
                                                       ['Total Index Encounters', 'Encounter Month'],
                                                       measure_nm, edac_readm_choice_num, edac_readm_month_list)
            return (results)

    elif cohort_nm == 'Large Specialized Complex Care Medical Center' and domain_nm == 'Efficiency':
        if measure_nm.split('_')[0] == 'DCOST':

            results = filter_func_all_cohort_hosps(excel_file_name, lookup_dict, keyname, lsccmc_cohort_list,
                                                   'Cases',
                                                   measure_nm, edac_readm_choice_num)
            return (results)

        elif measure_nm.split('_')[0] == 'LOS':

            results = filter_func_all_cohort_hosps(excel_file_name, lookup_dict, keyname, lsccmc_cohort_list,
                                                   'Cases', measure_nm, edac_readm_choice_num)
            return (results)

    elif cohort_nm == 'Large Specialized Complex Care Medical Center' and domain_nm == 'Safety':
        if measure_nm.split('_')[0] == 'PSI':
            if measure_nm == 'PSI_03':

                results = filter_func_all_cohort_hosps(excel_file_name, lookup_dict, keyname, lsccmc_cohort_list,
                                                       'AHRQ Safety Denominator',
                                                       measure_nm, edac_readm_choice_num)
                return (results)

            elif measure_nm == 'PSI_06':

                results = filter_func_all_cohort_hosps(excel_file_name, lookup_dict, keyname, lsccmc_cohort_list,
                                                       'AHRQ Safety Denominator',
                                                       measure_nm, edac_readm_choice_num)
                return (results)

            elif measure_nm == 'PSI_09':

                results = filter_func_all_cohort_hosps(excel_file_name, lookup_dict, keyname, lsccmc_cohort_list,
                                                       'AHRQ Safety Denominator',
                                                       measure_nm, edac_readm_choice_num)
                return (results)

            elif measure_nm == 'PSI_11':

                results = filter_func_all_cohort_hosps(excel_file_name, lookup_dict, keyname, lsccmc_cohort_list,
                                                       'AHRQ Safety Denominator',
                                                       measure_nm, edac_readm_choice_num)
                return (results)

            elif measure_nm == 'PSI_13':

                results = filter_func_all_cohort_hosps(excel_file_name, lookup_dict, keyname, lsccmc_cohort_list,
                                                       'AHRQ Safety Denominator',
                                                       measure_nm, edac_readm_choice_num)
                return (results)

        elif measure_nm == 'THK':

            results = filter_func_all_cohort_hosps(excel_file_name, lookup_dict, keyname, lsccmc_cohort_list,
                                                   'Total THK cases', measure_nm,
                                                   edac_readm_choice_num)
            return (results)



    elif cohort_nm == 'Large Specialized Complex Care Medical Center' and domain_nm == 'Outpatient':

        results = filter_func_all_cohort_hosps(excel_file_name, lookup_dict, keyname, lsccmc_cohort_list,
                                               'Readmit Rate Denom Cases',
                                               measure_nm, edac_readm_choice_num)
        return (results)

    # CCMC Logic:
    elif cohort_nm == 'Complex Care Medical Center' and domain_nm == 'Mortality':
        # print('got to ccmc')

        results = filter_func_all_cohort_hosps(excel_file_name, lookup_dict, keyname, ccmc_cohort_list, 'Cases',
                                               measure_nm, edac_readm_choice_num)
        return (results)

    elif cohort_nm == 'Complex Care Medical Center' and domain_nm == 'Effectiveness':

        if measure_nm.split('_')[0] == 'EDAC':

            if edac_readm_choice_num == 1:
                results = filter_func_all_cohort_hosps(excel_file_name, lookup_dict, keyname, ccmc_cohort_list,
                                                       ['Total Index Encounters', 'Encounter Month'], measure_nm,
                                                       edac_readm_choice_num)
                # 'Excess Days Per 100 Index Encounters', measure_nm)
            else:
                results = filter_func_all_cohort_hosps(excel_file_name, lookup_dict, keyname, ccmc_cohort_list,
                                                       ['Total Index Encounters', 'Encounter Month'], measure_nm,
                                                       edac_readm_choice_num, edac_readm_month_list)
                # 'Excess Days Per 100 Index Encounters', measure_nm)

            return (results)

        elif measure_nm.split('_')[0] == 'READM':

            if edac_readm_choice_num == 1:
                results = filter_func_all_cohort_hosps(excel_file_name, lookup_dict, keyname, ccmc_cohort_list,
                                                       ['Total Index Encounters', 'Encounter Month'],
                                                       measure_nm, edac_readm_choice_num)
            else:
                results = filter_func_all_cohort_hosps(excel_file_name, lookup_dict, keyname, ccmc_cohort_list,
                                                       ['Total Index Encounters', 'Encounter Month'],
                                                       measure_nm, edac_readm_choice_num, edac_readm_month_list)
            return (results)

    elif cohort_nm == 'Complex Care Medical Center' and domain_nm == 'Efficiency':
        if measure_nm.split('_')[0] == 'DCOST':

            results = filter_func_all_cohort_hosps(excel_file_name, lookup_dict, keyname, ccmc_cohort_list,
                                                   'Cases',
                                                   measure_nm, edac_readm_choice_num)
            return (results)

        elif measure_nm.split('_')[0] == 'LOS':

            results = filter_func_all_cohort_hosps(excel_file_name, lookup_dict, keyname, ccmc_cohort_list,
                                                   'Cases', measure_nm, edac_readm_choice_num)
            return (results)

    elif cohort_nm == 'Complex Care Medical Center' and domain_nm == 'Safety':
        if measure_nm.split('_')[0] == 'PSI':
            if measure_nm == 'PSI_03':

                results = filter_func_all_cohort_hosps(excel_file_name, lookup_dict, keyname, ccmc_cohort_list,
                                                       'AHRQ Safety Denominator',
                                                       measure_nm, edac_readm_choice_num)
                return (results)

            elif measure_nm == 'PSI_06':

                results = filter_func_all_cohort_hosps(excel_file_name, lookup_dict, keyname, ccmc_cohort_list,
                                                       'AHRQ Safety Denominator',
                                                       measure_nm, edac_readm_choice_num)
                return (results)

            elif measure_nm == 'PSI_09':

                results = filter_func_all_cohort_hosps(excel_file_name, lookup_dict, keyname, ccmc_cohort_list,
                                                       'AHRQ Safety Denominator',
                                                       measure_nm, edac_readm_choice_num)
                return (results)

            elif measure_nm == 'PSI_11':

                results = filter_func_all_cohort_hosps(excel_file_name, lookup_dict, keyname, ccmc_cohort_list,
                                                       'AHRQ Safety Denominator',
                                                       measure_nm, edac_readm_choice_num)
                return (results)

            elif measure_nm == 'PSI_13':

                results = filter_func_all_cohort_hosps(excel_file_name, lookup_dict, keyname, ccmc_cohort_list,
                                                       'AHRQ Safety Denominator',
                                                       measure_nm, edac_readm_choice_num)
                return (results)

        elif measure_nm == 'THK':

            results = filter_func_all_cohort_hosps(excel_file_name, lookup_dict, keyname, ccmc_cohort_list,
                                                   'Total THK cases', measure_nm,
                                                   edac_readm_choice_num)
            return (results)

    elif cohort_nm == 'Complex Care Medical Center' and domain_nm == 'Outpatient':

        results = filter_func_all_cohort_hosps(excel_file_name, lookup_dict, keyname, ccmc_cohort_list,
                                               'Readmit Rate Denom Cases',
                                               measure_nm, edac_readm_choice_num)
        return (results)

    # COMM Logic:
    elif cohort_nm == 'Community Medical Center' and domain_nm == 'Mortality':
        # print('got to ccmc')

        results = filter_func_all_cohort_hosps(excel_file_name, lookup_dict, keyname, comm_cohort_list, 'Cases',
                                               measure_nm, edac_readm_choice_num)
        return (results)

    elif cohort_nm == 'Community Medical Center' and domain_nm == 'Effectiveness':

        if measure_nm.split('_')[0] == 'EDAC':

            if edac_readm_choice_num == 1:
                results = filter_func_all_cohort_hosps(excel_file_name, lookup_dict, keyname, comm_cohort_list,
                                                       ['Total Index Encounters', 'Encounter Month'], measure_nm,
                                                       edac_readm_choice_num)
                # 'Excess Days Per 100 Index Encounters', measure_nm)
            else:
                results = filter_func_all_cohort_hosps(excel_file_name, lookup_dict, keyname, comm_cohort_list,
                                                       ['Total Index Encounters', 'Encounter Month'], measure_nm,
                                                       edac_readm_choice_num, edac_readm_month_list)
                # 'Excess Days Per 100 Index Encounters', measure_nm)

            return (results)

        elif measure_nm.split('_')[0] == 'READM':

            if edac_readm_choice_num == 1:
                results = filter_func_all_cohort_hosps(excel_file_name, lookup_dict, keyname, comm_cohort_list,
                                                       ['Total Index Encounters', 'Encounter Month'],
                                                       measure_nm, edac_readm_choice_num)
            else:
                results = filter_func_all_cohort_hosps(excel_file_name, lookup_dict, keyname, comm_cohort_list,
                                                       ['Total Index Encounters', 'Encounter Month'],
                                                       measure_nm, edac_readm_choice_num, edac_readm_month_list)
            return (results)

    elif cohort_nm == 'Community Medical Center' and domain_nm == 'Efficiency':
        if measure_nm.split('_')[0] == 'DCOST':

            results = filter_func_all_cohort_hosps(excel_file_name, lookup_dict, keyname, comm_cohort_list,
                                                   'Cases',
                                                   measure_nm, edac_readm_choice_num)
            return (results)

        elif measure_nm.split('_')[0] == 'LOS':

            results = filter_func_all_cohort_hosps(excel_file_name, lookup_dict, keyname, comm_cohort_list,
                                                   'Cases', measure_nm, edac_readm_choice_num)
            return (results)

    elif cohort_nm == 'Community Medical Center' and domain_nm == 'Safety':
        if measure_nm.split('_')[0] == 'PSI':
            if measure_nm == 'PSI_03':

                results = filter_func_all_cohort_hosps(excel_file_name, lookup_dict, keyname, comm_cohort_list,
                                                       'AHRQ Safety Denominator',
                                                       measure_nm, edac_readm_choice_num)
                return (results)

            elif measure_nm == 'PSI_06':

                results = filter_func_all_cohort_hosps(excel_file_name, lookup_dict, keyname, comm_cohort_list,
                                                       'AHRQ Safety Denominator',
                                                       measure_nm, edac_readm_choice_num)
                return (results)

            elif measure_nm == 'PSI_09':

                results = filter_func_all_cohort_hosps(excel_file_name, lookup_dict, keyname, comm_cohort_list,
                                                       'AHRQ Safety Denominator',
                                                       measure_nm, edac_readm_choice_num)
                return (results)

            elif measure_nm == 'PSI_11':

                results = filter_func_all_cohort_hosps(excel_file_name, lookup_dict, keyname, comm_cohort_list,
                                                       'AHRQ Safety Denominator',
                                                       measure_nm, edac_readm_choice_num)
                return (results)

            elif measure_nm == 'PSI_13':

                results = filter_func_all_cohort_hosps(excel_file_name, lookup_dict, keyname, comm_cohort_list,
                                                       'AHRQ Safety Denominator',
                                                       measure_nm, edac_readm_choice_num)
                return (results)

        elif measure_nm == 'THK':

            results = filter_func_all_cohort_hosps(excel_file_name, lookup_dict, keyname, comm_cohort_list,
                                                   'Total THK cases', measure_nm,
                                                   edac_readm_choice_num)
            return (results)

    elif cohort_nm == 'Community Medical Center' and domain_nm == 'Outpatient':

        results = filter_func_all_cohort_hosps(excel_file_name, lookup_dict, keyname, comm_cohort_list,
                                               'Readmit Rate Denom Cases',
                                               measure_nm, edac_readm_choice_num)
        return (results)

'''

#UL003

def file_filter_switcher_all_cohort_hosps_d_events(cohort_nm, domain_nm, measure_nm, excel_file_name, lookup_dict,
                                                   keyname, edac_readm_choice_num, camc_cohort_list, lsccmc_cohort_list,
                                                   ccmc_cohort_list, comm_cohort_list,casc_cohort_list, edac_readm_month_list='No List'):
    #print('got to n size switcher')

    if cohort_nm == 'Comprehensive Academic Medical Center' and domain_nm == 'Mortality':

        results = filter_func_all_cohort_hosps(excel_file_name, lookup_dict, keyname, camc_cohort_list, 'Cases',
                                               measure_nm, edac_readm_choice_num)
        return (results)

    elif cohort_nm == 'Comprehensive Academic Medical Center' and domain_nm == 'Effectiveness':

        if measure_nm.split('_')[0] == 'EDAC':

            if edac_readm_choice_num == 1:
                results = filter_func_all_cohort_hosps(excel_file_name, lookup_dict, keyname, camc_cohort_list,
                                                       ['Total Index Encounters', 'Encounter Month'], measure_nm,
                                                       edac_readm_choice_num)

            else:
                results = filter_func_all_cohort_hosps(excel_file_name, lookup_dict, keyname, camc_cohort_list,
                                                       ['Total Index Encounters', 'Encounter Month'], measure_nm,
                                                       edac_readm_choice_num, edac_readm_month_list)

            return (results)

        elif measure_nm.split('_')[0] == 'READM' and measure_nm.split('_')[1] not in ['ARTHRO','CHOL','COLON','URI']:

            if edac_readm_choice_num == 1:
                results = filter_func_all_cohort_hosps(excel_file_name, lookup_dict, keyname, camc_cohort_list,
                                                       ['Total Index Encounters', 'Encounter Month'],
                                                       measure_nm, edac_readm_choice_num)
            else:
                results = filter_func_all_cohort_hosps(excel_file_name, lookup_dict, keyname, camc_cohort_list,
                                                       ['Total Index Encounters', 'Encounter Month'],
                                                       measure_nm, edac_readm_choice_num, edac_readm_month_list)
            return (results)
        
        elif measure_nm.split('_')[0] == 'READM' and domain_nm == 'Effectiveness' and measure_nm.split('_')[1] in ['ARTHRO','CHOL','COLON','URI']:

            results = filter_func_all_cohort_hosps(excel_file_name, lookup_dict, keyname, camc_cohort_list,
                                               'Readmit Rate Denom Cases',
                                               measure_nm, edac_readm_choice_num)
            return (results)


    elif cohort_nm == 'Comprehensive Academic Medical Center' and domain_nm == 'Efficiency':
        if measure_nm.split('_')[0] == 'DCOST':

            results = filter_func_all_cohort_hosps(excel_file_name, lookup_dict, keyname, camc_cohort_list,
                                                   'Cases',
                                                   measure_nm, edac_readm_choice_num)
            return (results)

        elif measure_nm.split('_')[0] == 'LOS':

            results = filter_func_all_cohort_hosps(excel_file_name, lookup_dict, keyname, camc_cohort_list,
                                                   'Cases', measure_nm, edac_readm_choice_num)
            return (results)

    elif cohort_nm == 'Comprehensive Academic Medical Center' and domain_nm == 'Safety':
        if measure_nm.split('_')[0] == 'PSI':
            if measure_nm == 'PSI_03':

                results = filter_func_all_cohort_hosps(excel_file_name, lookup_dict, keyname, camc_cohort_list,
                                                       'AHRQ Safety Denominator',
                                                       measure_nm, edac_readm_choice_num)
                return (results)

            elif measure_nm == 'PSI_06':

                results = filter_func_all_cohort_hosps(excel_file_name, lookup_dict, keyname, camc_cohort_list,
                                                       'AHRQ Safety Denominator',
                                                       measure_nm, edac_readm_choice_num)
                return (results)

            elif measure_nm == 'PSI_09':

                results = filter_func_all_cohort_hosps(excel_file_name, lookup_dict, keyname, camc_cohort_list,
                                                       'AHRQ Safety Denominator',
                                                       measure_nm, edac_readm_choice_num)
                return (results)

            elif measure_nm == 'PSI_11':

                results = filter_func_all_cohort_hosps(excel_file_name, lookup_dict, keyname, camc_cohort_list,
                                                       'AHRQ Safety Denominator',
                                                       measure_nm, edac_readm_choice_num)
                return (results)

            elif measure_nm == 'PSI_13':

                results = filter_func_all_cohort_hosps(excel_file_name, lookup_dict, keyname, camc_cohort_list,
                                                       'AHRQ Safety Denominator',
                                                       measure_nm, edac_readm_choice_num)
                return (results)

        elif measure_nm == 'THK':

            results = filter_func_all_cohort_hosps(excel_file_name, lookup_dict, keyname, camc_cohort_list,
                                                   'Total THK cases', measure_nm,
                                                   edac_readm_choice_num)
            return (results)

    elif cohort_nm == 'Comprehensive Academic Medical Center' and domain_nm == 'Outpatient':

        results = filter_func_all_cohort_hosps(excel_file_name, lookup_dict, keyname, camc_cohort_list,
                                               'Readmit Rate Denom Cases',
                                               measure_nm, edac_readm_choice_num)
        return (results)
    
    # LSCCMC logic:
    elif cohort_nm == 'Large Specialized Complex Care Medical Center' and domain_nm == 'Mortality':

        results = filter_func_all_cohort_hosps(excel_file_name, lookup_dict, keyname, lsccmc_cohort_list,
                                               'Cases', measure_nm, edac_readm_choice_num)
        return (results)

    elif cohort_nm == 'Large Specialized Complex Care Medical Center' and domain_nm == 'Effectiveness':
        #print('got into LSCCMC effectiveness')
        if measure_nm.split('_')[0] == 'EDAC':
            print('got to EDAC')
            if edac_readm_choice_num == 1:
                results = filter_func_all_cohort_hosps(excel_file_name, lookup_dict, keyname, lsccmc_cohort_list,
                                                       ['Total Index Encounters', 'Encounter Month'], measure_nm,
                                                       edac_readm_choice_num)
                # 'Excess Days Per 100 Index Encounters', measure_nm)
            else:
                results = filter_func_all_cohort_hosps(excel_file_name, lookup_dict, keyname, lsccmc_cohort_list,
                                                       ['Total Index Encounters', 'Encounter Month'], measure_nm,
                                                       edac_readm_choice_num, edac_readm_month_list)
                # 'Excess Days Per 100 Index Encounters', measure_nm)

            return (results)

        elif measure_nm.split('_')[0] == 'READM' and measure_nm.split('_')[1] not in ['ARTHRO','CHOL','COLON','URI']:

            if edac_readm_choice_num == 1:
                results = filter_func_all_cohort_hosps(excel_file_name, lookup_dict, keyname, lsccmc_cohort_list,
                                                       ['Total Index Encounters', 'Encounter Month'],
                                                       measure_nm, edac_readm_choice_num)
            else:
                results = filter_func_all_cohort_hosps(excel_file_name, lookup_dict, keyname, lsccmc_cohort_list,
                                                       ['Total Index Encounters', 'Encounter Month'],
                                                       measure_nm, edac_readm_choice_num, edac_readm_month_list)
            return (results)
        
        elif measure_nm.split('_')[0] == 'READM' and domain_nm == 'Effectiveness' and measure_nm.split('_')[1] in ['ARTHRO','CHOL','COLON','URI']:

            results = filter_func_all_cohort_hosps(excel_file_name, lookup_dict, keyname, lsccmc_cohort_list,
                                               'Readmit Rate Denom Cases',
                                               measure_nm, edac_readm_choice_num)
            return (results)

    elif cohort_nm == 'Large Specialized Complex Care Medical Center' and domain_nm == 'Efficiency':
        if measure_nm.split('_')[0] == 'DCOST':

            results = filter_func_all_cohort_hosps(excel_file_name, lookup_dict, keyname, lsccmc_cohort_list,
                                                   'Cases',
                                                   measure_nm, edac_readm_choice_num)
            return (results)

        elif measure_nm.split('_')[0] == 'LOS':

            results = filter_func_all_cohort_hosps(excel_file_name, lookup_dict, keyname, lsccmc_cohort_list,
                                                   'Cases', measure_nm, edac_readm_choice_num)
            return (results)

    elif cohort_nm == 'Large Specialized Complex Care Medical Center' and domain_nm == 'Safety':
        if measure_nm.split('_')[0] == 'PSI':
            if measure_nm == 'PSI_03':

                results = filter_func_all_cohort_hosps(excel_file_name, lookup_dict, keyname, lsccmc_cohort_list,
                                                       'AHRQ Safety Denominator',
                                                       measure_nm, edac_readm_choice_num)
                return (results)

            elif measure_nm == 'PSI_06':

                results = filter_func_all_cohort_hosps(excel_file_name, lookup_dict, keyname, lsccmc_cohort_list,
                                                       'AHRQ Safety Denominator',
                                                       measure_nm, edac_readm_choice_num)
                return (results)

            elif measure_nm == 'PSI_09':

                results = filter_func_all_cohort_hosps(excel_file_name, lookup_dict, keyname, lsccmc_cohort_list,
                                                       'AHRQ Safety Denominator',
                                                       measure_nm, edac_readm_choice_num)
                return (results)

            elif measure_nm == 'PSI_11':

                results = filter_func_all_cohort_hosps(excel_file_name, lookup_dict, keyname, lsccmc_cohort_list,
                                                       'AHRQ Safety Denominator',
                                                       measure_nm, edac_readm_choice_num)
                return (results)

            elif measure_nm == 'PSI_13':

                results = filter_func_all_cohort_hosps(excel_file_name, lookup_dict, keyname, lsccmc_cohort_list,
                                                       'AHRQ Safety Denominator',
                                                       measure_nm, edac_readm_choice_num)
                return (results)

        elif measure_nm == 'THK':

            results = filter_func_all_cohort_hosps(excel_file_name, lookup_dict, keyname, lsccmc_cohort_list,
                                                   'Total THK cases', measure_nm,
                                                   edac_readm_choice_num)
            return (results)



    elif cohort_nm == 'Large Specialized Complex Care Medical Center' and domain_nm == 'Outpatient':

        results = filter_func_all_cohort_hosps(excel_file_name, lookup_dict, keyname, lsccmc_cohort_list,
                                               'Readmit Rate Denom Cases',
                                               measure_nm, edac_readm_choice_num)
        return (results)
    

    # CCMC Logic:
    elif cohort_nm == 'Complex Care Medical Center' and domain_nm == 'Mortality':
        # print('got to ccmc')

        results = filter_func_all_cohort_hosps(excel_file_name, lookup_dict, keyname, ccmc_cohort_list, 'Cases',
                                               measure_nm, edac_readm_choice_num)
        return (results)

    elif cohort_nm == 'Complex Care Medical Center' and domain_nm == 'Effectiveness':

        if measure_nm.split('_')[0] == 'EDAC':

            if edac_readm_choice_num == 1:
                results = filter_func_all_cohort_hosps(excel_file_name, lookup_dict, keyname, ccmc_cohort_list,
                                                       ['Total Index Encounters', 'Encounter Month'], measure_nm,
                                                       edac_readm_choice_num)
                # 'Excess Days Per 100 Index Encounters', measure_nm)
            else:
                results = filter_func_all_cohort_hosps(excel_file_name, lookup_dict, keyname, ccmc_cohort_list,
                                                       ['Total Index Encounters', 'Encounter Month'], measure_nm,
                                                       edac_readm_choice_num, edac_readm_month_list)
                # 'Excess Days Per 100 Index Encounters', measure_nm)

            return (results)

        elif measure_nm.split('_')[0] == 'READM' and measure_nm.split('_')[1] not in ['ARTHRO','CHOL','COLON','URI']:

            if edac_readm_choice_num == 1:
                results = filter_func_all_cohort_hosps(excel_file_name, lookup_dict, keyname, ccmc_cohort_list,
                                                       ['Total Index Encounters', 'Encounter Month'],
                                                       measure_nm, edac_readm_choice_num)
            else:
                results = filter_func_all_cohort_hosps(excel_file_name, lookup_dict, keyname, ccmc_cohort_list,
                                                       ['Total Index Encounters', 'Encounter Month'],
                                                       measure_nm, edac_readm_choice_num, edac_readm_month_list)
            return (results)
        
        elif measure_nm.split('_')[0] == 'READM' and domain_nm == 'Effectiveness' and measure_nm.split('_')[1] in ['ARTHRO','CHOL','COLON','URI']:

            results = filter_func_all_cohort_hosps(excel_file_name, lookup_dict, keyname, ccmc_cohort_list,
                                               'Readmit Rate Denom Cases',
                                               measure_nm, edac_readm_choice_num)
            return (results)

    elif cohort_nm == 'Complex Care Medical Center' and domain_nm == 'Efficiency':
        if measure_nm.split('_')[0] == 'DCOST':

            results = filter_func_all_cohort_hosps(excel_file_name, lookup_dict, keyname, ccmc_cohort_list,
                                                   'Cases',
                                                   measure_nm, edac_readm_choice_num)
            return (results)

        elif measure_nm.split('_')[0] == 'LOS':

            results = filter_func_all_cohort_hosps(excel_file_name, lookup_dict, keyname, ccmc_cohort_list,
                                                   'Cases', measure_nm, edac_readm_choice_num)
            return (results)

    elif cohort_nm == 'Complex Care Medical Center' and domain_nm == 'Safety':
        if measure_nm.split('_')[0] == 'PSI':
            if measure_nm == 'PSI_03':

                results = filter_func_all_cohort_hosps(excel_file_name, lookup_dict, keyname, ccmc_cohort_list,
                                                       'AHRQ Safety Denominator',
                                                       measure_nm, edac_readm_choice_num)
                return (results)

            elif measure_nm == 'PSI_06':

                results = filter_func_all_cohort_hosps(excel_file_name, lookup_dict, keyname, ccmc_cohort_list,
                                                       'AHRQ Safety Denominator',
                                                       measure_nm, edac_readm_choice_num)
                return (results)

            elif measure_nm == 'PSI_09':

                results = filter_func_all_cohort_hosps(excel_file_name, lookup_dict, keyname, ccmc_cohort_list,
                                                       'AHRQ Safety Denominator',
                                                       measure_nm, edac_readm_choice_num)
                return (results)

            elif measure_nm == 'PSI_11':

                results = filter_func_all_cohort_hosps(excel_file_name, lookup_dict, keyname, ccmc_cohort_list,
                                                       'AHRQ Safety Denominator',
                                                       measure_nm, edac_readm_choice_num)
                return (results)

            elif measure_nm == 'PSI_13':

                results = filter_func_all_cohort_hosps(excel_file_name, lookup_dict, keyname, ccmc_cohort_list,
                                                       'AHRQ Safety Denominator',
                                                       measure_nm, edac_readm_choice_num)
                return (results)

        elif measure_nm == 'THK':

            results = filter_func_all_cohort_hosps(excel_file_name, lookup_dict, keyname, ccmc_cohort_list,
                                                   'Total THK cases', measure_nm,
                                                   edac_readm_choice_num)
            return (results)

    elif cohort_nm == 'Complex Care Medical Center' and domain_nm == 'Outpatient':

        results = filter_func_all_cohort_hosps(excel_file_name, lookup_dict, keyname, ccmc_cohort_list,
                                               'Readmit Rate Denom Cases',
                                               measure_nm, edac_readm_choice_num)
        return (results)
    
    # COMM Logic:
    elif cohort_nm == 'Community Medical Center' and domain_nm == 'Mortality':
        # print('got to ccmc')

        results = filter_func_all_cohort_hosps(excel_file_name, lookup_dict, keyname, comm_cohort_list, 'Cases',
                                               measure_nm, edac_readm_choice_num)
        return (results)

    elif cohort_nm == 'Community Medical Center' and domain_nm == 'Effectiveness':

        if measure_nm.split('_')[0] == 'EDAC':

            if edac_readm_choice_num == 1:
                results = filter_func_all_cohort_hosps(excel_file_name, lookup_dict, keyname, comm_cohort_list,
                                                       ['Total Index Encounters', 'Encounter Month'], measure_nm,
                                                       edac_readm_choice_num)
                # 'Excess Days Per 100 Index Encounters', measure_nm)
            else:
                results = filter_func_all_cohort_hosps(excel_file_name, lookup_dict, keyname, comm_cohort_list,
                                                       ['Total Index Encounters', 'Encounter Month'], measure_nm,
                                                       edac_readm_choice_num, edac_readm_month_list)
                # 'Excess Days Per 100 Index Encounters', measure_nm)

            return (results)

        elif measure_nm.split('_')[0] == 'READM' and measure_nm.split('_')[1] not in ['ARTHRO','CHOL','COLON','URI']:

            if edac_readm_choice_num == 1:
                results = filter_func_all_cohort_hosps(excel_file_name, lookup_dict, keyname, comm_cohort_list,
                                                       ['Total Index Encounters', 'Encounter Month'],
                                                       measure_nm, edac_readm_choice_num)
            else:
                results = filter_func_all_cohort_hosps(excel_file_name, lookup_dict, keyname, comm_cohort_list,
                                                       ['Total Index Encounters', 'Encounter Month'],
                                                       measure_nm, edac_readm_choice_num, edac_readm_month_list)
            return (results)
        
        elif measure_nm.split('_')[0] == 'READM' and domain_nm == 'Effectiveness' and measure_nm.split('_')[1] in ['ARTHRO','CHOL','COLON','URI']:

            results = filter_func_all_cohort_hosps(excel_file_name, lookup_dict, keyname, comm_cohort_list,
                                               'Readmit Rate Denom Cases',
                                               measure_nm, edac_readm_choice_num)
            return (results)

    elif cohort_nm == 'Community Medical Center' and domain_nm == 'Efficiency':
        if measure_nm.split('_')[0] == 'DCOST':

            results = filter_func_all_cohort_hosps(excel_file_name, lookup_dict, keyname, comm_cohort_list,
                                                   'Cases',
                                                   measure_nm, edac_readm_choice_num)
            return (results)

        elif measure_nm.split('_')[0] == 'LOS':

            results = filter_func_all_cohort_hosps(excel_file_name, lookup_dict, keyname, comm_cohort_list,
                                                   'Cases', measure_nm, edac_readm_choice_num)
            return (results)

    elif cohort_nm == 'Community Medical Center' and domain_nm == 'Safety':
        if measure_nm.split('_')[0] == 'PSI':
            if measure_nm == 'PSI_03':

                results = filter_func_all_cohort_hosps(excel_file_name, lookup_dict, keyname, comm_cohort_list,
                                                       'AHRQ Safety Denominator',
                                                       measure_nm, edac_readm_choice_num)
                return (results)

            elif measure_nm == 'PSI_06':

                results = filter_func_all_cohort_hosps(excel_file_name, lookup_dict, keyname, comm_cohort_list,
                                                       'AHRQ Safety Denominator',
                                                       measure_nm, edac_readm_choice_num)
                return (results)

            elif measure_nm == 'PSI_09':

                results = filter_func_all_cohort_hosps(excel_file_name, lookup_dict, keyname, comm_cohort_list,
                                                       'AHRQ Safety Denominator',
                                                       measure_nm, edac_readm_choice_num)
                return (results)

            elif measure_nm == 'PSI_11':

                results = filter_func_all_cohort_hosps(excel_file_name, lookup_dict, keyname, comm_cohort_list,
                                                       'AHRQ Safety Denominator',
                                                       measure_nm, edac_readm_choice_num)
                return (results)

            elif measure_nm == 'PSI_13':

                results = filter_func_all_cohort_hosps(excel_file_name, lookup_dict, keyname, comm_cohort_list,
                                                       'AHRQ Safety Denominator',
                                                       measure_nm, edac_readm_choice_num)
                return (results)

        elif measure_nm == 'THK':

            results = filter_func_all_cohort_hosps(excel_file_name, lookup_dict, keyname, comm_cohort_list,
                                                   'Total THK cases', measure_nm,
                                                   edac_readm_choice_num)
            return (results)

    elif cohort_nm == 'Community Medical Center' and domain_nm == 'Outpatient':

        results = filter_func_all_cohort_hosps(excel_file_name, lookup_dict, keyname, comm_cohort_list,
                                               'Readmit Rate Denom Cases',
                                               measure_nm, edac_readm_choice_num)
        return (results)
    
    #UL003
    # CASC Logic:
    elif cohort_nm == 'Critical Access & Small Community' and domain_nm == 'Mortality':

        results = filter_func_all_cohort_hosps(excel_file_name, lookup_dict, keyname, casc_cohort_list, 'Cases',
                                               measure_nm, edac_readm_choice_num)
        return (results)
    
    elif cohort_nm == 'Critical Access & Small Community' and domain_nm == 'Efficiency':
        if measure_nm.split('_')[0] == 'DCOST':

            results = filter_func_all_cohort_hosps(excel_file_name, lookup_dict, keyname, casc_cohort_list,
                                                   'Cases',
                                                   measure_nm, edac_readm_choice_num)
            return (results)

        elif measure_nm.split('_')[0] == 'LOS':

            results = filter_func_all_cohort_hosps(excel_file_name, lookup_dict, keyname, casc_cohort_list,
                                                   'Cases', measure_nm, edac_readm_choice_num)
            return (results)
        
    elif cohort_nm == 'Critical Access & Small Community' and domain_nm == 'Effectiveness' and measure_nm.split('_')[0] not in ['ADE','EARLY']:

        if measure_nm.split('_')[0] == 'EDAC':

            if edac_readm_choice_num == 1:
                results = filter_func_all_cohort_hosps(excel_file_name, lookup_dict, keyname, casc_cohort_list,
                                                       ['Total Index Encounters', 'Encounter Month'], measure_nm,
                                                       edac_readm_choice_num)

            else:
                results = filter_func_all_cohort_hosps(excel_file_name, lookup_dict, keyname, casc_cohort_list,
                                                       ['Total Index Encounters', 'Encounter Month'], measure_nm,
                                                       edac_readm_choice_num, edac_readm_month_list)

            return (results)

        elif measure_nm.split('_')[0] == 'READM' and measure_nm.split('_')[1] not in ['ARTHRO','CHOL','COLON','URI']:

            if edac_readm_choice_num == 1:
                results = filter_func_all_cohort_hosps(excel_file_name, lookup_dict, keyname, casc_cohort_list,
                                                       ['Total Index Encounters', 'Encounter Month'],
                                                       measure_nm, edac_readm_choice_num)
            else:
                results = filter_func_all_cohort_hosps(excel_file_name, lookup_dict, keyname, casc_cohort_list,
                                                       ['Total Index Encounters', 'Encounter Month'],
                                                       measure_nm, edac_readm_choice_num, edac_readm_month_list)
            return (results)
        
        elif measure_nm.split('_')[0] == 'READM' and domain_nm == 'Effectiveness' and measure_nm.split('_')[1] in ['ARTHRO','CHOL','COLON','URI']:

            results = filter_func_all_cohort_hosps(excel_file_name, lookup_dict, keyname, casc_cohort_list,
                                               'Readmit Rate Denom Cases',
                                               measure_nm, edac_readm_choice_num)
            return (results)
        
        elif domain_nm == 'Effectiveness' and measure_nm.split('_')[1] in ['REVISIT','RETURN']:

            results = filter_func_all_cohort_hosps(excel_file_name, lookup_dict, keyname, casc_cohort_list,
                                               'Readmit Rate Denom Cases',
                                               measure_nm, edac_readm_choice_num)
            return (results)
        
    elif measure_nm.split('_')[0] in ['ADE','EARLY']:

        results = filter_func_all_cohort_hosps(excel_file_name, lookup_dict, keyname, casc_cohort_list,
                                               'Cases',
                                               measure_nm, edac_readm_choice_num)
        return (results)


################################################################################################

# subset the measure dataframes by domain, sort by measure name and then return the dataframe
def subset_sort_df(dframe, domain_nm):
    if dframe.columns[0] == 'Hospital':
        # print(domain_nm)
        # print('first column is "Hospital"')
        dframe[dframe['Domain'] == domain_nm].sort_values(by=['Measure'])
        dframe['Meas1'], dframe['Meas2'] = dframe['Measure'].str.split('-', 1).str
        dframe['Meas2'] = dframe['Meas2'].str.upper()
        dframe['Meas1'] = dframe['Meas1'].str.upper()
        dframe = dframe[dframe['Domain'] == domain_nm].sort_values(by=['Meas1', 'Meas2', 'Hospital'])
        dframe = dframe[['Hospital', 'Domain', 'Measure', 'Metric Value']]
        return (dframe)

    else:
        # print(domain_nm)
        dframe[dframe['Domain'] == domain_nm].sort_values(by=['Measure'])
        # print(dframe[dframe['Domain'] == domain_nm])
        dframe['Meas1'], dframe['Meas2'] = dframe['Measure'].str.split('-', 1).str
        dframe['Meas2'] = dframe['Meas2'].str.upper()
        dframe['Meas1'] = dframe['Meas1'].str.upper()
        dframe = dframe[dframe['Domain'] == domain_nm].sort_values(by=['Meas1', 'Meas2'])
        dframe = dframe[['Domain', 'Measure', 'Metric Value']]
        return (dframe)


################################################################################################


# This function puts together all the above data wrangling functions
# The goal of this function is to parse all Vizient Q&A files within a folder structure,
# extract the correct data column value per measure and cohort, sort the measure values by Domain
# then concatenate all domain values in the same order as the Vizient Calculator.
# Last step is to write to csv file.
def vizient_data_folder_walker():
    # Set the path of the folder structure we want to recursively walk through
    wd = input('Enter the path of the folder you want to parse.')
    wd = os.path.abspath(wd)
    # Set teh path of the folder structure where we want the final files
    wd_dest = input('Enter the path of the folder you want to final files to go in.')
    wd_dest = os.path.abspath(wd_dest)

    # step 2:  Import cohort data from the Vizient documentation file.
    cohort_helper_df = gather_cohort_data()

    # step 3:  Import the Vizient template hyperlink file and generate a helper dictionary from it.
    hyperlinks_helper_df = get_report_template_links()
    merged_hyperlink_helper_df = pd.merge(cohort_helper_df, hyperlinks_helper_df, on='Hospital')
    wrangled_hyperlinks_dict = create_hyperlink_dict_wrangle(merged_hyperlink_helper_df)

    # Check which type of sum the client wants for edac and readmission values
    edac_readm_sum_option = 'not answered'
    while edac_readm_sum_option not in [1, 2]:
        try:
            edac_readm_sum_option = int(input(
                "EDAC and Readmission values were pulled by quarter.  Do you want the sum of all months (choose 1) or do you want the sum of specific months? (choose 2)"))
        except:
            print('Not an integer.')
        if edac_readm_sum_option not in [1, 2]:
            print('Please choose 1 or 2')

    # if choice 2, then get a list of months to sum up.
    if edac_readm_sum_option == 2:
        try:
            print(
                'You chose "sum of specific months" (choice 2).  Please add all months in the quarter date range you would like to sum for EDAC and READMISSIONS.')
            edac_readm_month_list = []
            month = 'GO'
            r = re.compile('\d\d\d\d-\d\d')
            while month != 'STOP':
                month = input(
                    'ADD Month with "YYYY-MM" format.  Type "STOP" to stop adding months to custom month list.')
                if r.match(month) is not None:
                    print('Matches Format.  Added to month list')
                    edac_readm_month_list.append(month)
                elif month.upper() == 'STOP':
                    print('Stopping...')
                    print(edac_readm_month_list)
                    break
                else:
                    print('That does not match "YYYY-MM" format.  Try again.')
            # if input is not-integer, just print the list
        except:
            print(edac_readm_month_list)

            # Set counters which will act as proxy indexes in order to append rows to pandas dataframes
    CAMC_counter = 0
    LSCCMC_counter = 0
    CCMC_counter = 0
    # Create empty dataframes to store final values.  First dataframe stores the unsorted values, the _final dataframe
    # will store the sorted values and will be written to a csv file
    CAMC_df = pd.DataFrame(columns=['Domain', 'Measure', 'Metric Value'])
    LSCCMC_df = pd.DataFrame(columns=['Domain', 'Measure', 'Metric Value'])
    CCMC_df = pd.DataFrame(columns=['Hospital', 'Domain', 'Measure', 'Metric Value'])
    CAMC_df_final = pd.DataFrame(columns=['Domain', 'Measure', 'Metric Value'])
    LSCCMC_df_final = pd.DataFrame(columns=['Domain', 'Measure', 'Metric Value'])
    CCMC_df_final = pd.DataFrame(columns=['Hospital', 'Domain', 'Measure', 'Metric Value'])

    do_not_parse = ['ED_2B', 'ED_OP_18B']
    # iterate through folder structure.
    for path, dirs, files in os.walk(wd):
        # if files list is not empty, open it up
        if len(files) > 0:
            # parse folder name to get measure name
            path_tail = str((os.path.basename(path)))
            # parse parent folder to get cohort name
            path_head = os.path.dirname(path)
            path_second_tail = str(os.path.basename(path_head))
            # use cohort and measure composite key to access dictionary values
            path_key = (path_second_tail, path_tail)
            excel_file_data = open_excel_file(path, files)
            # print(path_key)
            # print(path_key[1])
            # print(wrangled_hyperlinks_dict[path_key][1])
            # Run file_filter_switcher

            if path_key[1] in do_not_parse:
                print('passed!')
                pass
            else:
                if edac_readm_sum_option == 1:
                    results = file_filter_switcher(path_second_tail, wrangled_hyperlinks_dict[path_key][1], path_tail,
                                                   excel_file_data, wrangled_hyperlinks_dict, path_key,
                                                   edac_readm_sum_option)
                else:
                    results = file_filter_switcher(path_second_tail, wrangled_hyperlinks_dict[path_key][1], path_tail,
                                                   excel_file_data, wrangled_hyperlinks_dict, path_key,
                                                   edac_readm_sum_option, edac_readm_month_list)
                # print(results)
                # print(path_second_tail)
                # print(CAMC_counter)

                if path_second_tail == 'Comprehensive Academic Medical Center':
                    CAMC_df.loc[CAMC_counter] = results
                    CAMC_counter += 1
                    # CAMC_df.loc[results[1]] = results[0]
                elif path_second_tail == 'Large Specialized Complex Care Medical Center':
                    LSCCMC_df.loc[LSCCMC_counter] = results
                    LSCCMC_counter += 1
                elif path_second_tail == 'Complex Care Medical Center':
                    for ii, item in enumerate(results):
                        # print(item)
                        CCMC_df.loc[CCMC_counter] = item
                        CCMC_counter += 1

    # list of domains in the correct order
    domain_list = ['Mortality', 'Efficiency', 'Safety', 'Effectiveness', 'Outpatient']
    result_dfs = [CAMC_df, LSCCMC_df, CCMC_df]
    final_dfs = [CAMC_df_final, LSCCMC_df_final, CCMC_df_final]
    # print(CAMC_df_final)
    for ii, item in enumerate(result_dfs):
        if item.empty == False:
            for jj, item2 in enumerate(domain_list):
                # sort and return the domain values
                try:
                    sub_df = subset_sort_df(item, item2)
                    # print('1')
                    # union all the domain values to the existing dataframe
                    frames = [final_dfs[ii], sub_df]
                    # print('2')
                    # print(final_dfs[ii])
                    concat_df = pd.concat(frames)
                    # print('3')
                    final_dfs[ii] = concat_df
                    # print('4')
                    # print('got to end')
                except:
                    print('Issue subsetting/sorting dataframe.')
                    pass
        elif item.empty == True:
            final_dfs[ii] = item

    os.chdir(wd_dest)
    if CAMC_df.empty == False:
        CAMC_df.to_csv('CAMC_df1.csv')
    else:
        pass
    if LSCCMC_df.empty == False:
        LSCCMC_df.to_csv('LSCCMC_df1.csv')
    else:
        pass
    if CCMC_df.empty == False:
        CCMC_df.to_csv('CCMC_df1.csv')
    else:
        pass
    if final_dfs[0].empty == False:
        final_dfs[0].to_csv('CAMC_df_sorted.csv')
    else:
        pass
    if final_dfs[1].empty == False:
        final_dfs[1].to_csv('LSCCMC_df_sorted.csv')
    else:
        pass
    if final_dfs[2].empty == False:
        final_dfs[2].to_csv('CCMC_df_sorted.csv')
    else:
        pass

    return ([CAMC_df, LSCCMC_df, CCMC_df])


##############################################################################################

def return_event_type(domain_nm, measure_nm):
    if domain_nm == 'Mortality':

        event_type = 'Deaths (Obs)'

        return (event_type)

    elif domain_nm == 'Effectiveness':

        if measure_nm.split('_')[0] == 'EDAC':

            event_type = 'Excess Days'
            return (event_type)
            
        #UL003
        #elif measure_nm.split('_')[0] == 'READM':
        elif measure_nm.split('_')[0] == 'READM' and measure_nm.split('_')[1] not in ['ARTHRO','CHOL','COLON','URI']:

            event_type = 'Revisit Inpatient Cases'

            return (event_type)
        
        #UL003
        elif measure_nm.split('_')[0] == 'READM' and domain_nm == 'Effectiveness' and measure_nm.split('_')[1] in ['ARTHRO','CHOL','COLON','URI']:
            
            event_type = 'Readmit Rate Num Cases (Readmit Cases)'

            return (event_type)
        
        #UL003
        elif measure_nm.split('_')[0] == 'EARLY':

            event_type = 'Cases'

            return (event_type)
        
        #UL003
        elif measure_nm.split('_')[1] in ['REVISIT','RETURN']:

            event_type = 'Readmit Rate Num Cases (Readmit Cases)'

            return (event_type)

    elif domain_nm == 'Efficiency':

        if measure_nm.split('_')[0] == 'DCOST':

            event_type = 'Mean Direct Cost (Obs)'

            return (event_type)

        elif measure_nm.split('_')[0] == 'LOS':

            event_type = 'Mean LOS (Obs)'

            return (event_type)

    elif domain_nm == 'Safety':

        if measure_nm.split('_')[0] == 'PSI':

            event_type = 'AHRQ Safety Numerator'

            return (event_type)

        elif measure_nm == 'THK':

            event_type = 'Total THK cases with Complications'

            return (event_type)
        
        #UL003
        elif measure_nm.split('_')[0] == 'ADE':

            event_type = 'Cases'

            return (event_type)

    elif domain_nm == 'Outpatient':

        event_type = 'Readmit Rate Num Cases (Readmit Cases)'

        return (event_type)


##############################################################################################

# same function as above for denominators.

def return_denom_type(domain_nm, measure_nm):
    if domain_nm == 'Mortality':

        event_type = 'Cases'

        return (event_type)

    elif domain_nm == 'Effectiveness':

        if measure_nm.split('_')[0] == 'EDAC':

            event_type = 'Total Index Encounters'
            return (event_type)

        #UL003
        #elif measure_nm.split('_')[0] == 'READM':
        elif measure_nm.split('_')[0] == 'READM' and measure_nm.split('_')[1] not in ['ARTHRO','CHOL','COLON','URI']:

            event_type = 'Total Index Encounters'

            return (event_type)
        
        #UL003
        elif measure_nm.split('_')[0] == 'READM' and domain_nm == 'Effectiveness' and measure_nm.split('_')[1] in ['ARTHRO','CHOL','COLON','URI']:
            
            event_type = 'Readmit Rate Denom Cases'

            return (event_type)
        
        #UL003
        elif measure_nm.split('_')[0] == 'EARLY':

            event_type = 'Cases'

            return (event_type)
        
        #UL003
        elif measure_nm.split('_')[1] in ['REVISIT','RETURN']:

            event_type = 'Readmit Rate Denom Cases'

            return (event_type)

    elif domain_nm == 'Efficiency':

        if measure_nm.split('_')[0] == 'DCOST':

            event_type = 'Cases'

            return (event_type)

        elif measure_nm.split('_')[0] == 'LOS':

            event_type = 'Cases'

            return (event_type)

    elif domain_nm == 'Safety':

        if measure_nm.split('_')[0] == 'PSI':

            event_type = 'AHRQ Safety Denominator'

            return (event_type)

        elif measure_nm == 'THK':

            event_type = 'Total THK cases'

            return (event_type)
        
        #UL003
        elif measure_nm.split('_')[0] == 'ADE':

            event_type = 'Cases'

            return (event_type)

    elif domain_nm == 'Outpatient':

        event_type = 'Readmit Rate Denom Cases'

        return (event_type)

###############################################################################################
'''
def vizient_data_folder_walker_and_prep_for_db_inserts():
    # This function puts together all the above data wrangling functions
    # The goal of this function is to parse all Vizient Q&A files within a folder structure,
    # extract the correct data column value per measure and cohort, sort the measure values by Domain
    # then concatenate all domain values in the same order as the Vizient Calculator.
    # Last step is to write to csv file.
    # Set the path of the folder structure we want to recursively walk through
    wd = input('Enter the path of the folder you want to parse.')
    wd = os.path.abspath(wd)
    # Set teh path of the folder structure where we want the final files
    wd_dest = input('Enter the path of the folder you want to final files to go in.')
    wd_dest = os.path.abspath(wd_dest)

    # step 2:  Import cohort data from the Vizient documentation file.
    cohort_helper_df = gather_cohort_data()

    # step 3:  Import the Vizient template hyperlink file and generate a helper dictionary from it.
    hyperlinks_helper_df = get_report_template_links()

    merged_hyperlink_helper_df = pd.merge(cohort_helper_df, hyperlinks_helper_df, on='Hospital')
    wrangled_hyperlinks_dict = create_hyperlink_dict_wrangle(merged_hyperlink_helper_df)

    # Check which type of sum the client wants for edac and readmission values
    edac_readm_sum_option = 'not answered'
    while edac_readm_sum_option not in [1, 2]:
        try:
            edac_readm_sum_option = int(input(
                "EDAC and Readmission values were pulled by quarter.  Do you want the sum of all months (choose 1) or do you want the sum of specific months? (choose 2)"))
        except:
            print('Not an integer.')
        if edac_readm_sum_option not in [1, 2]:
            print('Please choose 1 or 2')

    # if choice 2, then get a list of months to sum up.
    if edac_readm_sum_option == 2:
        try:
            print(
                'You chose "sum of specific months" (choice 2).  Please add all months in the quarter date range you would like to sum for EDAC and READMISSIONS.')
            edac_readm_month_list = []
            month = 'GO'
            r = re.compile('\d\d\d\d-\d\d')
            while month != 'STOP':
                month = input(
                    'ADD Month with "YYYY-MM" format.  Type "STOP" to stop adding months to custom month list.')
                if r.match(month) is not None:
                    print('Matches Format.  Added to month list')
                    edac_readm_month_list.append(month)
                elif month.upper() == 'STOP':
                    print('Stopping...')
                    print(edac_readm_month_list)
                    break
                else:
                    print('That does not match "YYYY-MM" format.  Try again.')
            # if input is not-integer, just print the list
        except:
            print(edac_readm_month_list)

    conn = pyodbc.connect('Driver={SQL Server};'
                          'Server=edw00pd05wva.corp.nm.org\EDWIDS1;'
                          'Database=clarity;'
                          'Trusted_Connection=yes;')

    # Set counters which will act as proxy indexes in order to append rows to pandas dataframes
    CAMC_counter = 0
    LSCCMC_counter = 0
    CCMC_counter = 0

    CAMC_events_counter = 0
    LSCCMC_events_counter = 0
    CCMC_events_counter = 0

    # Create empty dataframes to store final values.  First dataframe stores the unsorted values, the _final dataframe
    # will store the sorted values and will be written to a csv file
    CAMC_df = pd.DataFrame(columns=['Hospital', 'Domain', 'Measure', 'Metric Value'])
    LSCCMC_df = pd.DataFrame(columns=['Hospital', 'Domain', 'Measure', 'Metric Value'])
    CCMC_df = pd.DataFrame(columns=['Hospital', 'Domain', 'Measure', 'Metric Value'])

    CAMC_events_df = pd.DataFrame(columns=['Hospital', 'Domain', 'Measure', 'N Events', 'event_type_nm'])
    LSCCMC_events_df = pd.DataFrame(columns=['Hospital', 'Domain', 'Measure', 'N Events', 'event_type_nm'])
    CCMC_events_df = pd.DataFrame(columns=['Hospital', 'Domain', 'Measure', 'N Events', 'event_type_nm'])

    # Query the NM_Analytics.vizient_qa database to grab all hospitals within each Vizient cohort.
    camc_cohort_hosp = """
    SELECT
    concat(hospital_medicare_id,' ',hospital_name) as hospital_name
    FROM
    NM_Analytics_prototype.vizient_qa.hospitals
    where
    hospital_cohort_id = 1  --Comprehensive Academic Medical Center
    """

    lsccmc_cohort_hosp = """
    SELECT
    concat(hospital_medicare_id,' ',hospital_name) as hospital_name
    FROM
    NM_Analytics_prototype.vizient_qa.hospitals
    where
    hospital_cohort_id = 2  --Large Specialized Complex Care Medical Center
    """

    ccmc_cohort_hosp = """
    SELECT
    concat(hospital_medicare_id,' ',hospital_name) as hospital_name
    FROM
    NM_Analytics_prototype.vizient_qa.hospitals
    where
    hospital_cohort_id = 3	--Complex Care Medical Center
    """

    comm_cohort_hosp = """
    SELECT
    concat(hospital_medicare_id,' ',hospital_name) as hospital_name
    FROM
    NM_Analytics_prototype.vizient_qa.hospitals
    where
    hospital_cohort_id = 4	--Community
    """
    # Execute the query and store the values in pandas dataframes
    camc_cohort = pd.DataFrame(pd.read_sql(camc_cohort_hosp, conn))
    lsccmc_cohort = pd.DataFrame(pd.read_sql(lsccmc_cohort_hosp, conn))
    ccmc_cohort = pd.DataFrame(pd.read_sql(ccmc_cohort_hosp, conn))
    comm_cohort = pd.DataFrame(pd.read_sql(comm_cohort_hosp, conn))

    # Convert the dataframes into lists in order to use the list in the dataframe filter functions
    camc_cohort_list = camc_cohort['hospital_name'].values.tolist()
    lsccmc_cohort_list = lsccmc_cohort['hospital_name'].values.tolist()
    ccmc_cohort_list = ccmc_cohort['hospital_name'].values.tolist()
    comm_cohort_list = comm_cohort['hospital_name'].values.tolist()

    conn.close()

    do_not_parse = ['ED_2B', 'ED_OP_18B']
    # iterate through folder structure.
    for path, dirs, files in os.walk(wd):
        # if files list is not empty, open it up
        if len(files) > 0:
            # parse folder name to get measure name
            path_tail = str((os.path.basename(path)))
            # parse parent folder to get cohort name
            path_head = os.path.dirname(path)
            path_second_tail = str(os.path.basename(path_head))
            # use cohort and measure composite key to access dictionary values
            path_key = (path_second_tail, path_tail)
            excel_file_data = open_excel_file(path, files)
            if path_key[1] in do_not_parse:
                print('passed!')
                pass
            else:
                if edac_readm_sum_option == 1:
                    # collect/calculate measure
                    results = file_filter_switcher_all_cohort_hosps(path_second_tail,
                                                                    wrangled_hyperlinks_dict[path_key][1], path_tail,
                                                                    excel_file_data, wrangled_hyperlinks_dict, path_key,
                                                                    edac_readm_sum_option, camc_cohort_list,
                                                                    lsccmc_cohort_list, ccmc_cohort_list,
                                                                    comm_cohort_list)

                else:
                    # collect/calculate measure
                    results = file_filter_switcher_all_cohort_hosps(path_second_tail,
                                                                    wrangled_hyperlinks_dict[path_key][1], path_tail,
                                                                    excel_file_data, wrangled_hyperlinks_dict, path_key,
                                                                    edac_readm_sum_option, camc_cohort_list,
                                                                    lsccmc_cohort_list, ccmc_cohort_list,
                                                                    comm_cohort_list, edac_readm_month_list)

                #print('Just before n-size section')
                if edac_readm_sum_option == 1:
                    # collect/calculate n events
                    results_n_events = file_filter_switcher_all_cohort_hosps_n_events(path_second_tail,
                                                                                      wrangled_hyperlinks_dict[
                                                                                          path_key][1], path_tail,
                                                                                      excel_file_data,
                                                                                      wrangled_hyperlinks_dict,
                                                                                      path_key, edac_readm_sum_option,
                                                                                      camc_cohort_list,
                                                                                      lsccmc_cohort_list,
                                                                                      ccmc_cohort_list,
                                                                                      comm_cohort_list)
                else:
                    # collect/calculate n events
                    results_n_events = file_filter_switcher_all_cohort_hosps_n_events(path_second_tail,
                                                                                      wrangled_hyperlinks_dict[
                                                                                          path_key][1], path_tail,
                                                                                      excel_file_data,
                                                                                      wrangled_hyperlinks_dict,
                                                                                      path_key, edac_readm_sum_option,
                                                                                      camc_cohort_list,
                                                                                      lsccmc_cohort_list,
                                                                                      ccmc_cohort_list,
                                                                                      comm_cohort_list,
                                                                                      edac_readm_month_list)

                # store measure values into dataframes
                if path_second_tail == 'Comprehensive Academic Medical Center':
                    for ii, item in enumerate(results):
                        # print(item)
                        CAMC_df.loc[CAMC_counter] = item
                        CAMC_counter += 1
                elif path_second_tail == 'Large Specialized Complex Care Medical Center':
                    for ii, item in enumerate(results):
                        # print(item)
                        LSCCMC_df.loc[LSCCMC_counter] = item
                        LSCCMC_counter += 1
                    # LSCCMC_df.loc[LSCCMC_counter] = results
                    # LSCCMC_counter += 1
                elif path_second_tail == 'Complex Care Medical Center':
                    for ii, item in enumerate(results):
                        # print(item)
                        CCMC_df.loc[CCMC_counter] = item
                        CCMC_counter += 1

                # get name of event type
                event_type_nm = return_event_type(wrangled_hyperlinks_dict[path_key][1], path_key[1])

                #print('event_type: ', event_type_nm)
                if path_second_tail == 'Comprehensive Academic Medical Center':
                    for ii, item_n in enumerate(results_n_events):
                        # append event type to data list
                        item_n.append(event_type_nm)
                        CAMC_events_df.loc[CAMC_events_counter] = item_n
                        CAMC_events_counter += 1
                elif path_second_tail == 'Large Specialized Complex Care Medical Center':
                    for ii, item_n in enumerate(results_n_events):
                        # append event type to data list
                        item_n.append(event_type_nm)
                        LSCCMC_events_df.loc[LSCCMC_events_counter] = item_n
                        LSCCMC_events_counter += 1

                elif path_second_tail == 'Complex Care Medical Center':
                    for ii, item_n in enumerate(results_n_events):
                        # append event type to data list
                        item_n.append(event_type_nm)
                        CCMC_events_df.loc[CCMC_events_counter] = item_n
                        CCMC_events_counter += 1

    CAMC_df_merged = pd.merge(CAMC_df, CAMC_events_df, how='left', on=['Hospital', 'Domain', 'Measure'])
    LSCCMC_df_merged = pd.merge(LSCCMC_df, LSCCMC_events_df, how='left', on=['Hospital', 'Domain', 'Measure'])
    CCMC_df_merged = pd.merge(CCMC_df, CCMC_events_df, how='left', on=['Hospital', 'Domain', 'Measure'])

    frames = [CAMC_df_merged, LSCCMC_df_merged, CCMC_df_merged]

    result_df = pd.concat(frames)

    final_result_path = os.path.join(wd_dest, 'custom_time_period_data.csv')
    result_df.to_csv(final_result_path)

    return (result_df)
'''


'''
def vizient_data_folder_walker_and_prep_for_db_inserts(cc_calc_id):
    # This function puts together all the above data wrangling functions
    # The goal of this function is to parse all Vizient Q&A files within a folder structure,
    # extract the correct data column value per measure and cohort, sort the measure values by Domain
    # then concatenate all domain values in the same order as the Vizient Calculator.
    # Last step is to write to csv file.
    # Set the path of the folder structure we want to recursively walk through
    wd = input('Enter the path of the folder you want to parse.')
    wd = os.path.abspath(wd)
    # Set teh path of the folder structure where we want the final files
    wd_dest = input('Enter the path of the folder you want to final files to go in.')
    wd_dest = os.path.abspath(wd_dest)

    # step 2:  Import cohort data from the Vizient documentation file.
    cohort_helper_df = gather_cohort_data()

    # step 3:  Import the Vizient template hyperlink file and generate a helper dictionary from it.
    hyperlinks_helper_df = get_report_template_links()

    merged_hyperlink_helper_df = pd.merge(cohort_helper_df, hyperlinks_helper_df, on='Hospital')
    wrangled_hyperlinks_dict = create_hyperlink_dict_wrangle(merged_hyperlink_helper_df)

    # Check which type of sum the client wants for edac and readmission values
    edac_readm_sum_option = 'not answered'
    while edac_readm_sum_option not in [1, 2]:
        try:
            edac_readm_sum_option = int(input(
                "EDAC and Readmission values were pulled by quarter.  Do you want the sum of all months (choose 1) or do you want the sum of specific months? (choose 2)"))
        except:
            print('Not an integer.')
        if edac_readm_sum_option not in [1, 2]:
            print('Please choose 1 or 2')

    # if choice 2, then get a list of months to sum up.
    if edac_readm_sum_option == 2:
        try:
            print(
                'You chose "sum of specific months" (choice 2).  Please add all months in the quarter date range you would like to sum for EDAC and READMISSIONS.')
            edac_readm_month_list = []
            month = 'GO'
            r = re.compile('\d\d\d\d-\d\d')
            while month != 'STOP':
                month = input(
                    'ADD Month with "YYYY-MM" format.  Type "STOP" to stop adding months to custom month list.')
                if r.match(month) is not None:
                    print('Matches Format.  Added to month list')
                    edac_readm_month_list.append(month)
                elif month.upper() == 'STOP':
                    print('Stopping...')
                    print(edac_readm_month_list)
                    break
                else:
                    print('That does not match "YYYY-MM" format.  Try again.')
            # if input is not-integer, just print the list
        except:
            print(edac_readm_month_list)

    conn = pyodbc.connect('Driver={SQL Server};'
                          'Server=edw00pd05wva.corp.nm.org\EDWIDS1;'
                          'Database=clarity;'
                          'Trusted_Connection=yes;')

    # Set counters which will act as proxy indexes in order to append rows to pandas dataframes
    CAMC_counter = 0
    LSCCMC_counter = 0
    CCMC_counter = 0
    COMM_counter = 0

    CAMC_events_counter = 0
    LSCCMC_events_counter = 0
    CCMC_events_counter = 0
    COMM_events_counter = 0

    CAMC_denom_counter = 0
    LSCCMC_denom_counter = 0
    CCMC_denom_counter = 0
    COMM_denom_counter = 0

    # Create empty dataframes to store final values.  First dataframe stores the unsorted values, the _final dataframe
    # will store the sorted values and will be written to a csv file
    CAMC_df = pd.DataFrame(columns=['Hospital', 'Domain', 'Measure', 'Metric Value'])
    LSCCMC_df = pd.DataFrame(columns=['Hospital', 'Domain', 'Measure', 'Metric Value'])
    CCMC_df = pd.DataFrame(columns=['Hospital', 'Domain', 'Measure', 'Metric Value'])
    COMM_df = pd.DataFrame(columns=['Hospital', 'Domain', 'Measure', 'Metric Value'])

    CAMC_events_df = pd.DataFrame(columns=['Hospital', 'Domain', 'Measure', 'N Events', 'event_type_nm'])
    LSCCMC_events_df = pd.DataFrame(columns=['Hospital', 'Domain', 'Measure', 'N Events', 'event_type_nm'])
    CCMC_events_df = pd.DataFrame(columns=['Hospital', 'Domain', 'Measure', 'N Events', 'event_type_nm'])
    COMM_events_df = pd.DataFrame(columns=['Hospital', 'Domain', 'Measure', 'N Events', 'event_type_nm'])

    CAMC_denom_df = pd.DataFrame(columns=['Hospital', 'Domain', 'Measure', 'denominator', 'denominator_event_type_nm'])
    LSCCMC_denom_df = pd.DataFrame(
        columns=['Hospital', 'Domain', 'Measure', 'denominator', 'denominator_event_type_nm'])
    CCMC_denom_df = pd.DataFrame(columns=['Hospital', 'Domain', 'Measure', 'denominator', 'denominator_event_type_nm'])
    COMM_denom_df = pd.DataFrame(columns=['Hospital', 'Domain', 'Measure', 'denominator', 'denominator_event_type_nm'])

    # Query the NM_Analytics.vizient_qa database to grab all hospitals within each Vizient cohort.
    camc_cohort_hosp = """
    SELECT
    distinct
    concat(hospital_medicare_id,' ',hospital_name) as hospital_name
    FROM
    NM_Analytics_prototype.vizient_qa.hospitals as h
    join NM_Analytics_Prototype.vizient_qa.calc_hospital_cohort as chc
    on chc.hospital_id = h.hospital_id
    where
    chc.hospital_cohort_id = 1  --Comprehensive Academic Medical Center
    """

    lsccmc_cohort_hosp = """
    SELECT
    distinct
    concat(hospital_medicare_id,' ',hospital_name) as hospital_name
    FROM
    NM_Analytics_prototype.vizient_qa.hospitals as h
    join NM_Analytics_Prototype.vizient_qa.calc_hospital_cohort as chc
    on chc.hospital_id = h.hospital_id
    where
    chc.hospital_cohort_id = 2  --Large Specialized Complex Care Medical Center
    """

    ccmc_cohort_hosp = """
    SELECT
    distinct
    concat(hospital_medicare_id,' ',hospital_name) as hospital_name
    FROM
    NM_Analytics_prototype.vizient_qa.hospitals as h
    join NM_Analytics_Prototype.vizient_qa.calc_hospital_cohort as chc
    on chc.hospital_id = h.hospital_id
    where
    chc.hospital_cohort_id = 3	--Complex Care Medical Center
    """

    comm_cohort_hosp = """
    SELECT
    distinct
    concat(hospital_medicare_id,' ',hospital_name) as hospital_name
    FROM
    NM_Analytics_prototype.vizient_qa.hospitals as h
    join NM_Analytics_Prototype.vizient_qa.calc_hospital_cohort as chc
    on chc.hospital_id = h.hospital_id
    where
    chc.hospital_cohort_id = 4	--Community
    """
    # Execute the query and store the values in pandas dataframes
    camc_cohort = pd.DataFrame(pd.read_sql(camc_cohort_hosp, conn))
    lsccmc_cohort = pd.DataFrame(pd.read_sql(lsccmc_cohort_hosp, conn))
    ccmc_cohort = pd.DataFrame(pd.read_sql(ccmc_cohort_hosp, conn))
    comm_cohort = pd.DataFrame(pd.read_sql(comm_cohort_hosp, conn))

    # Convert the dataframes into lists in order to use the list in the dataframe filter functions
    camc_cohort_list = camc_cohort['hospital_name'].values.tolist()
    lsccmc_cohort_list = lsccmc_cohort['hospital_name'].values.tolist()
    ccmc_cohort_list = ccmc_cohort['hospital_name'].values.tolist()
    comm_cohort_list = comm_cohort['hospital_name'].values.tolist()

    conn.close()

    do_not_parse = ['ED_2B', 'ED_OP_18B']
    # iterate through folder structure.
    for path, dirs, files in os.walk(wd):
        # if files list is not empty, open it up
        if len(files) > 0:
            # parse folder name to get measure name
            path_tail = str((os.path.basename(path)))
            # parse parent folder to get cohort name
            path_head = os.path.dirname(path)
            path_second_tail = str(os.path.basename(path_head))
            # use cohort and measure composite key to access dictionary values
            path_key = (path_second_tail, path_tail)
            print(files)
            excel_file_data = open_excel_file(path, files)
            if path_key[1] in do_not_parse:
                print('passed!')
                pass
            else:
                if edac_readm_sum_option == 1:
                    # collect/calculate measure
                    results = file_filter_switcher_all_cohort_hosps(path_second_tail,
                                                                    wrangled_hyperlinks_dict[path_key][1], path_tail,
                                                                    excel_file_data, wrangled_hyperlinks_dict, path_key,
                                                                    edac_readm_sum_option, camc_cohort_list,
                                                                    lsccmc_cohort_list, ccmc_cohort_list,
                                                                    comm_cohort_list)

                else:
                    # collect/calculate measure
                    results = file_filter_switcher_all_cohort_hosps(path_second_tail,
                                                                    wrangled_hyperlinks_dict[path_key][1], path_tail,
                                                                    excel_file_data, wrangled_hyperlinks_dict, path_key,
                                                                    edac_readm_sum_option, camc_cohort_list,
                                                                    lsccmc_cohort_list, ccmc_cohort_list,
                                                                    comm_cohort_list, edac_readm_month_list)

                #print('Just before n-size section')
                if edac_readm_sum_option == 1:
                    # collect/calculate n events
                    results_n_events = file_filter_switcher_all_cohort_hosps_n_events(path_second_tail,
                                                                                      wrangled_hyperlinks_dict[
                                                                                          path_key][1], path_tail,
                                                                                      excel_file_data,
                                                                                      wrangled_hyperlinks_dict,
                                                                                      path_key, edac_readm_sum_option,
                                                                                      camc_cohort_list,
                                                                                      lsccmc_cohort_list,
                                                                                      ccmc_cohort_list,
                                                                                      comm_cohort_list)
                else:
                    # collect/calculate n events
                    results_n_events = file_filter_switcher_all_cohort_hosps_n_events(path_second_tail,
                                                                                      wrangled_hyperlinks_dict[
                                                                                          path_key][1], path_tail,
                                                                                      excel_file_data,
                                                                                      wrangled_hyperlinks_dict,
                                                                                      path_key, edac_readm_sum_option,
                                                                                      camc_cohort_list,
                                                                                      lsccmc_cohort_list,
                                                                                      ccmc_cohort_list,
                                                                                      comm_cohort_list,
                                                                                      edac_readm_month_list)

                #print('Just before denominator section')
                if edac_readm_sum_option == 1:
                    # collect/calculate n events
                    results_d_events = file_filter_switcher_all_cohort_hosps_d_events(path_second_tail,
                                                                                      wrangled_hyperlinks_dict[
                                                                                          path_key][1], path_tail,
                                                                                      excel_file_data,
                                                                                      wrangled_hyperlinks_dict,
                                                                                      path_key, edac_readm_sum_option,
                                                                                      camc_cohort_list,
                                                                                      lsccmc_cohort_list,
                                                                                      ccmc_cohort_list,
                                                                                      comm_cohort_list)
                else:
                    # collect/calculate n events
                    results_d_events = file_filter_switcher_all_cohort_hosps_d_events(path_second_tail,
                                                                                      wrangled_hyperlinks_dict[
                                                                                          path_key][1], path_tail,
                                                                                      excel_file_data,
                                                                                      wrangled_hyperlinks_dict,
                                                                                      path_key, edac_readm_sum_option,
                                                                                      camc_cohort_list,
                                                                                      lsccmc_cohort_list,
                                                                                      ccmc_cohort_list,
                                                                                      comm_cohort_list,
                                                                                      edac_readm_month_list)
                #print(results)
                    # store measure values into dataframes
                if path_second_tail == 'Comprehensive Academic Medical Center':
                    for ii, item in enumerate(results):
                        # print(item)
                        CAMC_df.loc[CAMC_counter] = item
                        CAMC_counter += 1
                elif path_second_tail == 'Large Specialized Complex Care Medical Center':
                    for ii, item in enumerate(results):
                        # print(item)
                        LSCCMC_df.loc[LSCCMC_counter] = item
                        LSCCMC_counter += 1
                    # LSCCMC_df.loc[LSCCMC_counter] = results
                    # LSCCMC_counter += 1
                elif path_second_tail == 'Complex Care Medical Center':
                    for ii, item in enumerate(results):
                        # print(item)
                        CCMC_df.loc[CCMC_counter] = item
                        CCMC_counter += 1
                elif path_second_tail == 'Community Medical Center':
                    for ii, item in enumerate(results):
                        # print(item)
                        COMM_df.loc[COMM_counter] = item
                        COMM_counter += 1

                # get name of event type
                event_type_nm = return_event_type(wrangled_hyperlinks_dict[path_key][1], path_key[1])

                print('event_type: ', event_type_nm)
                if path_second_tail == 'Comprehensive Academic Medical Center':
                    for ii, item_n in enumerate(results_n_events):
                        # append event type to data list
                        item_n.append(event_type_nm)
                        CAMC_events_df.loc[CAMC_events_counter] = item_n
                        CAMC_events_counter += 1
                elif path_second_tail == 'Large Specialized Complex Care Medical Center':
                    for ii, item_n in enumerate(results_n_events):
                        # append event type to data list
                        item_n.append(event_type_nm)
                        LSCCMC_events_df.loc[LSCCMC_events_counter] = item_n
                        LSCCMC_events_counter += 1

                elif path_second_tail == 'Complex Care Medical Center':
                    for ii, item_n in enumerate(results_n_events):
                        # append event type to data list
                        item_n.append(event_type_nm)
                        CCMC_events_df.loc[CCMC_events_counter] = item_n
                        CCMC_events_counter += 1
                elif path_second_tail == 'Community Medical Center':
                    for ii, item_n in enumerate(results_n_events):
                        # append event type to data list
                        item_n.append(event_type_nm)
                        COMM_events_df.loc[COMM_events_counter] = item_n
                        COMM_events_counter += 1

                # get name of event type
                denom_type_nm = return_denom_type(wrangled_hyperlinks_dict[path_key][1], path_key[1])

                print('event_type: ', denom_type_nm)
                if path_second_tail == 'Comprehensive Academic Medical Center':
                    for ii, item_d in enumerate(results_d_events):
                        # append event type to data list
                        item_d.append(denom_type_nm)
                        CAMC_denom_df.loc[CAMC_denom_counter] = item_d
                        CAMC_denom_counter += 1
                elif path_second_tail == 'Large Specialized Complex Care Medical Center':
                    for ii, item_d in enumerate(results_d_events):
                        # append event type to data list
                        item_d.append(denom_type_nm)
                        LSCCMC_denom_df.loc[LSCCMC_denom_counter] = item_d
                        LSCCMC_denom_counter += 1

                elif path_second_tail == 'Complex Care Medical Center':
                    for ii, item_d in enumerate(results_d_events):
                        # append event type to data list
                        item_d.append(denom_type_nm)
                        CCMC_denom_df.loc[CCMC_denom_counter] = item_d
                        CCMC_denom_counter += 1
                elif path_second_tail == 'Community Medical Center':
                    for ii, item_d in enumerate(results_d_events):
                        # append event type to data list
                        item_d.append(denom_type_nm)
                        COMM_denom_df.loc[COMM_denom_counter] = item_d
                        COMM_denom_counter += 1

    CAMC_df_merged = pd.merge(CAMC_df, CAMC_events_df, how='left', on=['Hospital', 'Domain', 'Measure'])
    LSCCMC_df_merged = pd.merge(LSCCMC_df, LSCCMC_events_df, how='left', on=['Hospital', 'Domain', 'Measure'])
    CCMC_df_merged = pd.merge(CCMC_df, CCMC_events_df, how='left', on=['Hospital', 'Domain', 'Measure'])
    COMM_df_merged = pd.merge(COMM_df, COMM_events_df, how='left', on=['Hospital', 'Domain', 'Measure'])

    CAMC_df_merged2 = pd.merge(CAMC_df_merged, CAMC_denom_df, how='left', on=['Hospital', 'Domain', 'Measure'])
    LSCCMC_df_merged2 = pd.merge(LSCCMC_df_merged, LSCCMC_denom_df, how='left', on=['Hospital', 'Domain', 'Measure'])
    CCMC_df_merged2 = pd.merge(CCMC_df_merged, CCMC_denom_df, how='left', on=['Hospital', 'Domain', 'Measure'])
    COMM_df_merged2 = pd.merge(COMM_df_merged, COMM_denom_df, how='left', on=['Hospital', 'Domain', 'Measure'])

    frames = [CAMC_df_merged2, LSCCMC_df_merged2, CCMC_df_merged2,COMM_df_merged2]

    result_df = pd.concat(frames,sort=True)
    result_df.drop_duplicates(inplace=True)
    final_result_path = os.path.join(wd_dest, 'custom_time_period_data.csv')
    result_df.to_csv(final_result_path)

    return (result_df)

'''
#UL003
#updated vizient folder walker for fy21 2020 period 2 calculators
###############################################################################
###############################################################################

def vizient_data_folder_walker_and_prep_for_db_inserts(cc_calc_id):
    # This function puts together all the above data wrangling functions
    # The goal of this function is to parse all Vizient Q&A files within a folder structure,
    # extract the correct data column value per measure and cohort, sort the measure values by Domain
    # then concatenate all domain values in the same order as the Vizient Calculator.
    # Last step is to write to csv file.
    # Set the path of the folder structure we want to recursively walk through
    wd = input('Enter the path of the folder you want to parse.')
    wd = os.path.abspath(wd)
    # Set teh path of the folder structure where we want the final files
    wd_dest = input('Enter the path of the folder you want to final files to go in.')
    wd_dest = os.path.abspath(wd_dest)

    # step 2:  Import cohort data from the Vizient documentation file.
    cohort_helper_df = gather_cohort_data()

    # step 3:  Import the Vizient template hyperlink file and generate a helper dictionary from it.
    hyperlinks_helper_df = get_report_template_links()

    merged_hyperlink_helper_df = pd.merge(cohort_helper_df, hyperlinks_helper_df, on='Hospital')
    wrangled_hyperlinks_dict = create_hyperlink_dict_wrangle(merged_hyperlink_helper_df)

    # Check which type of sum the client wants for edac and readmission values
    edac_readm_sum_option = 'not answered'
    
    
    while edac_readm_sum_option not in [1, 2]:
        try:
            edac_readm_sum_option = int(input(
                "EDAC and Readmission values were pulled by quarter.  Do you want the sum of all months (choose 1) or do you want the sum of specific months? (choose 2)"))
        except:
            print('Not an integer.')
        if edac_readm_sum_option not in [1, 2]:
            print('Please choose 1 or 2')

    # if choice 2, then get a list of months to sum up.
    if edac_readm_sum_option == 2:
        try:
            print(
                'You chose "sum of specific months" (choice 2).  Please add all months in the quarter date range you would like to sum for EDAC and READMISSIONS.')
            edac_readm_month_list = []
            month = 'GO'
            r = re.compile('\d\d\d\d-\d\d')
            while month != 'STOP':
                month = input(
                    'ADD Month with "YYYY-MM" format.  Type "STOP" to stop adding months to custom month list.')
                if r.match(month) is not None:
                    print('Matches Format.  Added to month list')
                    edac_readm_month_list.append(month)
                elif month.upper() == 'STOP':
                    print('Stopping...')
                    print(edac_readm_month_list)
                    break
                else:
                    print('That does not match "YYYY-MM" format.  Try again.')
            # if input is not-integer, just print the list
        except:
            print(edac_readm_month_list)
            
        # Set counters which will act as proxy indexes in order to append rows to pandas dataframes
    CAMC_counter = 0
    LSCCMC_counter = 0
    CCMC_counter = 0
    COMM_counter = 0
    #UL003
    CASC_counter= 0

    CAMC_events_counter = 0
    LSCCMC_events_counter = 0
    CCMC_events_counter = 0
    COMM_events_counter = 0
    #UL003
    CASC_events_counter = 0

    CAMC_denom_counter = 0
    LSCCMC_denom_counter = 0
    CCMC_denom_counter = 0
    COMM_denom_counter = 0
    #UL003
    CASC_denom_counter = 0

    # Create empty dataframes to store final values.  First dataframe stores the unsorted values, the _final dataframe
    # will store the sorted values and will be written to a csv file
    CAMC_df = pd.DataFrame(columns=['Hospital', 'Domain', 'Measure', 'Metric Value'])
    LSCCMC_df = pd.DataFrame(columns=['Hospital', 'Domain', 'Measure', 'Metric Value'])
    CCMC_df = pd.DataFrame(columns=['Hospital', 'Domain', 'Measure', 'Metric Value'])
    COMM_df = pd.DataFrame(columns=['Hospital', 'Domain', 'Measure', 'Metric Value'])
    #UL003
    CASC_df = pd.DataFrame(columns=['Hospital', 'Domain', 'Measure', 'Metric Value'])

    CAMC_events_df = pd.DataFrame(columns=['Hospital', 'Domain', 'Measure', 'N Events', 'event_type_nm'])
    LSCCMC_events_df = pd.DataFrame(columns=['Hospital', 'Domain', 'Measure', 'N Events', 'event_type_nm'])
    CCMC_events_df = pd.DataFrame(columns=['Hospital', 'Domain', 'Measure', 'N Events', 'event_type_nm'])
    COMM_events_df = pd.DataFrame(columns=['Hospital', 'Domain', 'Measure', 'N Events', 'event_type_nm'])
    #UL003
    CASC_events_df = pd.DataFrame(columns=['Hospital', 'Domain', 'Measure', 'N Events', 'event_type_nm'])

    CAMC_denom_df = pd.DataFrame(columns=['Hospital', 'Domain', 'Measure', 'denominator', 'denominator_event_type_nm'])
    LSCCMC_denom_df = pd.DataFrame(
        columns=['Hospital', 'Domain', 'Measure', 'denominator', 'denominator_event_type_nm'])
    CCMC_denom_df = pd.DataFrame(columns=['Hospital', 'Domain', 'Measure', 'denominator', 'denominator_event_type_nm'])
    COMM_denom_df = pd.DataFrame(columns=['Hospital', 'Domain', 'Measure', 'denominator', 'denominator_event_type_nm'])
    #UL003
    CASC_denom_df = pd.DataFrame(columns=['Hospital', 'Domain', 'Measure', 'denominator', 'denominator_event_type_nm'])

        
        
    conn = pyodbc.connect('Driver={SQL Server};'
                          'Server=edw00pd05wva.corp.nm.org\EDWIDS1;'
                          'Database=clarity;'
                          'Trusted_Connection=yes;')

        # Query the NM_Analytics.vizient_qa database to grab all hospitals within each Vizient cohort.
    camc_cohort_hosp = """
    SELECT
    distinct
    concat(hospital_medicare_id,' ',hospital_name) as hospital_name
    FROM
    NM_Analytics_prototype.vizient_qa.hospitals as h
    join NM_Analytics_Prototype.vizient_qa.calc_hospital_cohort as chc
    on chc.hospital_id = h.hospital_id
    where
    chc.hospital_cohort_id = 1  --Comprehensive Academic Medical Center
    and
    chc.calc_id = %s
    """ % cc_calc_id

    lsccmc_cohort_hosp = """
    SELECT
    distinct
    concat(hospital_medicare_id,' ',hospital_name) as hospital_name
    FROM
    NM_Analytics_prototype.vizient_qa.hospitals as h
    join NM_Analytics_Prototype.vizient_qa.calc_hospital_cohort as chc
    on chc.hospital_id = h.hospital_id
    where
    chc.hospital_cohort_id = 2  --Large Specialized Complex Care Medical Center
    and
    chc.calc_id = %s
    """ % cc_calc_id

    ccmc_cohort_hosp = """
    SELECT
    distinct
    concat(hospital_medicare_id,' ',hospital_name) as hospital_name
    FROM
    NM_Analytics_prototype.vizient_qa.hospitals as h
    join NM_Analytics_Prototype.vizient_qa.calc_hospital_cohort as chc
    on chc.hospital_id = h.hospital_id
    where
    chc.hospital_cohort_id = 3	--Complex Care Medical Center
    and
    chc.calc_id = %s
    """ % cc_calc_id

    comm_cohort_hosp = """
    SELECT
    distinct
    concat(hospital_medicare_id,' ',hospital_name) as hospital_name
    FROM
    NM_Analytics_prototype.vizient_qa.hospitals as h
    join NM_Analytics_Prototype.vizient_qa.calc_hospital_cohort as chc
    on chc.hospital_id = h.hospital_id
    where
    chc.hospital_cohort_id = 4	--Community
    and
    chc.calc_id = %s
    """  % cc_calc_id

    #UL003
    casc_cohort_hosp = """
    SELECT
    distinct
    concat(hospital_medicare_id,' ',hospital_name) as hospital_name
    FROM
    NM_Analytics_prototype.vizient_qa.hospitals as h
    join NM_Analytics_Prototype.vizient_qa.calc_hospital_cohort as chc
    on chc.hospital_id = h.hospital_id
    where
    chc.hospital_cohort_id = 5 --Critical Access & Small Community
    and
    chc.calc_id = %s
    """  % cc_calc_id

    # Execute the query and store the values in pandas dataframes
    camc_cohort = pd.DataFrame(pd.read_sql(camc_cohort_hosp, conn))
    lsccmc_cohort = pd.DataFrame(pd.read_sql(lsccmc_cohort_hosp, conn))
    ccmc_cohort = pd.DataFrame(pd.read_sql(ccmc_cohort_hosp, conn))
    comm_cohort = pd.DataFrame(pd.read_sql(comm_cohort_hosp, conn))
    #UL003
    casc_cohort = pd.DataFrame(pd.read_sql(casc_cohort_hosp, conn))

    # Convert the dataframes into lists in order to use the list in the dataframe filter functions
    camc_cohort_list = camc_cohort['hospital_name'].values.tolist()
    lsccmc_cohort_list = lsccmc_cohort['hospital_name'].values.tolist()
    ccmc_cohort_list = ccmc_cohort['hospital_name'].values.tolist()
    comm_cohort_list = comm_cohort['hospital_name'].values.tolist()
    #UL003
    casc_cohort_list = casc_cohort['hospital_name'].values.tolist()

    conn.close()

    do_not_parse = ['ED_2B', 'ED_OP_18B']

    for path, dirs, files in os.walk(wd):
            # if files list is not empty, open it up
            if len(files) > 0:
                # parse folder name to get measure name
                path_tail = str((os.path.basename(path)))
                # parse parent folder to get cohort name
                path_head = os.path.dirname(path)
                path_second_tail = str(os.path.basename(path_head))
                # use cohort and measure composite key to access dictionary values

                #UL003
                #Critical Access ADE & % Early Transfers Out require 2 files, a numerator and denominator.
                #Since these file names no longer match the hyperlink file names (ADE,EARLY_TRANS),
                #we need to convert these names to use in the dictionary.

                if path_tail.split('_')[0] == 'ADE':
                    path_tail2 = 'ADE'
                elif path_tail.split('_')[0] == 'EARLY':
                    path_tail2 = 'EARLY_TRANS'
                else:
                    path_tail2 = path_tail

                path_key = (path_second_tail, path_tail2)
                print(files)
                excel_file_data = open_excel_file(path, files)
                if path_key[1] in do_not_parse:
                    print('passed!')
                    pass
                else:
                    if edac_readm_sum_option == 1:
                        # collect/calculate measure
                        results = file_filter_switcher_all_cohort_hosps(path_second_tail,
                                                                        wrangled_hyperlinks_dict[path_key][1], path_tail,
                                                                        excel_file_data, wrangled_hyperlinks_dict, path_key,
                                                                        edac_readm_sum_option, camc_cohort_list,
                                                                        lsccmc_cohort_list, ccmc_cohort_list,
                                                                        comm_cohort_list,casc_cohort_list)

                    else:
                        # collect/calculate measure
                        results = file_filter_switcher_all_cohort_hosps(path_second_tail,
                                                                        wrangled_hyperlinks_dict[path_key][1], path_tail,
                                                                        excel_file_data, wrangled_hyperlinks_dict, path_key,
                                                                        edac_readm_sum_option, camc_cohort_list,
                                                                        lsccmc_cohort_list, ccmc_cohort_list,
                                                                        comm_cohort_list,casc_cohort_list, edac_readm_month_list)

                    #print('Just before n-size section')
                    if edac_readm_sum_option == 1:
                        # collect/calculate n events
                        results_n_events = file_filter_switcher_all_cohort_hosps_n_events(path_second_tail,
                                                                                          wrangled_hyperlinks_dict[
                                                                                              path_key][1], path_tail,
                                                                                          excel_file_data,
                                                                                          wrangled_hyperlinks_dict,
                                                                                          path_key, edac_readm_sum_option,
                                                                                          camc_cohort_list,
                                                                                          lsccmc_cohort_list,
                                                                                          ccmc_cohort_list,
                                                                                          comm_cohort_list,casc_cohort_list)
                    else:
                        # collect/calculate n events
                        results_n_events = file_filter_switcher_all_cohort_hosps_n_events(path_second_tail,
                                                                                          wrangled_hyperlinks_dict[
                                                                                              path_key][1], path_tail,
                                                                                          excel_file_data,
                                                                                          wrangled_hyperlinks_dict,
                                                                                          path_key, edac_readm_sum_option,
                                                                                          camc_cohort_list,
                                                                                          lsccmc_cohort_list,
                                                                                          ccmc_cohort_list,
                                                                                          comm_cohort_list,casc_cohort_list,
                                                                                          edac_readm_month_list)

                    #print('Just before denominator section')
                    if edac_readm_sum_option == 1:
                        # collect/calculate n events
                        results_d_events = file_filter_switcher_all_cohort_hosps_d_events(path_second_tail,
                                                                                          wrangled_hyperlinks_dict[
                                                                                              path_key][1], path_tail,
                                                                                          excel_file_data,
                                                                                          wrangled_hyperlinks_dict,
                                                                                          path_key, edac_readm_sum_option,
                                                                                          camc_cohort_list,
                                                                                          lsccmc_cohort_list,
                                                                                          ccmc_cohort_list,
                                                                                          comm_cohort_list,casc_cohort_list)
                    else:
                        # collect/calculate n events
                        results_d_events = file_filter_switcher_all_cohort_hosps_d_events(path_second_tail,
                                                                                          wrangled_hyperlinks_dict[
                                                                                              path_key][1], path_tail,
                                                                                          excel_file_data,
                                                                                          wrangled_hyperlinks_dict,
                                                                                          path_key, edac_readm_sum_option,
                                                                                          camc_cohort_list,
                                                                                          lsccmc_cohort_list,
                                                                                          ccmc_cohort_list,
                                                                                          comm_cohort_list,casc_cohort_list,
                                                                                          edac_readm_month_list)
                    #print(results)
                        # store measure values into dataframes
                    if path_second_tail == 'Comprehensive Academic Medical Center':
                        for ii, item in enumerate(results):
                            # print(item)
                            CAMC_df.loc[CAMC_counter] = item
                            CAMC_counter += 1
                    elif path_second_tail == 'Large Specialized Complex Care Medical Center':
                        for ii, item in enumerate(results):
                            # print(item)
                            LSCCMC_df.loc[LSCCMC_counter] = item
                            LSCCMC_counter += 1
                        # LSCCMC_df.loc[LSCCMC_counter] = results
                        # LSCCMC_counter += 1
                    elif path_second_tail == 'Complex Care Medical Center':
                        for ii, item in enumerate(results):
                            # print(item)
                            CCMC_df.loc[CCMC_counter] = item
                            CCMC_counter += 1
                    elif path_second_tail == 'Community Medical Center':
                        for ii, item in enumerate(results):
                            # print(item)
                            COMM_df.loc[COMM_counter] = item
                            COMM_counter += 1
                    #UL003
                    elif path_second_tail == 'Critical Access & Small Community':
                        #print("KEY:",path_tail)
                        #print(results)
                        for ii, item in enumerate(results):
                            print(item)
                            CASC_df.loc[CASC_counter] = item
                            CASC_counter += 1

                    # get name of event type
                    event_type_nm = return_event_type(wrangled_hyperlinks_dict[path_key][1], path_key[1])

                    print('event_type: ', event_type_nm)
                    if path_second_tail == 'Comprehensive Academic Medical Center':
                        for ii, item_n in enumerate(results_n_events):
                            # append event type to data list
                            item_n.append(event_type_nm)
                            CAMC_events_df.loc[CAMC_events_counter] = item_n
                            CAMC_events_counter += 1
                    elif path_second_tail == 'Large Specialized Complex Care Medical Center':
                        for ii, item_n in enumerate(results_n_events):
                            # append event type to data list
                            item_n.append(event_type_nm)
                            LSCCMC_events_df.loc[LSCCMC_events_counter] = item_n
                            LSCCMC_events_counter += 1

                    elif path_second_tail == 'Complex Care Medical Center':
                        for ii, item_n in enumerate(results_n_events):
                            # append event type to data list
                            item_n.append(event_type_nm)
                            CCMC_events_df.loc[CCMC_events_counter] = item_n
                            CCMC_events_counter += 1
                    elif path_second_tail == 'Community Medical Center':
                        for ii, item_n in enumerate(results_n_events):
                            # append event type to data list
                            item_n.append(event_type_nm)
                            COMM_events_df.loc[COMM_events_counter] = item_n
                            COMM_events_counter += 1
                    #UL003
                    elif path_second_tail == 'Critical Access & Small Community':
                        for ii, item_n in enumerate(results_n_events):
                            # append event type to data list
                            item_n.append(event_type_nm)
                            CASC_events_df.loc[CASC_events_counter] = item_n
                            CASC_events_counter += 1


                    # get name of event type
                    denom_type_nm = return_denom_type(wrangled_hyperlinks_dict[path_key][1], path_key[1])

                    print('event_type: ', denom_type_nm)
                    if path_second_tail == 'Comprehensive Academic Medical Center':
                        for ii, item_d in enumerate(results_d_events):
                            # append event type to data list
                            item_d.append(denom_type_nm)
                            CAMC_denom_df.loc[CAMC_denom_counter] = item_d
                            CAMC_denom_counter += 1
                    elif path_second_tail == 'Large Specialized Complex Care Medical Center':
                        for ii, item_d in enumerate(results_d_events):
                            # append event type to data list
                            item_d.append(denom_type_nm)
                            LSCCMC_denom_df.loc[LSCCMC_denom_counter] = item_d
                            LSCCMC_denom_counter += 1

                    elif path_second_tail == 'Complex Care Medical Center':
                        for ii, item_d in enumerate(results_d_events):
                            # append event type to data list
                            item_d.append(denom_type_nm)
                            CCMC_denom_df.loc[CCMC_denom_counter] = item_d
                            CCMC_denom_counter += 1
                    elif path_second_tail == 'Community Medical Center':
                        for ii, item_d in enumerate(results_d_events):
                            # append event type to data list
                            item_d.append(denom_type_nm)
                            COMM_denom_df.loc[COMM_denom_counter] = item_d
                            COMM_denom_counter += 1
                    #UL003
                    elif path_second_tail == 'Critical Access & Small Community':
                        for ii, item_d in enumerate(results_d_events):
                            # append event type to data list
                            item_d.append(denom_type_nm)
                            CASC_denom_df.loc[CASC_denom_counter] = item_d
                            CASC_denom_counter += 1


    CAMC_df_merged = pd.merge(CAMC_df, CAMC_events_df, how='left', on=['Hospital', 'Domain', 'Measure'])
    LSCCMC_df_merged = pd.merge(LSCCMC_df, LSCCMC_events_df, how='left', on=['Hospital', 'Domain', 'Measure'])
    CCMC_df_merged = pd.merge(CCMC_df, CCMC_events_df, how='left', on=['Hospital', 'Domain', 'Measure'])
    COMM_df_merged = pd.merge(COMM_df, COMM_events_df, how='left', on=['Hospital', 'Domain', 'Measure'])
    #UL003
    CASC_df_merged = pd.merge(CASC_df, CASC_events_df, how='left', on=['Hospital', 'Domain', 'Measure'])

    CAMC_df_merged2 = pd.merge(CAMC_df_merged, CAMC_denom_df, how='left', on=['Hospital', 'Domain', 'Measure'])
    LSCCMC_df_merged2 = pd.merge(LSCCMC_df_merged, LSCCMC_denom_df, how='left', on=['Hospital', 'Domain', 'Measure'])
    CCMC_df_merged2 = pd.merge(CCMC_df_merged, CCMC_denom_df, how='left', on=['Hospital', 'Domain', 'Measure'])
    COMM_df_merged2 = pd.merge(COMM_df_merged, COMM_denom_df, how='left', on=['Hospital', 'Domain', 'Measure'])
    #UL003
    CASC_df_merged2 = pd.merge(CASC_df_merged, CASC_denom_df, how='left', on=['Hospital', 'Domain', 'Measure'])


    #UL003  ADE & % EARLY TRANSFERS need to now calculate true metric value by num/denominator
    #isolate everything other than ADE & % EARLY TRANSFERS
    CASC_df_merged3 = CASC_df_merged2[~CASC_df_merged2['Measure'].isin(['ADE_NUM','ADE_DENOM','EARLY_TRANS_DENOM','EARLY_TRANS_NUM'])]

    #isolate numerator per hospital and metric
    CASC_df_merged2_num = CASC_df_merged2[CASC_df_merged2['Measure'].isin(['ADE_NUM','EARLY_TRANS_NUM'])][['Hospital','Domain','N Events','event_type_nm']]

    #create placeholder measure column
    CASC_df_merged2_num['Measure'] = None

    #Conditionally update the Measure column with the correct measure name.['Adverse Drug Events Rate','% Early Transfers Out']
    CASC_df_merged2_num['Measure'] = np.where(CASC_df_merged2_num['Domain']=='Safety', 'Adverse Drug Events Rate', '% Early Transfers Out')

    #isolate denominator per hospital and metric
    #[['Hospital','Domain','N Events','event_type_nm']]
    CASC_df_merged2_denom = CASC_df_merged2[CASC_df_merged2['Measure'].isin(['ADE_DENOM','EARLY_TRANS_DENOM'])][['Hospital','Domain','denominator','denominator_event_type_nm']]

    #same thing.  for the denominator df, create the measure column.
    CASC_df_merged2_denom['Measure'] = None
    #Conditionally update the Measure column with the correct measure name.['Adverse Drug Events Rate','% Early Transfers Out']
    CASC_df_merged2_denom['Measure'] = np.where(CASC_df_merged2_denom['Domain']=='Safety', 'Adverse Drug Events Rate', '% Early Transfers Out')

    #join the numerator and denominator dataframes with inner join.
    CASC_ADE_EARLY = pd.merge(CASC_df_merged2_num, CASC_df_merged2_denom, how='inner', on=['Hospital', 'Domain', 'Measure'])

    #some hospitals have 'missing values.'  Cannot do math on a string value so we need to isolate these rows.
    CASC_ADE_EARLY_missing = CASC_ADE_EARLY[(CASC_ADE_EARLY['N Events'] == 'Missing') | (CASC_ADE_EARLY['denominator'] == 'Missing')]

    #Do the opposite.  Isolate the rows that are actual numbers we can do math with.
    CASC_ADE_EARLY_not_missing = CASC_ADE_EARLY[(CASC_ADE_EARLY['N Events'] != 'Missing') & (CASC_ADE_EARLY['denominator'] != 'Missing')]

    #rows with values, now calculate the o/e ratio.
    CASC_ADE_EARLY_not_missing['Metric Value'] = ((CASC_ADE_EARLY_not_missing['N Events'].astype('float')/CASC_ADE_EARLY_not_missing['denominator'].astype('float')) * 100).astype('str')

    #sort the columns to match other dataframes.
    CASC_ADE_EARLY_not_missing = CASC_ADE_EARLY_not_missing[['Hospital', 'Domain', 'Measure', 'Metric Value', 'N Events','event_type_nm', 'denominator', 'denominator_event_type_nm']]

    #the rows with missing numerators, we still need a Metric Value column.  Make this 'Missing' as well.
    CASC_ADE_EARLY_missing['Metric Value'] = 'Missing'
    #sort the columns
    CASC_ADE_EARLY_missing = CASC_ADE_EARLY_missing[['Hospital', 'Domain', 'Measure', 'Metric Value', 'N Events','event_type_nm', 'denominator', 'denominator_event_type_nm']]

    #put all the dataframes all back together
    casc_frames = [CASC_df_merged3,CASC_ADE_EARLY_not_missing,CASC_ADE_EARLY_missing]
    CASC_df_merged4 = pd.concat(casc_frames,sort=True)

    #UL003
    #frames = [CAMC_df_merged2, LSCCMC_df_merged2, CCMC_df_merged2,COMM_df_merged2]
    frames = [CAMC_df_merged2, LSCCMC_df_merged2, CCMC_df_merged2,COMM_df_merged2,CASC_df_merged4]


    result_df = pd.concat(frames,sort=True)
    result_df.drop_duplicates(inplace=True)
    final_result_path = os.path.join(wd_dest, 'custom_time_period_data.csv')
    result_df.to_csv(final_result_path)
    
    return (result_df)

################################################################################################
################################################################################################
################################################################################################

# Functions to parse the calculator files and insert the data into a staging table called
# vizient_qa.datadump


def parse_calc_excel_for_datadump(period_lbl, calc_nm):
    path_obj = input('Enter file path of the calculator you want to parse and dump.')
    file_obj = input('Enter file name.')
    # Create filename path in order to open the excel file
    file_loc = os.path.join(os.path.abspath(path_obj), file_obj)
    # save the excel workbook object in a variable.  Read only data and not formulas.
    wb = openpyxl.load_workbook(file_loc, data_only=True)
    # save the excel worksheet object in a variable
    ws = wb['Calculator']
    # Create empty dataframe to store values in.
    dump_df = pd.DataFrame(columns=['dd_desc', 'dd_x', 'dd_y', 'dd_value'])
    # set counter to create dataframe row indexes
    counter = 0
    # create description column value so the sql query can parse correctly
    description = period_lbl + '|' + calc_nm
    # iterate over worksheet rows and store the values and coordinates into the dataframe
    for row in ws.iter_rows():
        for cell in row:
            cellContent = str(cell.value)
            cellRow = str(cell.row)
            cellColumn = str(cell.column)
            if cellContent is not None and cellContent != 'None':
                dump_df.loc[counter] = [description, cellRow, cellColumn, cellContent]
                counter += 1
    # return the df
    print('done.')
    return (dump_df)

################################################################################################

# takes the resulting dataframe from parse_calc_excel_for_datadump function and inserts those values
# into vizient_qa.datadump
def insert_datadump_df(df):
    # connect to the NM_Analytics database
    conn = pyodbc.connect('Driver={SQL Server};'
                          'Server=edw00pd05wva.corp.nm.org\EDWIDS1;'
                          'Database=NM_Analytics_Prototype;'
                          'Trusted_Connection=yes;')

    cursor = conn.cursor()
    # iterate over df rows and insert into NM_Analytics_Prototype.vizient_qa.datadump
    for index, row in df.iterrows():
        cursor.execute(
            "INSERT INTO NM_Analytics_Prototype.vizient_qa.datadump([dd_desc],[dd_x],[dd_y],[dd_value]) values (?,?,?,?)",
            row['dd_desc'], row['dd_x'], row['dd_y'], row['dd_value'])
        conn.commit()

    cursor.close()
    conn.close()
    print('done inserting rows.')


################################################################################################
#####################################################################################################################
#####################################################################################################################
#####################################################################################################################
# Function to validate the downloaded Vizient Q&A files.
# Dependencies:
#   This file requires an excel file called validation_file/vizient_qa_validation.xlsx
#   Inside this file, there should be the following two columns called 'validation_item' and 'value'.
#   Format should be similar to as below:
#validation_item	                    value
#CAMC/LSCCMC Risk Adjustment Model	    2018 Risk Model (AMC)
#CCMC Risk Adjustment Model	            2018 Risk Model (Community)
#CAMC/LSCCMC AHRQ Version	            8.0 (CMS Safety)
#CCMC AHRQ Version	                    8.0 (CMS Safety)
#Standard Time Period	                Aug 2018 to Apr 2019
#THK Time Period	                    Jul 2018 to Mar 2019
#READM/EDAC Time Period	2019            Quarter 2,2019 Quarter 1,2018 Quarter 4,2018 Quarter 3
#Focus Hospital	                        NORTHWESTERN_MEMORIAL 140281

'''
def validate_downloaded_files():
    # Set the path of the folder to parse and validate
    wd = input('Enter the path of the folder you want to validate.')
    wd = os.path.abspath(wd)
    # Set the path of the folder structure where we validation file to go
    wd_dest = input('Enter the path of the folder you want the validation file to go in.')
    wd_dest = os.path.abspath(wd_dest)
    # Open the validation key file.
    validation_file_name = 'validation_file/vizient_qa_validation.xlsx'
    validation_file_loc = os.path.join(os.path.abspath(wd_dest), validation_file_name)

    # Read in the validation excel file
    validation_xlsx_file = pd.DataFrame(pd.read_excel(validation_file_loc, sheet_name='Values'))

    # Store the validation values into variables
    camc_lsccmc_risk_model = \
    validation_xlsx_file[validation_xlsx_file['validation_item'] == 'CAMC/LSCCMC Risk Adjustment Model']['value']
    ccmc_risk_model = validation_xlsx_file[validation_xlsx_file['validation_item'] == 'CCMC Risk Adjustment Model'][
        'value']
    camc_lsccmc_ahrq_model = \
    validation_xlsx_file[validation_xlsx_file['validation_item'] == 'CAMC/LSCCMC AHRQ Version']['value']
    ccmc_ahrq_model = validation_xlsx_file[validation_xlsx_file['validation_item'] == 'CCMC AHRQ Version']['value']
    standard_time_period = validation_xlsx_file[validation_xlsx_file['validation_item'] == 'Standard Time Period'][
        'value']
    thk_time_period = validation_xlsx_file[validation_xlsx_file['validation_item'] == 'THK Time Period']['value']
    readm_edac_time_period = validation_xlsx_file[validation_xlsx_file['validation_item'] == 'READM/EDAC Time Period'][
        'value']
    focus_hospital = validation_xlsx_file[validation_xlsx_file['validation_item'] == 'Focus Hospital']['value']
    compare_hospital = validation_xlsx_file[validation_xlsx_file['validation_item'] == 'Compare Hospitals']['value']
    print('compare hospital:', compare_hospital.values[0])

    # Set counters which will act as proxy indexes in order to append rows to pandas dataframes
    validation_counter = 0
    # Create empty dataframes to store final validation result values.
    validation_results_df = pd.DataFrame(
        columns=['Cohort', 'Measure', 'File', 'File Opened', 'Risk Model', 'AHRQ Model', 'Time Period',
                 'Focus Hospital', 'Compare Hospitals', 'Group By Month', 'Validation Result'])

    # Do not need to parse ED-2B or ED OP 18B
    do_not_parse = ['ED_2B', 'ED_OP_18B']
    # iterate through folder structure.
    for path, dirs, files in os.walk(wd):
        # if files list is not empty, open it up
        if len(files) > 0:
            # parse folder name to get measure name
            path_tail = str((os.path.basename(path)))
            # parse parent folder to get cohort name
            path_head = os.path.dirname(path)
            path_second_tail = str(os.path.basename(path_head))
            # use cohort and measure composite key to access dictionary values
            path_key = (path_second_tail, path_tail)
            # set default variables.  If nothing is found, then PASS.  If an issue is found then set to FAIL.
            file_opened_result = 'PASS'
            risk_model_result = 'PASS'
            ahrq_result = 'PASS'
            time_period_result = 'PASS'
            focus_hosp_result = 'PASS'
            validation_result = 'PASS'
            compare_hosp_result = 'PASS'
            group_by_month_result = 'N/A'
            # collect file name, cohort name and measure name into variables for reporting in the results file.
            file_name = files[0]
            cohort_name = path_key[0]
            measure_name = path_key[1]
            print('path_key: ',path_key)
            # skip ED-1B and ED-OP-18B
            if path_key[1] in do_not_parse:
                # print('passed!')
                pass
            else:

                file_loc = os.path.join(os.path.abspath(path), files[0])

                try:
                    # try to open the excel workbook
                    data_wb = openpyxl.load_workbook(file_loc)
                except:
                    # if unable to open the excel workbook, it is likely corrupted.  Automatic validation fail.
                    # Set all other validation variables to 'N/A' and set final validation_result to 'FAIL'.
                    # skip subsequent operations in the loop
                    validation_result = 'FAIL'
                    file_opened_result = 'FAIL'
                    risk_model_result = 'N/A'
                    ahrq_result = 'N/A'
                    focus_hosp_result = 'N/A'
                    time_period_result = 'N/A'
                    compare_hosp_result = 'N/A'
                    group_by_month_result = 'N/A'
                    # create the row list object
                    result_list = [cohort_name, measure_name, file_name, file_opened_result, risk_model_result,
                                   ahrq_result, time_period_result, focus_hosp_result, compare_hosp_result,
                                   group_by_month_result, validation_result]
                    # append the row to the results dataframe
                    validation_results_df.loc[validation_counter] = result_list
                    validation_counter += 1
                    print('Issue opening the excel file.')
                    continue

                # if excel file can be opened, continue onto validating the parameter settings.
                if file_opened_result == 'PASS':
                    data_wb_sheetnames = data_wb.sheetnames
                    validation_ws = data_wb[data_wb_sheetnames[0]]
                    first_validation_col = find_first_ws_col(validation_ws)
                    header_row = find_ws_header_row(validation_ws, first_validation_col)
                    print('first col & header row: ',first_validation_col,', ',header_row)
                    xlsx_file = pd.DataFrame(
                        pd.read_excel(file_loc, sheet_name=data_wb_sheetnames[0], usecols=[first_validation_col - 1]).iloc[
                        :header_row])
                    print(xlsx_file.head())
                    if path_key[0] in ['Large Specialized Complex Care Medical Center',
                                       'Comprehensive Academic Medical Center']:
                        # print('AMC model')
                        #print('test...')
                        #print(xlsx_file[xlsx_file.columns[0]])
                        risk_model_check = \
                        xlsx_file[xlsx_file[xlsx_file.columns[0]].str.contains('Risk Adjustment Model') == True].values[
                            0][0].split(":", 1)[1].strip() == camc_lsccmc_risk_model.values[0]
                        if risk_model_check == True:
                            risk_model_result = 'PASS'
                        else:
                            risk_model_result = 'FAIL'

                        ahrq_model_check = \
                        xlsx_file[xlsx_file[xlsx_file.columns[0]].str.contains('AHRQ Version') == True].values[0][
                            0].split(": :", 1)[1].strip() == camc_lsccmc_ahrq_model.values[0]

                        if ahrq_model_check == True:
                            ahrq_result = 'PASS'
                        else:
                            ahrq_result = 'FAIL'

                        focus_hosp_check = \
                        xlsx_file[xlsx_file[xlsx_file.columns[0]].str.contains('Focus Hospital') == True].values[0][
                            0].split(":", 1)[1].strip() == focus_hospital.values[0]

                        if focus_hosp_check == True:
                            focus_hosp_result = 'PASS'
                        else:
                            focus_hosp_result = 'FAIL'

                        compare_hosp_check = compare_hospital.values[0] in xlsx_file[
                            xlsx_file[xlsx_file.columns[0]].str.contains('Compare Hospitals') == True].values[0][
                            0].split(":", 1)[1].strip()

                        if compare_hosp_check == True:
                            compare_hosp_result = 'PASS'
                        else:
                            compare_hosp_result = 'FAIL'

                        if (path_key[1].split('_')[0] in ['EDAC', 'READM']) and (
                                path_key[1] not in ['READM_ARTHRO', 'READM_CHOL', 'READM_COLON', 'READM_URI']):

                            time_period_check = \
                            xlsx_file[xlsx_file[xlsx_file.columns[0]].str.contains('Time Period') == True].values[0][0].split(":", 1)[1].strip() == readm_edac_time_period.values[0]

                            if time_period_check == True:
                                time_period_result = 'PASS'
                            else:
                                time_period_result = 'FAIL'

                            # print('test test: ',xlsx_file[xlsx_file[xlsx_file.columns[0]].str.contains('Encounter Month') == True].values[0][0])
                            # check whether or not the report was grouped by month by checking for existence of 'Encounter Month' column
                            try:
                                group_by_month_check = xlsx_file[xlsx_file[xlsx_file.columns[0]].str.contains(
                                    'Encounter Month') == True].values[0][0].strip() == 'Encounter Month'

                                if group_by_month_check == True:
                                    group_by_month_result = 'PASS'
                                else:
                                    group_by_month_result = 'FAIL'
                            except:
                                group_by_month_result = 'FAIL'

                        elif path_key[1] in ['THK']:

                            time_period_check = \
                            xlsx_file[xlsx_file[xlsx_file.columns[0]].str.contains('Time Period') == True].values[0][
                                0].split(":", 1)[1].strip() == thk_time_period.values[0]
                            if time_period_check == True:
                                time_period_result = 'PASS'
                            else:
                                time_period_result = 'FAIL'
                        else:
                            time_period_check = \
                            xlsx_file[xlsx_file[xlsx_file.columns[0]].str.contains('Time Period') == True].values[0][
                                0].split(":", 1)[1].strip() == standard_time_period.values[0]
                            if time_period_check == True:
                                time_period_result = 'PASS'
                            else:
                                time_period_result = 'FAIL'

                        if compare_hosp_result == 'FAIL' or group_by_month_result == 'FAIL' or file_opened_result == 'FAIL' or risk_model_result == 'FAIL' or ahrq_result == 'FAIL' or time_period_result == 'FAIL' or focus_hosp_result == 'FAIL':
                            validation_result = 'FAIL'
                        else:
                            pass

                        result_list = [cohort_name, measure_name, file_name, file_opened_result, risk_model_result,
                                       ahrq_result, time_period_result, focus_hosp_result, compare_hosp_result,
                                       group_by_month_result, validation_result]
                        validation_results_df.loc[validation_counter] = result_list
                        validation_counter += 1
                        # print('test: ',compare_hosp_result)

                    elif path_key[0] in ['Complex Care Medical Center','Community Medical Center','Critical Access & Small Community']:
                        # print('Comm Model')

                        risk_model_check = \
                        xlsx_file[xlsx_file[xlsx_file.columns[0]].str.contains('Risk Adjustment Model') == True].values[
                            0][0].split(":", 1)[1].strip() == ccmc_risk_model.values[0]

                        if risk_model_check == True:
                            risk_model_result = 'PASS'
                        else:
                            risk_model_result = 'FAIL'

                        ahrq_model_check = \
                        xlsx_file[xlsx_file[xlsx_file.columns[0]].str.contains('AHRQ Version') == True].values[0][
                            0].split(": :", 1)[1].strip() == ccmc_ahrq_model.values[0]

                        if ahrq_model_check == True:
                            ahrq_result = 'PASS'
                        else:
                            ahrq_result = 'FAIL'

                        focus_hosp_check = \
                        xlsx_file[xlsx_file[xlsx_file.columns[0]].str.contains('Focus Hospital') == True].values[0][
                            0].split(":", 1)[1].strip() == focus_hospital.values[0]

                        if focus_hosp_check == True:
                            focus_hosp_result = 'PASS'
                        else:
                            focus_hosp_result = 'FAIL'

                        compare_hosp_check = compare_hospital.values[0] in xlsx_file[
                            xlsx_file[xlsx_file.columns[0]].str.contains('Compare Hospitals') == True].values[0][
                            0].split(":", 1)[1].strip()

                        if compare_hosp_check == True:
                            compare_hosp_result = 'PASS'
                        else:
                            compare_hosp_result = 'FAIL'

                        if (path_key[1].split('_')[0] in ['EDAC', 'READM','Readm']) and (
                                path_key[1] not in ['READM_ARTHRO', 'READM_CHOL', 'READM_COLON', 'READM_URI']):

                            time_period_check = \
                            xlsx_file[xlsx_file[xlsx_file.columns[0]].str.contains('Time Period') == True].values[0][
                                0].split(":", 1)[1].strip() == readm_edac_time_period.values[0]

                            if time_period_check == True:
                                time_period_result = 'PASS'
                            else:
                                time_period_result = 'FAIL'

                            # print('test test: ',xlsx_file[xlsx_file[xlsx_file.columns[0]].str.contains('Encounter Month') == True].values[0][0])
                            # check whether or not the report was grouped by month by checking for existence of 'Encounter Month' column
                            try:
                                group_by_month_check = xlsx_file[xlsx_file[xlsx_file.columns[0]].str.contains(
                                    'Encounter Month') == True].values[0][0].strip() == 'Encounter Month'

                                if group_by_month_check == True:
                                    group_by_month_result = 'PASS'
                                else:
                                    group_by_month_result = 'FAIL'
                            except:
                                group_by_month_result = 'FAIL'

                        elif path_key[1] in ['THK']:

                            time_period_check = \
                            xlsx_file[xlsx_file[xlsx_file.columns[0]].str.contains('Time Period') == True].values[0][
                                0].split(":", 1)[1].strip() == thk_time_period.values[0]
                            if time_period_check == True:
                                time_period_result = 'PASS'
                            else:
                                time_period_result = 'FAIL'
                        else:
                            time_period_check = \
                            xlsx_file[xlsx_file[xlsx_file.columns[0]].str.contains('Time Period') == True].values[0][
                                0].split(":", 1)[1].strip() == standard_time_period.values[0]
                            if time_period_check == True:
                                time_period_result = 'PASS'
                            else:
                                time_period_result = 'FAIL'

                        if compare_hosp_result == 'FAIL' or group_by_month_result == 'FAIL' or file_opened_result == 'FAIL' or risk_model_result == 'FAIL' or ahrq_result == 'FAIL' or time_period_result == 'FAIL' or focus_hosp_result == 'FAIL':
                            validation_result = 'FAIL'
                        else:
                            pass
                        result_list = [cohort_name, measure_name, file_name, file_opened_result, risk_model_result,
                                       ahrq_result, time_period_result, focus_hosp_result, compare_hosp_result,
                                       group_by_month_result, validation_result]
                        validation_results_df.loc[validation_counter] = result_list
                        validation_counter += 1

    validation_results_df.to_csv(os.path.join(os.path.abspath(wd_dest), "validation_results.csv"), index=False)
    print('Done.')
'''

##UL004  Updated validation script to check for FY21 Covid restrictions
'''
def validate_downloaded_files():
    pd.set_option('display.max_colwidth', None)
    # Set the path of the folder to parse and validate
    wd = input('Enter the path of the folder you want to validate.')
    wd = os.path.abspath(wd)
    # Set the path of the folder structure where we validation file to go
    wd_dest = input('Enter the path of the folder you want the validation file to go in.')
    wd_dest = os.path.abspath(wd_dest)
    # Open the validation key file.
    validation_file_name = 'validation_file/vizient_qa_validation.xlsx'
    validation_file_loc = os.path.join(os.path.abspath(wd_dest), validation_file_name)

    # Read in the validation excel file
    validation_xlsx_file = pd.DataFrame(pd.read_excel(validation_file_loc, sheet_name='Values'))

    # Store the validation values into variables
    camc_lsccmc_risk_model = \
    validation_xlsx_file[validation_xlsx_file['validation_item'] == 'CAMC/LSCCMC Risk Adjustment Model']['value']
    ccmc_risk_model = validation_xlsx_file[validation_xlsx_file['validation_item'] == 'CCMC Risk Adjustment Model'][
        'value']
    camc_lsccmc_ahrq_model = \
    validation_xlsx_file[validation_xlsx_file['validation_item'] == 'CAMC/LSCCMC AHRQ Version']['value']
    ccmc_ahrq_model = validation_xlsx_file[validation_xlsx_file['validation_item'] == 'CCMC AHRQ Version']['value']
    standard_time_period = validation_xlsx_file[validation_xlsx_file['validation_item'] == 'Standard Time Period'][
        'value']
    thk_time_period = validation_xlsx_file[validation_xlsx_file['validation_item'] == 'THK Time Period']['value']
    readm_edac_time_period = validation_xlsx_file[validation_xlsx_file['validation_item'] == 'READM/EDAC Time Period'][
        'value']
    focus_hospital = validation_xlsx_file[validation_xlsx_file['validation_item'] == 'Focus Hospital']['value']
    compare_hospital = validation_xlsx_file[validation_xlsx_file['validation_item'] == 'Compare Hospitals']['value']
    print('compare hospital:', compare_hospital.values[0])

    # Set counters which will act as proxy indexes in order to append rows to pandas dataframes
    validation_counter = 0
    # Create empty dataframes to store final validation result values.
    validation_results_df = pd.DataFrame(
        columns=['Cohort', 'Measure', 'File', 'File Opened', 'Risk Model', 'AHRQ Model', 'Time Period',
                 'Focus Hospital', 'Compare Hospitals', 'Group By Month','No Records In Study',\
                 'Group By Disch Month Custom','Custom Group Bys',\
                 'Discharge Month Restriction',\
                 'COVID Grouper','COVID Restriction','Validation Result'])

    # Do not need to parse ED-2B or ED OP 18B
    do_not_parse = ['ED_2B', 'ED_OP_18B']
    # iterate through folder structure.
    for path, dirs, files in os.walk(wd):
        # if files list is not empty, open it up
        if len(files) > 0:
            # parse folder name to get measure name
            path_tail = str((os.path.basename(path)))
            # parse parent folder to get cohort name
            path_head = os.path.dirname(path)
            path_second_tail = str(os.path.basename(path_head))
            # use cohort and measure composite key to access dictionary values
            path_key = (path_second_tail, path_tail)
            # set default variables.  If nothing is found, then PASS.  If an issue is found then set to FAIL.
            file_opened_result = 'PASS'
            risk_model_result = 'PASS'
            ahrq_result = 'PASS'
            time_period_result = 'PASS'
            focus_hosp_result = 'PASS'
            validation_result = 'PASS'
            compare_hosp_result = 'PASS'
            group_by_month_result = 'N/A'
            #UL004  Vizient added new filters and group bys to the report templates to handle COVID.
            #       we need to check whether those filters were removed or not.
            #       If all the covid filters are applied, then we just won't get data and the phrase No Reocords In Study Population
            #       will appear instead.  Check for this phrase.
            no_records_in_study_pop_result = 'PASS'
            
            #UL004  Look for custom group bys.  This means Vizient changed some stuff related to covid.
            #       We only want reports grouped by Hospital/Hospital System OR Discharge Month 1st Admit
            #       If the report is grouped by Discharge Month (Custom), then we didn't remove that parameter and we FAIL.
            group_by_discharge_month_custom_result = 'PASS'
            
            #UL004
            # Also need to check for any custom group bys.  If there are any, then fail.
            custom_group_by_result = 'PASS'
            
            #UL004 Also need to look for discharge month restriction and make sure we removed it.
            discharge_month_restriction_result = 'PASS'
            
            #Hospital / Hospital System
            #Discharge Month 1st Admit
            # collect file name, cohort name and measure name into variables for reporting in the results file.
            file_name = files[0]
            cohort_name = path_key[0]
            measure_name = path_key[1]
            print('path_key: ',path_key)
            # skip ED-1B and ED-OP-18B
            if path_key[1] in do_not_parse:
                # print('passed!')
                pass
            else:

                file_loc = os.path.join(os.path.abspath(path), files[0])

                try:
                    # try to open the excel workbook
                    data_wb = openpyxl.load_workbook(file_loc)
                except:
                    # if unable to open the excel workbook, it is likely corrupted.  Automatic validation fail.
                    # Set all other validation variables to 'N/A' and set final validation_result to 'FAIL'.
                    # skip subsequent operations in the loop
                    validation_result = 'FAIL'
                    file_opened_result = 'FAIL'
                    risk_model_result = 'N/A'
                    ahrq_result = 'N/A'
                    focus_hosp_result = 'N/A'
                    time_period_result = 'N/A'
                    compare_hosp_result = 'N/A'
                    group_by_month_result = 'N/A'
                    #UL004
                    no_records_in_study_pop_result = 'N/A'
                    group_by_discharge_month_custom_result = 'N/A'
                    custom_group_by_result = 'N/A'
                    discharge_month_restriction_result = 'N/A'
                    covid_grouper_result = 'N/A'
                    covid_restriction_result = 'N/A'
                    # create the row list object
                    result_list = [cohort_name, measure_name, file_name, file_opened_result, risk_model_result,
                                   ahrq_result, time_period_result, focus_hosp_result, compare_hosp_result,
                                   group_by_month_result,no_records_in_study_pop_result,\
                                   group_by_discharge_month_custom_result,custom_group_by_result,\
                                   discharge_month_restriction_result,\
                                   covid_grouper_result,covid_restriction_result,validation_result]
                    # append the row to the results dataframe
                    validation_results_df.loc[validation_counter] = result_list
                    validation_counter += 1
                    print('Issue opening the excel file.')
                    continue

                # if excel file can be opened, continue onto validating the parameter settings.
                if file_opened_result == 'PASS':
                    data_wb_sheetnames = data_wb.sheetnames
                    validation_ws = data_wb[data_wb_sheetnames[0]]
                    first_validation_col = find_first_ws_col(validation_ws)
                    header_row = find_ws_header_row(validation_ws, first_validation_col)
                    #print('first col & header row: ',first_validation_col,', ',header_row)
                    xlsx_file = pd.DataFrame(
                        pd.read_excel(file_loc, sheet_name=data_wb_sheetnames[0], usecols=[first_validation_col - 1]).iloc[
                        :header_row])
                    
                    #print(xlsx_file.head(n=50))
                    if path_key[0] in ['Large Specialized Complex Care Medical Center',
                                       'Comprehensive Academic Medical Center']:
                        #UL004
                        #look for the phrase 'No Records in Study Population.'  This means the whole thing failed.
                        #if present, then FAIL
                        if xlsx_file[xlsx_file[xlsx_file.columns[0]].str.upper().str.replace(' ','').str.replace('.','').str.contains('NORECORDSINSTUDYPOPULATION') == True].empty == False:
                            no_records_in_study_pop_result = 'FAIL'
                        else:
                            no_records_in_study_pop_result = 'PASS'
                            
                            
                        #UL004
                        #look for the phrase 'By Discharge Month (Custom) or BY AHRQ SAFETY / DISCHARGE MONTH.'  This means we failed to change the 
                        #group by back to default.
                        #print(xlsx_file[xlsx_file[xlsx_file.columns[0]].str.upper().str.replace(' ','').str.replace('.','').str.contains('COVID') == True])
                        #print(xlsx_file[[xlsx_file.columns[0]]].head())
                        #print(xlsx_file[(xlsx_file[xlsx_file.columns[0]].str.upper().str.replace(' ','').str.replace('(','').str.contains('BYDISCHARGEMONTHCUSTOM') == True) | (xlsx_file[xlsx_file.columns[0]].str.upper().str.replace(' ','').str.replace('(','').str.contains('BYAHRQSAFETY/DISCHARGEMONTHCUSTOM') == True)])
                        if xlsx_file[(xlsx_file[xlsx_file.columns[0]].str.upper().str.replace(' ','').str.replace('(','').str.contains('BYDISCHARGEMONTHCUSTOM') == True) | (xlsx_file[xlsx_file.columns[0]].str.upper().str.replace(' ','').str.replace('(','').str.contains('BYAHRQSAFETY/DISCHARGEMONTHCUSTOM') == True)].empty == False:
                            group_by_discharge_month_custom_result = 'FAIL'
                        else:
                            group_by_discharge_month_custom_result = 'PASS'
                            
                            
                            
                        #UL004  Look out for custom group bys.  This also means we failed to get all the covid groupings.
                        
                        if xlsx_file[xlsx_file[xlsx_file.columns[0]].str.upper().str.replace(' ','').str.contains('CUSTOMGROUPBYS') == True].empty == False:
                            custom_group_by_result = 'FAIL'
                        else:
                            custom_group_by_result = 'PASS'
                            
                        
                        if xlsx_file[xlsx_file[xlsx_file.columns[0]].str.upper().str.replace(' ','').str.contains('ADVANCEDRESTRICTIONS') == True].empty == False:
                            adv_rest_index = xlsx_file[xlsx_file[xlsx_file.columns[0]].str.upper().str.replace(' ','').str.contains('ADVANCEDRESTRICTIONS') == True].index.tolist()[0]
                            #print('got it...')
                            #print(adv_rest_index)
                            
                        #isolate before and after advanced restrictions to narrow down search   
                        just_adv_rest = xlsx_file[xlsx_file.columns[0]].iloc[adv_rest_index:]
                        
                        just_adv_rest = just_adv_rest.to_frame()
                        
                        before_adv_rest = xlsx_file[xlsx_file.columns[0]].iloc[:adv_rest_index]
                        before_adv_rest = before_adv_rest.to_frame()
                         
                        #print(xlsx_file[xlsx_file[xlsx_file.columns[0]].str.upper().str.replace(' ','').str.contains('ADVANCEDRESTRICTIONS') == True].index.tolist()[0])
                        #UL004 Look out for discharge month restrictions.  This was used to restriction pre/post covid periods.
                        #only search for 'discharge month' after the advanced restrictions.
                        if just_adv_rest[just_adv_rest[just_adv_rest.columns[0]].str.upper().str.replace(' ','').str.contains('DISCHARGEMONTH:') == True].empty == False:
                            discharge_month_restriction_result = 'FAIL'
                        else:
                            discharge_month_restriction_result = 'PASS'

                        
                        #UL004 Look out for COVID text before and after advanced restrictions.  Splitting this up to help narrow
                        #      the search later.
                        
                        if just_adv_rest[just_adv_rest[just_adv_rest.columns[0]].str.upper().str.replace(' ','').str.contains('COVID') == True].empty == False:
                            covid_restriction_result = 'FAIL'
                        else:
                            covid_restriction_result = 'PASS'
                            
                        if before_adv_rest[before_adv_rest[before_adv_rest.columns[0]].str.upper().str.replace(' ','').str.contains('COVID') == True].empty == False:
                            covid_grouper_result = 'FAIL'
                        else:
                            covid_grouper_result = 'PASS' 
                            
                        
                        
                        risk_model_check = \
                        xlsx_file[xlsx_file[xlsx_file.columns[0]].str.contains('Risk Adjustment Model') == True].values[
                            0][0].split(":", 1)[1].strip() == camc_lsccmc_risk_model.values[0]
                        if risk_model_check == True:
                            risk_model_result = 'PASS'
                        else:
                            risk_model_result = 'FAIL'

                        ahrq_model_check = \
                        xlsx_file[xlsx_file[xlsx_file.columns[0]].str.contains('AHRQ Version') == True].values[0][
                            0].split(": :", 1)[1].strip() == camc_lsccmc_ahrq_model.values[0]

                        if ahrq_model_check == True:
                            ahrq_result = 'PASS'
                        else:
                            ahrq_result = 'FAIL'

                        focus_hosp_check = \
                        xlsx_file[xlsx_file[xlsx_file.columns[0]].str.contains('Focus Hospital') == True].values[0][
                            0].split(":", 1)[1].strip() == focus_hospital.values[0]

                        if focus_hosp_check == True:
                            focus_hosp_result = 'PASS'
                        else:
                            focus_hosp_result = 'FAIL'

                        compare_hosp_check = compare_hospital.values[0] in xlsx_file[
                            xlsx_file[xlsx_file.columns[0]].str.contains('Compare Hospitals') == True].values[0][
                            0].split(":", 1)[1].strip()

                        if compare_hosp_check == True:
                            compare_hosp_result = 'PASS'
                        else:
                            compare_hosp_result = 'FAIL'

                        if (path_key[1].split('_')[0] in ['EDAC', 'READM']) and (
                                path_key[1] not in ['READM_ARTHRO', 'READM_CHOL', 'READM_COLON', 'READM_URI']):

                            time_period_check = \
                            xlsx_file[xlsx_file[xlsx_file.columns[0]].str.contains('Time Period') == True].values[0][0].split(":", 1)[1].strip() == readm_edac_time_period.values[0]

                            if time_period_check == True:
                                time_period_result = 'PASS'
                            else:
                                time_period_result = 'FAIL'

                            # print('test test: ',xlsx_file[xlsx_file[xlsx_file.columns[0]].str.contains('Encounter Month') == True].values[0][0])
                            # check whether or not the report was grouped by month by checking for existence of 'Encounter Month' column
                            try:
                                group_by_month_check = xlsx_file[xlsx_file[xlsx_file.columns[0]].str.contains(
                                    'Encounter Month') == True].values[0][0].strip() == 'Encounter Month'

                                if group_by_month_check == True:
                                    group_by_month_result = 'PASS'
                                else:
                                    group_by_month_result = 'FAIL'
                            except:
                                group_by_month_result = 'FAIL'

                        elif path_key[1] in ['THK']:

                            time_period_check = \
                            xlsx_file[xlsx_file[xlsx_file.columns[0]].str.contains('Time Period') == True].values[0][
                                0].split(":", 1)[1].strip() == thk_time_period.values[0]
                            if time_period_check == True:
                                time_period_result = 'PASS'
                            else:
                                time_period_result = 'FAIL'
                        else:
                            time_period_check = \
                            xlsx_file[xlsx_file[xlsx_file.columns[0]].str.contains('Time Period') == True].values[0][
                                0].split(":", 1)[1].strip() == standard_time_period.values[0]
                            if time_period_check == True:
                                time_period_result = 'PASS'
                            else:
                                time_period_result = 'FAIL'

                        if compare_hosp_result == 'FAIL' or group_by_month_result == 'FAIL' \
                        or file_opened_result == 'FAIL' or risk_model_result == 'FAIL' or ahrq_result == 'FAIL' \
                        or time_period_result == 'FAIL' or focus_hosp_result == 'FAIL' \
                        or no_records_in_study_pop_result == 'FAIL' or group_by_discharge_month_custom_result == 'FAIL'\
                        or custom_group_by_result == 'FAIL' or discharge_month_restriction_result == 'FAIL'\
                        or covid_grouper_result == 'FAIL' or covid_restriction_result == 'FAIL':
                            validation_result = 'FAIL'
                        else:
                            pass

                        result_list = [cohort_name, measure_name, file_name, file_opened_result, risk_model_result,
                                       ahrq_result, time_period_result, focus_hosp_result, compare_hosp_result,
                                       group_by_month_result,no_records_in_study_pop_result,\
                                       group_by_discharge_month_custom_result,custom_group_by_result,\
                                       discharge_month_restriction_result,\
                                       covid_grouper_result,covid_restriction_result,validation_result]
                        
                        
                        validation_results_df.loc[validation_counter] = result_list
                        validation_counter += 1
                        # print('test: ',compare_hosp_result)

                    elif path_key[0] in ['Complex Care Medical Center','Community Medical Center','Critical Access & Small Community']:
                        
                        
                        #print(xlsx_file[xlsx_file[xlsx_file.columns[0]].str.upper().str.replace(' ','').str.replace('.','').str.contains('COVID') == True])
                        #print(xlsx_file[[xlsx_file.columns[0]]].head())
                        #print(xlsx_file[[xlsx_file.columns[0]]])
                        # print('Comm Model')
                        #UL004
                        #look for the phrase 'No Records in Study Population.'  This means the whole thing failed.
                        #if present, then FAIL
                        if xlsx_file[xlsx_file[xlsx_file.columns[0]].str.upper().str.replace(' ','').str.replace('.','').str.contains('NORECORDSINSTUDYPOPULATION') == True].empty == False:
                            no_records_in_study_pop_result = 'FAIL'
                        else:
                            no_records_in_study_pop_result = 'PASS'
                            
                        #UL004
                        #look for the phrase 'By Discharge Month (Custom).'  This means we failed to change the 
                        #group by back to default.
                        
                        if xlsx_file[(xlsx_file[xlsx_file.columns[0]].str.upper().str.replace(' ','').str.replace('(','').str.contains('BYDISCHARGEMONTHCUSTOM') == True) | (xlsx_file[xlsx_file.columns[0]].str.upper().str.replace(' ','').str.replace('(','').str.contains('BYAHRQSAFETY/DISCHARGEMONTHCUSTOM') == True)].empty == False:
                            group_by_discharge_month_custom_result = 'FAIL'
                        else:
                            group_by_discharge_month_custom_result = 'PASS'
                            
                        #UL004  Look out for custom group bys.  This also means we failed to get all the covid groupings.
                        
                        if xlsx_file[xlsx_file[xlsx_file.columns[0]].str.upper().str.replace(' ','').str.contains('CUSTOMGROUPBYS') == True].empty == False:
                            custom_group_by_result = 'FAIL'
                        else:
                            custom_group_by_result = 'PASS'
                            
                        
                        
                        #UL004 Look out for discharge month restrictions.  This was used to restriction pre/post covid periods.
                        
                        if xlsx_file[xlsx_file[xlsx_file.columns[0]].str.upper().str.replace(' ','').str.contains('ADVANCEDRESTRICTIONS') == True].empty == False:
                            adv_rest_index = xlsx_file[xlsx_file[xlsx_file.columns[0]].str.upper().str.replace(' ','').str.contains('ADVANCEDRESTRICTIONS') == True].index.tolist()[0]
                            
                        #isolate before and after advanced restrictions to narrow down search    
                        just_adv_rest = xlsx_file[xlsx_file.columns[0]].iloc[adv_rest_index:]

                        just_adv_rest = just_adv_rest.to_frame() 
                        
                        before_adv_rest = xlsx_file[xlsx_file.columns[0]].iloc[:adv_rest_index]
                        before_adv_rest = before_adv_rest.to_frame()
                            
                        #only search for 'discharge month' after the advanced restrictions.
                        if just_adv_rest[just_adv_rest[just_adv_rest.columns[0]].str.upper().str.replace(' ','').str.contains('DISCHARGEMONTH:') == True].empty == False:
                            discharge_month_restriction_result = 'FAIL'
                        else:
                            discharge_month_restriction_result = 'PASS'
                            
                            
                        #UL004 Look out for COVID text before and after advanced restrictions.  Splitting this up to help narrow
                        #      the search later.
                        
                        if just_adv_rest[just_adv_rest[just_adv_rest.columns[0]].str.upper().str.replace(' ','').str.contains('COVID') == True].empty == False:
                            covid_restriction_result = 'FAIL'
                        else:
                            covid_restriction_result = 'PASS'
                            
                        if before_adv_rest[before_adv_rest[before_adv_rest.columns[0]].str.upper().str.replace(' ','').str.contains('COVID') == True].empty == False:
                            covid_grouper_result = 'FAIL'
                        else:
                            covid_grouper_result = 'PASS' 
                            
            

                        risk_model_check = \
                        xlsx_file[xlsx_file[xlsx_file.columns[0]].str.contains('Risk Adjustment Model') == True].values[
                            0][0].split(":", 1)[1].strip() == ccmc_risk_model.values[0]

                        if risk_model_check == True:
                            risk_model_result = 'PASS'
                        else:
                            risk_model_result = 'FAIL'

                        ahrq_model_check = \
                        xlsx_file[xlsx_file[xlsx_file.columns[0]].str.contains('AHRQ Version') == True].values[0][
                            0].split(": :", 1)[1].strip() == ccmc_ahrq_model.values[0]

                        if ahrq_model_check == True:
                            ahrq_result = 'PASS'
                        else:
                            ahrq_result = 'FAIL'

                        focus_hosp_check = \
                        xlsx_file[xlsx_file[xlsx_file.columns[0]].str.contains('Focus Hospital') == True].values[0][
                            0].split(":", 1)[1].strip() == focus_hospital.values[0]

                        if focus_hosp_check == True:
                            focus_hosp_result = 'PASS'
                        else:
                            focus_hosp_result = 'FAIL'

                        compare_hosp_check = compare_hospital.values[0] in xlsx_file[
                            xlsx_file[xlsx_file.columns[0]].str.contains('Compare Hospitals') == True].values[0][
                            0].split(":", 1)[1].strip()

                        if compare_hosp_check == True:
                            compare_hosp_result = 'PASS'
                        else:
                            compare_hosp_result = 'FAIL'

                        if (path_key[1].split('_')[0] in ['EDAC', 'READM','Readm']) and (
                                path_key[1] not in ['READM_ARTHRO', 'READM_CHOL', 'READM_COLON', 'READM_URI']):

                            time_period_check = \
                            xlsx_file[xlsx_file[xlsx_file.columns[0]].str.contains('Time Period') == True].values[0][
                                0].split(":", 1)[1].strip() == readm_edac_time_period.values[0]

                            if time_period_check == True:
                                time_period_result = 'PASS'
                            else:
                                time_period_result = 'FAIL'

                            # print('test test: ',xlsx_file[xlsx_file[xlsx_file.columns[0]].str.contains('Encounter Month') == True].values[0][0])
                            # check whether or not the report was grouped by month by checking for existence of 'Encounter Month' column
                            try:
                                group_by_month_check = xlsx_file[xlsx_file[xlsx_file.columns[0]].str.contains(
                                    'Encounter Month') == True].values[0][0].strip() == 'Encounter Month'

                                if group_by_month_check == True:
                                    group_by_month_result = 'PASS'
                                else:
                                    group_by_month_result = 'FAIL'
                            except:
                                group_by_month_result = 'FAIL'

                        elif path_key[1] in ['THK']:

                            time_period_check = \
                            xlsx_file[xlsx_file[xlsx_file.columns[0]].str.contains('Time Period') == True].values[0][
                                0].split(":", 1)[1].strip() == thk_time_period.values[0]
                            if time_period_check == True:
                                time_period_result = 'PASS'
                            else:
                                time_period_result = 'FAIL'
                        else:
                            time_period_check = \
                            xlsx_file[xlsx_file[xlsx_file.columns[0]].str.contains('Time Period') == True].values[0][
                                0].split(":", 1)[1].strip() == standard_time_period.values[0]
                            if time_period_check == True:
                                time_period_result = 'PASS'
                            else:
                                time_period_result = 'FAIL'

                        if compare_hosp_result == 'FAIL' or group_by_month_result == 'FAIL' \
                        or file_opened_result == 'FAIL' or risk_model_result == 'FAIL' or ahrq_result == 'FAIL' \
                        or time_period_result == 'FAIL' or focus_hosp_result == 'FAIL' \
                        or no_records_in_study_pop_result == 'FAIL' or group_by_discharge_month_custom_result == 'FAIL'\
                        or custom_group_by_result == 'FAIL' or discharge_month_restriction_result == 'FAIL'\
                        or covid_grouper_result == 'FAIL' or covid_restriction_result == 'FAIL':
                            validation_result = 'FAIL'
                        else:
                            pass
                        
                        result_list = [cohort_name, measure_name, file_name, file_opened_result, risk_model_result,
                                       ahrq_result, time_period_result, focus_hosp_result, compare_hosp_result,
                                       group_by_month_result,no_records_in_study_pop_result,\
                                       group_by_discharge_month_custom_result,custom_group_by_result,\
                                       discharge_month_restriction_result,\
                                       covid_grouper_result,covid_restriction_result,validation_result]
                        
                        
                        validation_results_df.loc[validation_counter] = result_list
                        validation_counter += 1

    validation_results_df.to_csv(os.path.join(os.path.abspath(wd_dest), "validation_results.csv"), index=False)
    print('Done.')
'''
##UL006 updated to handle u071 diagnosis exclusion check
'''
def validate_downloaded_files(remove_covid_pats = False):
    pd.set_option('display.max_colwidth', None)
    # Set the path of the folder to parse and validate
    wd = input('Enter the path of the folder you want to validate.')
    wd = os.path.abspath(wd)
    # Set the path of the folder structure where we validation file to go
    wd_dest = input('Enter the path of the folder you want the validation file to go in.')
    wd_dest = os.path.abspath(wd_dest)
    # Open the validation key file.
    validation_file_name = 'validation_file/vizient_qa_validation.xlsx'
    validation_file_loc = os.path.join(os.path.abspath(wd_dest), validation_file_name)

    # Read in the validation excel file
    #UL007
    validation_xlsx_file = pd.DataFrame(pd.read_excel(validation_file_loc, sheet_name='Values',engine='openpyxl'))
    
    # Store the validation values into variables
    camc_lsccmc_risk_model = \
    validation_xlsx_file[validation_xlsx_file['validation_item'] == 'CAMC/LSCCMC Risk Adjustment Model']['value']
    ccmc_risk_model = validation_xlsx_file[validation_xlsx_file['validation_item'] == 'CCMC Risk Adjustment Model'][
        'value']
    camc_lsccmc_ahrq_model = \
    validation_xlsx_file[validation_xlsx_file['validation_item'] == 'CAMC/LSCCMC AHRQ Version']['value']
    ccmc_ahrq_model = validation_xlsx_file[validation_xlsx_file['validation_item'] == 'CCMC AHRQ Version']['value']
    standard_time_period = validation_xlsx_file[validation_xlsx_file['validation_item'] == 'Standard Time Period'][
        'value']
    thk_time_period = validation_xlsx_file[validation_xlsx_file['validation_item'] == 'THK Time Period']['value']
    readm_edac_time_period = validation_xlsx_file[validation_xlsx_file['validation_item'] == 'READM/EDAC Time Period'][
        'value']
    focus_hospital = validation_xlsx_file[validation_xlsx_file['validation_item'] == 'Focus Hospital']['value']
    compare_hospital = validation_xlsx_file[validation_xlsx_file['validation_item'] == 'Compare Hospitals']['value']
    print('compare hospital:', compare_hospital.values[0])

    # Set counters which will act as proxy indexes in order to append rows to pandas dataframes
    validation_counter = 0
    # Create empty dataframes to store final validation result values.
    validation_results_df = pd.DataFrame(
        columns=['Cohort', 'Measure', 'File', 'File Opened', 'Risk Model', 'AHRQ Model', 'Time Period',
                 'Focus Hospital', 'Compare Hospitals', 'Group By Month','No Records In Study',\
                 'Group By Disch Month Custom','Custom Group Bys',\
                 'Discharge Month Restriction',\
                 'COVID Grouper','COVID Restriction','U071 Check','Validation Result'])

    # Do not need to parse ED-2B or ED OP 18B
    do_not_parse = ['ED_2B', 'ED_OP_18B']
    # iterate through folder structure.
    for path, dirs, files in os.walk(wd):
        # if files list is not empty, open it up
        if len(files) > 0:
            # parse folder name to get measure name
            path_tail = str((os.path.basename(path)))
            # parse parent folder to get cohort name
            path_head = os.path.dirname(path)
            path_second_tail = str(os.path.basename(path_head))
            # use cohort and measure composite key to access dictionary values
            path_key = (path_second_tail, path_tail)
            # set default variables.  If nothing is found, then PASS.  If an issue is found then set to FAIL.
            file_opened_result = 'PASS'
            risk_model_result = 'PASS'
            ahrq_result = 'PASS'
            time_period_result = 'PASS'
            focus_hosp_result = 'PASS'
            validation_result = 'PASS'
            compare_hosp_result = 'PASS'
            group_by_month_result = 'N/A'
            #UL004  Vizient added new filters and group bys to the report templates to handle COVID.
            #       we need to check whether those filters were removed or not.
            #       If all the covid filters are applied, then we just won't get data and the phrase No Reocords In Study Population
            #       will appear instead.  Check for this phrase.
            no_records_in_study_pop_result = 'PASS'
            
            #UL004  Look for custom group bys.  This means Vizient changed some stuff related to covid.
            #       We only want reports grouped by Hospital/Hospital System OR Discharge Month 1st Admit
            #       If the report is grouped by Discharge Month (Custom), then we didn't remove that parameter and we FAIL.
            group_by_discharge_month_custom_result = 'PASS'
            
            #UL004
            # Also need to check for any custom group bys.  If there are any, then fail.
            custom_group_by_result = 'PASS'
            
            #UL004 Also need to look for discharge month restriction and make sure we removed it.
            discharge_month_restriction_result = 'PASS'
            
            u071_check = 'PASS'
            
            #Hospital / Hospital System
            #Discharge Month 1st Admit
            # collect file name, cohort name and measure name into variables for reporting in the results file.
            file_name = files[0]
            cohort_name = path_key[0]
            measure_name = path_key[1]
            print('path_key: ',path_key)
            # skip ED-1B and ED-OP-18B
            if path_key[1] in do_not_parse:
                # print('passed!')
                pass
            else:

                file_loc = os.path.join(os.path.abspath(path), files[0])

                try:
                    # try to open the excel workbook
                    data_wb = openpyxl.load_workbook(file_loc)
                except:
                    # if unable to open the excel workbook, it is likely corrupted.  Automatic validation fail.
                    # Set all other validation variables to 'N/A' and set final validation_result to 'FAIL'.
                    # skip subsequent operations in the loop
                    validation_result = 'FAIL'
                    file_opened_result = 'FAIL'
                    risk_model_result = 'N/A'
                    ahrq_result = 'N/A'
                    focus_hosp_result = 'N/A'
                    time_period_result = 'N/A'
                    compare_hosp_result = 'N/A'
                    group_by_month_result = 'N/A'
                    #UL004
                    no_records_in_study_pop_result = 'N/A'
                    group_by_discharge_month_custom_result = 'N/A'
                    custom_group_by_result = 'N/A'
                    discharge_month_restriction_result = 'N/A'
                    covid_grouper_result = 'N/A'
                    covid_restriction_result = 'N/A'
                    #UL006
                    u071_check = 'N/A'
                    # create the row list object
                    result_list = [cohort_name, measure_name, file_name, file_opened_result, risk_model_result,
                                   ahrq_result, time_period_result, focus_hosp_result, compare_hosp_result,
                                   group_by_month_result,no_records_in_study_pop_result,\
                                   group_by_discharge_month_custom_result,custom_group_by_result,\
                                   discharge_month_restriction_result,\
                                   covid_grouper_result,covid_restriction_result,u071_check,validation_result]
                    # append the row to the results dataframe
                    validation_results_df.loc[validation_counter] = result_list
                    validation_counter += 1
                    print('Issue opening the excel file.')
                    continue

                # if excel file can be opened, continue onto validating the parameter settings.
                if file_opened_result == 'PASS':
                    data_wb_sheetnames = data_wb.sheetnames
                    validation_ws = data_wb[data_wb_sheetnames[0]]
                    first_validation_col = find_first_ws_col(validation_ws)
                    header_row = find_ws_header_row(validation_ws, first_validation_col)
                    #print('first col & header row: ',first_validation_col,', ',header_row)
                    #UL007
                    xlsx_file = pd.DataFrame(
                        pd.read_excel(file_loc, sheet_name=data_wb_sheetnames[0], usecols=[first_validation_col - 1],engine='openpyxl').iloc[
                        :header_row])
                    
                    #print(xlsx_file.head(n=50))
                    if path_key[0] in ['Large Specialized Complex Care Medical Center',
                                       'Comprehensive Academic Medical Center']:
                        #UL004
                        #look for the phrase 'No Records in Study Population.'  This means the whole thing failed.
                        #if present, then FAIL
                        if xlsx_file[xlsx_file[xlsx_file.columns[0]].str.upper().str.replace(' ','').str.replace('.','').str.contains('NORECORDSINSTUDYPOPULATION') == True].empty == False:
                            no_records_in_study_pop_result = 'FAIL'
                        else:
                            no_records_in_study_pop_result = 'PASS'
                            
                            
                        #UL004
                        #look for the phrase 'By Discharge Month (Custom) or BY AHRQ SAFETY / DISCHARGE MONTH.'  This means we failed to change the 
                        #group by back to default.
                        #print(xlsx_file[xlsx_file[xlsx_file.columns[0]].str.upper().str.replace(' ','').str.replace('.','').str.contains('COVID') == True])
                        #print(xlsx_file[[xlsx_file.columns[0]]].head())
                        #print(xlsx_file[(xlsx_file[xlsx_file.columns[0]].str.upper().str.replace(' ','').str.replace('(','').str.contains('BYDISCHARGEMONTHCUSTOM') == True) | (xlsx_file[xlsx_file.columns[0]].str.upper().str.replace(' ','').str.replace('(','').str.contains('BYAHRQSAFETY/DISCHARGEMONTHCUSTOM') == True)])
                        if xlsx_file[(xlsx_file[xlsx_file.columns[0]].str.upper().str.replace(' ','').str.replace('(','').str.contains('BYDISCHARGEMONTHCUSTOM') == True) | (xlsx_file[xlsx_file.columns[0]].str.upper().str.replace(' ','').str.replace('(','').str.contains('BYAHRQSAFETY/DISCHARGEMONTHCUSTOM') == True)].empty == False:
                            group_by_discharge_month_custom_result = 'FAIL'
                        else:
                            group_by_discharge_month_custom_result = 'PASS'
                            
                            
                            
                        #UL004  Look out for custom group bys.  This also means we failed to get all the covid groupings.
                        
                        if xlsx_file[xlsx_file[xlsx_file.columns[0]].str.upper().str.replace(' ','').str.contains('CUSTOMGROUPBYS') == True].empty == False:
                            custom_group_by_result = 'FAIL'
                        else:
                            custom_group_by_result = 'PASS'
                            
                        
                        if xlsx_file[xlsx_file[xlsx_file.columns[0]].str.upper().str.replace(' ','').str.contains('ADVANCEDRESTRICTIONS') == True].empty == False:
                            adv_rest_index = xlsx_file[xlsx_file[xlsx_file.columns[0]].str.upper().str.replace(' ','').str.contains('ADVANCEDRESTRICTIONS') == True].index.tolist()[0]
                            #print('got it...')
                            #print(adv_rest_index)
                            
                        #isolate before and after advanced restrictions to narrow down search   
                        just_adv_rest = xlsx_file[xlsx_file.columns[0]].iloc[adv_rest_index:]
                        
                        just_adv_rest = just_adv_rest.to_frame()
                        
                        before_adv_rest = xlsx_file[xlsx_file.columns[0]].iloc[:adv_rest_index]
                        before_adv_rest = before_adv_rest.to_frame()
                         
                        #print(xlsx_file[xlsx_file[xlsx_file.columns[0]].str.upper().str.replace(' ','').str.contains('ADVANCEDRESTRICTIONS') == True].index.tolist()[0])
                        #UL004 Look out for discharge month restrictions.  This was used to restriction pre/post covid periods.
                        #only search for 'discharge month' after the advanced restrictions.
                        if just_adv_rest[just_adv_rest[just_adv_rest.columns[0]].str.upper().str.replace(' ','').str.contains('DISCHARGEMONTH:') == True].empty == False:
                            discharge_month_restriction_result = 'FAIL'
                        else:
                            discharge_month_restriction_result = 'PASS'

                        
                        #UL004 Look out for COVID text before and after advanced restrictions.  Splitting this up to help narrow
                        #      the search later.
                        
                        if just_adv_rest[just_adv_rest[just_adv_rest.columns[0]].str.upper().str.replace(' ','').str.contains('COVID') == True].empty == False:
                            covid_restriction_result = 'FAIL'
                        else:
                            covid_restriction_result = 'PASS'
                            
                        if before_adv_rest[before_adv_rest[before_adv_rest.columns[0]].str.upper().str.replace(' ','').str.contains('COVID') == True].empty == False:
                            covid_grouper_result = 'FAIL'
                        else:
                            covid_grouper_result = 'PASS' 
                            
                        #UL006
                        
                        if remove_covid_pats == True:
                                
                            if xlsx_file[xlsx_file[xlsx_file.columns[0]].str.upper().str.replace(' ','').str.replace(':','').str.replace('(','').str.replace(')','').str.contains('ANYDIAGNOSISNOTICD-10U071') == True].empty == False:
                                u071_check = 'PASS'
                                print('u071 pass')
                            else:
                                u071_check = 'FAIL'
                        else:
                            u071_check = 'N/A'
                        
                        risk_model_check = \
                        xlsx_file[xlsx_file[xlsx_file.columns[0]].str.contains('Risk Adjustment Model') == True].values[
                            0][0].split(":", 1)[1].strip() == camc_lsccmc_risk_model.values[0]
                        if risk_model_check == True:
                            risk_model_result = 'PASS'
                        else:
                            risk_model_result = 'FAIL'
                        
                        try:
                            ahrq_model_check = \
                            xlsx_file[xlsx_file[xlsx_file.columns[0]].str.contains('AHRQ Version') == True].values[0][
                            0].split(": :", 1)[1].strip() == camc_lsccmc_ahrq_model.values[0]
                            
                        except:
                            ahrq_model_check = \
                            xlsx_file[xlsx_file[xlsx_file.columns[0]].str.contains('AHRQ Version') == True].values[0][
                            0].split(":", 1)[1].strip().split(":", 1)[1].strip() == camc_lsccmc_ahrq_model.values[0]
                            
        
        

                        if ahrq_model_check == True:
                            ahrq_result = 'PASS'
                        else:
                            ahrq_result = 'FAIL'

                        focus_hosp_check = \
                        xlsx_file[xlsx_file[xlsx_file.columns[0]].str.contains('Focus Hospital') == True].values[0][
                            0].split(":", 1)[1].strip() == focus_hospital.values[0]

                        if focus_hosp_check == True:
                            focus_hosp_result = 'PASS'
                        else:
                            focus_hosp_result = 'FAIL'

                        compare_hosp_check = compare_hospital.values[0] in xlsx_file[
                            xlsx_file[xlsx_file.columns[0]].str.contains('Compare Hospitals') == True].values[0][
                            0].split(":", 1)[1].strip()

                        if compare_hosp_check == True:
                            compare_hosp_result = 'PASS'
                        else:
                            compare_hosp_result = 'FAIL'

                        if (path_key[1].split('_')[0] in ['EDAC', 'READM']) and (
                                path_key[1] not in ['READM_ARTHRO', 'READM_CHOL', 'READM_COLON', 'READM_URI']):

                            time_period_check = \
                            xlsx_file[xlsx_file[xlsx_file.columns[0]].str.contains('Time Period') == True].values[0][0].split(":", 1)[1].strip() == readm_edac_time_period.values[0]

                            if time_period_check == True:
                                time_period_result = 'PASS'
                            else:
                                time_period_result = 'FAIL'

                            # print('test test: ',xlsx_file[xlsx_file[xlsx_file.columns[0]].str.contains('Encounter Month') == True].values[0][0])
                            # check whether or not the report was grouped by month by checking for existence of 'Encounter Month' column
                            try:
                                group_by_month_check = xlsx_file[xlsx_file[xlsx_file.columns[0]].str.contains(
                                    'Encounter Month') == True].values[0][0].strip() == 'Encounter Month'

                                if group_by_month_check == True:
                                    group_by_month_result = 'PASS'
                                else:
                                    group_by_month_result = 'FAIL'
                            except:
                                group_by_month_result = 'FAIL'

                        elif path_key[1] in ['THK']:

                            time_period_check = \
                            xlsx_file[xlsx_file[xlsx_file.columns[0]].str.contains('Time Period') == True].values[0][
                                0].split(":", 1)[1].strip() == thk_time_period.values[0]
                            if time_period_check == True:
                                time_period_result = 'PASS'
                            else:
                                time_period_result = 'FAIL'
                        else:
                            time_period_check = \
                            xlsx_file[xlsx_file[xlsx_file.columns[0]].str.contains('Time Period') == True].values[0][
                                0].split(":", 1)[1].strip() == standard_time_period.values[0]
                            if time_period_check == True:
                                time_period_result = 'PASS'
                            else:
                                time_period_result = 'FAIL'

                        if compare_hosp_result == 'FAIL' or group_by_month_result == 'FAIL' \
                        or file_opened_result == 'FAIL' or risk_model_result == 'FAIL' or ahrq_result == 'FAIL' \
                        or time_period_result == 'FAIL' or focus_hosp_result == 'FAIL' \
                        or no_records_in_study_pop_result == 'FAIL' or group_by_discharge_month_custom_result == 'FAIL'\
                        or custom_group_by_result == 'FAIL' or discharge_month_restriction_result == 'FAIL'\
                        or covid_grouper_result == 'FAIL' or covid_restriction_result == 'FAIL' or u071_check == 'FAIL':
                            validation_result = 'FAIL'
                        else:
                            pass
                        

                        result_list = [cohort_name, measure_name, file_name, file_opened_result, risk_model_result,
                                       ahrq_result, time_period_result, focus_hosp_result, compare_hosp_result,
                                       group_by_month_result,no_records_in_study_pop_result,\
                                       group_by_discharge_month_custom_result,custom_group_by_result,\
                                       discharge_month_restriction_result,\
                                       covid_grouper_result,covid_restriction_result,u071_check,validation_result]
                        
                        
                        validation_results_df.loc[validation_counter] = result_list
                        validation_counter += 1
                        # print('test: ',compare_hosp_result)

                    elif path_key[0] in ['Complex Care Medical Center','Community Medical Center','Critical Access & Small Community']:
                        
                        
                        #print(xlsx_file[xlsx_file[xlsx_file.columns[0]].str.upper().str.replace(' ','').str.replace('.','').str.contains('COVID') == True])
                        #print(xlsx_file[[xlsx_file.columns[0]]].head())
                        #print(xlsx_file[[xlsx_file.columns[0]]])
                        # print('Comm Model')
                        #UL004
                        #look for the phrase 'No Records in Study Population.'  This means the whole thing failed.
                        #if present, then FAIL
                        if xlsx_file[xlsx_file[xlsx_file.columns[0]].str.upper().str.replace(' ','').str.replace('.','').str.contains('NORECORDSINSTUDYPOPULATION') == True].empty == False:
                            no_records_in_study_pop_result = 'FAIL'
                        else:
                            no_records_in_study_pop_result = 'PASS'
                            
                        #UL004
                        #look for the phrase 'By Discharge Month (Custom).'  This means we failed to change the 
                        #group by back to default.
                        
                        if xlsx_file[(xlsx_file[xlsx_file.columns[0]].str.upper().str.replace(' ','').str.replace('(','').str.contains('BYDISCHARGEMONTHCUSTOM') == True) | (xlsx_file[xlsx_file.columns[0]].str.upper().str.replace(' ','').str.replace('(','').str.contains('BYAHRQSAFETY/DISCHARGEMONTHCUSTOM') == True)].empty == False:
                            group_by_discharge_month_custom_result = 'FAIL'
                        else:
                            group_by_discharge_month_custom_result = 'PASS'
                            
                        #UL004  Look out for custom group bys.  This also means we failed to get all the covid groupings.
                        
                        if xlsx_file[xlsx_file[xlsx_file.columns[0]].str.upper().str.replace(' ','').str.contains('CUSTOMGROUPBYS') == True].empty == False:
                            custom_group_by_result = 'FAIL'
                        else:
                            custom_group_by_result = 'PASS'
                            
                        
                        
                        #UL004 Look out for discharge month restrictions.  This was used to restriction pre/post covid periods.
                        
                        if xlsx_file[xlsx_file[xlsx_file.columns[0]].str.upper().str.replace(' ','').str.contains('ADVANCEDRESTRICTIONS') == True].empty == False:
                            adv_rest_index = xlsx_file[xlsx_file[xlsx_file.columns[0]].str.upper().str.replace(' ','').str.contains('ADVANCEDRESTRICTIONS') == True].index.tolist()[0]
                            
                        #isolate before and after advanced restrictions to narrow down search    
                        just_adv_rest = xlsx_file[xlsx_file.columns[0]].iloc[adv_rest_index:]

                        just_adv_rest = just_adv_rest.to_frame() 
                        
                        before_adv_rest = xlsx_file[xlsx_file.columns[0]].iloc[:adv_rest_index]
                        before_adv_rest = before_adv_rest.to_frame()
                            
                        #only search for 'discharge month' after the advanced restrictions.
                        if just_adv_rest[just_adv_rest[just_adv_rest.columns[0]].str.upper().str.replace(' ','').str.contains('DISCHARGEMONTH:') == True].empty == False:
                            discharge_month_restriction_result = 'FAIL'
                        else:
                            discharge_month_restriction_result = 'PASS'
                            
                            
                        #UL004 Look out for COVID text before and after advanced restrictions.  Splitting this up to help narrow
                        #      the search later.
                        
                        if just_adv_rest[just_adv_rest[just_adv_rest.columns[0]].str.upper().str.replace(' ','').str.contains('COVID') == True].empty == False:
                            covid_restriction_result = 'FAIL'
                        else:
                            covid_restriction_result = 'PASS'
                            
                        if before_adv_rest[before_adv_rest[before_adv_rest.columns[0]].str.upper().str.replace(' ','').str.contains('COVID') == True].empty == False:
                            covid_grouper_result = 'FAIL'
                        else:
                            covid_grouper_result = 'PASS' 
                        
                        
                        #UL006
                        
                        
                        if remove_covid_pats == True:
                            print('yep')
                            if xlsx_file[xlsx_file[xlsx_file.columns[0]].str.upper().str.replace(' ','').str.replace(':','').str.replace('(','').str.replace(')','').str.contains('ANYDIAGNOSISNOTICD-10U071') == True].empty == False:
                                u071_check = 'PASS'
                                print('u071 pass')
                            else:
                                u071_check = 'FAIL'
                        else:
                            u071_check = 'N/A'
            

                        risk_model_check = \
                        xlsx_file[xlsx_file[xlsx_file.columns[0]].str.contains('Risk Adjustment Model') == True].values[
                            0][0].split(":", 1)[1].strip() == ccmc_risk_model.values[0]

                        if risk_model_check == True:
                            risk_model_result = 'PASS'
                        else:
                            risk_model_result = 'FAIL'
                        
                        try:
                            ahrq_model_check = \
                            xlsx_file[xlsx_file[xlsx_file.columns[0]].str.contains('AHRQ Version') == True].values[0][
                            0].split(": :", 1)[1].strip() == ccmc_ahrq_model.values[0]
                            
                        except:
                            ahrq_model_check = \
                            xlsx_file[xlsx_file[xlsx_file.columns[0]].str.contains('AHRQ Version') == True].values[0][
                            0].split(":", 1)[1].strip().split(":", 1)[1].strip() == ccmc_ahrq_model.values[0]
                            
                            
                        if ahrq_model_check == True:
                            ahrq_result = 'PASS'
                        else:
                            ahrq_result = 'FAIL'

                        focus_hosp_check = \
                        xlsx_file[xlsx_file[xlsx_file.columns[0]].str.contains('Focus Hospital') == True].values[0][
                            0].split(":", 1)[1].strip() == focus_hospital.values[0]

                        if focus_hosp_check == True:
                            focus_hosp_result = 'PASS'
                        else:
                            focus_hosp_result = 'FAIL'

                        compare_hosp_check = compare_hospital.values[0] in xlsx_file[
                            xlsx_file[xlsx_file.columns[0]].str.contains('Compare Hospitals') == True].values[0][
                            0].split(":", 1)[1].strip()

                        if compare_hosp_check == True:
                            compare_hosp_result = 'PASS'
                        else:
                            compare_hosp_result = 'FAIL'

                        if (path_key[1].split('_')[0] in ['EDAC', 'READM','Readm']) and (
                                path_key[1] not in ['READM_ARTHRO', 'READM_CHOL', 'READM_COLON', 'READM_URI']):

                            time_period_check = \
                            xlsx_file[xlsx_file[xlsx_file.columns[0]].str.contains('Time Period') == True].values[0][
                                0].split(":", 1)[1].strip() == readm_edac_time_period.values[0]

                            if time_period_check == True:
                                time_period_result = 'PASS'
                            else:
                                time_period_result = 'FAIL'

                            # print('test test: ',xlsx_file[xlsx_file[xlsx_file.columns[0]].str.contains('Encounter Month') == True].values[0][0])
                            # check whether or not the report was grouped by month by checking for existence of 'Encounter Month' column
                            try:
                                group_by_month_check = xlsx_file[xlsx_file[xlsx_file.columns[0]].str.contains(
                                    'Encounter Month') == True].values[0][0].strip() == 'Encounter Month'

                                if group_by_month_check == True:
                                    group_by_month_result = 'PASS'
                                else:
                                    group_by_month_result = 'FAIL'
                            except:
                                group_by_month_result = 'FAIL'

                        elif path_key[1] in ['THK']:

                            time_period_check = \
                            xlsx_file[xlsx_file[xlsx_file.columns[0]].str.contains('Time Period') == True].values[0][
                                0].split(":", 1)[1].strip() == thk_time_period.values[0]
                            if time_period_check == True:
                                time_period_result = 'PASS'
                            else:
                                time_period_result = 'FAIL'
                        else:
                            time_period_check = \
                            xlsx_file[xlsx_file[xlsx_file.columns[0]].str.contains('Time Period') == True].values[0][
                                0].split(":", 1)[1].strip() == standard_time_period.values[0]
                            if time_period_check == True:
                                time_period_result = 'PASS'
                            else:
                                time_period_result = 'FAIL'

                        if compare_hosp_result == 'FAIL' or group_by_month_result == 'FAIL' \
                        or file_opened_result == 'FAIL' or risk_model_result == 'FAIL' or ahrq_result == 'FAIL' \
                        or time_period_result == 'FAIL' or focus_hosp_result == 'FAIL' \
                        or no_records_in_study_pop_result == 'FAIL' or group_by_discharge_month_custom_result == 'FAIL'\
                        or custom_group_by_result == 'FAIL' or discharge_month_restriction_result == 'FAIL'\
                        or covid_grouper_result == 'FAIL' or covid_restriction_result == 'FAIL' or u071_check == 'FAIL':
                            validation_result = 'FAIL'
                        else:
                            pass
                        
                        
                        result_list = [cohort_name, measure_name, file_name, file_opened_result, risk_model_result,
                                       ahrq_result, time_period_result, focus_hosp_result, compare_hosp_result,
                                       group_by_month_result,no_records_in_study_pop_result,\
                                       group_by_discharge_month_custom_result,custom_group_by_result,\
                                       discharge_month_restriction_result,\
                                       covid_grouper_result,covid_restriction_result,u071_check,validation_result]
                        
                        
                        validation_results_df.loc[validation_counter] = result_list
                        validation_counter += 1

    validation_results_df.to_csv(os.path.join(os.path.abspath(wd_dest), "validation_results.csv"), index=False)
    print('Done.')
'''
#UL011  Checking for standard restrictions and advanced restrictions

def validate_downloaded_files(remove_covid_pats = False):
    pd.set_option('display.max_colwidth', None)
    # Set the path of the folder to parse and validate
    wd = input('Enter the path of the folder you want to validate.')
    wd = os.path.abspath(wd)
    # Set the path of the folder structure where we validation file to go
    wd_dest = input('Enter the path of the folder you want the validation file to go in.')
    wd_dest = os.path.abspath(wd_dest)
    # Open the validation key file.
    validation_file_name = 'validation_file/vizient_qa_validation.xlsx'
    validation_file_loc = os.path.join(os.path.abspath(wd_dest), validation_file_name)

    # Read in the validation excel file
    #UL007
    validation_xlsx_file = pd.DataFrame(pd.read_excel(validation_file_loc, sheet_name='Values',engine='openpyxl'))
    
    # Store the validation values into variables
    camc_lsccmc_risk_model = \
    validation_xlsx_file[validation_xlsx_file['validation_item'] == 'CAMC/LSCCMC Risk Adjustment Model']['value']
    ccmc_risk_model = validation_xlsx_file[validation_xlsx_file['validation_item'] == 'CCMC Risk Adjustment Model'][
        'value']
    camc_lsccmc_ahrq_model = \
    validation_xlsx_file[validation_xlsx_file['validation_item'] == 'CAMC/LSCCMC AHRQ Version']['value']
    ccmc_ahrq_model = validation_xlsx_file[validation_xlsx_file['validation_item'] == 'CCMC AHRQ Version']['value']
    standard_time_period = validation_xlsx_file[validation_xlsx_file['validation_item'] == 'Standard Time Period'][
        'value']
    thk_time_period = validation_xlsx_file[validation_xlsx_file['validation_item'] == 'THK Time Period']['value']
    readm_edac_time_period = validation_xlsx_file[validation_xlsx_file['validation_item'] == 'READM/EDAC Time Period'][
        'value']
    focus_hospital = validation_xlsx_file[validation_xlsx_file['validation_item'] == 'Focus Hospital']['value']
    compare_hospital = validation_xlsx_file[validation_xlsx_file['validation_item'] == 'Compare Hospitals']['value']
    print('compare hospital:', compare_hospital.values[0])

    # Set counters which will act as proxy indexes in order to append rows to pandas dataframes
    validation_counter = 0
    # Create empty dataframes to store final validation result values.
    validation_results_df = pd.DataFrame(
        columns=['Cohort', 'Measure', 'File', 'File Opened', 'Risk Model', 'AHRQ Model', 'Time Period',
                 'Focus Hospital', 'Compare Hospitals', 'Group By Month','No Records In Study',\
                 'Group By Disch Month Custom','Custom Group Bys',\
                 'Discharge Month Restriction',\
                 'COVID Grouper','COVID Restriction','U071 Check','Standard Restriction Check','Advanced Restriction Check','Validation Result'])

    # Do not need to parse ED-2B or ED OP 18B
    do_not_parse = ['ED_2B', 'ED_OP_18B']
    # iterate through folder structure.
    for path, dirs, files in os.walk(wd):
        # if files list is not empty, open it up
        if len(files) > 0:
            # parse folder name to get measure name
            path_tail = str((os.path.basename(path)))
            # parse parent folder to get cohort name
            path_head = os.path.dirname(path)
            path_second_tail = str(os.path.basename(path_head))
            # use cohort and measure composite key to access dictionary values
            path_key = (path_second_tail, path_tail)
            # set default variables.  If nothing is found, then PASS.  If an issue is found then set to FAIL.
            file_opened_result = 'PASS'
            risk_model_result = 'PASS'
            ahrq_result = 'PASS'
            time_period_result = 'PASS'
            focus_hosp_result = 'PASS'
            validation_result = 'PASS'
            compare_hosp_result = 'PASS'
            group_by_month_result = 'N/A'
            #UL004  Vizient added new filters and group bys to the report templates to handle COVID.
            #       we need to check whether those filters were removed or not.
            #       If all the covid filters are applied, then we just won't get data and the phrase No Reocords In Study Population
            #       will appear instead.  Check for this phrase.
            no_records_in_study_pop_result = 'PASS'
            
            #UL004  Look for custom group bys.  This means Vizient changed some stuff related to covid.
            #       We only want reports grouped by Hospital/Hospital System OR Discharge Month 1st Admit
            #       If the report is grouped by Discharge Month (Custom), then we didn't remove that parameter and we FAIL.
            group_by_discharge_month_custom_result = 'PASS'
            
            #UL004
            # Also need to check for any custom group bys.  If there are any, then fail.
            custom_group_by_result = 'PASS'
            
            #UL004 Also need to look for discharge month restriction and make sure we removed it.
            discharge_month_restriction_result = 'PASS'
            
            u071_check = 'PASS'
            #UL011
            standard_rest_check = 'PASS'
            advanced_rest_check = 'PASS'
            
            #Hospital / Hospital System
            #Discharge Month 1st Admit
            # collect file name, cohort name and measure name into variables for reporting in the results file.
            file_name = files[0]
            cohort_name = path_key[0]
            measure_name = path_key[1]
            print('path_key: ',path_key)
            
            no_restriction_meas_list = ['PSI_03','PSI_06','PSI_09','PSI_11','PSI_13','THK','ADE_DENOM','DCOST_OVERALL','EARLY_TRANS_DENOM']
            
            
            # skip ED-1B and ED-OP-18B
            if path_key[1] in do_not_parse:
                # print('passed!')
                pass
            else:

                file_loc = os.path.join(os.path.abspath(path), files[0])

                try:
                    # try to open the excel workbook
                    data_wb = openpyxl.load_workbook(file_loc)
                except:
                    # if unable to open the excel workbook, it is likely corrupted.  Automatic validation fail.
                    # Set all other validation variables to 'N/A' and set final validation_result to 'FAIL'.
                    # skip subsequent operations in the loop
                    validation_result = 'FAIL'
                    file_opened_result = 'FAIL'
                    risk_model_result = 'N/A'
                    ahrq_result = 'N/A'
                    focus_hosp_result = 'N/A'
                    time_period_result = 'N/A'
                    compare_hosp_result = 'N/A'
                    group_by_month_result = 'N/A'
                    #UL004
                    no_records_in_study_pop_result = 'N/A'
                    group_by_discharge_month_custom_result = 'N/A'
                    custom_group_by_result = 'N/A'
                    discharge_month_restriction_result = 'N/A'
                    covid_grouper_result = 'N/A'
                    covid_restriction_result = 'N/A'
                    #UL006
                    u071_check = 'N/A'
                    #UL011
                    standard_rest_check = 'N/A'
                    advanced_rest_check = 'N/A'
                    # create the row list object
                    result_list = [cohort_name, measure_name, file_name, file_opened_result, risk_model_result,
                                   ahrq_result, time_period_result, focus_hosp_result, compare_hosp_result,
                                   group_by_month_result,no_records_in_study_pop_result,\
                                   group_by_discharge_month_custom_result,custom_group_by_result,\
                                   discharge_month_restriction_result,\
                                   covid_grouper_result,covid_restriction_result,u071_check,\
                                   standard_rest_check,advanced_rest_check,validation_result]
                    # append the row to the results dataframe
                    validation_results_df.loc[validation_counter] = result_list
                    validation_counter += 1
                    print('Issue opening the excel file.')
                    continue

                # if excel file can be opened, continue onto validating the parameter settings.
                if file_opened_result == 'PASS':
                    data_wb_sheetnames = data_wb.sheetnames
                    validation_ws = data_wb[data_wb_sheetnames[0]]
                    first_validation_col = find_first_ws_col(validation_ws)
                    header_row = find_ws_header_row(validation_ws, first_validation_col)
                    #print('first col & header row: ',first_validation_col,', ',header_row)
                    #UL007
                    xlsx_file = pd.DataFrame(
                        pd.read_excel(file_loc, sheet_name=data_wb_sheetnames[0], usecols=[first_validation_col - 1],engine='openpyxl').iloc[
                        :header_row])
                    
                    #print(xlsx_file.head(n=50))
                    if path_key[0] in ['Large Specialized Complex Care Medical Center',
                                       'Comprehensive Academic Medical Center']:
                        #UL004
                        #look for the phrase 'No Records in Study Population.'  This means the whole thing failed.
                        #if present, then FAIL
                        if xlsx_file[xlsx_file[xlsx_file.columns[0]].str.upper().str.replace(' ','').str.replace('.','').str.contains('NORECORDSINSTUDYPOPULATION') == True].empty == False:
                            no_records_in_study_pop_result = 'FAIL'
                        else:
                            no_records_in_study_pop_result = 'PASS'
                            
                            
                        #UL004
                        #look for the phrase 'By Discharge Month (Custom) or BY AHRQ SAFETY / DISCHARGE MONTH.'  This means we failed to change the 
                        #group by back to default.
                        #print(xlsx_file[xlsx_file[xlsx_file.columns[0]].str.upper().str.replace(' ','').str.replace('.','').str.contains('COVID') == True])
                        #print(xlsx_file[[xlsx_file.columns[0]]].head())
                        #print(xlsx_file[(xlsx_file[xlsx_file.columns[0]].str.upper().str.replace(' ','').str.replace('(','').str.contains('BYDISCHARGEMONTHCUSTOM') == True) | (xlsx_file[xlsx_file.columns[0]].str.upper().str.replace(' ','').str.replace('(','').str.contains('BYAHRQSAFETY/DISCHARGEMONTHCUSTOM') == True)])
                        if xlsx_file[(xlsx_file[xlsx_file.columns[0]].str.upper().str.replace(' ','').str.replace('(','').str.contains('BYDISCHARGEMONTHCUSTOM') == True) | (xlsx_file[xlsx_file.columns[0]].str.upper().str.replace(' ','').str.replace('(','').str.contains('BYAHRQSAFETY/DISCHARGEMONTHCUSTOM') == True)].empty == False:
                            group_by_discharge_month_custom_result = 'FAIL'
                        else:
                            group_by_discharge_month_custom_result = 'PASS'
                            
                            
                            
                        #UL004  Look out for custom group bys.  This also means we failed to get all the covid groupings.
                        
                        if xlsx_file[xlsx_file[xlsx_file.columns[0]].str.upper().str.replace(' ','').str.contains('CUSTOMGROUPBYS') == True].empty == False:
                            custom_group_by_result = 'FAIL'
                        else:
                            custom_group_by_result = 'PASS'
                            
                        
                        if xlsx_file[xlsx_file[xlsx_file.columns[0]].str.upper().str.replace(' ','').str.contains('ADVANCEDRESTRICTIONS') == True].empty == False:
                            adv_rest_index = xlsx_file[xlsx_file[xlsx_file.columns[0]].str.upper().str.replace(' ','').str.contains('ADVANCEDRESTRICTIONS') == True].index.tolist()[0]
                            #print('got it...')
                            #print(adv_rest_index)
                            
                        #isolate before and after advanced restrictions to narrow down search   
                        just_adv_rest = xlsx_file[xlsx_file.columns[0]].iloc[adv_rest_index:]
                        
                        just_adv_rest = just_adv_rest.to_frame()
                        
                        before_adv_rest = xlsx_file[xlsx_file.columns[0]].iloc[:adv_rest_index]
                        before_adv_rest = before_adv_rest.to_frame()
                         
                        #print(xlsx_file[xlsx_file[xlsx_file.columns[0]].str.upper().str.replace(' ','').str.contains('ADVANCEDRESTRICTIONS') == True].index.tolist()[0])
                        #UL004 Look out for discharge month restrictions.  This was used to restriction pre/post covid periods.
                        #only search for 'discharge month' after the advanced restrictions.
                        if just_adv_rest[just_adv_rest[just_adv_rest.columns[0]].str.upper().str.replace(' ','').str.contains('DISCHARGEMONTH:') == True].empty == False:
                            discharge_month_restriction_result = 'FAIL'
                        else:
                            discharge_month_restriction_result = 'PASS'

                        
                        #UL004 Look out for COVID text before and after advanced restrictions.  Splitting this up to help narrow
                        #      the search later.
                        
                        if just_adv_rest[just_adv_rest[just_adv_rest.columns[0]].str.upper().str.replace(' ','').str.contains('COVID') == True].empty == False:
                            covid_restriction_result = 'FAIL'
                        else:
                            covid_restriction_result = 'PASS'
                            
                        if before_adv_rest[before_adv_rest[before_adv_rest.columns[0]].str.upper().str.replace(' ','').str.contains('COVID') == True].empty == False:
                            covid_grouper_result = 'FAIL'
                        else:
                            covid_grouper_result = 'PASS' 
                            
                        #UL006
                        
                        if remove_covid_pats == True:
                                
                            if xlsx_file[xlsx_file[xlsx_file.columns[0]].str.upper().str.replace(' ','').str.replace(':','').str.replace('(','').str.replace(')','').str.contains('ANYDIAGNOSISNOTICD-10U071') == True].empty == False:
                                u071_check = 'PASS'
                                print('u071 pass')
                            else:
                                u071_check = 'FAIL'
                        else:
                            u071_check = 'N/A'
                        
                        risk_model_check = \
                        xlsx_file[xlsx_file[xlsx_file.columns[0]].str.contains('Risk Adjustment Model') == True].values[
                            0][0].split(":", 1)[1].strip() == camc_lsccmc_risk_model.values[0]
                        if risk_model_check == True:
                            risk_model_result = 'PASS'
                        else:
                            risk_model_result = 'FAIL'
                        
                        try:
                            ahrq_model_check = \
                            xlsx_file[xlsx_file[xlsx_file.columns[0]].str.contains('AHRQ Version') == True].values[0][
                            0].split(": :", 1)[1].strip() == camc_lsccmc_ahrq_model.values[0]
                            
                        except:
                            ahrq_model_check = \
                            xlsx_file[xlsx_file[xlsx_file.columns[0]].str.contains('AHRQ Version') == True].values[0][
                            0].split(":", 1)[1].strip().split(":", 1)[1].strip() == camc_lsccmc_ahrq_model.values[0]
                            
        
        

                        if ahrq_model_check == True:
                            ahrq_result = 'PASS'
                        else:
                            ahrq_result = 'FAIL'

                        focus_hosp_check = \
                        xlsx_file[xlsx_file[xlsx_file.columns[0]].str.contains('Focus Hospital') == True].values[0][
                            0].split(":", 1)[1].strip() == focus_hospital.values[0]

                        if focus_hosp_check == True:
                            focus_hosp_result = 'PASS'
                        else:
                            focus_hosp_result = 'FAIL'

                        compare_hosp_check = compare_hospital.values[0] in xlsx_file[
                            xlsx_file[xlsx_file.columns[0]].str.contains('Compare Hospitals') == True].values[0][
                            0].split(":", 1)[1].strip()

                        if compare_hosp_check == True:
                            compare_hosp_result = 'PASS'
                        else:
                            compare_hosp_result = 'FAIL'

                        if (path_key[1].split('_')[0] in ['EDAC', 'READM']) and (
                                path_key[1] not in ['READM_ARTHRO', 'READM_CHOL', 'READM_COLON', 'READM_URI']):

                            time_period_check = \
                            xlsx_file[xlsx_file[xlsx_file.columns[0]].str.contains('Time Period') == True].values[0][0].split(":", 1)[1].strip() == readm_edac_time_period.values[0]

                            if time_period_check == True:
                                time_period_result = 'PASS'
                            else:
                                time_period_result = 'FAIL'

                            # print('test test: ',xlsx_file[xlsx_file[xlsx_file.columns[0]].str.contains('Encounter Month') == True].values[0][0])
                            # check whether or not the report was grouped by month by checking for existence of 'Encounter Month' column
                            try:
                                group_by_month_check = xlsx_file[xlsx_file[xlsx_file.columns[0]].str.contains(
                                    'Encounter Month') == True].values[0][0].strip() == 'Encounter Month'

                                if group_by_month_check == True:
                                    group_by_month_result = 'PASS'
                                else:
                                    group_by_month_result = 'FAIL'
                            except:
                                group_by_month_result = 'FAIL'

                        elif path_key[1] in ['THK']:

                            time_period_check = \
                            xlsx_file[xlsx_file[xlsx_file.columns[0]].str.contains('Time Period') == True].values[0][
                                0].split(":", 1)[1].strip() == thk_time_period.values[0]
                            if time_period_check == True:
                                time_period_result = 'PASS'
                            else:
                                time_period_result = 'FAIL'
                        else:
                            time_period_check = \
                            xlsx_file[xlsx_file[xlsx_file.columns[0]].str.contains('Time Period') == True].values[0][
                                0].split(":", 1)[1].strip() == standard_time_period.values[0]
                            if time_period_check == True:
                                time_period_result = 'PASS'
                            else:
                                time_period_result = 'FAIL'
                                
                        #UL011

                        if xlsx_file[xlsx_file[xlsx_file.columns[0]].str.upper().str.replace(' ','').str.replace('.','').str.contains('STANDARDRESTRICTIONS:') == True].empty == False:
                            standard_rest_check = 'PASS'
                        else:
                            standard_rest_check = 'FAIL'
                            
                        if measure_name in no_restriction_meas_list:
                            advanced_rest_check = 'N/A'    
                        elif xlsx_file[xlsx_file[xlsx_file.columns[0]].str.upper().str.replace(' ','').str.replace('.','').str.contains('ADVANCEDRESTRICTIONS:') == True].empty == False:
                            advanced_rest_check = 'PASS'
                        else:
                            advanced_rest_check = 'FAIL'
                            
                            

                        if compare_hosp_result == 'FAIL' or group_by_month_result == 'FAIL' \
                        or file_opened_result == 'FAIL' or risk_model_result == 'FAIL' or ahrq_result == 'FAIL' \
                        or time_period_result == 'FAIL' or focus_hosp_result == 'FAIL' \
                        or no_records_in_study_pop_result == 'FAIL' or group_by_discharge_month_custom_result == 'FAIL'\
                        or custom_group_by_result == 'FAIL' or discharge_month_restriction_result == 'FAIL'\
                        or covid_grouper_result == 'FAIL' or covid_restriction_result == 'FAIL' or u071_check == 'FAIL' \
                        or standard_rest_check == 'FAIL' or advanced_rest_check == 'FAIL':
                            validation_result = 'FAIL'
                        else:
                            pass
                        

                        result_list = [cohort_name, measure_name, file_name, file_opened_result, risk_model_result,
                                       ahrq_result, time_period_result, focus_hosp_result, compare_hosp_result,
                                       group_by_month_result,no_records_in_study_pop_result,\
                                       group_by_discharge_month_custom_result,custom_group_by_result,\
                                       discharge_month_restriction_result,\
                                       covid_grouper_result,covid_restriction_result,u071_check,\
                                       standard_rest_check,advanced_rest_check,validation_result]
                        
                        
                        validation_results_df.loc[validation_counter] = result_list
                        validation_counter += 1
                        # print('test: ',compare_hosp_result)

                    elif path_key[0] in ['Complex Care Medical Center','Community Medical Center','Critical Access & Small Community']:
                        
                        
                        #print(xlsx_file[xlsx_file[xlsx_file.columns[0]].str.upper().str.replace(' ','').str.replace('.','').str.contains('COVID') == True])
                        #print(xlsx_file[[xlsx_file.columns[0]]].head())
                        #print(xlsx_file[[xlsx_file.columns[0]]])
                        # print('Comm Model')
                        #UL004
                        #look for the phrase 'No Records in Study Population.'  This means the whole thing failed.
                        #if present, then FAIL
                        if xlsx_file[xlsx_file[xlsx_file.columns[0]].str.upper().str.replace(' ','').str.replace('.','').str.contains('NORECORDSINSTUDYPOPULATION') == True].empty == False:
                            no_records_in_study_pop_result = 'FAIL'
                        else:
                            no_records_in_study_pop_result = 'PASS'
                            
                        #UL004
                        #look for the phrase 'By Discharge Month (Custom).'  This means we failed to change the 
                        #group by back to default.
                        
                        if xlsx_file[(xlsx_file[xlsx_file.columns[0]].str.upper().str.replace(' ','').str.replace('(','').str.contains('BYDISCHARGEMONTHCUSTOM') == True) | (xlsx_file[xlsx_file.columns[0]].str.upper().str.replace(' ','').str.replace('(','').str.contains('BYAHRQSAFETY/DISCHARGEMONTHCUSTOM') == True)].empty == False:
                            group_by_discharge_month_custom_result = 'FAIL'
                        else:
                            group_by_discharge_month_custom_result = 'PASS'
                            
                        #UL004  Look out for custom group bys.  This also means we failed to get all the covid groupings.
                        
                        if xlsx_file[xlsx_file[xlsx_file.columns[0]].str.upper().str.replace(' ','').str.contains('CUSTOMGROUPBYS') == True].empty == False:
                            custom_group_by_result = 'FAIL'
                        else:
                            custom_group_by_result = 'PASS'
                            
                        
                        
                        #UL004 Look out for discharge month restrictions.  This was used to restriction pre/post covid periods.
                        
                        if xlsx_file[xlsx_file[xlsx_file.columns[0]].str.upper().str.replace(' ','').str.contains('ADVANCEDRESTRICTIONS') == True].empty == False:
                            adv_rest_index = xlsx_file[xlsx_file[xlsx_file.columns[0]].str.upper().str.replace(' ','').str.contains('ADVANCEDRESTRICTIONS') == True].index.tolist()[0]
                            
                        #isolate before and after advanced restrictions to narrow down search    
                        just_adv_rest = xlsx_file[xlsx_file.columns[0]].iloc[adv_rest_index:]

                        just_adv_rest = just_adv_rest.to_frame() 
                        
                        before_adv_rest = xlsx_file[xlsx_file.columns[0]].iloc[:adv_rest_index]
                        before_adv_rest = before_adv_rest.to_frame()
                            
                        #only search for 'discharge month' after the advanced restrictions.
                        if just_adv_rest[just_adv_rest[just_adv_rest.columns[0]].str.upper().str.replace(' ','').str.contains('DISCHARGEMONTH:') == True].empty == False:
                            discharge_month_restriction_result = 'FAIL'
                        else:
                            discharge_month_restriction_result = 'PASS'
                            
                            
                        #UL004 Look out for COVID text before and after advanced restrictions.  Splitting this up to help narrow
                        #      the search later.
                        
                        if just_adv_rest[just_adv_rest[just_adv_rest.columns[0]].str.upper().str.replace(' ','').str.contains('COVID') == True].empty == False:
                            covid_restriction_result = 'FAIL'
                        else:
                            covid_restriction_result = 'PASS'
                            
                        if before_adv_rest[before_adv_rest[before_adv_rest.columns[0]].str.upper().str.replace(' ','').str.contains('COVID') == True].empty == False:
                            covid_grouper_result = 'FAIL'
                        else:
                            covid_grouper_result = 'PASS' 
                        
                        
                        #UL006
                        
                        
                        if remove_covid_pats == True:
                            print('yep')
                            if xlsx_file[xlsx_file[xlsx_file.columns[0]].str.upper().str.replace(' ','').str.replace(':','').str.replace('(','').str.replace(')','').str.contains('ANYDIAGNOSISNOTICD-10U071') == True].empty == False:
                                u071_check = 'PASS'
                                print('u071 pass')
                            else:
                                u071_check = 'FAIL'
                        else:
                            u071_check = 'N/A'
            

                        risk_model_check = \
                        xlsx_file[xlsx_file[xlsx_file.columns[0]].str.contains('Risk Adjustment Model') == True].values[
                            0][0].split(":", 1)[1].strip() == ccmc_risk_model.values[0]

                        if risk_model_check == True:
                            risk_model_result = 'PASS'
                        else:
                            risk_model_result = 'FAIL'
                        
                        try:
                            ahrq_model_check = \
                            xlsx_file[xlsx_file[xlsx_file.columns[0]].str.contains('AHRQ Version') == True].values[0][
                            0].split(": :", 1)[1].strip() == ccmc_ahrq_model.values[0]
                            
                        except:
                            ahrq_model_check = \
                            xlsx_file[xlsx_file[xlsx_file.columns[0]].str.contains('AHRQ Version') == True].values[0][
                            0].split(":", 1)[1].strip().split(":", 1)[1].strip() == ccmc_ahrq_model.values[0]
                            
                            
                        if ahrq_model_check == True:
                            ahrq_result = 'PASS'
                        else:
                            ahrq_result = 'FAIL'

                        focus_hosp_check = \
                        xlsx_file[xlsx_file[xlsx_file.columns[0]].str.contains('Focus Hospital') == True].values[0][
                            0].split(":", 1)[1].strip() == focus_hospital.values[0]

                        if focus_hosp_check == True:
                            focus_hosp_result = 'PASS'
                        else:
                            focus_hosp_result = 'FAIL'

                        compare_hosp_check = compare_hospital.values[0] in xlsx_file[
                            xlsx_file[xlsx_file.columns[0]].str.contains('Compare Hospitals') == True].values[0][
                            0].split(":", 1)[1].strip()

                        if compare_hosp_check == True:
                            compare_hosp_result = 'PASS'
                        else:
                            compare_hosp_result = 'FAIL'

                        if (path_key[1].split('_')[0] in ['EDAC', 'READM','Readm']) and (
                                path_key[1] not in ['READM_ARTHRO', 'READM_CHOL', 'READM_COLON', 'READM_URI']):

                            time_period_check = \
                            xlsx_file[xlsx_file[xlsx_file.columns[0]].str.contains('Time Period') == True].values[0][
                                0].split(":", 1)[1].strip() == readm_edac_time_period.values[0]

                            if time_period_check == True:
                                time_period_result = 'PASS'
                            else:
                                time_period_result = 'FAIL'

                            # print('test test: ',xlsx_file[xlsx_file[xlsx_file.columns[0]].str.contains('Encounter Month') == True].values[0][0])
                            # check whether or not the report was grouped by month by checking for existence of 'Encounter Month' column
                            try:
                                group_by_month_check = xlsx_file[xlsx_file[xlsx_file.columns[0]].str.contains(
                                    'Encounter Month') == True].values[0][0].strip() == 'Encounter Month'

                                if group_by_month_check == True:
                                    group_by_month_result = 'PASS'
                                else:
                                    group_by_month_result = 'FAIL'
                            except:
                                group_by_month_result = 'FAIL'

                        elif path_key[1] in ['THK']:

                            time_period_check = \
                            xlsx_file[xlsx_file[xlsx_file.columns[0]].str.contains('Time Period') == True].values[0][
                                0].split(":", 1)[1].strip() == thk_time_period.values[0]
                            if time_period_check == True:
                                time_period_result = 'PASS'
                            else:
                                time_period_result = 'FAIL'
                        else:
                            time_period_check = \
                            xlsx_file[xlsx_file[xlsx_file.columns[0]].str.contains('Time Period') == True].values[0][
                                0].split(":", 1)[1].strip() == standard_time_period.values[0]
                            if time_period_check == True:
                                time_period_result = 'PASS'
                            else:
                                time_period_result = 'FAIL'
                                
                                
                        #UL011
                        if xlsx_file[xlsx_file[xlsx_file.columns[0]].str.upper().str.replace(' ','').str.replace('.','').str.contains('STANDARDRESTRICTIONS:') == True].empty == False:
                            standard_rest_check = 'PASS'
                        else:
                            standard_rest_check = 'FAIL'
                            
                            
                        if measure_name in no_restriction_meas_list:
                            advanced_rest_check = 'N/A'    
                        elif xlsx_file[xlsx_file[xlsx_file.columns[0]].str.upper().str.replace(' ','').str.replace('.','').str.contains('ADVANCEDRESTRICTIONS:') == True].empty == False:
                            advanced_rest_check = 'PASS'
                        else:
                            advanced_rest_check = 'FAIL'

                        if compare_hosp_result == 'FAIL' or group_by_month_result == 'FAIL' \
                        or file_opened_result == 'FAIL' or risk_model_result == 'FAIL' or ahrq_result == 'FAIL' \
                        or time_period_result == 'FAIL' or focus_hosp_result == 'FAIL' \
                        or no_records_in_study_pop_result == 'FAIL' or group_by_discharge_month_custom_result == 'FAIL'\
                        or custom_group_by_result == 'FAIL' or discharge_month_restriction_result == 'FAIL'\
                        or covid_grouper_result == 'FAIL' or covid_restriction_result == 'FAIL' or u071_check == 'FAIL'\
                        or standard_rest_check == 'FAIL' or advanced_rest_check == 'FAIL':
                            validation_result = 'FAIL'
                        else:
                            pass
                        
                        
                        result_list = [cohort_name, measure_name, file_name, file_opened_result, risk_model_result,
                                       ahrq_result, time_period_result, focus_hosp_result, compare_hosp_result,
                                       group_by_month_result,no_records_in_study_pop_result,\
                                       group_by_discharge_month_custom_result,custom_group_by_result,\
                                       discharge_month_restriction_result,\
                                       covid_grouper_result,covid_restriction_result,u071_check,\
                                       standard_rest_check,advanced_rest_check, validation_result]
                        
                        
                        validation_results_df.loc[validation_counter] = result_list
                        validation_counter += 1

    validation_results_df.to_csv(os.path.join(os.path.abspath(wd_dest), "validation_results.csv"), index=False)
    print('Done.')

######################################################################################################################
#####################################################################################################################
#####################################################################################################################
#####################################################################################################################
# function to convert results of vizient_data_folder_walker_and_prep_for_db_inserts()
# into same form as NM_Analytics.vizient_qa.measure_values table.

def convert_metric_df_to_insert_measure_value_df(df_nm, calc_nm, period_type, period_end_dts):
    # set up connection and sql queries
    conn = pyodbc.connect('Driver={SQL Server};'
                          'Server=edw00pd05wva.corp.nm.org\EDWIDS1;'
                          'Database=clarity;'
                          'Trusted_Connection=yes;')

    calc_id = """
    SELECT
    c.calc_id
    FROM
    NM_Analytics_Prototype.vizient_qa.calculator as c
    where
    c.calc_nm = '%s'
    """ % calc_nm

    calc_id_result = pd.DataFrame(pd.read_sql(calc_id, conn))['calc_id'].values[0]

    period_type_id = """
    SELECT
    pt.period_type_id
    FROM
    NM_Analytics_Prototype.vizient_qa.period_types as pt
    WHERE
    pt.period_type_nm = '%s'
    """ % period_type

    period_type_id_result = pd.DataFrame(pd.read_sql(period_type_id, conn))['period_type_id'].values[0]

    period_id = '''
    SELECT 
    p.period_id 
    FROM NM_Analytics_Prototype.vizient_qa.periods as p 
    WHERE p.period_end_date = '%s'
    and p.period_type = '%s'
    ''' % (period_end_dts, period_type_id_result)

    period_id_result = pd.DataFrame(pd.read_sql(period_id, conn))['period_id'].values[0]

    all_hospitals = """
    SELECT
    hospital_id
    ,hospital_medicare_id
    --,hospital_medicare_id + ' ' + hospital_name as hospital_name
    from
    NM_Analytics_Prototype.vizient_qa.hospitals
    """

    all_measures = '''
    SELECT
    measure_id
    ,measure_name
    FROM
    NM_Analytics_Prototype.vizient_qa.measure
    '''

    num_event_types = '''
    SELECT
    event_type_id as numerator_event_type_id
    ,event_type_nm as num_event_type_nm
    from
    NM_Analytics_Prototype.vizient_qa.event_types
    '''

    denom_event_types = '''
    SELECT
    event_type_id as denominator_event_type_id
    ,event_type_nm as denom_event_type_nm
    from
    NM_Analytics_Prototype.vizient_qa.event_types
    '''

    # query the DB to get correct values and IDs
    calc_id_result = pd.DataFrame(pd.read_sql(calc_id, conn))['calc_id'].values[0]
    hospital_df = pd.DataFrame(pd.read_sql(all_hospitals, conn))
    measure_df = pd.DataFrame(pd.read_sql(all_measures, conn))
    num_event_type_df = pd.DataFrame(pd.read_sql(num_event_types, conn))
    denom_event_type_df = pd.DataFrame(pd.read_sql(denom_event_types, conn))

    # close SQL DB connection
    conn.close()

    #create a new column of just the medicare ID
    new = df_nm["Hospital"].str.split(" ", n = 1, expand = True)
    df_nm["hospital_medicare_id"] = new[0]

    # add calc_id to dataframe
    df_nm['calc_id'] = calc_id_result
    # add period id to dataframe
    df_nm['period_id'] = period_id_result
    # values will always be Metric Values so add measure_value_id = 1
    df_nm['measure_value_id'] = 1
    # rename columns to match measure_values table column names
    df_nm = df_nm.rename(
        columns={'Metric Value': 'measure_value', 'N Events': 'numerator', 'event_type_nm': 'numerator_event_type_nm',
                 'denominator': 'denominator', 'denominator_event_type_nm': 'denominator_event_type_nm'})
    # left join to get hospital, measure and event_type IDs
    df_nm2 = pd.merge(df_nm, hospital_df, how='left', left_on=['hospital_medicare_id'], right_on=['hospital_medicare_id'])
    #df_nm2 = pd.merge(df_nm, hospital_df, how='left', left_on=['Hospital'], right_on=['hospital_name'])
    df_nm3 = pd.merge(df_nm2, measure_df, how='left', left_on=['Measure'], right_on=['measure_name'])
    df_nm4 = pd.merge(df_nm3, num_event_type_df, how='left', left_on=['numerator_event_type_nm'],
                      right_on=['num_event_type_nm'])
    df_nm5 = pd.merge(df_nm4, denom_event_type_df, how='left', left_on=['denominator_event_type_nm'],
                      right_on=['denom_event_type_nm'])

    # subset and reorder columns
    cols = ['calc_id', 'hospital_id', 'measure_id', 'period_id', 'measure_value_id', 'measure_value', 'numerator',
            'numerator_event_type_id', 'denominator', 'denominator_event_type_id']
    df_nm6 = df_nm5[cols]

    return (df_nm6)

#####################################################################################################################
#####################################################################################################################
#####################################################################################################################
#####################################################################################################################

def insert_measure_values_from_reports_df(df):
    # connect to the NM_Analytics database
    conn = pyodbc.connect('Driver={SQL Server};'
                          'Server=edw00pd05wva.corp.nm.org\EDWIDS1;'
                          'Database=NM_Analytics_Prototype;'
                          'Trusted_Connection=yes;')

    cursor = conn.cursor()
    # iterate over df rows and insert into NM_Analytics_Prototype.vizient_qa.datadump
    for index, row in df.iterrows():
        if row['measure_value'] != 'Missing':
            cursor.execute(
                "INSERT INTO NM_Analytics_Prototype.vizient_qa.measure_values([calc_id],[hospital_id],[measure_id],[period_id],[measure_value_id],[measure_value],[numerator],[numerator_event_type_id],[denominator],[denominator_event_type_id]) values (?,?,?,?,?,?,?,?,?,?)",
                row['calc_id'], row['hospital_id'], row['measure_id'], row['period_id'], row['measure_value_id'],
                row['measure_value'], row['numerator'], row['numerator_event_type_id'], row['denominator'],
                row['denominator_event_type_id'])
            conn.commit()

    cursor.close()
    conn.close()
    print('done inserting rows.')



##########################################################################################################################

# Function to open core measure excel files (ED-2B & ED-OP-18B)

def open_core_measure_excel_file(path_obj, file_obj):
    file_loc2 = os.path.join(path_obj,file_obj)
    #UL007
    opened_excel_file = pd.read_excel(file_loc2, sheet_name=0,engine='openpyxl')
    return(opened_excel_file)

##########################################################################################################################

# Function to parse the folders/files of the core measures and return a tidy dataset (ED-2B & ED-OP-18B)

def core_measure_data_folder_walker_and_prep_for_db_inserts():
    # set the path of the folder structure you want to parse
    wd = input('Enter the path of the folder you want to parse.')
    wd = os.path.abspath(wd)
    # Set the path of the folder structure where we want the final files
    wd_dest = input('Enter the path of the folder you want to final files to go in.')
    wd_dest = os.path.abspath(wd_dest)

    # initialize empty list to store the core measure dataframes in.  Later this
    # will be used to quickly union the dataframes together.
    frames_list = []

    # only parse folders with ED_2B or ED_OP_18B in the name
    yes_parse = ['ED_2B', 'ED_OP_18B']
    # iterate through folder structure.
    for path, dirs, files in os.walk(wd):
        # if files list is not empty, and file is one of core measures open it up
        if len(files) > 0:
            # parse folder name to get measure name
            path_tail = str((os.path.basename(path)))
            # parse parent folder to get cohort name
            path_head = os.path.dirname(path)
            path_second_tail = str(os.path.basename(path_head))
            # use cohort and measure composite key to access dictionary values
            path_key = (path_second_tail, path_tail)
            if path_key[1] in yes_parse:
                # excel_file_data = open_excel_file(path, files)
                ed_df = open_core_measure_excel_file(path, files[0])
                # append core measure dataframe to the frames list
                frames_list.append(ed_df)

    # Once all the excel files have been read and the dataframes appended to the container list,
    # union them all together
    result_df = pd.concat(frames_list)

    # rename columns to match output of other measures so the convert_metric_df_to_insert_measure_value_df()
    # can be recycled to work on the result of this function as well.
    result_df = result_df.rename(columns={'Num Size': 'N Events', 'Denom Size': 'denominator'})

    # Core measure numerator event type is conditional.
    # ED-2B:  Sum of Admit Decision Time to ED Departure Time for Admitted Patients
    # ED-OP-18B: Sum of Time from ED Arrival to ED Departure for Discharged ED Patients

    result_df['event_type_nm'] = np.where(result_df['Measure'] == 'ED-2B',
                                          'Sum of Admit Decision Time to ED Departure Time for Admitted Patients',
                                          'Sum of Time from ED Arrival to ED Departure for Discharged ED Patients')

    # Core measure denominator event type is always 'Core Measure Population.'  Add this constant to the right column.
    result_df['denominator_event_type_nm'] = 'Core Measure Population'

    # remove duplicate rows
    result_df.drop_duplicates(inplace=True)

    # write a copy of the dataframe to the destination folder path
    final_result_path = os.path.join(wd_dest, 'core_meas_custom_time_period_data.csv')
    result_df.to_csv(final_result_path)
    return (result_df)

#####################################################################################################################
#####################################################################################################################
#####################################################################################################################
#####################################################################################################################

#function to query db and get measure ids and measure names.
def get_all_measures():
    # define the query
    all_measures = '''
        SELECT
        measure_id
        ,measure_name
        FROM
        NM_Analytics_Prototype.vizient_qa.measure
        '''
    # create the connection to the ms sql db
    conn = pyodbc.connect('Driver={SQL Server};'
                          'Server=edw00pd05wva.corp.nm.org\EDWIDS1;'
                          'Database=clarity;'
                          'Trusted_Connection=yes;')

    # read query results into pandas dataframe
    measure_df = pd.DataFrame(pd.read_sql(all_measures, conn))
    # close the db connection
    conn.close()
    # return the results dataframe of measure ids and names
    return (measure_df)

#####################################################################################################################

# function to parse hyperlink files and return just the measure name and keywords

def get_report_measure_keywords():
    #find the folder with Vizient calculator template hyperlinks and put file names in a list
    try:
        file_names = gather_hyperlink_files()
    except:
        print('Problem gathering hyperlink files.')
        return
    #print(file_names)
    #Get all unique report templates for all hospitals
    #initialize empty dataframe to store hyperlinks
    all_measures = pd.DataFrame()
    #iterate through list of files obtain from Vizient calculators and store measure names and keywords
    for ii, item in enumerate(file_names[0]):
        #open first sheet
        #UL007
        dataframe_ob = pd.DataFrame(pd.read_excel(item,sheet_name="Sheet1",engine='openpyxl'))
        #subset columns to only include measure formal name and keyword
        dataframe_ob = pd.DataFrame(dataframe_ob,columns=['Formal Name','Keyword/Metric'])
        #remove rows where formal name is 0
        dataframe_ob_indices1 =  dataframe_ob['Formal Name'] != 0
        dataframe_ob = dataframe_ob[dataframe_ob_indices1]
        all_measures = pd.concat([all_measures, dataframe_ob])
    #remove null rows
    all_measures_indices2 =  all_measures['Formal Name'].notnull()
    all_measures = all_measures[all_measures_indices2]
    #remove duplicate rows
    all_measures = all_measures.drop_duplicates()
    return(all_measures)

#####################################################################################################################

#Find the index of the header row of the Metric Weights sheet.
#takes an excel file path/filename as input and returns the index of the header row.
def find_ws_measure_weight_xls_header_row(xl_file_obj):
    #UL007
    xl_file_df = pd.read_excel(xl_file_obj, sheet_name="Metric Weights",header=None,engine='openpyxl')
    #get the name of the first column
    first_col_name = xl_file_df.columns[0]
    #get the index of the first row equal to 'Metric'
    first_row_loc = xl_file_df.index[xl_file_df[first_col_name] == 'Metric'].tolist()
    #return index
    return(first_row_loc[0])

#####################################################################################################################

# function to open the calculator files and read metric weights sheet
'''
def open_calc_measure_weight(path_obj, file_obj):
    #join the file path and the file name
    file_loc2 = os.path.abspath(os.path.join(path_obj,file_obj))
    #fine the index of the header row in order to know how many rows to skip
    measure_weight_header_row = find_ws_measure_weight_xls_header_row(file_loc2)
    #open the excel file 'Metric Weights' sheet.
    opened_excel_file = pd.read_excel(file_loc2, sheetname="Metric Weights",skiprows =measure_weight_header_row)
    #subset columns to only measure weights.
    opened_excel_file = opened_excel_file[['Metric', 'Metric weight \n(% of domain weight)','Metric weight \n(% of overall)']]
    return(opened_excel_file)
'''
#UL003

def open_calc_measure_weight(path_obj, file_obj):
    #join the file path and the file name
    file_loc2 = os.path.abspath(os.path.join(path_obj,file_obj))
    #fine the index of the header row in order to know how many rows to skip
    measure_weight_header_row = find_ws_measure_weight_xls_header_row(file_loc2)
    #open the excel file 'Metric Weights' sheet.
    #UL007
    opened_excel_file = pd.read_excel(file_loc2, sheet_name="Metric Weights",skiprows =measure_weight_header_row,engine='openpyxl')
    #subset columns to only measure weights.
    #find 3 columns:  Metric, Metric Weight (% of domain weight), Metric Weight (% of overall)
    #UL003
    keep_cols = [i for i in opened_excel_file.columns if (i.upper().startswith('METRIC') and 'DIRECTION' not in i.upper())]
    
    opened_excel_file = opened_excel_file[keep_cols]
    
    return(opened_excel_file)


#####################################################################################################################

# function to iterate a folder with Q&A calculators and return the file path and list of file names

def gather_calculator_path_and_files():
    try:
        wd = input('Enter file path for calculator files.')
        # change directory to directory with file.  abspath function normalizes the directory path.
        os.chdir(os.path.abspath(wd))
    except:
        print('Something is wrong with Vizient hyperlink excel file path.')
        return
    #gather folder files into a list.
    files = os.listdir(os.curdir)
    # Filter folder files to only include ''QACalculator_' excel files.
    files = [ii for ii in files if 'QACalculator_' in ii]
    #sometimes there is a strange prefix 'ghost' file in the folder.  Remove this if it exists.
    prefixes = ('~')
    for filename in files:
        if filename.startswith(prefixes):
            files.remove(filename)
    return(wd,files)

#####################################################################################################################

# function to iterate over calculator files, read in the measure weights, join to db ids and return a dataframe
# that matches the format of the measure_weight table in the db.
'''
def parse_calculator_measure_weights():
    #function takes the path of calculator file folder and returns the path and list of all file names
    path_and_files = gather_calculator_path_and_files()

    #query the database and get all measure ids
    measure_df = get_all_measures()

    #get measure keywords from the hyperlink files because the measure weights sheet only has keywords
    #while the database only has formal names of measures.
    all_measure_names = get_report_measure_keywords()

    #join the measure ids to the measure keyword dataframes
    measure_names_keys_ids = pd.merge(measure_df, all_measure_names, how='left', left_on=['measure_name'], right_on=['Formal Name'])

    #remove rows that didn't join.  This should only be measures not in use or 'informational only' measures.
    measure_names_keys_ids = measure_names_keys_ids[pd.notnull(measure_names_keys_ids['Formal Name'])]

    #store the file path in a variable
    calc_file_path = path_and_files[0]

    #empty list to store measure weight dataframes in.
    frames_list = []

    #iterate over calculator file names
    for i, item in enumerate(path_and_files[1]):
        #look up hospital ID in the database.
        #step one.  Get the medicare ID from the file name
        hospital_medicare_id = item.split('_')[1]
        #get period name from calc file name
        period_nm = item.split('_')[2].replace(" ","%")
        #get year from calc file name
        year = item.split('_')[3].split('.')[0]

        #define the calculator id query
        calc_id_query = "SELECT calc_id,calc_nm from NM_Analytics_Prototype.vizient_qa.calculator where calc_nm like '%" + year + "%" + period_nm + "'"
        #create connection to the ms sql database
        conn = pyodbc.connect('Driver={SQL Server};'
                              'Server=edw00pd05wva.corp.nm.org\EDWIDS1;'
                              'Database=clarity;'
                              'Trusted_Connection=yes;')
        #define the hospital query
        find_hosp_by_medicare_id_query = 'SELECT hospital_id, hospital_medicare_id from NM_Analytics_Prototype.vizient_qa.hospitals WHERE hospital_medicare_id = ' + hospital_medicare_id

        #query the database
        hospital_ids = pd.DataFrame(pd.read_sql(find_hosp_by_medicare_id_query ,conn))
        calc_ids = pd.DataFrame(pd.read_sql(calc_id_query ,conn))

        #close the connection
        conn.close()
        #store the hospital id in a variable
        current_hospital_id = hospital_ids['hospital_id'][0]
        current_calc_id = calc_ids['calc_id'][0]

        #open the calculator file to the 'Metric Weights' sheet.
        measure_weight_excel_file = open_calc_measure_weight(calc_file_path,item)

        #join calculator measure weights dataframe with dataframe with measure names and ids from db
        calc_weights_with_measure_ids = pd.merge(measure_weight_excel_file, measure_names_keys_ids, how='left', left_on=['Metric'], right_on=['Keyword/Metric'])

        #rename two columns to match the db
        calc_weights_with_measure_ids = calc_weights_with_measure_ids.rename(columns={'Metric weight \n(% of domain weight)': 'measure_wgt', 'Metric weight \n(% of overall)': 'measure_overall_wgt'})

        #append calc and hospital ids to the dataframe to match the db table
        calc_weights_with_measure_ids['calc_id'] = current_calc_id
        calc_weights_with_measure_ids['hosp_id'] = current_hospital_id

        #subset only columns needed for db inserts
        calc_weights_with_measure_ids = calc_weights_with_measure_ids[['calc_id', 'hosp_id','measure_id','measure_wgt','measure_overall_wgt']]

        #reorder columns to match db
        cols = ['calc_id', 'hosp_id','measure_id','measure_wgt','measure_overall_wgt']
        calc_weights_with_measure_ids = calc_weights_with_measure_ids[cols]

        #append dataframe to frames_list to union later
        frames_list.append(calc_weights_with_measure_ids)

    result_df = pd.concat(frames_list)
    result_df = result_df.drop_duplicates()
    return(result_df)
'''
#UL003
def parse_calculator_measure_weights():
    #function takes the path of calculator file folder and returns the path and list of all file names
    path_and_files = gather_calculator_path_and_files()

    #query the database and get all measure ids
    measure_df = get_all_measures()

    #get measure keywords from the hyperlink files because the measure weights sheet only has keywords
    #while the database only has formal names of measures.
    #all_measure_names = get_report_measure_keywords()

    #join the measure ids to the measure keyword dataframes
    #measure_names_keys_ids = pd.merge(measure_df, all_measure_names, how='left', left_on=['measure_name'], right_on=['Formal Name'])

    #remove rows that didn't join.  This should only be measures not in use or 'informational only' measures.
    #measure_names_keys_ids = measure_names_keys_ids[pd.notnull(measure_names_keys_ids['Formal Name'])]

    #store the file path in a variable
    calc_file_path = path_and_files[0]

    #empty list to store measure weight dataframes in.
    frames_list = []

    #iterate over calculator file names
    for i, item in enumerate(path_and_files[1]):
        
        #UL003
        #Adding additional logic to handle Critical Access calculators.  For some reason, Vizient decided to name the
        #Critical Access calculators completely differently so we need to use a different method.
        #first split up the file name, then iterate over the subsections of the name to find the part with the medicare id
        #All NM hospital medicare ids start with 140 or 141.
        for i, subitem in enumerate(item.split('_')):
            if '140' in subitem or '141' in subitem or '149' in subitem:
                hospital_medicare_id_index = i
            else:
                pass

        # look up hospital ID in the database.
        # step one.  Get the medicare ID from the file name
        hospital_medicare_id = item.split('_')[hospital_medicare_id_index]

        for i, subitem in enumerate(item.split('_')):
            if 'PERIOD' in subitem.upper() or 'ANNUAL' in subitem.upper():
                period_nm_index = i
            else:
                pass

        #have to split the period name because in 2019 there is a space between period and number.
        #In 2020, there 
        period_nm_split = re.split('(\d+)',item.split('_')[period_nm_index])

        if period_nm_split[0].replace(" ","") == 'Annual':
            period_str = 'Period'
            period_num = '4'
        else:
            period_str = period_nm_split[0].replace(" ","")
            period_num = period_nm_split[1].replace(" ","")

        #['Period ', '3', '']
        # get period name from calc file name
        #period_nm = item.split('_')[period_nm_index].replace(" ", "%")

        period_nm = period_str + '%' + period_num


        # get year from calc file name
        # The section containing the year will start with 20...for the next 80 years...
        # and it will end with .xlsm because it contains the file extension.
        for i, subitem in enumerate(item.split('_')):

            if subitem.startswith('20') and subitem.upper().endswith('.XLSM'):
                year_index = i
            else:
                pass

        
        year = item.split('_')[year_index].split('.')[0]

        '''

        hospital_medicare_id = item.split('_')[1]
        # get period name from calc file name
        period_nm = item.split('_')[2].replace(" ", "%")
        # get year from calc file name
        year = item.split('_')[3].split('.')[0]

        print(hospital_medicare_id)
        print(period_nm)
        print(year)
        '''

        # define the calculator id query
        calc_id_query = "SELECT calc_id,calc_nm from NM_Analytics_Prototype.vizient_qa.calculator where calc_nm like '%" + year + "%" + period_nm + "'"
        # create connection to the ms sql database
        conn = pyodbc.connect('Driver={SQL Server};'
                              'Server=edw00pd05wva.corp.nm.org\EDWIDS1;'
                              'Database=clarity;'
                              'Trusted_Connection=yes;')
        # define the hospital query
        find_hosp_by_medicare_id_query = 'SELECT hospital_id, hospital_medicare_id from NM_Analytics_Prototype.vizient_qa.hospitals WHERE hospital_medicare_id = ' + hospital_medicare_id

        # query the database
        hospital_ids = pd.DataFrame(pd.read_sql(find_hosp_by_medicare_id_query, conn))
        calc_ids = pd.DataFrame(pd.read_sql(calc_id_query, conn))

        # close the connection
        conn.close()


        # store the hospital id in a variable
        current_hospital_id = hospital_ids['hospital_id'][0]
        current_calc_id = calc_ids['calc_id'][0]
        
        ###############################################################
        #UL003
        #Since the metric weights sheet does not have the metric formal name, it only has the metric keyword,
        #we need a way to join back to the database (which only stores the formal name).  In the calculator
        #worksheet, there is a hidden column called 'Metric', which has the keyword and a column called 'Measure'
        #which has the formal name.  Use this to get the mapping.
        
        
        # open the calculator file and isolate the 'Calculator' sheet.
        ws = grab_calc_worksheet_calc_sheet(calc_file_path, item)

        #find the coordinates of the 'Domain' header cell of the 'Calculator' worksheet
        domain_coord = find_calc_phrase_cell_coords(ws,'Domain')
        
        #store worksheet object as pandas dataframe
        calc_df = pd.DataFrame(ws.values)

        #rename column headers using the 'Domain' header row coordinates
        calc_df = calc_df.rename(columns=calc_df.iloc[domain_coord[0]-1])

        #use row coordinates of "Domain" header row to drop all rows before it.
        row_drop_list = [i for i in range(domain_coord[0])]
        calc_df = calc_df.drop(row_drop_list)

        #drop colum where all values are NA
        calc_df = calc_df.dropna(axis='columns', how="all")

        #drop rows where all values are NA
        calc_df = calc_df.dropna(how="all")
        
        #remove copyright row at the bottom.
        calc_df = calc_df[~calc_df['Domain'].str.contains('Copyright', na=False)]
        
        #only need 2 columns
        calc_df = calc_df[['Metric',"Measure"]]
        
        #clean things up to ensure best join
        calc_df['Metric'] = calc_df['Metric'].str.replace(' ','').str.upper()
        calc_df['Measure'] = calc_df['Measure'].str.replace(' ','').str.replace('-','_').str.upper()
        calc_df.loc[calc_df["Measure"] == 'COLONSCOPYREVISITSWITHIN7_DAYS', "Measure"] = 'COLONOSCOPYREVISITSWITHIN7_DAYS'
        measure_df['measure_name'] = measure_df['measure_name'].str.replace(' ','').str.replace('-','_').str.upper()
        #join to measure_df dataframe to get db ids
        measure_names_keys_ids = pd.merge(measure_df, calc_df, how='inner', left_on=['measure_name'], right_on=['Measure'])
        
    
        #open the calculator file to the 'Metric Weights' sheet.
        measure_weight_excel_file = open_calc_measure_weight(calc_file_path,item)
        
        #UL003  cleaning metric names because Vizient's file sometimes uses all uppercase and sometimes capital-case.
        
        measure_weight_excel_file['Metric'] = measure_weight_excel_file['Metric'].str.replace(' ','').str.upper()
        

        #join calculator measure weights dataframe with dataframe with measure names and ids from db
        calc_weights_with_measure_ids = pd.merge(measure_weight_excel_file, measure_names_keys_ids, how='left', left_on=['Metric'], right_on=['Metric'])
        
        #rename two columns to match the db
        #calc_weights_with_measure_ids = calc_weights_with_measure_ids.rename(columns={'Metric weight \n(% of domain weight)': 'measure_wgt', 'Metric weight \n(% of overall)': 'measure_overall_wgt'})
        
        #UL003  different column naming conventions so we must rename all columns to match database names.
        calc_weights_with_measure_ids = calc_weights_with_measure_ids.rename(columns={calc_weights_with_measure_ids.columns[1]: 'measure_wgt', calc_weights_with_measure_ids.columns[2]: 'measure_overall_wgt'})
        #append calc and hospital ids to the dataframe to match the db table
        calc_weights_with_measure_ids['calc_id'] = current_calc_id
        calc_weights_with_measure_ids['hosp_id'] = current_hospital_id

        #subset only columns needed for db inserts
        calc_weights_with_measure_ids = calc_weights_with_measure_ids[['calc_id', 'hosp_id','measure_id','measure_wgt','measure_overall_wgt']]

        #reorder columns to match db
        cols = ['calc_id', 'hosp_id','measure_id','measure_wgt','measure_overall_wgt']
        calc_weights_with_measure_ids = calc_weights_with_measure_ids[cols]

        #append dataframe to frames_list to union later
        frames_list.append(calc_weights_with_measure_ids)

    result_df = pd.concat(frames_list)
    
    result_df = result_df[result_df['measure_id'].notnull()]
    
    result_df = result_df.drop_duplicates()
    return(result_df)

#####################################################################################################################

# function takes the results of parse_calculator_measure_weights() and
# inserts them into the measure_weight table.

def insert_measure_weights(df):
    # connect to the NM_Analytics database
    conn = pyodbc.connect('Driver={SQL Server};'
                          'Server=edw00pd05wva.corp.nm.org\EDWIDS1;'
                          'Database=NM_Analytics_Prototype;'
                          'Trusted_Connection=yes;')

    cursor = conn.cursor()
    # iterate over df rows and insert into NM_Analytics_Prototype.vizient_qa.datadump
    for index, row in df.iterrows():
        if row['measure_id'] != 'Missing':
            cursor.execute(
                "INSERT INTO NM_Analytics_Prototype.vizient_qa.measure_weight([calc_id],[hosp_id],[measure_id],[measure_wgt],[measure_overall_wgt]) values (?,?,?,?,?)",
                row['calc_id'], row['hosp_id'], row['measure_id'], row['measure_wgt'], row['measure_overall_wgt'])
            conn.commit()

    cursor.close()
    conn.close()
    print('done inserting rows.')


#####################################################################################################################
#####################################################################################################################
#####################################################################################################################
#####################################################################################################################

# Parse domain weights

# function to query domain ids and domain names

def get_all_domains():
    # define the query
    all_domains = '''
         SELECT
        domain_id
        ,domain_nm
        FROM
        NM_Analytics_Prototype.vizient_qa.domain
        '''
    # create the connection to the ms sql db
    conn = pyodbc.connect('Driver={SQL Server};'
                          'Server=edw00pd05wva.corp.nm.org\EDWIDS1;'
                          'Database=clarity;'
                          'Trusted_Connection=yes;')

    # read query results into pandas dataframe
    domain_df = pd.DataFrame(pd.read_sql(all_domains, conn))
    # close the db connection
    conn.close()
    # return the results dataframe of domain ids and names
    return (domain_df)

#####################################################################################################################

# function to open calculator 'Metric Weights' sheet and return domain and domain weight
'''
def open_calc_domain_weight(path_obj, file_obj):
    #join the file path and the file name
    file_loc2 = os.path.abspath(os.path.join(path_obj,file_obj))
    #fine the index of the header row in order to know how many rows to skip
    measure_weight_header_row = find_ws_measure_weight_xls_header_row(file_loc2)
    #open the excel file 'Metric Weights' sheet.
    opened_excel_file = pd.read_excel(file_loc2, sheetname="Metric Weights",skiprows =measure_weight_header_row)
    #subset columns to only measure weights.
    opened_excel_file = opened_excel_file[['Domain','Domain weight']]
    return(opened_excel_file)

'''

def open_calc_domain_weight(path_obj, file_obj):
    #join the file path and the file name
    file_loc2 = os.path.abspath(os.path.join(path_obj,file_obj))
    #fine the index of the header row in order to know how many rows to skip
    measure_weight_header_row = find_ws_measure_weight_xls_header_row(file_loc2)
    #open the excel file 'Metric Weights' sheet.
    #UL007
    opened_excel_file = pd.read_excel(file_loc2, sheet_name="Metric Weights",skiprows =measure_weight_header_row,engine='openpyxl')
    #subset columns to only measure weights.
    #UL003  Vizient changed column header in 2020 by using uppercase 'Weight' e.i. 'Domain Weight.'  Making this
    #more robust to look for all column headers starting with 'Domain'
    #opened_excel_file = opened_excel_file[['Domain','Domain weight']]
    keep_cols = [i for i in opened_excel_file.columns if i.upper().startswith('DOMAIN')]
    opened_excel_file = opened_excel_file[keep_cols]
    return(opened_excel_file)


#####################################################################################################################

# function to iterate over calculators and return domain weight data for db inserts
'''
def parse_calculator_domain_weights():
    #function takes the path of calculator file folder and returns the path and list of all file names
    path_and_files = gather_calculator_path_and_files()

    #query the database and get all domain ids
    domain_df = get_all_domains()

    #store the file path in a variable
    calc_file_path = path_and_files[0]

    #empty list to store measure weight dataframes in.
    frames_list = []

    #iterate over calculator file names
    for i, item in enumerate(path_and_files[1]):
        #look up hospital ID in the database.
        #step one.  Get the medicare ID from the file name
        hospital_medicare_id = item.split('_')[1]
        #get period name from calc file name
        period_nm = item.split('_')[2].replace(" ","%")
        #get year from calc file name
        year = item.split('_')[3].split('.')[0]

        #define the calculator id query
        calc_id_query = "SELECT calc_id,calc_nm from NM_Analytics_Prototype.vizient_qa.calculator where calc_nm like '%" + year + "%" + period_nm + "'"
        #create connection to the ms sql database
        conn = pyodbc.connect('Driver={SQL Server};'
                              'Server=edw00pd05wva.corp.nm.org\EDWIDS1;'
                              'Database=clarity;'
                              'Trusted_Connection=yes;')
        #define the hospital query
        find_hosp_by_medicare_id_query = 'SELECT hospital_id, hospital_medicare_id from NM_Analytics_Prototype.vizient_qa.hospitals WHERE hospital_medicare_id = ' + hospital_medicare_id

        #query the database
        hospital_ids = pd.DataFrame(pd.read_sql(find_hosp_by_medicare_id_query ,conn))
        calc_ids = pd.DataFrame(pd.read_sql(calc_id_query ,conn))

        #close the connection
        conn.close()
        #store the hospital id in a variable
        current_hospital_id = hospital_ids['hospital_id'][0]
        current_calc_id = calc_ids['calc_id'][0]

        #open the calculator file to the 'Metric Weights' sheet.
        domain_weight_excel_file = open_calc_domain_weight(calc_file_path,item)

        #join calculator domain weights dataframe with dataframe with domain names and ids from db
        calc_weights_with_domain_ids = pd.merge(domain_weight_excel_file, domain_df, how='left', left_on=['Domain'], right_on=['domain_nm'])

        #rename domain weight column to match the db
        calc_weights_with_domain_ids = calc_weights_with_domain_ids.rename(columns={'Domain weight': 'domain_wgt'})

        #append calc and hospital ids to the dataframe to match the db table
        calc_weights_with_domain_ids['calc_id'] = current_calc_id
        calc_weights_with_domain_ids['hosp_id'] = current_hospital_id

        #subset only columns needed for db inserts
        calc_weights_with_domain_ids = calc_weights_with_domain_ids[['calc_id', 'hosp_id','domain_id','domain_wgt']]

        #reorder columns to match db
        cols = ['calc_id', 'hosp_id','domain_id','domain_wgt']
        calc_weights_with_domain_ids = calc_weights_with_domain_ids[cols]

        #append dataframe to frames_list to union later
        frames_list.append(calc_weights_with_domain_ids)

    result_df = pd.concat(frames_list)
    result_df = result_df.drop_duplicates()
    return(result_df)
'''

def parse_calculator_domain_weights():
    #function takes the path of calculator file folder and returns the path and list of all file names
    path_and_files = gather_calculator_path_and_files()

    #query the database and get all domain ids
    domain_df = get_all_domains()

    #store the file path in a variable
    calc_file_path = path_and_files[0]

    #empty list to store measure weight dataframes in.
    frames_list = []

    #iterate over calculator file names
    for i, item in enumerate(path_and_files[1]):
        #UL003
        #Adding additional logic to handle Critical Access calculators.  For some reason, Vizient decided to name the
        #Critical Access calculators completely differently so we need to use a different method.
        #first split up the file name, then iterate over the subsections of the name to find the part with the medicare id
        #All NM hospital medicare ids start with 140 or 141.
        for i, subitem in enumerate(item.split('_')):
            if '140' in subitem or '141' in subitem or '149' in subitem:
                hospital_medicare_id_index = i
            else:
                pass

        # look up hospital ID in the database.
        # step one.  Get the medicare ID from the file name
        hospital_medicare_id = item.split('_')[hospital_medicare_id_index]

        for i, subitem in enumerate(item.split('_')):
            if 'PERIOD' in subitem.upper() or 'ANNUAL' in subitem.upper():
                period_nm_index = i
            else:
                pass

        #have to split the period name because in 2019 there is a space between period and number.
        #In 2020, there 
        period_nm_split = re.split('(\d+)',item.split('_')[period_nm_index])

        if period_nm_split[0].replace(" ","") == 'Annual':
            period_str = 'Period'
            period_num = '4'
        else:
            period_str = period_nm_split[0].replace(" ","")
            period_num = period_nm_split[1].replace(" ","")

        #['Period ', '3', '']
        # get period name from calc file name
        #period_nm = item.split('_')[period_nm_index].replace(" ", "%")

        period_nm = period_str + '%' + period_num


        # get year from calc file name
        # The section containing the year will start with 20...for the next 80 years...
        # and it will end with .xlsm because it contains the file extension.
        for i, subitem in enumerate(item.split('_')):

            if subitem.startswith('20') and subitem.upper().endswith('.XLSM'):
                year_index = i
            else:
                pass

        
        year = item.split('_')[year_index].split('.')[0]

        '''

        hospital_medicare_id = item.split('_')[1]
        # get period name from calc file name
        period_nm = item.split('_')[2].replace(" ", "%")
        # get year from calc file name
        year = item.split('_')[3].split('.')[0]

        print(hospital_medicare_id)
        print(period_nm)
        print(year)
        '''

        # define the calculator id query
        calc_id_query = "SELECT calc_id,calc_nm from NM_Analytics_Prototype.vizient_qa.calculator where calc_nm like '%" + year + "%" + period_nm + "'"
        # create connection to the ms sql database
        conn = pyodbc.connect('Driver={SQL Server};'
                              'Server=edw00pd05wva.corp.nm.org\EDWIDS1;'
                              'Database=clarity;'
                              'Trusted_Connection=yes;')
        # define the hospital query
        find_hosp_by_medicare_id_query = 'SELECT hospital_id, hospital_medicare_id from NM_Analytics_Prototype.vizient_qa.hospitals WHERE hospital_medicare_id = ' + hospital_medicare_id

        # query the database
        hospital_ids = pd.DataFrame(pd.read_sql(find_hosp_by_medicare_id_query, conn))
        calc_ids = pd.DataFrame(pd.read_sql(calc_id_query, conn))

        # close the connection
        conn.close()


        # store the hospital id in a variable
        current_hospital_id = hospital_ids['hospital_id'][0]
        current_calc_id = calc_ids['calc_id'][0]

        #open the calculator file to the 'Metric Weights' sheet.
        domain_weight_excel_file = open_calc_domain_weight(calc_file_path,item)

        #join calculator domain weights dataframe with dataframe with domain names and ids from db
        calc_weights_with_domain_ids = pd.merge(domain_weight_excel_file, domain_df, how='left', left_on=['Domain'], right_on=['domain_nm'])
        
    
        #rename domain weight column to match the db
        #calc_weights_with_domain_ids = calc_weights_with_domain_ids.rename(columns={'Domain weight': 'domain_wgt'})
        #UL003 commented out above because Vizient 2020 uses uppercase 'Weight' in header name so I'm making this
        #more robust.
        domain_weight_col_header = [i for i in calc_weights_with_domain_ids.columns if i.upper().startswith('DOMAIN') and i.upper().endswith('WEIGHT')][0]
        
        calc_weights_with_domain_ids = calc_weights_with_domain_ids.rename(columns={domain_weight_col_header: 'domain_wgt'})
        
        #append calc and hospital ids to the dataframe to match the db table
        calc_weights_with_domain_ids['calc_id'] = current_calc_id
        calc_weights_with_domain_ids['hosp_id'] = current_hospital_id

        #subset only columns needed for db inserts
        calc_weights_with_domain_ids = calc_weights_with_domain_ids[['calc_id', 'hosp_id','domain_id','domain_wgt']]

        #reorder columns to match db
        cols = ['calc_id', 'hosp_id','domain_id','domain_wgt']
        calc_weights_with_domain_ids = calc_weights_with_domain_ids[cols]

        #append dataframe to frames_list to union later
        frames_list.append(calc_weights_with_domain_ids)

    result_df = pd.concat(frames_list)
    
    result_df = result_df[result_df['domain_id'].notnull()]
    
    result_df = result_df.drop_duplicates()
    return(result_df)


#####################################################################################################################

def insert_domain_weights(df):
    # connect to the NM_Analytics database
    conn = pyodbc.connect('Driver={SQL Server};'
                          'Server=edw00pd05wva.corp.nm.org\EDWIDS1;'
                          'Database=NM_Analytics_Prototype;'
                          'Trusted_Connection=yes;')

    cursor = conn.cursor()

    for index, row in df.iterrows():
        if row['domain_id'] != 'Missing':
            cursor.execute(
                "INSERT INTO NM_Analytics_Prototype.vizient_qa.domain_weight([calc_id],[hosp_id],[domain_id],[domain_wgt]) values (?,?,?,?)",
                row['calc_id'], row['hosp_id'], row['domain_id'], row['domain_wgt'])
            conn.commit()

    cursor.close()
    conn.close()
    print('done inserting rows.')


#####################################################################################################################

def grab_calc_worksheet_calc_sheet(path_obj,file_obj):
    file_loc = os.path.abspath(os.path.join(path_obj,file_obj))
    wb = openpyxl.load_workbook(file_loc, data_only = True)
    ws = wb['Calculator']
    return(ws)


#####################################################################################################################

#function takes an openpyxl worksheet object and returns the coordinates of
#the  cell value. It finds the header phrase provided.
def find_calc_phrase_cell_coords(worksheet_obj,phrase):
    #Find phrase cell coordinates.
    #iterate over all worksheet cells until you find the desired phrase
    for row in worksheet_obj.iter_rows():
        for cell in row:
                #iterate over worksheet cells until you find the first instance
                #of the phrase you are looking for.
                if cell.value == phrase:
                    return([cell.row,cell.column,phrase])


#####################################################################################################################
#UL003  New function
#this version of the above function will return a list of lists 
#of all instances of the desired phrase in a dataset.  Not just the first one.
def find_calc_phrase_cell_coords_multiple(worksheet_obj,phrase):
    #Find phrase cell coordinates.
    #iterate over all worksheet cells until you find the desired phrase
    cell_coord_list = []
    for row in worksheet_obj.iter_rows():
        for cell in row:
                #iterate over worksheet cells until you find the first instance
                #of the phrase you are looking for.
                if cell.value == phrase:
                    cell_coord_list.append([cell.row,cell.column,phrase])
    return(cell_coord_list)        

#####################################################################################################################

# function returns the value of a calculator hospital rank.
# function takes an openpyxl calculator worksheet object
# iterates over rows and columns until it finds desired value.  Then
# returns the desired hospital rank.  If rank type is 'Possible Rank'
# then function returns value at one row below the header phrase.  If rank type is
# 'Target Ranking: ', then it returns the value at one column to the right
# of the header phrase.
# ranks_coord_list should have 3 values:  rank header cell row, rank header cell column, rank header phrase.
def find_hosp_rank_value(worksheet_obj, rank_coords_list):
    if rank_coords_list[2] == 'Possible Rank':
        # if looking for 'Possible Rank', return value at 'Possible Rank' cell row + 1 (one below)
        return (worksheet_obj.cell(row=rank_coords_list[0] + 1, column=rank_coords_list[1]).value)

    elif rank_coords_list[2] == 'Target Ranking: ':
        # if looking for 'Target Ranking: ', return value at 'Target Ranking: ' cell column + 1 (one to the right)
        return (worksheet_obj.cell(row=rank_coords_list[0], column=rank_coords_list[1] + 1).value)


#####################################################################################################################

#function to parse calculators and return dataframe for db inserts of hospital ranks
'''
def parse_calculator_hosp_ranks(hosp_rank_phase, period_type, period_end_dts):
    # assign db measure value name based on calculator hospital rank search phrase
    try:
        if hosp_rank_phase == 'Possible Rank':
            measure_value_type = 'Hospital Rank'
        elif hosp_rank_phase == 'Target Ranking: ':
            measure_value_type = 'Top Decile Target Hospital Ranking'
    except:
        print('Hospital Rank phrase does not match.')
        return

    # function takes the path of calculator file folder and returns the path and list of all file names
    path_and_files = gather_calculator_path_and_files()

    # store the file path in a variable
    calc_file_path = path_and_files[0]

    # empty list to store measure weight dataframes in.
    frames_list = []

    # query the db to get period id
    conn = pyodbc.connect('Driver={SQL Server};'
                          'Server=edw00pd05wva.corp.nm.org\EDWIDS1;'
                          'Database=clarity;'
                          'Trusted_Connection=yes;')
    period_type_id = """
    SELECT
    pt.period_type_id
    FROM
    NM_Analytics_Prototype.vizient_qa.period_types as pt
    WHERE
    pt.period_type_nm = '%s'
    """ % period_type

    period_type_id_result = pd.DataFrame(pd.read_sql(period_type_id, conn))['period_type_id'].values[0]

    period_id = '''''''
    SELECT 
    p.period_id 
    FROM NM_Analytics_Prototype.vizient_qa.periods as p 
    WHERE p.period_end_date = '%s'
    and p.period_type = '%s'
    ''' '''% (period_end_dts, period_type_id_result)

    period_id_result = pd.DataFrame(pd.read_sql(period_id, conn))['period_id'].values[0]

    measure_value_id = """
    SELECT
    measure_value_id
    from
    NM_Analytics_Prototype.vizient_qa.measure_value_types
    WHERE
    measure_value_name = '%s'
    """ % measure_value_type

    measure_value_id_result = pd.DataFrame(pd.read_sql(measure_value_id, conn))['measure_value_id'].values[0]

    conn.close()

    # iterate over calculator file names
    for i, item in enumerate(path_and_files[1]):
        # look up hospital ID in the database.
        # step one.  Get the medicare ID from the file name
        hospital_medicare_id = item.split('_')[1]
        # get period name from calc file name
        period_nm = item.split('_')[2].replace(" ", "%")
        # get year from calc file name
        year = item.split('_')[3].split('.')[0]

        # define the calculator id query
        calc_id_query = "SELECT calc_id,calc_nm from NM_Analytics_Prototype.vizient_qa.calculator where calc_nm like '%" + year + "%" + period_nm + "'"
        # create connection to the ms sql database
        conn = pyodbc.connect('Driver={SQL Server};'
                              'Server=edw00pd05wva.corp.nm.org\EDWIDS1;'
                              'Database=clarity;'
                              'Trusted_Connection=yes;')
        # define the hospital query
        find_hosp_by_medicare_id_query = 'SELECT hospital_id, hospital_medicare_id from NM_Analytics_Prototype.vizient_qa.hospitals WHERE hospital_medicare_id = ' + hospital_medicare_id

        # query the database
        hospital_ids = pd.DataFrame(pd.read_sql(find_hosp_by_medicare_id_query, conn))
        calc_ids = pd.DataFrame(pd.read_sql(calc_id_query, conn))

        # close the connection
        conn.close()
        # store the hospital id in a variable
        current_hospital_id = hospital_ids['hospital_id'][0]
        current_calc_id = calc_ids['calc_id'][0]

        # open the calculator file and isolate the 'Calculator' sheet.
        ws = grab_calc_worksheet_calc_sheet(calc_file_path, item)

        # find and return the cell row and cell column coordinates of the hospital rank search phrase
        rank_header_loc = find_calc_phrase_cell_coords(ws, hosp_rank_phase)

        # using result of find_calc_phrase_cell_coords() function, return hospital rank value
        rank_val = find_hosp_rank_value(ws, rank_header_loc)

        # create result list of all db-ready ids plus hospital rank value
        row_list = [current_calc_id, current_hospital_id, period_id_result, measure_value_id_result, rank_val]
        row_df = pd.DataFrame(columns=['calc_id', 'hospital_id', 'period_id', 'measure_value_id', 'hospital_value'])
        row_df.loc[0] = row_list
        frames_list.append(row_df)

    result_df = pd.concat(frames_list)
    result_df = result_df.drop_duplicates()
    return (result_df)
'''

#UL003  Updated the above function to handle strange critical access calculators.
# Basically, the critical access hospital calculators do not display possible rank
# or target rank so you have to conditionally look these up a hidden worksheet.

#updated function  UL003 udpate.
#updated function  UL003 udpate.
def parse_calculator_hosp_ranks(hosp_rank_phase, period_type, period_end_dts):
    # assign db measure value name based on calculator hospital rank search phrase
    try:
        if hosp_rank_phase == 'Possible Rank':
            measure_value_type = 'Hospital Rank'
        elif hosp_rank_phase == 'Target Ranking: ':
            measure_value_type = 'Top Decile Target Hospital Ranking'
    except:
        print('Hospital Rank phrase does not match.')
        #return

    # function takes the path of calculator file folder and returns the path and list of all file names
    path_and_files = gather_calculator_path_and_files()

    # store the file path in a variable
    calc_file_path = path_and_files[0]
    
    # empty list to store measure weight dataframes in.
    frames_list = []

    # query the db to get period id
    conn = pyodbc.connect('Driver={SQL Server};'
                          'Server=edw00pd05wva.corp.nm.org\EDWIDS1;'
                          'Database=clarity;'
                          'Trusted_Connection=yes;')
    period_type_id = """
    SELECT
    pt.period_type_id
    FROM
    NM_Analytics_Prototype.vizient_qa.period_types as pt
    WHERE
    pt.period_type_nm = '%s'
    """ % period_type

    period_type_id_result = pd.DataFrame(pd.read_sql(period_type_id, conn))['period_type_id'].values[0]

    period_id = '''
    SELECT 
    p.period_id 
    FROM NM_Analytics_Prototype.vizient_qa.periods as p 
    WHERE p.period_end_date = '%s'
    and p.period_type = '%s'
    ''' % (period_end_dts, period_type_id_result)

    period_id_result = pd.DataFrame(pd.read_sql(period_id, conn))['period_id'].values[0]

    measure_value_id = """
    SELECT
    measure_value_id
    from
    NM_Analytics_Prototype.vizient_qa.measure_value_types
    WHERE
    measure_value_name = '%s'
    """ % measure_value_type

    measure_value_id_result = pd.DataFrame(pd.read_sql(measure_value_id, conn))['measure_value_id'].values[0]

    conn.close()
    
    # iterate over calculator file names
    for i, item in enumerate(path_and_files[1]):

        #UL003
        #Adding additional logic to handle Critical Access calculators.  For some reason, Vizient decided to name the
        #Critical Access calculators completely differently so we need to use a different method.
        #first split up the file name, then iterate over the subsections of the name to find the part with the medicare id
        #All NM hospital medicare ids start with 140 or 141.
        for i, subitem in enumerate(item.split('_')):
            if '140' in subitem or '141' in subitem or '149' in subitem:
                hospital_medicare_id_index = i
            else:
                pass

        # look up hospital ID in the database.
        # step one.  Get the medicare ID from the file name
        hospital_medicare_id = item.split('_')[hospital_medicare_id_index]

        for i, subitem in enumerate(item.split('_')):
            if 'PERIOD' in subitem.upper() or 'ANNUAL' in subitem.upper():
                period_nm_index = i
            else:
                pass

        #have to split the period name because in 2019 there is a space between period and number.
        #In 2020, there 
        period_nm_split = re.split('(\d+)',item.split('_')[period_nm_index])
        

        if period_nm_split[0].replace(" ","").upper() == 'ANNUAL':
            period_str = 'Period'
            period_num = '4'
        else:
            period_str = period_nm_split[0].replace(" ","")
            period_num = period_nm_split[1].replace(" ","")

        #['Period ', '3', '']
        # get period name from calc file name
        #period_nm = item.split('_')[period_nm_index].replace(" ", "%")

        period_nm = period_str + '%' + period_num


        # get year from calc file name
        # The section containing the year will start with 20...for the next 80 years...
        # and it will end with .xlsm because it contains the file extension.
        for i, subitem in enumerate(item.split('_')):
            if subitem.startswith('20') and subitem.upper().endswith('.XLSM'):
                year_index = i
            else:
                pass


        year = item.split('_')[year_index].split('.')[0]

        '''

        hospital_medicare_id = item.split('_')[1]
        # get period name from calc file name
        period_nm = item.split('_')[2].replace(" ", "%")
        # get year from calc file name
        year = item.split('_')[3].split('.')[0]

        print(hospital_medicare_id)
        print(period_nm)
        print(year)
        '''

        # define the calculator id query
        calc_id_query = "SELECT calc_id,calc_nm from NM_Analytics_Prototype.vizient_qa.calculator where calc_nm like '%" + year + "%" + period_nm + "'"
        # create connection to the ms sql database
        conn = pyodbc.connect('Driver={SQL Server};'
                              'Server=edw00pd05wva.corp.nm.org\EDWIDS1;'
                              'Database=clarity;'
                              'Trusted_Connection=yes;')
        # define the hospital query
        find_hosp_by_medicare_id_query = 'SELECT hospital_id, hospital_medicare_id from NM_Analytics_Prototype.vizient_qa.hospitals WHERE hospital_medicare_id = ' + hospital_medicare_id

        # query the database
        hospital_ids = pd.DataFrame(pd.read_sql(find_hosp_by_medicare_id_query, conn))
        calc_ids = pd.DataFrame(pd.read_sql(calc_id_query, conn))

        # close the connection
        conn.close()


        # store the hospital id in a variable
        current_hospital_id = hospital_ids['hospital_id'][0]
        current_calc_id = calc_ids['calc_id'][0]


        # open the calculator file and isolate the 'Calculator' sheet.
        ws = grab_calc_worksheet_calc_sheet(calc_file_path, item)

        file_item = item

        #Valley West calculators do not have Possible Rank or Target Rank on the main page for some reason.  Thanks Vizient.
        #Therefore, we need to sum up the metric score and lookup the closes rank on the hidden ranks worksheet.
        if current_hospital_id == 356:  #Valley West db ID.

            if hosp_rank_phase == 'Target Ranking: ':
                ca_file_loc = os.path.abspath(os.path.join(calc_file_path,file_item))
                ca_wb = openpyxl.load_workbook(ca_file_loc, data_only = True)
                ranks_df = pd.DataFrame(ca_wb['CurrentQA Cumulative-Rank'].values)
                ##in openpyxl, the headers are listed as numbers.  Renaming the headers using the first row values.
                ranks_df = ranks_df.rename(columns=ranks_df.iloc[0])
                ##drop the first row because they are now the header
                ranks_df = ranks_df.drop([0])
                
                if 'parent_id' in ranks_df.columns:
                    ranks_df = ranks_df.rename(columns={"parent_id": "Parent_ID"})
                
                ##drop all rows that do not have a hospital value
                df_all_row_indices =  ranks_df['Parent_ID'].notnull()
                ranks_df = ranks_df[df_all_row_indices]

                ##only save specific columns we need
                ranks_df = ranks_df[['Parent_ID', 'HCO_SHORT_NAME', 'final_score','final_Rank']]

                #count the number of hospitals in the cohort

                cohort_num = len(ranks_df['Parent_ID'].unique())
                #target ranking is always top decile.  So, take the cohort number divide by 10 and round the result.
                top_decile_target_ranking = round(((cohort_num*1.0)/10.0))

                #make sure the dataframe is still sort in ascending order.
                ranks_df = ranks_df.sort_values(by=['final_score'],ascending=False)

                rank_val = ranks_df['final_Rank'].iloc[top_decile_target_ranking-1]

                # create result list of all db-ready ids plus hospital rank value
                row_list = [current_calc_id, current_hospital_id, period_id_result, measure_value_id_result, rank_val]
                row_df = pd.DataFrame(columns=['calc_id', 'hospital_id', 'period_id', 'measure_value_id', 'hospital_value'])
                row_df.loc[0] = row_list
                frames_list.append(row_df)
            else:
                #find the coordinates of the 'Domain' header cell of the 'Calculator' worksheet
                domain_coord = find_calc_phrase_cell_coords(ws,'Domain')
                #find the coordinates of the 'Target Domain Ranking' header cell of the 'Calculator' worksheet
                what_if_metric_value_coord = find_calc_phrase_cell_coords(ws,' "What if" Performance Evaluation')

                #set row coordinate to 2 below actual in order to locate the 'metric value' header.
                what_if_metric_value_coord[0] = what_if_metric_value_coord[0] + 2

                #Find the LAST instance of 'Domain Rank'.  Get a list of all instances, then take last index.
                domain_rank_coord = find_calc_phrase_cell_coords_multiple(ws,'Domain Rank')
                domain_rank_coord = domain_rank_coord[-1]
                #store worksheet object as pandas dataframe
                m_df = pd.DataFrame(ws.values)

                #use column coordinates of "Domain" header to drop all columns to the left of it.
                col_drop_list1 = [i for i in range(domain_rank_coord[1],m_df.shape[1])]
                m_df = m_df.drop(col_drop_list1,axis=1)

                #use column coordinates of "Domain" header to drop all columns to the left of it.
                col_drop_list = [i for i in range(domain_coord[1]-1)]
                m_df = m_df.drop(col_drop_list,axis=1)

                #rename column headers using the 'Domain' header row coordinates
                m_df = m_df.rename(columns=m_df.iloc[domain_coord[0]-1])

                #use row coordinates of "Domain" header row to drop all rows before it.
                row_drop_list = [i for i in range(domain_coord[0])]
                m_df = m_df.drop(row_drop_list)

                #drop colum where all values are NA
                m_df = m_df.dropna(axis='columns', how="all")

                #drop rows where all values are NA
                m_df = m_df.dropna(how="all")

                domain_and_measure_indices = [0,1]

                #indexes have now changed.  Find the last Domain Rank column header.
                domain_rank_list = []
                for i, thing in enumerate(m_df.columns):
                    if thing == 'Domain Rank':
                        domain_rank_list.append(i)
                #find last 'Domain Rank.'
                new_domain_rank_coord = domain_rank_list[-1]
                #keep_cols = [i for i in range(df.columns.get_loc(' Domain Rank Result')-4,df.columns.get_loc('Target Domain Ranking')+1)]
                keep_cols = [i for i in range(new_domain_rank_coord-4,new_domain_rank_coord+1)]

                #subset columns to keep only the ones we care about.
                final_keep_cols = domain_and_measure_indices + keep_cols
                m_df = m_df.iloc[:,final_keep_cols]

                #remove copyright row at the bottom.
                m_df = m_df[~m_df['Domain'].str.contains('Copyright', na=False)]

                #domain names and ranks are in merged excel rows so we now need to forward fill them so we have a value every row.
                m_df['Domain'] = m_df['Domain'].fillna(method='ffill')
                m_df['Domain Rank'] = m_df['Domain Rank'].fillna(method='ffill')

                #remove 'LV' rows.
                m_df = m_df[m_df['% of Overall Score'] != '-']

                overall_score = m_df['% of Overall Score'].astype('float').sum()
                overall_score = round(overall_score,6)
                #Now that we have the overall score from the calculator worksheet, we have to 
                #Perform a rank lookup on the CurrentQA Cumulative-Rank worksheet to see 
                #what rank the score would correspond to.  We only have to do this for the Critical Access
                #calculators for now because Vizient is being a pain in the ass.  They may activate the macros/vlookups
                #in the future and this step will not be needed anymore.

                ca_file_loc = os.path.abspath(os.path.join(calc_file_path,file_item))
                ca_wb = openpyxl.load_workbook(ca_file_loc, data_only = True)
                ranks_df = pd.DataFrame(ca_wb['CurrentQA Cumulative-Rank'].values)
                ##in openpyxl, the headers are listed as numbers.  Renaming the headers using the first row values.
                ranks_df = ranks_df.rename(columns=ranks_df.iloc[0])
                ##drop the first row because they are now the header
                ranks_df = ranks_df.drop([0])
                
                
                if 'parent_id' in ranks_df.columns:
                    ranks_df = ranks_df.rename(columns={"parent_id": "Parent_ID"})
                
                ##drop all rows that do not have a hospital value
                df_all_row_indices =  ranks_df['Parent_ID'].notnull()
                ranks_df = ranks_df[df_all_row_indices]
                ##only save specific columns we need
                ranks_df = ranks_df[['Parent_ID', 'HCO_SHORT_NAME', 'final_score','final_Rank']]
                print('overall_score:',overall_score)
                
                ranks_df['final_score'] = ranks_df['final_score'].astype('float')
                ranks_df['final_score'] = ranks_df['final_score'].round(6)
                #find the first index row that has a value less than or equal to the current overall score.
                #because, in order to get a rank, you have to beat or match another hospital.
                score_index = ranks_df[ranks_df['final_score'].le(overall_score)].index[0]
                #now retrieve the final rank associated with the score index.
                rank_val = ranks_df['final_Rank'].iloc[score_index-1]
                #print(ranks_df['final_score'].head(n=15))
                # create result list of all db-ready ids plus hospital rank value
                row_list = [current_calc_id, current_hospital_id, period_id_result, measure_value_id_result, rank_val]
                row_df = pd.DataFrame(columns=['calc_id', 'hospital_id', 'period_id', 'measure_value_id', 'hospital_value'])
                row_df.loc[0] = row_list
                frames_list.append(row_df)
        else:
            # find and return the cell row and cell column coordinates of the hospital rank search phrase
            rank_header_loc = find_calc_phrase_cell_coords(ws, hosp_rank_phase)

            # using result of find_calc_phrase_cell_coords() function, return hospital rank value
            rank_val = find_hosp_rank_value(ws, rank_header_loc)

            # create result list of all db-ready ids plus hospital rank value
            row_list = [current_calc_id, current_hospital_id, period_id_result, measure_value_id_result, rank_val]
            row_df = pd.DataFrame(columns=['calc_id', 'hospital_id', 'period_id', 'measure_value_id', 'hospital_value'])
            row_df.loc[0] = row_list
            frames_list.append(row_df)
    result_df = pd.concat(frames_list)
    result_df = result_df.drop_duplicates()
    return (result_df)


#############################################################################################################################

def insert_hospital_ranks(df):
    # connect to the NM_Analytics database
    conn = pyodbc.connect('Driver={SQL Server};'
                          'Server=edw00pd05wva.corp.nm.org\EDWIDS1;'
                          'Database=NM_Analytics_Prototype;'
                          'Trusted_Connection=yes;')

    cursor = conn.cursor()

    for index, row in df.iterrows():
        if row['hospital_value'] != 'Missing' and row['hospital_value'] is not None:
            cursor.execute(
                "INSERT INTO NM_Analytics_Prototype.vizient_qa.calc_hospital_values([calc_id],[hospital_id],[period_id],[measure_value_id],[hospital_value]) values (?,?,?,?,?)",
                row['calc_id'], row['hospital_id'], row['period_id'], row['measure_value_id'], row['hospital_value'])
            conn.commit()

    cursor.close()
    conn.close()
    print('done inserting rows.')



##############################################################################################################################


# function used to parse domain ranks and measure values from the calculators.

# function that takes a calculator file path and name
# and returns a pandas dataframe of just the domain name, measure name,
# "What if" section, "Current ranking" section and "Target Performance Evaluation" section
# Basically, this is used to put the calculator into an easy-to-wrangle format which
# only includes the sections we care about.

def parse_calculator_and_return_clean_dataframe(path_obj, file_obj):
    # take the path and file name, join them together and isolate the 'Calculator' worksheet
    ws = grab_calc_worksheet_calc_sheet(path_obj, file_obj)

    # find the coordinates of the 'Domain' header cell of the 'Calculator' worksheet
    domain_coord = find_calc_phrase_cell_coords(ws, 'Domain')

    # find the coordinates of the 'Target Domain Ranking' header cell of the 'Calculator' worksheet
    target_domain_coord = find_calc_phrase_cell_coords(ws, 'Target Domain Ranking')

    # store worksheet object as pandas dataframe
    df = pd.DataFrame(ws.values)

    # use column coordinates of "Domain" header to drop all columns to the left of it.
    col_drop_list = [i for i in range(domain_coord[1] - 1)]
    df = df.drop(col_drop_list, axis=1)

    # rename column headers using the 'Domain' header row coordinates
    df = df.rename(columns=df.iloc[domain_coord[0] - 1])

    # use row coordinates of "Domain" header row to drop all rows before it.
    row_drop_list = [i for i in range(domain_coord[0])]
    df = df.drop(row_drop_list)

    # drop colum where all values are NA
    df = df.dropna(axis='columns', how="all")

    # drop rows where all values are NA
    df = df.dropna(how="all")

    # forward fill domain name, domain rank result and target domain ranking columns because
    # the values only appear on one row per domain group causing NAs.  For easier parsing, just repeat for each line.
    df['Domain'] = df['Domain'].fillna(method='ffill')
    df[' Domain Rank Result'] = df[' Domain Rank Result'].fillna(method='ffill')
    df['Target Domain Ranking'] = df['Target Domain Ranking'].fillna(method='ffill')
    df['Current Domain Ranking'] = df['Current Domain Ranking'].fillna(method='ffill')

    # find the range of columns between "Target Domain Ranking" column and the "Metric Value" column
    # in the "What if" Performance Evaluation section.  Unfortunately, there are several "Metric Value"
    # columns so, we must first find the index of "Domain Rank Result", then subtract 4 columns positions.
    # want to keep first two columns so "Domain", and "Measure" are grabbed
    domain_and_measure_indices = [0, 1]
    keep_cols = [i for i in
                 range(df.columns.get_loc(' Domain Rank Result') - 4, df.columns.get_loc('Target Domain Ranking') + 1)]
    final_keep_cols = domain_and_measure_indices + keep_cols
    #df = df[final_keep_cols]
    df = df.iloc[:,final_keep_cols]

    #print(df.columns)
    # remove copyright row at the bottom.
    df = df[~df['Domain'].str.contains('Copyright', na=False)]

    #UL003  
    # remove rows that have "No Rank" in the domain rank section.  This means they are greyed out and not used.
    #df = df[df[' Domain Rank Result'] != 'No Rank']
    return (df)

##############################################################################################################################
##############################################################################################################################


#function to run final parse of domain ranks from the calculators.
# this can be used to get the hospital's "what if" section domain ranks or the target hospital's ranks.

# function takes a
'''
def parse_calculator_domain_ranks(domain_rank_phase, period_type, period_end_dts):
    # assign db measure value name based on calculator hospital rank search phrase
    try:
        if domain_rank_phase == 'Domain Rank Result':
            measure_value_type = 'Domain Rank'
        elif domain_rank_phase == 'Target Domain Ranking':
            measure_value_type = 'Top Decile Target Domain Ranking'
    except:
        print('Domain Rank phrase does not match.')
        return

    # query the database and get all domain ids
    domain_df = get_all_domains()

    # function takes the path of calculator file folder and returns the path and list of all file names
    path_and_files = gather_calculator_path_and_files()

    # store the file path in a variable
    calc_file_path = path_and_files[0]

    # empty list to store measure weight dataframes in.
    frames_list = []

    # query the db to get period id
    conn = pyodbc.connect('Driver={SQL Server};'
                          'Server=edw00pd05wva.corp.nm.org\EDWIDS1;'
                          'Database=clarity;'
                          'Trusted_Connection=yes;')
    period_type_id = """
    SELECT
    pt.period_type_id
    FROM
    NM_Analytics_Prototype.vizient_qa.period_types as pt
    WHERE
    pt.period_type_nm = '%s'
    """ % period_type

    period_type_id_result = pd.DataFrame(pd.read_sql(period_type_id, conn))['period_type_id'].values[0]

    period_id = '''''''
    SELECT 
    p.period_id 
    FROM NM_Analytics_Prototype.vizient_qa.periods as p 
    WHERE p.period_end_date = '%s'
    and p.period_type = '%s'
    ''''''' % (period_end_dts, period_type_id_result)

    period_id_result = pd.DataFrame(pd.read_sql(period_id, conn))['period_id'].values[0]

    measure_value_id = """
    SELECT
    measure_value_id
    from
    NM_Analytics_Prototype.vizient_qa.measure_value_types
    WHERE
    measure_value_name = '%s'
    """ % measure_value_type

    measure_value_id_result = pd.DataFrame(pd.read_sql(measure_value_id, conn))['measure_value_id'].values[0]

    conn.close()

    # iterate over calculator file names
    for i, item in enumerate(path_and_files[1]):
        # look up hospital ID in the database.
        # step one.  Get the medicare ID from the file name
        hospital_medicare_id = item.split('_')[1]
        # get period name from calc file name
        period_nm = item.split('_')[2].replace(" ", "%")
        # get year from calc file name
        year = item.split('_')[3].split('.')[0]

        # define the calculator id query
        calc_id_query = "SELECT calc_id,calc_nm from NM_Analytics_Prototype.vizient_qa.calculator where calc_nm like '%" + year + "%" + period_nm + "'"
        # create connection to the ms sql database
        conn = pyodbc.connect('Driver={SQL Server};'
                              'Server=edw00pd05wva.corp.nm.org\EDWIDS1;'
                              'Database=clarity;'
                              'Trusted_Connection=yes;')
        # define the hospital query
        find_hosp_by_medicare_id_query = 'SELECT hospital_id, hospital_medicare_id from NM_Analytics_Prototype.vizient_qa.hospitals WHERE hospital_medicare_id = ' + hospital_medicare_id

        # query the database
        hospital_ids = pd.DataFrame(pd.read_sql(find_hosp_by_medicare_id_query, conn))
        calc_ids = pd.DataFrame(pd.read_sql(calc_id_query, conn))

        # close the connection
        conn.close()
        # store the hospital id in a variable
        current_hospital_id = hospital_ids['hospital_id'][0]
        current_calc_id = calc_ids['calc_id'][0]

        # function that takes a calculator file path and name
        # and returns a pandas dataframe of just the domain name, measure name,
        # "What if" section, "Current ranking" section and "Target Performance Evaluation" section
        clean_df = parse_calculator_and_return_clean_dataframe(calc_file_path, item)

        # isolate necessary domain rank column

        # assign db measure value name based on calculator hospital rank search phrase
        # We have to do this separately because Vizient's headers have strange spacing.
        try:
            if domain_rank_phase == 'Domain Rank Result':
                keep_cols = ['Domain', ' Domain Rank Result']
            elif domain_rank_phase == 'Target Domain Ranking':
                keep_cols = ['Domain', 'Target Domain Ranking']
        except:
            print('Domain Rank phrase does not match.  Could not isolate column header.')
            return

        clean_df = clean_df[keep_cols]

        # remove duplicates
        clean_df = clean_df.drop_duplicates()

        # join calculator domain ranks dataframe with dataframe with domain names and ids from db
        clean_df = pd.merge(clean_df, domain_df, how='left', left_on=['Domain'], right_on=['domain_nm'])

        clean_df['calc_id'] = current_calc_id
        clean_df['hospital_id'] = current_hospital_id
        clean_df['period_id'] = period_id_result
        clean_df['measure_value_id'] = measure_value_id_result

        try:
            if domain_rank_phase == 'Domain Rank Result':
                clean_df = clean_df.rename(columns={' Domain Rank Result': 'domain_value'})
            elif domain_rank_phase == 'Target Domain Ranking':
                clean_df = clean_df.rename(columns={'Target Domain Ranking': 'domain_value'})
        except:
            print('Unable to rename domain_value column.')
            return

            # reorder columns to match db
        cols = ['calc_id', 'hospital_id', 'domain_id', 'period_id', 'measure_value_id', 'domain_value']
        clean_df = clean_df[cols]

        # append result dataframe to frames_list in order to union later.
        frames_list.append(clean_df)

    result_df = pd.concat(frames_list)
    result_df = result_df.drop_duplicates()
    return (result_df)
'''
#UL003

def parse_calculator_domain_ranks(domain_rank_phase, period_type, period_end_dts):
    # assign db measure value name based on calculator hospital rank search phrase
    try:
        if domain_rank_phase == 'Domain Rank Result':
            measure_value_type = 'Domain Rank'
        elif domain_rank_phase == 'Target Domain Ranking':
            measure_value_type = 'Top Decile Target Domain Ranking'
    except:
        print('Domain Rank phrase does not match.')
        return
    
    # query the database and get all domain ids
    domain_df = get_all_domains()

    # function takes the path of calculator file folder and returns the path and list of all file names
    path_and_files = gather_calculator_path_and_files()

    # store the file path in a variable
    calc_file_path = path_and_files[0]

    # empty list to store measure weight dataframes in.
    frames_list = []

    # query the db to get period id
    conn = pyodbc.connect('Driver={SQL Server};'
                          'Server=edw00pd05wva.corp.nm.org\EDWIDS1;'
                          'Database=clarity;'
                          'Trusted_Connection=yes;')
    period_type_id = """
    SELECT
    pt.period_type_id
    FROM
    NM_Analytics_Prototype.vizient_qa.period_types as pt
    WHERE
    pt.period_type_nm = '%s'
    """ % period_type

    period_type_id_result = pd.DataFrame(pd.read_sql(period_type_id, conn))['period_type_id'].values[0]

    period_id = '''
    SELECT 
    p.period_id 
    FROM NM_Analytics_Prototype.vizient_qa.periods as p 
    WHERE p.period_end_date = '%s'
    and p.period_type = '%s'
    ''' % (period_end_dts, period_type_id_result)

    period_id_result = pd.DataFrame(pd.read_sql(period_id, conn))['period_id'].values[0]

    measure_value_id = """
    SELECT
    measure_value_id
    from
    NM_Analytics_Prototype.vizient_qa.measure_value_types
    WHERE
    measure_value_name = '%s'
    """ % measure_value_type

    measure_value_id_result = pd.DataFrame(pd.read_sql(measure_value_id, conn))['measure_value_id'].values[0]

    conn.close()

    # iterate over calculator file names
    for i, item in enumerate(path_and_files[1]):
        
        #UL003
        #Adding additional logic to handle Critical Access calculators.  For some reason, Vizient decided to name the
        #Critical Access calculators completely differently so we need to use a different method.
        #first split up the file name, then iterate over the subsections of the name to find the part with the medicare id
        #All NM hospital medicare ids start with 140 or 141.
        for i, subitem in enumerate(item.split('_')):
            if '140' in subitem or '141' in subitem or '149' in subitem:
                hospital_medicare_id_index = i
            else:
                pass

        # look up hospital ID in the database.
        # step one.  Get the medicare ID from the file name
        hospital_medicare_id = item.split('_')[hospital_medicare_id_index]

        for i, subitem in enumerate(item.split('_')):
            if 'PERIOD' in subitem.upper() or 'ANNUAL' in subitem.upper():
                period_nm_index = i
            else:
                pass

        #have to split the period name because in 2019 there is a space between period and number.
        #In 2020, there 
        period_nm_split = re.split('(\d+)',item.split('_')[period_nm_index])

        if period_nm_split[0].replace(" ","") == 'Annual':
            period_str = 'Period'
            period_num = '4'
        else:
            period_str = period_nm_split[0].replace(" ","")
            period_num = period_nm_split[1].replace(" ","")

        #['Period ', '3', '']
        # get period name from calc file name
        #period_nm = item.split('_')[period_nm_index].replace(" ", "%")

        period_nm = period_str + '%' + period_num


        # get year from calc file name
        # The section containing the year will start with 20...for the next 80 years...
        # and it will end with .xlsm because it contains the file extension.
        for i, subitem in enumerate(item.split('_')):

            if subitem.startswith('20') and subitem.upper().endswith('.XLSM'):
                year_index = i
            else:
                pass

        
        year = item.split('_')[year_index].split('.')[0]

        '''

        hospital_medicare_id = item.split('_')[1]
        # get period name from calc file name
        period_nm = item.split('_')[2].replace(" ", "%")
        # get year from calc file name
        year = item.split('_')[3].split('.')[0]

        print(hospital_medicare_id)
        print(period_nm)
        print(year)
        '''

        # define the calculator id query
        calc_id_query = "SELECT calc_id,calc_nm from NM_Analytics_Prototype.vizient_qa.calculator where calc_nm like '%" + year + "%" + period_nm + "'"
        # create connection to the ms sql database
        conn = pyodbc.connect('Driver={SQL Server};'
                              'Server=edw00pd05wva.corp.nm.org\EDWIDS1;'
                              'Database=clarity;'
                              'Trusted_Connection=yes;')
        # define the hospital query
        find_hosp_by_medicare_id_query = 'SELECT hospital_id, hospital_medicare_id from NM_Analytics_Prototype.vizient_qa.hospitals WHERE hospital_medicare_id = ' + hospital_medicare_id

        # query the database
        hospital_ids = pd.DataFrame(pd.read_sql(find_hosp_by_medicare_id_query, conn))
        calc_ids = pd.DataFrame(pd.read_sql(calc_id_query, conn))

        # close the connection
        conn.close()


        # store the hospital id in a variable
        current_hospital_id = hospital_ids['hospital_id'][0]
        current_calc_id = calc_ids['calc_id'][0]

        #Valley West calculators do not have Possible Rank or Target Rank on the main page for some reason.  Thanks Vizient.
        #Therefore, we need to sum up the metric score and lookup the closes rank on the hidden ranks worksheet.
        if current_hospital_id == 356:  #Valley West db ID.
            
            if domain_rank_phase == 'Target Domain Ranking':
            
                ca_file_loc = os.path.abspath(os.path.join(calc_file_path, item))
                ca_wb = openpyxl.load_workbook(ca_file_loc, data_only = True)
                ranks_df = pd.DataFrame(ca_wb['CurrentQA Cumulative-Rank'].values)
                ##in openpyxl, the headers are listed as numbers.  Renaming the headers using the first row values.
                ranks_df = ranks_df.rename(columns=ranks_df.iloc[0])
                ##drop the first row because they are now the header
                ranks_df = ranks_df.drop([0])
                
                if 'parent_id' in ranks_df.columns:
                    ranks_df = ranks_df.rename(columns={"parent_id": "Parent_ID"})
                
                ##drop all rows that do not have a hospital value
                df_all_row_indices =  ranks_df['Parent_ID'].notnull()
                ranks_df = ranks_df[df_all_row_indices]
                
                ##only save specific columns we need
                ranks_df = ranks_df[['Parent_ID', 'HCO_SHORT_NAME', \
                                     'final_score',\
                                     'final_Rank',\
                                     'mort_Rank',\
                                     'eff_Rank','safety_Rank',\
                                    'effect_Rank',\
                                    'Patct_Rank',\
                                    'eq_rank']]
                
                
                                                                  
                #count the number of hospitals in the cohort
                cohort_num = len(ranks_df['Parent_ID'].unique())
                #target ranking is always top decile.  So, take the cohort number divide by 10 and round the result.
                top_decile_target_ranking = round(((cohort_num*1.0)/10.0))

                #make sure the dataframe is still sort in ascending order.
                ranks_df = ranks_df.sort_values(by=['final_score'],ascending=False)
                #create empty dataframe to store values.
                target_df = pd.DataFrame(columns = ['Domain', 'domain_value'])
                rank_val = ranks_df['final_Rank'].iloc[top_decile_target_ranking-1]
                
                #locate each domain's rank at the index of the target hospital rank.
                mort_Rank = ranks_df['mort_Rank'].iloc[top_decile_target_ranking-1]
                
                # append rows to an empty DataFrame 
                if mort_Rank is not None:
                    target_df = target_df.append({'Domain' : 'Mortality', 'domain_value' : mort_Rank},  ignore_index = True) 
                else:
                    pass
                eff_Rank = ranks_df['eff_Rank'].iloc[top_decile_target_ranking-1]
                
                # append rows 
                if eff_Rank is not None:
                    target_df = target_df.append({'Domain' : 'Efficiency', 'domain_value' : eff_Rank},  ignore_index = True) 
                else:
                    pass
                safety_Rank = ranks_df['safety_Rank'].iloc[top_decile_target_ranking-1]
                
                # append rows 
                if safety_Rank is not None:
                    target_df = target_df.append({'Domain' : 'Safety', 'domain_value' : safety_Rank},  ignore_index = True) 
                else:
                    pass
                
                effect_Rank = ranks_df['effect_Rank'].iloc[top_decile_target_ranking-1]
                # append rows 
                if effect_Rank is not None:
                    target_df = target_df.append({'Domain' : 'Effectiveness', 'domain_value' : effect_Rank},  ignore_index = True) 
                else:
                    pass
                
                Patct_Rank = ranks_df['Patct_Rank'].iloc[top_decile_target_ranking-1]
                                              
                # append rows
                if Patct_Rank is not None:
                    target_df = target_df.append({'Domain' : 'Patient Centeredness', 'domain_value' : Patct_Rank},  ignore_index = True) 
                else:
                    pass
                
                eq_rank = ranks_df['eq_rank'].iloc[top_decile_target_ranking-1]
                                              
                # append rows
                if eq_rank is not None:
                    target_df = target_df.append({'Domain' : 'Equity', 'domain_value' : eq_rank},  ignore_index = True) 
                else:
                    pass
                
                clean_df = target_df
                
                # join calculator domain ranks dataframe with dataframe with domain names and ids from db
                clean_df = pd.merge(clean_df, domain_df, how='left', left_on=['Domain'], right_on=['domain_nm'])

                clean_df['calc_id'] = current_calc_id
                clean_df['hospital_id'] = current_hospital_id
                clean_df['period_id'] = period_id_result
                clean_df['measure_value_id'] = measure_value_id_result

                try:
                    if domain_rank_phase == 'Domain Rank Result':
                        clean_df = clean_df.rename(columns={'Domain Rank': 'domain_value'})
                    elif domain_rank_phase == 'Target Domain Ranking':
                        clean_df = clean_df.rename(columns={'Target Domain Ranking': 'domain_value'})
                except:
                    print('Unable to rename domain_value column.')
                    return

                    # reorder columns to match db
                cols = ['calc_id', 'hospital_id', 'domain_id', 'period_id', 'measure_value_id', 'domain_value']
                clean_df = clean_df[cols]
                 
                # append result dataframe to frames_list in order to union later.
                frames_list.append(clean_df)
                
            else:
                
                #if looking for 'domain rank result', then we check the 'what if' section of the critical access
                #calculator and just take the modified 'Domain Rank' column.
                
                #Again, Vizient is awesome and gave the domain rank column different headers for the critical access 
                #cohort only.  Don't ask me why...smh...
                
                # open the calculator file and isolate the 'Calculator' sheet.
                ws = grab_calc_worksheet_calc_sheet(calc_file_path, item)
                
                #find the coordinates of the 'Domain' header cell of the 'Calculator' worksheet
                domain_coord = find_calc_phrase_cell_coords(ws,'Domain')
                #find the coordinates of the 'Target Domain Ranking' header cell of the 'Calculator' worksheet
                what_if_metric_value_coord = find_calc_phrase_cell_coords(ws,' "What if" Performance Evaluation')

                #set row coordinate to 2 below actual in order to locate the 'metric value' header.
                what_if_metric_value_coord[0] = what_if_metric_value_coord[0] + 2

                #Find the LAST instance of 'Domain Rank'.  Get a list of all instances, then take last index.
                domain_rank_coord = find_calc_phrase_cell_coords_multiple(ws,'Domain Rank')
                domain_rank_coord = domain_rank_coord[-1]
                #store worksheet object as pandas dataframe
                m_df = pd.DataFrame(ws.values)

                #use column coordinates of "Domain" header to drop all columns to the left of it.
                col_drop_list1 = [i for i in range(domain_rank_coord[1],m_df.shape[1])]
                m_df = m_df.drop(col_drop_list1,axis=1)

                #use column coordinates of "Domain" header to drop all columns to the left of it.
                col_drop_list = [i for i in range(domain_coord[1]-1)]
                m_df = m_df.drop(col_drop_list,axis=1)

                #rename column headers using the 'Domain' header row coordinates
                m_df = m_df.rename(columns=m_df.iloc[domain_coord[0]-1])

                #use row coordinates of "Domain" header row to drop all rows before it.
                row_drop_list = [i for i in range(domain_coord[0])]
                m_df = m_df.drop(row_drop_list)

                #drop colum where all values are NA
                m_df = m_df.dropna(axis='columns', how="all")

                #drop rows where all values are NA
                m_df = m_df.dropna(how="all")

                domain_and_measure_indices = [0,1]

                #indexes have now changed.  Find the last Domain Rank column header.
                domain_rank_list = []
                for i, thing in enumerate(m_df.columns):
                    if thing == 'Domain Rank':
                        domain_rank_list.append(i)
                #find last 'Domain Rank.'
                new_domain_rank_coord = domain_rank_list[-1]
                #keep_cols = [i for i in range(df.columns.get_loc(' Domain Rank Result')-4,df.columns.get_loc('Target Domain Ranking')+1)]
                keep_cols = [i for i in range(new_domain_rank_coord-4,new_domain_rank_coord+1)]

                #subset columns to keep only the ones we care about.
                final_keep_cols = domain_and_measure_indices + keep_cols
                m_df = m_df.iloc[:,final_keep_cols]

                #remove copyright row at the bottom.
                m_df = m_df[~m_df['Domain'].str.contains('Copyright', na=False)]

                #domain names and ranks are in merged excel rows so we now need to forward fill them so we have a value every row.
                m_df['Domain'] = m_df['Domain'].fillna(method='ffill')
                m_df['Domain Rank'] = m_df['Domain Rank'].fillna(method='ffill')

                clean_df = m_df[['Domain','Domain Rank']]
                
                #UL003
                #conditionally remove rank rows with 'No Rank' phrase.  Sometimes 
                #current rank will be no rank while target rank does have a rank so we have to check
                #selection before removing rows.
                try:
                    if domain_rank_phase == 'Domain Rank Result':
                        clean_df = clean_df[clean_df['Domain Rank'] != 'No Rank']
                    elif domain_rank_phase == 'Target Domain Ranking':
                        clean_df = clean_df[clean_df['Target Domain Ranking'] != 'No Rank']
                except:
                    continue
                
                # remove duplicates
                clean_df = clean_df.drop_duplicates()

                # join calculator domain ranks dataframe with dataframe with domain names and ids from db
                clean_df = pd.merge(clean_df, domain_df, how='left', left_on=['Domain'], right_on=['domain_nm'])

                clean_df['calc_id'] = current_calc_id
                clean_df['hospital_id'] = current_hospital_id
                clean_df['period_id'] = period_id_result
                clean_df['measure_value_id'] = measure_value_id_result

                try:
                    if domain_rank_phase == 'Domain Rank Result':
                        clean_df = clean_df.rename(columns={'Domain Rank': 'domain_value'})
                    elif domain_rank_phase == 'Target Domain Ranking':
                        clean_df = clean_df.rename(columns={'Target Domain Ranking': 'domain_value'})
                except:
                    print('Unable to rename domain_value column.')
                    return

                    # reorder columns to match db
                cols = ['calc_id', 'hospital_id', 'domain_id', 'period_id', 'measure_value_id', 'domain_value']
                clean_df = clean_df[cols]

                # append result dataframe to frames_list in order to union later.
                frames_list.append(clean_df)
  
            
        #all other hospital calculators, use legacy parsing logic.    
        else:
        
            # function that takes a calculator file path and name
            # and returns a pandas dataframe of just the domain name, measure name,
            # "What if" section, "Current ranking" section and "Target Performance Evaluation" section
            clean_df = parse_calculator_and_return_clean_dataframe(calc_file_path, item)
            
            #UL003
            #conditionally remove rank rows with 'No Rank' phrase.  Sometimes 
            #current rank will be no rank while target rank does have a rank so we have to check
            #selection before removing rows.
            try:
                if domain_rank_phase == 'Domain Rank Result':
                    clean_df = clean_df[clean_df[' Domain Rank Result'] != 'No Rank']
                elif domain_rank_phase == 'Target Domain Ranking':
                    clean_df = clean_df[clean_df['Target Domain Ranking'] != 'No Rank']
            except:
                continue
            
            # isolate necessary domain rank column

            # assign db measure value name based on calculator hospital rank search phrase
            # We have to do this separately because Vizient's headers have strange spacing.
            try:
                if domain_rank_phase == 'Domain Rank Result':
                    keep_cols = ['Domain', ' Domain Rank Result']
                elif domain_rank_phase == 'Target Domain Ranking':
                    keep_cols = ['Domain', 'Target Domain Ranking']
            except:
                print('Domain Rank phrase does not match.  Could not isolate column header.')
                return

            clean_df = clean_df[keep_cols]

            # remove duplicates
            clean_df = clean_df.drop_duplicates()

            # join calculator domain ranks dataframe with dataframe with domain names and ids from db
            clean_df = pd.merge(clean_df, domain_df, how='left', left_on=['Domain'], right_on=['domain_nm'])

            clean_df['calc_id'] = current_calc_id
            clean_df['hospital_id'] = current_hospital_id
            clean_df['period_id'] = period_id_result
            clean_df['measure_value_id'] = measure_value_id_result

            try:
                if domain_rank_phase == 'Domain Rank Result':
                    clean_df = clean_df.rename(columns={' Domain Rank Result': 'domain_value'})
                elif domain_rank_phase == 'Target Domain Ranking':
                    clean_df = clean_df.rename(columns={'Target Domain Ranking': 'domain_value'})
            except:
                print('Unable to rename domain_value column.')
                return

                # reorder columns to match db
            cols = ['calc_id', 'hospital_id', 'domain_id', 'period_id', 'measure_value_id', 'domain_value']
            clean_df = clean_df[cols]

            # append result dataframe to frames_list in order to union later.
            frames_list.append(clean_df)

    result_df = pd.concat(frames_list)
    result_df = result_df.drop_duplicates()
    #result_df['domain_value'] = pd.to_numeric(result_df['domain_value'])
    
    #in 2021, Vizient deactivated THK and, for some reason, created a domain rank cell for just THK.
    #and set that cell to 'NA.'  We cannot use the value 'NA,' so we need to exclude.
    result_df = result_df[result_df['domain_value'] != 'NA']
    
    result_df['domain_value'] = result_df['domain_value'].astype('float')
    #result_df.to_csv('test_mf.csv')
    return (result_df)


##############################################################################################################################

# function takes the resulting dataframe from parse_calculator_domain_ranks()
# and performs inserts into the database.

def insert_domain_ranks(df):
    # connect to the NM_Analytics database
    conn = pyodbc.connect('Driver={SQL Server};'
                          'Server=edw00pd05wva.corp.nm.org\EDWIDS1;'
                          'Database=NM_Analytics_Prototype;'
                          'Trusted_Connection=yes;')

    cursor = conn.cursor()

    for index, row in df.iterrows():
        if row['domain_value'] != 'No Rank':
            cursor.execute(
                "INSERT INTO NM_Analytics_Prototype.vizient_qa.calc_domain_values([calc_id],[hospital_id],[domain_id],[period_id],[measure_value_id],[domain_value]) values (?,?,?,?,?,?)",
                row['calc_id'], row['hospital_id'], row['domain_id'], row['period_id'], row['measure_value_id'],
                row['domain_value'])
            conn.commit()

    cursor.close()
    conn.close()
    print('done inserting rows.')

##############################################################################################################################
    
#UL003
#New helper function for FY21 to convert domain and what if section measure values z-scores.
#Vizient added text values to their z-score column so now we need
#to check for those and replace them with 0.00 (mean) z-score.    
    
    
def clean_measure_values(x):
    if isinstance(x, str):
        x = 0.0000
        return(x)
    return(x)   

##############################################################################################################################
    
#UL003
#We cannot store Equity domain text values 'Pass','Fail','Warning' 
#in the measure_values table so we convert them into a number code with this function.
    
def clean_equity_domain_values(x):
    if (x == 'Fail') or (x == 'Unequal'):
        x = 0
        return(x)
    elif (x == 'Warning') or (x == 'Caution'):
        x = 1
        return(x)
    elif (x == 'Pass') or (x == 'Equal'):
        x = 2
        return(x)
    return(x)

##############################################################################################################################

#functions to parse and insert values from the calculator 'What-if' section

'''
def parse_calculator_what_if_section(period_type, period_end_dts):
    measure_value_type1 = 'Metric Value'
    measure_value_type2 = '% of Domain Score'
    measure_value_type3 = 'Z-Score'
    measure_value_type4 = '% of Overall Score'

    # query the database and get all measure ids
    measure_df = get_all_measures()

    # function takes the path of calculator file folder and returns the path and list of all file names
    path_and_files = gather_calculator_path_and_files()

    # store the file path in a variable
    calc_file_path = path_and_files[0]

    # empty list to store measure weight dataframes in.
    frames_list = []

    # query the db to get period id
    conn = pyodbc.connect('Driver={SQL Server};'
                          'Server=edw00pd05wva.corp.nm.org\EDWIDS1;'
                          'Database=clarity;'
                          'Trusted_Connection=yes;')
    period_type_id = """
    SELECT
    pt.period_type_id
    FROM
    NM_Analytics_Prototype.vizient_qa.period_types as pt
    WHERE
    pt.period_type_nm = '%s'
    """ % period_type

    period_type_id_result = pd.DataFrame(pd.read_sql(period_type_id, conn))['period_type_id'].values[0]

    period_id = '''''''
    SELECT 
    p.period_id 
    FROM NM_Analytics_Prototype.vizient_qa.periods as p 
    WHERE p.period_end_date = '%s'
    and p.period_type = '%s'
    ''''''' % (period_end_dts, period_type_id_result)

    period_id_result = pd.DataFrame(pd.read_sql(period_id, conn))['period_id'].values[0]

    measure_value_id = """
    SELECT
    measure_value_id
    ,measure_value_name
    from
    NM_Analytics_Prototype.vizient_qa.measure_value_types
    WHERE
    measure_value_name in ('%s','%s','%s','%s')
    """ % (measure_value_type1, measure_value_type2, measure_value_type3, measure_value_type4)

    measure_value_id_result = pd.DataFrame(pd.read_sql(measure_value_id, conn))

    conn.close()

    # iterate over calculator file names
    for i, item in enumerate(path_and_files[1]):
        # look up hospital ID in the database.
        # step one.  Get the medicare ID from the file name
        hospital_medicare_id = item.split('_')[1]
        # get period name from calc file name
        period_nm = item.split('_')[2].replace(" ", "%")
        # get year from calc file name
        year = item.split('_')[3].split('.')[0]

        # define the calculator id query
        calc_id_query = "SELECT calc_id,calc_nm from NM_Analytics_Prototype.vizient_qa.calculator where calc_nm like '%" + year + "%" + period_nm + "'"
        # create connection to the ms sql database
        conn = pyodbc.connect('Driver={SQL Server};'
                              'Server=edw00pd05wva.corp.nm.org\EDWIDS1;'
                              'Database=clarity;'
                              'Trusted_Connection=yes;')
        # define the hospital query
        find_hosp_by_medicare_id_query = 'SELECT hospital_id, hospital_medicare_id from NM_Analytics_Prototype.vizient_qa.hospitals WHERE hospital_medicare_id = ' + hospital_medicare_id

        # query the database
        hospital_ids = pd.DataFrame(pd.read_sql(find_hosp_by_medicare_id_query, conn))
        calc_ids = pd.DataFrame(pd.read_sql(calc_id_query, conn))

        # close the connection
        conn.close()
        # store the hospital id in a variable
        current_hospital_id = hospital_ids['hospital_id'][0]
        current_calc_id = calc_ids['calc_id'][0]

        # function that takes a calculator file path and name
        # and returns a pandas dataframe of just the domain name, measure name,
        # "What if" section, "Current ranking" section and "Target Performance Evaluation" section
        clean_df = parse_calculator_and_return_clean_dataframe(calc_file_path, item)

        # remove rows with LV marker or '-' marker

        clean_df = clean_df[clean_df['Metric Value'] != 'LV']
        clean_df = clean_df[clean_df['Metric Value'] != '-']

        # drop blank rows
        clean_df = clean_df[pd.notnull(clean_df['Metric Value'])]

        # assign db measure value name based on calculator hospital rank search phrase
        # We have to do this separately because Vizient's headers have strange spacing.
        try:
            if period_type == 'VIZIENT_CALC_PLACEHOLDER':
                # if parsing baseline period calculator, grab the metric value, z-score, etc. Do not keep the other columns.
                keep_cols = ['Measure', 'Metric Value', ' Z-Score', '% of Domain Score', '% of Overall Score']
                clean_df = clean_df[keep_cols]
                # Z-score column has a space...rename it to remove the space in front.
                clean_df = clean_df.rename(columns={' Z-Score': 'Z-Score'})
                # drop any duplicates.
                clean_df = clean_df.drop_duplicates()
                # unpivot the value columns so you can join to the measure_value_type ids.
                # this converts to long format.
                clean_df = pd.melt(clean_df, id_vars=['Measure'], var_name='measure_value_name',
                                   value_name='measure_value')

            else:
                # if parsing performance close period calculator, grab the z-score, etc. Do not keep the other columns.
                # No need to pull the Metric Value column because the metric values were already stored
                # when the reports were pulled and organized.
                keep_cols = ['Measure', ' Z-Score', '% of Domain Score', '% of Overall Score']
                clean_df = clean_df[keep_cols]
                # Z-score column has a space...rename it to remove the space in front.
                clean_df = clean_df.rename(columns={' Z-Score': 'Z-Score'})
                # drop any duplicates.
                clean_df = clean_df.drop_duplicates()
                # unpivot the value columns so you can join to the measure_value_type ids.
                # this converts to long format.
                clean_df = pd.melt(clean_df, id_vars=['Measure'], var_name='measure_value_name',
                                   value_name='measure_value')

        except:
            print('What if Section headers do not match.  Could not isolate column header.')
            return

            # print(clean_df.head())

        # join calculator domain ranks dataframe with dataframe with measure names and ids from db
        clean_df = pd.merge(clean_df, measure_df, how='left', left_on=['Measure'], right_on=['measure_name'])

        clean_df['calc_id'] = current_calc_id
        clean_df['hospital_id'] = current_hospital_id
        clean_df['period_id'] = period_id_result

        # join calculator measure value dataframe with dataframe with measure_value_types and ids from db
        clean_df = pd.merge(clean_df, measure_value_id_result, how='left', left_on=['measure_value_name'],
                            right_on=['measure_value_name'])

        # Because these values do not have n events or event_type_id, we must create NULL placeholders

        clean_df["numerator"] = 0
        clean_df["numerator_event_type_id"] = 0
        clean_df["denominator"] = 0
        clean_df["denominator_event_type_id"] = 0

        # clean_df["n_events"] = clean_df["n_events"].replace(0,None)
        # clean_df["event_type_id"] = clean_df["event_type_id"].replace(0,None)

        # clean_df["n_events"] = clean_df["n_events"].replace({0: None})

        # reorder columns to match db
        cols = ['calc_id', 'hospital_id', 'measure_id', 'period_id', 'measure_value_id', 'measure_value', 'numerator',
                'numerator_event_type_id','denominator','denominator_event_type_id']
        clean_df = clean_df[cols]

        # append result dataframe to frames_list in order to union later.
        frames_list.append(clean_df)

    result_df = pd.concat(frames_list)
    result_df = result_df.drop_duplicates()
    #make sure measure_value column is numeric
    result_df['measure_value'] = pd.to_numeric(result_df['measure_value'])

    return (result_df)
'''
#UL003

def parse_calculator_what_if_section(period_type, period_end_dts):
    measure_value_type1 = 'Metric Value'
    measure_value_type2 = '% of Domain Score'
    measure_value_type3 = 'Z-Score'
    measure_value_type4 = '% of Overall Score'
    
    # query the database and get all measure ids
    measure_df = get_all_measures()
    
    # function takes the path of calculator file folder and returns the path and list of all file names
    path_and_files = gather_calculator_path_and_files()
    
    # store the file path in a variable
    calc_file_path = path_and_files[0]
    
    # empty list to store measure weight dataframes in.
    frames_list = []
    
    # query the db to get period id
    conn = pyodbc.connect('Driver={SQL Server};'
                          'Server=edw00pd05wva.corp.nm.org\EDWIDS1;'
                          'Database=clarity;'
                          'Trusted_Connection=yes;')
    period_type_id = """
    SELECT
    pt.period_type_id
    FROM
    NM_Analytics_Prototype.vizient_qa.period_types as pt
    WHERE
    pt.period_type_nm = '%s'
    """ % period_type
    
    period_type_id_result = pd.DataFrame(pd.read_sql(period_type_id, conn))['period_type_id'].values[0]
    
    period_id = '''
    SELECT 
    p.period_id 
    FROM NM_Analytics_Prototype.vizient_qa.periods as p 
    WHERE p.period_end_date = '%s'
    and p.period_type = '%s'
    ''' % (period_end_dts, period_type_id_result)
    
    period_id_result = pd.DataFrame(pd.read_sql(period_id, conn))['period_id'].values[0]
    
    measure_value_id = """
    SELECT
    measure_value_id
    ,measure_value_name
    from
    NM_Analytics_Prototype.vizient_qa.measure_value_types
    WHERE
    measure_value_name in ('%s','%s','%s','%s')
    """ % (measure_value_type1, measure_value_type2, measure_value_type3, measure_value_type4)
    
    measure_value_id_result = pd.DataFrame(pd.read_sql(measure_value_id, conn))
    
    conn.close()

    # iterate over calculator file names
    for i, item in enumerate(path_and_files[1]):
        #UL003
        #Adding additional logic to handle Critical Access calculators.  For some reason, Vizient decided to name the
        #Critical Access calculators completely differently so we need to use a different method.
        #first split up the file name, then iterate over the subsections of the name to find the part with the medicare id
        #All NM hospital medicare ids start with 140 or 141.
        print('item:',item)
        for i, subitem in enumerate(item.split('_')):
            if '140' in subitem or '141' in subitem or '149' in subitem:
                hospital_medicare_id_index = i
            else:
                pass
    
        # look up hospital ID in the database.
        # step one.  Get the medicare ID from the file name
        hospital_medicare_id = item.split('_')[hospital_medicare_id_index]
    
        for i, subitem in enumerate(item.split('_')):
            if 'PERIOD' in subitem.upper() or 'ANNUAL' in subitem.upper():
                period_nm_index = i
            else:
                pass
    
        #have to split the period name because in 2019 there is a space between period and number.
        #In 2020, there 
        period_nm_split = re.split('(\d+)',item.split('_')[period_nm_index])
    
        if period_nm_split[0].replace(" ","") == 'Annual':
            period_str = 'Period'
            period_num = '4'
        else:
            period_str = period_nm_split[0].replace(" ","")
            period_num = period_nm_split[1].replace(" ","")
    
        #['Period ', '3', '']
        # get period name from calc file name
        #period_nm = item.split('_')[period_nm_index].replace(" ", "%")
    
        period_nm = period_str + '%' + period_num
    
    
        # get year from calc file name
        # The section containing the year will start with 20...for the next 80 years...
        # and it will end with .xlsm because it contains the file extension.
        for i, subitem in enumerate(item.split('_')):
    
            if subitem.startswith('20') and subitem.upper().endswith('.XLSM'):
                year_index = i
            else:
                pass
    
        
        year = item.split('_')[year_index].split('.')[0]
    
        '''
    
        hospital_medicare_id = item.split('_')[1]
        # get period name from calc file name
        period_nm = item.split('_')[2].replace(" ", "%")
        # get year from calc file name
        year = item.split('_')[3].split('.')[0]
    
        print(hospital_medicare_id)
        print(period_nm)
        print(year)
        '''
    
        # define the calculator id query
        calc_id_query = "SELECT calc_id,calc_nm from NM_Analytics_Prototype.vizient_qa.calculator where calc_nm like '%" + year + "%" + period_nm + "'"
        # create connection to the ms sql database
        conn = pyodbc.connect('Driver={SQL Server};'
                              'Server=edw00pd05wva.corp.nm.org\EDWIDS1;'
                              'Database=clarity;'
                              'Trusted_Connection=yes;')
        # define the hospital query
        find_hosp_by_medicare_id_query = 'SELECT hospital_id, hospital_medicare_id from NM_Analytics_Prototype.vizient_qa.hospitals WHERE hospital_medicare_id = ' + hospital_medicare_id
    
        # query the database
        hospital_ids = pd.DataFrame(pd.read_sql(find_hosp_by_medicare_id_query, conn))
        calc_ids = pd.DataFrame(pd.read_sql(calc_id_query, conn))
    
        # close the connection
        conn.close()
    
    
        # store the hospital id in a variable
        current_hospital_id = hospital_ids['hospital_id'][0]
        current_calc_id = calc_ids['calc_id'][0]
            
        
        #UL003
        #Valley West calculators do not have Possible Rank or Target Rank on the main page for some reason.  Thanks Vizient.
        #Therefore, we need to sum up the metric score and lookup the closes rank on the hidden ranks worksheet.
        if current_hospital_id == 356:  #Valley West db ID.
            #UL003
            #Critical Access calculators are different, of course.
    
            #If the hospital is Valley West, we need to subset the 'what if' section in 
            #a different way.  The column headers are also slightly different
            #and, for some reason, the Z-Score column is completely different.  It mixes
            #actual z-scores with text-based phrases indicating statistical significance.
            #since we cannot store a text phrase in our database column for z-scores, we
            #will convert any non-number value z-score to median z-score (0).
    
    
            # open the calculator file and isolate the 'Calculator' sheet.
            ws = grab_calc_worksheet_calc_sheet(calc_file_path, item)

            #find the coordinates of the 'Domain' header cell of the 'Calculator' worksheet
            domain_coord = find_calc_phrase_cell_coords(ws,'Domain')
            #find the coordinates of the 'Target Domain Ranking' header cell of the 'Calculator' worksheet
            what_if_metric_value_coord = find_calc_phrase_cell_coords(ws,' "What if" Performance Evaluation')
    
            #set row coordinate to 2 below actual in order to locate the 'metric value' header.
            what_if_metric_value_coord[0] = what_if_metric_value_coord[0] + 2
    
            #Find the LAST instance of 'Domain Rank'.  Get a list of all instances, then take last index.
            domain_rank_coord = find_calc_phrase_cell_coords_multiple(ws,'Domain Rank')
            domain_rank_coord = domain_rank_coord[-1]
            #store worksheet object as pandas dataframe
            m_df = pd.DataFrame(ws.values)
    
            #use column coordinates of "Domain" header to drop all columns to the left of it.
            col_drop_list1 = [i for i in range(domain_rank_coord[1],m_df.shape[1])]
            m_df = m_df.drop(col_drop_list1,axis=1)
    
            #use column coordinates of "Domain" header to drop all columns to the left of it.
            col_drop_list = [i for i in range(domain_coord[1]-1)]
            m_df = m_df.drop(col_drop_list,axis=1)
    
            #rename column headers using the 'Domain' header row coordinates
            m_df = m_df.rename(columns=m_df.iloc[domain_coord[0]-1])
    
            #use row coordinates of "Domain" header row to drop all rows before it.
            row_drop_list = [i for i in range(domain_coord[0])]
            m_df = m_df.drop(row_drop_list)
    
            #drop colum where all values are NA
            m_df = m_df.dropna(axis='columns', how="all")
    
            #drop rows where all values are NA
            m_df = m_df.dropna(how="all")
    
            domain_and_measure_indices = [0,1]
    
            #indexes have now changed.  Find the last Domain Rank column header.
            domain_rank_list = []
            for i, thing in enumerate(m_df.columns):
                if thing == 'Domain Rank':
                    domain_rank_list.append(i)
            #find last 'Domain Rank.'
            new_domain_rank_coord = domain_rank_list[-1]
            #keep_cols = [i for i in range(df.columns.get_loc(' Domain Rank Result')-4,df.columns.get_loc('Target Domain Ranking')+1)]
            keep_cols = [i for i in range(new_domain_rank_coord-4,new_domain_rank_coord+1)]
    
            #subset columns to keep only the ones we care about.
            final_keep_cols = domain_and_measure_indices + keep_cols
            m_df = m_df.iloc[:,final_keep_cols]
    
            #remove copyright row at the bottom.
            m_df = m_df[~m_df['Domain'].str.contains('Copyright', na=False)]
    
            #domain names and ranks are in merged excel rows so we now need to forward fill them so we have a value every row.
            m_df['Domain'] = m_df['Domain'].fillna(method='ffill')
            m_df['Domain Rank'] = m_df['Domain Rank'].fillna(method='ffill')
            
            clean_df = m_df
            
            clean_df = clean_df[clean_df['Domain Rank'].astype(str) != 'No Rank']
    
            # remove duplicates
            clean_df = clean_df.drop_duplicates()
    
            # remove rows with LV marker or '-' marker
    
            clean_df = clean_df[clean_df['Metric Value'] != 'LV']
            clean_df = clean_df[clean_df['Metric Value'] != '-']
    
            # drop blank rows
            clean_df = clean_df[pd.notnull(clean_df['Metric Value'])]
            
            #make sure all column names match the other cohort calculator column headers
            clean_df = clean_df.rename(columns={'Metric Z score or Statistical Significance': 'Z-Score',\
                                                '% of Domain Score':'% of Domain Score',\
                                                '% of Overall Score': '% of Overall Score'})
            
            try:
                if period_type == 'VIZIENT_CALC_PLACEHOLDER':
                    # if parsing baseline period calculator, grab the metric value, z-score, etc. Do not keep the other columns.
                    keep_cols = ['Measure', 'Metric Value', 'Z-Score', '% of Domain Score', '% of Overall Score']
                    clean_df = clean_df[keep_cols]
                    # drop any duplicates.
                    clean_df = clean_df.drop_duplicates()
                    
                    # unpivot the value columns so you can join to the measure_value_type ids.
                    # this converts to long format.
                    clean_df = pd.melt(clean_df, id_vars=['Measure'], var_name='measure_value_name',
                                       value_name='measure_value')
    
                else:
                    # if parsing performance close period calculator, grab the z-score, etc. Do not keep the other columns.
                    # No need to pull the Metric Value column because the metric values were already stored
                    # when the reports were pulled and organized.
                    keep_cols = ['Measure', 'Z-Score', '% of Domain Score', '% of Overall Score']
                    clean_df = clean_df[keep_cols]
                    # Z-score column has a space...rename it to remove the space in front.
                    #clean_df = clean_df.rename(columns={' Z-Score': 'Z-Score'})
                    # drop any duplicates.
                    clean_df = clean_df.drop_duplicates()
                    # unpivot the value columns so you can join to the measure_value_type ids.
                    # this converts to long format.
                    clean_df = pd.melt(clean_df, id_vars=['Measure'], var_name='measure_value_name',
                                       value_name='measure_value')
            except:
                print('What if Section headers do not match.  Could not isolate column header.')
                return
    
            #In the Critical Access calculators, Vizient has also included text phrases e.g. 'Not Different','Better'.
            #We cannot use phrases as a z-score, so I will set any phrases in this column to 0.0 (mean).
            #if value is all alphabetical, then replace it with '0.00'
            
            clean_df.loc[clean_df.Measure == 'Colonscopy Revisits within 7-days', 'Measure'] = 'Colonoscopy Revisits within 7-days'
            
            clean_df['measure_value'] = clean_df['measure_value'].apply(clean_measure_values).astype(float)
            
            print('TEST clean_df before left join:',clean_df)
            
            # join calculator domain ranks dataframe with dataframe with measure names and ids from db
            
            clean_df = pd.merge(clean_df, measure_df, how='left', left_on=['Measure'], right_on=['measure_name'])
    
            clean_df['calc_id'] = current_calc_id
            clean_df['hospital_id'] = current_hospital_id
            clean_df['period_id'] = period_id_result
    
            # join calculator measure value dataframe with dataframe with measure_value_types and ids from db
            clean_df = pd.merge(clean_df, measure_value_id_result, how='left', left_on=['measure_value_name'],
                                right_on=['measure_value_name'])
    
            # Because these values do not have n events or event_type_id, we must create NULL placeholders
    
            clean_df["numerator"] = 0
            clean_df["numerator_event_type_id"] = 0
            clean_df["denominator"] = 0
            clean_df["denominator_event_type_id"] = 0
    
            # clean_df["n_events"] = clean_df["n_events"].replace(0,None)
            # clean_df["event_type_id"] = clean_df["event_type_id"].replace(0,None)
    
            # clean_df["n_events"] = clean_df["n_events"].replace({0: None})
    
            # reorder columns to match db
            cols = ['calc_id', 'hospital_id', 'measure_id', 'period_id', 'measure_value_id', 'measure_value', 'numerator',
                    'numerator_event_type_id','denominator','denominator_event_type_id']
            clean_df = clean_df[cols]
    
            # append result dataframe to frames_list in order to union later.
            frames_list.append(clean_df)
            
        else:
        
            # function that takes a calculator file path and name
            # and returns a pandas dataframe of just the domain name, measure name,
            # "What if" section, "Current ranking" section and "Target Performance Evaluation" section
            clean_df = parse_calculator_and_return_clean_dataframe(calc_file_path, item)
            #UL003
            clean_df = clean_df[clean_df[' Domain Rank Result'] != 'No Rank']

            # remove rows with LV marker or '-' marker

            clean_df = clean_df[clean_df['Metric Value'] != 'LV']
            clean_df = clean_df[clean_df['Metric Value'] != '-']

            # drop blank rows
            clean_df = clean_df[pd.notnull(clean_df['Metric Value'])]

            # assign db measure value name based on calculator hospital rank search phrase
            # We have to do this separately because Vizient's headers have strange spacing.
            try:
                if period_type == 'VIZIENT_CALC_PLACEHOLDER':
                    # if parsing baseline period calculator, grab the metric value, z-score, etc. Do not keep the other columns.
                    keep_cols = ['Measure', 'Metric Value', ' Z-Score', '% of Domain Score', '% of Overall Score']
                    clean_df = clean_df[keep_cols]
                    # Z-score column has a space...rename it to remove the space in front.
                    clean_df = clean_df.rename(columns={' Z-Score': 'Z-Score'})
                    # drop any duplicates.
                    clean_df = clean_df.drop_duplicates()
                    # unpivot the value columns so you can join to the measure_value_type ids.
                    # this converts to long format.
                    clean_df = pd.melt(clean_df, id_vars=['Measure'], var_name='measure_value_name',
                                       value_name='measure_value')

                else:
                    # if parsing performance close period calculator, grab the z-score, etc. Do not keep the other columns.
                    # No need to pull the Metric Value column because the metric values were already stored
                    # when the reports were pulled and organized.
                    keep_cols = ['Measure', ' Z-Score', '% of Domain Score', '% of Overall Score']
                    clean_df = clean_df[keep_cols]
                    # Z-score column has a space...rename it to remove the space in front.
                    clean_df = clean_df.rename(columns={' Z-Score': 'Z-Score'})
                    # drop any duplicates.
                    clean_df = clean_df.drop_duplicates()
                    # unpivot the value columns so you can join to the measure_value_type ids.
                    # this converts to long format.
                    clean_df = pd.melt(clean_df, id_vars=['Measure'], var_name='measure_value_name',
                                       value_name='measure_value')

            except:
                print('What if Section headers do not match.  Could not isolate column header.')
                return

                # print(clean_df.head())
            
            
            
            # join calculator domain ranks dataframe with dataframe with measure names and ids from db
            clean_df = pd.merge(clean_df, measure_df, how='left', left_on=['Measure'], right_on=['measure_name'])

            clean_df['calc_id'] = current_calc_id
            clean_df['hospital_id'] = current_hospital_id
            clean_df['period_id'] = period_id_result

            # join calculator measure value dataframe with dataframe with measure_value_types and ids from db
            clean_df = pd.merge(clean_df, measure_value_id_result, how='left', left_on=['measure_value_name'],
                                right_on=['measure_value_name'])

            # Because these values do not have n events or event_type_id, we must create NULL placeholders

            clean_df["numerator"] = 0
            clean_df["numerator_event_type_id"] = 0
            clean_df["denominator"] = 0
            clean_df["denominator_event_type_id"] = 0

            # clean_df["n_events"] = clean_df["n_events"].replace(0,None)
            # clean_df["event_type_id"] = clean_df["event_type_id"].replace(0,None)

            # clean_df["n_events"] = clean_df["n_events"].replace({0: None})

            # reorder columns to match db
            cols = ['calc_id', 'hospital_id', 'measure_id', 'period_id', 'measure_value_id', 'measure_value', 'numerator',
                    'numerator_event_type_id','denominator','denominator_event_type_id']
            clean_df = clean_df[cols]

            # append result dataframe to frames_list in order to union later.
            frames_list.append(clean_df)

    result_df = pd.concat(frames_list)
    result_df = result_df.drop_duplicates()
    
    #if loading original calculator release values, we will 
    #also be loading the Metric Values.  Because of this, we need to
    #convert equity domain text values into numeric value codes.  Fail = 0, Warning = 1, Pass = 2
    result_df['measure_value'] = result_df['measure_value'].apply(clean_equity_domain_values)
    
    #UL003
    #In the 2020 scorecards, Vizient has kept the z-scores as '-' for all Equity Domain metrics even
    #if you have a score so, basically, they do not provide a z-score for this domain.
    #So, we need to drop '-' rows again before converting the measure values to numeric.
    result_df = result_df[result_df['measure_value'] != '-']
    
    #print(result_df[result_df['measure_value'] == 'Error'])
    #make sure measure_value column is numeric
    result_df['measure_value'] = pd.to_numeric(result_df['measure_value'])

    return (result_df)


##############################################################################################################################

#function takes the results of parse_calculator_what_if_section() and inserts them into the measure_value table

def insert_what_if_section(df):

    # connect to the NM_Analytics database
    conn = pyodbc.connect('Driver={SQL Server};'
                          'Server=edw00pd05wva.corp.nm.org\EDWIDS1;'
                          'Database=NM_Analytics_Prototype;'
                          'Trusted_Connection=yes;')

    cursor = conn.cursor()

    for index, row in df.iterrows():
        if row['measure_value'] is not None:
            cursor.execute(
                "INSERT INTO NM_Analytics_Prototype.vizient_qa.measure_values([calc_id],[hospital_id],[measure_id],[period_id],[measure_value_id],[measure_value],[numerator],[numerator_event_type_id],[denominator],[denominator_event_type_id]) values (?,?,?,?,?,?,?,?,?,?)",
                row['calc_id'], row['hospital_id'], row['measure_id'], row['period_id'], row['measure_value_id'],
                row['measure_value'], row['numerator'], row['numerator_event_type_id'],row['denominator'],row['denominator_event_type_id'])
            conn.commit()
            #print('insert successful.')

    cursor.close()
    conn.close()
    print('done inserting rows.')


##############################################################################################################################
##############################################################################################################################
##############################################################################################################################
#functions used to create the calculator value architecture.  These functions will
#populate the calc_measure_values table.  The main purpose of this table is to
#store lookup values for which values appear on each hospital's calculator along with it's report hyperlink
#and a flag to indicate whether the value was 'LV' at the time of baseline.

##############################################################################################################################
#this function takes a file path and file name then returns data from one of the _links.xlsx files
#that store each hospital calculator TC sheet hyperlink data.
def get_hosp_meas_hyperlinks(path_obj, file_obj):
    file_loc = os.path.abspath(os.path.join(path_obj, file_obj))
    #UL007
    dataframe_ob = pd.DataFrame(pd.read_excel(file_loc, sheet_name="Sheet1",engine='openpyxl'))

    keep_cols = ['Formal Name', 'Hyperlink']

    dataframe_ob = dataframe_ob[keep_cols]

    dataframe_ob = dataframe_ob[dataframe_ob['Hyperlink'] != 0]

    return (dataframe_ob)

##############################################################################################################################

#this function takes a hospital medicare id string and returns a standard hyperlink file name for that hospital.
def select_link_file(hosp_medicare_id):
    if hosp_medicare_id == '140130':
        link_file_name = 'lfh_links.xlsx'
    elif hosp_medicare_id == '140211':
        link_file_name = 'dch_links.xlsx'
    elif hosp_medicare_id == '140242':
        link_file_name = 'cdh_links.xlsx'
    elif hosp_medicare_id == '140281':
        link_file_name = 'nmh_links.xlsx'
    elif hosp_medicare_id == '140286':
        link_file_name = 'kish_links.xlsx'
    elif hosp_medicare_id == '141340':
        link_file_name = 'vwh_links.xlsx'
    #UL003  fy21 adding mchenry and huntley
    elif hosp_medicare_id == '149916':
        link_file_name = 'hh_links.xlsx'
    elif hosp_medicare_id == '140116':
        link_file_name = 'mch_links.xlsx'
    elif hosp_medicare_id == '140062':
        link_file_name = 'palos_links.xlsx'
    return(link_file_name)


##############################################################################################################################

#this function is used to dynamically fill a pandas dataframe column depending on whether a metric value is 'LV' or not.
#this is used to create the baseline_lv_flg column in the calc_measure_values table.
def lv_flg(c):
  if c['Metric Value'] == 'LV':
    return 1
  else:
    return 0

##############################################################################################################################
#UL003

#New function to help parse calculator calc_measure_values percentile distribution.    

def find_ws_perc_dist_xls_header_row(xl_file_obj):
    #UL007
    xl_file_df = pd.read_excel(xl_file_obj, sheet_name="Percentile Distribution",header=None,engine='openpyxl')
    #get the name of the first column
    first_col_name = xl_file_df.columns[0]
    #get the index of the first row equal to 'Keyword' or 'Domain'
    
    if 'Critical' not in xl_file_obj:
        first_row_loc = xl_file_df.index[xl_file_df[first_col_name].str.lower() == 'keyword'].tolist()
        #print(first_row_loc)
    else:
        first_row_loc = xl_file_df.index[xl_file_df[first_col_name].str.lower() == 'domain'].tolist()
        #print(first_row_loc)
    #return index
    return(first_row_loc[0])


##############################################################################################################################
    
#UL003

#New function to help parse calculator calc_measure_values percentile distribution.   

def open_calc_perc_dist(path_obj, file_obj):
    #join the file path and the file name
    file_loc2 = os.path.abspath(os.path.join(path_obj,file_obj))
    #fine the index of the header row in order to know how many rows to skip
    
    perc_dist_header_row = find_ws_perc_dist_xls_header_row(file_loc2)
    #open the excel file 'Metric Weights' sheet.
    #UL007
    opened_excel_file = pd.read_excel(file_loc2, sheet_name="Percentile Distribution",skiprows =perc_dist_header_row,engine='openpyxl')
    
    print(opened_excel_file.head())
    
    if 'Critical' in file_obj and ('2019' in file_obj or '2020' in file_obj):
        opened_excel_file = opened_excel_file[['Metric','score_10','score_50','score_90']]
        opened_excel_file = opened_excel_file.rename(columns={'Metric': 'Keyword','score_10':'P10','score_50':'P50','score_90':'P90'})
        opened_excel_file['Keyword'] = opened_excel_file['Keyword'].str.replace(' ','').str.upper()
        
    if 'Critical' in file_obj:
        opened_excel_file = opened_excel_file[['Metric','P10','P50','P90']]
        opened_excel_file = opened_excel_file.rename(columns={'Metric': 'Keyword'})
        opened_excel_file['Keyword'] = opened_excel_file['Keyword'].str.replace(' ','').str.upper()
    else:
        if opened_excel_file.columns[0] == 'keyword':
            opened_excel_file = opened_excel_file.rename(columns={'keyword': 'Keyword'})
        opened_excel_file = opened_excel_file[['Keyword','P10','P50','P90']]
        opened_excel_file['Keyword'] = opened_excel_file['Keyword'].str.replace(' ','').str.upper()
    return(opened_excel_file)


##############################################################################################################################



# This is the core function that parses a calculator file and returns a cleaned dataframe ready for db inserts into the
# calc_measure_values table.
'''
def parse_calculator_measures_and_links(period_type, period_end_dts):
    hyperlink_loc = input("Enter path of hyperlink files.")

    # query the database and get all measure ids
    measure_df = get_all_measures()

    # function takes the path of calculator file folder and returns the path and list of all file names
    path_and_files = gather_calculator_path_and_files()

    # store the file path in a variable
    calc_file_path = path_and_files[0]

    # empty list to store measure weight dataframes in.
    frames_list = []

    # query the db to get period id
    conn = pyodbc.connect('Driver={SQL Server};'
                          'Server=edw00pd05wva.corp.nm.org\EDWIDS1;'
                          'Database=clarity;'
                          'Trusted_Connection=yes;')
    period_type_id = """
    SELECT
    pt.period_type_id
    FROM
    NM_Analytics_Prototype.vizient_qa.period_types as pt
    WHERE
    pt.period_type_nm = '%s'
    """ % period_type

    period_type_id_result = pd.DataFrame(pd.read_sql(period_type_id, conn))['period_type_id'].values[0]

    period_id = '''''''
    SELECT 
    p.period_id 
    FROM NM_Analytics_Prototype.vizient_qa.periods as p 
    WHERE p.period_end_date = '%s'
    and p.period_type = '%s'
    ''''''' % (period_end_dts, period_type_id_result)

    period_id_result = pd.DataFrame(pd.read_sql(period_id, conn))['period_id'].values[0]

    conn.close()

    # iterate over calculator file names
    for i, item in enumerate(path_and_files[1]):
        # look up hospital ID in the database.
        # step one.  Get the medicare ID from the file name
        hospital_medicare_id = item.split('_')[1]
        # get period name from calc file name
        period_nm = item.split('_')[2].replace(" ", "%")
        # get year from calc file name
        year = item.split('_')[3].split('.')[0]

        # define the calculator id query
        calc_id_query = "SELECT calc_id,calc_nm from NM_Analytics_Prototype.vizient_qa.calculator where calc_nm like '%" + year + "%" + period_nm + "'"
        # create connection to the ms sql database
        conn = pyodbc.connect('Driver={SQL Server};'
                              'Server=edw00pd05wva.corp.nm.org\EDWIDS1;'
                              'Database=clarity;'
                              'Trusted_Connection=yes;')
        # define the hospital query
        find_hosp_by_medicare_id_query = 'SELECT hospital_id, hospital_medicare_id from NM_Analytics_Prototype.vizient_qa.hospitals WHERE hospital_medicare_id = ' + hospital_medicare_id

        # query the database
        hospital_ids = pd.DataFrame(pd.read_sql(find_hosp_by_medicare_id_query, conn))
        calc_ids = pd.DataFrame(pd.read_sql(calc_id_query, conn))

        # close the connection
        conn.close()
        # store the hospital id in a variable
        current_hospital_id = hospital_ids['hospital_id'][0]
        current_calc_id = calc_ids['calc_id'][0]

        # get hyperlink file based on hospital medicare id
        link_file_name = select_link_file(hospital_medicare_id)

        # open the hyperlink excel file and return only the necessary columns
        hosp_hyperlinks = get_hosp_meas_hyperlinks(hyperlink_loc, link_file_name)

        # function that takes a calculator file path and name
        # and returns a pandas dataframe of just the domain name, measure name,
        # "What if" section, "Current ranking" section and "Target Performance Evaluation" section
        clean_df = parse_calculator_and_return_clean_dataframe(calc_file_path, item)

        # only need the measure name and metric value
        keep_cols = ['Measure', 'Metric Value']
        clean_df = clean_df[keep_cols]

        # join calculator domain ranks dataframe with dataframe with measure names and ids from db
        clean_df = pd.merge(clean_df, measure_df, how='left', left_on=['Measure'], right_on=['measure_name'])

        # store the calc_id, hospital_id and period_id
        clean_df['calc_id'] = current_calc_id
        clean_df['hospital_id'] = current_hospital_id
        clean_df['period_id'] = period_id_result

        # join calculator measure value dataframe with dataframe with measure_value_types and ids from db
        clean_df = pd.merge(clean_df, hosp_hyperlinks, how='left', left_on=['Measure'], right_on=['Formal Name'])

        # if the calculator value is 'LV', store a flag for later reporting
        clean_df['baseline_lv_flg'] = clean_df.apply(lv_flg, axis=1)

        # fill nas in hyperlink column
        clean_df['Hyperlink'].fillna('None', inplace=True)

        # rename the hyperlink column to match the db
        clean_df = clean_df.rename(columns={'Hyperlink': 'report_hyperlink'})

        # reorder columns to match db
        cols = ['calc_id', 'hospital_id', 'measure_id', 'period_id', 'baseline_lv_flg', 'report_hyperlink']
        clean_df = clean_df[cols]

        # append result dataframe to frames_list in order to union later.
        frames_list.append(clean_df)

    result_df = pd.concat(frames_list)
    result_df = result_df.drop_duplicates()

    # result_df['measure_value'] = pd.to_numeric(result_df['measure_value'])

    # pd.set_option('display.max_rows', 20)
    # print(result_df.dtypes)
    # result_df.to_csv('C:/Users/NM184423/Desktop/final_hyperlinks_section.csv')
    return (result_df)
'''

#UL003
def parse_calculator_measures_and_links(period_type, period_end_dts):
    hyperlink_loc = input("Enter path of hyperlink files.")

    # query the database and get all measure ids
    measure_df = get_all_measures()
    
    #get measure keywords from the hyperlink files because the measure weights sheet only has keywords
    #while the database only has formal names of measures.
    #all_measure_names = get_report_measure_keywords()
    
    #join the measure ids to the measure keyword dataframes
    #measure_names_keys_ids = pd.merge(measure_df, all_measure_names, how='left', left_on=['measure_name'], right_on=['Formal Name'])

    #remove rows that didn't join.  This should only be measures not in use or 'informational only' measures.
    #measure_names_keys_ids = measure_names_keys_ids[pd.notnull(measure_names_keys_ids['Formal Name'])]
    
    #measure_names_keys_ids['Keyword/Metric'] = measure_names_keys_ids['Keyword/Metric'].str.replace(' ','').str.upper()
    
    
    # function takes the path of calculator file folder and returns the path and list of all file names
    path_and_files = gather_calculator_path_and_files()

    # store the file path in a variable
    calc_file_path = path_and_files[0]

    # empty list to store measure weight dataframes in.
    frames_list = []

    # query the db to get period id
    conn = pyodbc.connect('Driver={SQL Server};'
                          'Server=edw00pd05wva.corp.nm.org\EDWIDS1;'
                          'Database=clarity;'
                          'Trusted_Connection=yes;')
    period_type_id = """
    SELECT
    pt.period_type_id
    FROM
    NM_Analytics_Prototype.vizient_qa.period_types as pt
    WHERE
    pt.period_type_nm = '%s'
    """ % period_type

    period_type_id_result = pd.DataFrame(pd.read_sql(period_type_id, conn))['period_type_id'].values[0]

    period_id = '''
    SELECT 
    p.period_id 
    FROM NM_Analytics_Prototype.vizient_qa.periods as p 
    WHERE p.period_end_date = '%s'
    and p.period_type = '%s'
    ''' % (period_end_dts, period_type_id_result)

    period_id_result = pd.DataFrame(pd.read_sql(period_id, conn))['period_id'].values[0]

    conn.close()

    # iterate over calculator file names
    for i, item in enumerate(path_and_files[1]):
        #UL003
        #Adding additional logic to handle Critical Access calculators.  For some reason, Vizient decided to name the
        #Critical Access calculators completely differently so we need to use a different method.
        #first split up the file name, then iterate over the subsections of the name to find the part with the medicare id
        #All NM hospital medicare ids start with 140 or 141.
        for i, subitem in enumerate(item.split('_')):
            if '140' in subitem or '141' in subitem or '149' in subitem:
                hospital_medicare_id_index = i
            else:
                pass

        # look up hospital ID in the database.
        # step one.  Get the medicare ID from the file name
        hospital_medicare_id = item.split('_')[hospital_medicare_id_index]

        for i, subitem in enumerate(item.split('_')):
            if 'PERIOD' in subitem.upper() or 'ANNUAL' in subitem.upper():
                period_nm_index = i
            else:
                pass

        #have to split the period name because in 2019 there is a space between period and number.
        #In 2020, there 
        period_nm_split = re.split('(\d+)',item.split('_')[period_nm_index])

        if period_nm_split[0].replace(" ","") == 'Annual':
            period_str = 'Period'
            period_num = '4'
        else:
            period_str = period_nm_split[0].replace(" ","")
            period_num = period_nm_split[1].replace(" ","")

        #['Period ', '3', '']
        # get period name from calc file name
        #period_nm = item.split('_')[period_nm_index].replace(" ", "%")

        period_nm = period_str + '%' + period_num


        # get year from calc file name
        # The section containing the year will start with 20...for the next 80 years...
        # and it will end with .xlsm because it contains the file extension.
        for i, subitem in enumerate(item.split('_')):

            if subitem.startswith('20') and subitem.upper().endswith('.XLSM'):
                year_index = i
            else:
                pass

        
        year = item.split('_')[year_index].split('.')[0]

        '''

        hospital_medicare_id = item.split('_')[1]
        # get period name from calc file name
        period_nm = item.split('_')[2].replace(" ", "%")
        # get year from calc file name
        year = item.split('_')[3].split('.')[0]

        print(hospital_medicare_id)
        print(period_nm)
        print(year)
        '''

        # define the calculator id query
        calc_id_query = "SELECT calc_id,calc_nm from NM_Analytics_Prototype.vizient_qa.calculator where calc_nm like '%" + year + "%" + period_nm + "'"
        # create connection to the ms sql database
        conn = pyodbc.connect('Driver={SQL Server};'
                              'Server=edw00pd05wva.corp.nm.org\EDWIDS1;'
                              'Database=clarity;'
                              'Trusted_Connection=yes;')
        # define the hospital query
        find_hosp_by_medicare_id_query = 'SELECT hospital_id, hospital_medicare_id from NM_Analytics_Prototype.vizient_qa.hospitals WHERE hospital_medicare_id = ' + hospital_medicare_id

        # query the database
        hospital_ids = pd.DataFrame(pd.read_sql(find_hosp_by_medicare_id_query, conn))
        calc_ids = pd.DataFrame(pd.read_sql(calc_id_query, conn))

        # close the connection
        conn.close()


        # store the hospital id in a variable
        current_hospital_id = hospital_ids['hospital_id'][0]
        current_calc_id = calc_ids['calc_id'][0]

        # get hyperlink file based on hospital medicare id
        link_file_name = select_link_file(hospital_medicare_id)

        # open the hyperlink excel file and return only the necessary columns
        hosp_hyperlinks = get_hosp_meas_hyperlinks(hyperlink_loc, link_file_name)
        #clean up formal name column for better joins
        hosp_hyperlinks['Formal Name'] = hosp_hyperlinks['Formal Name'].str.replace(' ','').str.replace('-','_').str.upper()
        hosp_hyperlinks.loc[hosp_hyperlinks["Formal Name"] == 'COLONSCOPYREVISITSWITHIN7_DAYS', "Formal Name"] = 'COLONOSCOPYREVISITSWITHIN7_DAYS'
        #print(current_hospital_id)
        #UL003
        #Valley West calculators do not have Possible Rank or Target Rank on the main page for some reason.  Thanks Vizient.
        #Therefore, we need to sum up the metric score and lookup the closes rank on the hidden ranks worksheet.
        if current_hospital_id == 356:  #Valley West db ID.
            
            #UL003
            #Since the metric weights sheet does not have the metric formal name, it only has the metric keyword,
            #we need a way to join back to the database (which only stores the formal name).  In the calculator
            #worksheet, there is a hidden column called 'Metric', which has the keyword and a column called 'Measure'
            #which has the formal name.  Use this to get the mapping.


            # open the calculator file and isolate the 'Calculator' sheet.
            ws = grab_calc_worksheet_calc_sheet(calc_file_path, item)

            #find the coordinates of the 'Domain' header cell of the 'Calculator' worksheet
            domain_coord = find_calc_phrase_cell_coords(ws,'Domain')

            #store worksheet object as pandas dataframe
            calc_df = pd.DataFrame(ws.values)

            #rename column headers using the 'Domain' header row coordinates
            calc_df = calc_df.rename(columns=calc_df.iloc[domain_coord[0]-1])

            #use row coordinates of "Domain" header row to drop all rows before it.
            row_drop_list = [i for i in range(domain_coord[0])]
            calc_df = calc_df.drop(row_drop_list)

            #drop colum where all values are NA
            calc_df = calc_df.dropna(axis='columns', how="all")

            #drop rows where all values are NA
            calc_df = calc_df.dropna(how="all")

            #remove copyright row at the bottom.
            calc_df = calc_df[~calc_df['Domain'].str.contains('Copyright', na=False)]

            #only need 2 columns
            calc_df = calc_df[['Metric',"Measure"]]
            
            #clean things up to ensure best join
            calc_df['Metric'] = calc_df['Metric'].str.replace(' ','').str.upper()
            calc_df['Measure'] = calc_df['Measure'].str.replace(' ','').str.replace('-','_').str.upper()
            #Critical Access calcs still have a type-o.  For 3+ years...
            calc_df.loc[calc_df["Measure"] == 'COLONSCOPYREVISITSWITHIN7_DAYS', "Measure"] = 'COLONOSCOPYREVISITSWITHIN7_DAYS'
            
            measure_df['measure_name'] = measure_df['measure_name'].str.replace(' ','').str.replace('-','_').str.upper()
            #join to measure_df dataframe to get db ids
            measure_names_keys_ids = pd.merge(measure_df, calc_df, how='inner', left_on=['measure_name'], right_on=['Measure'])
            
            #UL003
            #Critical Access calculators are different, of course.

            #If the hospital is Valley West, we need to subset the 'what if' section in 
            #a different way.  The column headers are also slightly different
            #and, for some reason, the Z-Score column is completely different.  It mixes
            #actual z-scores with text-based phrases indicating statistical significance.
            #since we cannot store a text phrase in our database column for z-scores, we
            #will convert any non-number value z-score to median z-score (0).


            # open the calculator file and isolate the 'Calculator' sheet.
            ws = grab_calc_worksheet_calc_sheet(calc_file_path, item)

            #find the coordinates of the 'Domain' header cell of the 'Calculator' worksheet
            domain_coord = find_calc_phrase_cell_coords(ws,'Domain')
            #find the coordinates of the 'Target Domain Ranking' header cell of the 'Calculator' worksheet
            what_if_metric_value_coord = find_calc_phrase_cell_coords(ws,' "What if" Performance Evaluation')

            #set row coordinate to 2 below actual in order to locate the 'metric value' header.
            what_if_metric_value_coord[0] = what_if_metric_value_coord[0] + 2

            #Find the LAST instance of 'Domain Rank'.  Get a list of all instances, then take last index.
            domain_rank_coord = find_calc_phrase_cell_coords_multiple(ws,'Domain Rank')
            domain_rank_coord = domain_rank_coord[-1]
            #store worksheet object as pandas dataframe
            m_df = pd.DataFrame(ws.values)

            #use column coordinates of "Domain" header to drop all columns to the left of it.
            col_drop_list1 = [i for i in range(domain_rank_coord[1],m_df.shape[1])]
            m_df = m_df.drop(col_drop_list1,axis=1)

            #use column coordinates of "Domain" header to drop all columns to the left of it.
            col_drop_list = [i for i in range(domain_coord[1]-1)]
            m_df = m_df.drop(col_drop_list,axis=1)

            #rename column headers using the 'Domain' header row coordinates
            m_df = m_df.rename(columns=m_df.iloc[domain_coord[0]-1])

            #use row coordinates of "Domain" header row to drop all rows before it.
            row_drop_list = [i for i in range(domain_coord[0])]
            m_df = m_df.drop(row_drop_list)

            #drop colum where all values are NA
            m_df = m_df.dropna(axis='columns', how="all")

            #drop rows where all values are NA
            m_df = m_df.dropna(how="all")

            domain_and_measure_indices = [0,1]

            #indexes have now changed.  Find the last Domain Rank column header.
            domain_rank_list = []
            for i, thing in enumerate(m_df.columns):
                if thing == 'Domain Rank':
                    domain_rank_list.append(i)
            #find last 'Domain Rank.'
            new_domain_rank_coord = domain_rank_list[-1]
            #keep_cols = [i for i in range(df.columns.get_loc(' Domain Rank Result')-4,df.columns.get_loc('Target Domain Ranking')+1)]
            keep_cols = [i for i in range(new_domain_rank_coord-4,new_domain_rank_coord+1)]

            #subset columns to keep only the ones we care about.
            final_keep_cols = domain_and_measure_indices + keep_cols
            m_df = m_df.iloc[:,final_keep_cols]

            #remove copyright row at the bottom.
            m_df = m_df[~m_df['Domain'].str.contains('Copyright', na=False)]

            #domain names and ranks are in merged excel rows so we now need to forward fill them so we have a value every row.
            m_df['Domain'] = m_df['Domain'].fillna(method='ffill')
            m_df['Domain Rank'] = m_df['Domain Rank'].fillna(method='ffill')
            
            clean_df = m_df
            
            clean_df = clean_df[clean_df['Domain Rank'] != 'No Rank']
            
            # only need the measure name and metric value
            keep_cols = ['Measure', 'Metric Value']
            clean_df = clean_df[keep_cols]
            
            clean_df['Measure'] = clean_df['Measure'].str.replace(' ','').str.replace('-','_').str.upper()
            
            #get Percentile Distribution sheet info.  
            
            perc_dist_df = open_calc_perc_dist(calc_file_path, item)
            
            clean_df.loc[clean_df["Measure"] == "COLONSCOPYREVISITSWITHIN7_DAYS", "Measure"] = 'COLONOSCOPYREVISITSWITHIN7_DAYS'

            

            # join calculator domain ranks dataframe with dataframe with measure names and ids from db
            clean_df = pd.merge(clean_df, measure_df, how='left', left_on=['Measure'], right_on=['measure_name'])
            
            perc_dist_df = pd.merge(perc_dist_df, measure_names_keys_ids, how='left', left_on=['Keyword'], right_on=['Metric'])
            
            perc_dist_df = perc_dist_df[['Keyword','P10','P50','P90','measure_id']]
            
            # store the calc_id, hospital_id and period_id
            clean_df['calc_id'] = current_calc_id
            clean_df['hospital_id'] = current_hospital_id
            clean_df['period_id'] = period_id_result

            # join calculator measure value dataframe with dataframe with measure_value_types and ids from db
            clean_df = pd.merge(clean_df, hosp_hyperlinks, how='left', left_on=['Measure'], right_on=['Formal Name'])
            
            
            #add percentile distribution
            clean_df = pd.merge(clean_df, perc_dist_df, how='left', left_on=['measure_id'], right_on=['measure_id']) 

            # if the calculator value is 'LV', store a flag for later reporting
            clean_df['baseline_lv_flg'] = clean_df.apply(lv_flg, axis=1)

            # fill nas in hyperlink column
            clean_df['Hyperlink'].fillna('None', inplace=True)

            # rename the hyperlink column to match the db
            clean_df = clean_df.rename(columns={'Hyperlink': 'report_hyperlink'})

            # reorder columns to match db
            cols = ['calc_id', 'hospital_id','Measure','measure_id', 'period_id', 'baseline_lv_flg', 'report_hyperlink','P10','P50','P90']
            clean_df = clean_df[cols]

            # append result dataframe to frames_list in order to union later.
            frames_list.append(clean_df)
            
        else:
            
            #UL003
            #Since the metric weights sheet does not have the metric formal name, it only has the metric keyword,
            #we need a way to join back to the database (which only stores the formal name).  In the calculator
            #worksheet, there is a hidden column called 'Metric', which has the keyword and a column called 'Measure'
            #which has the formal name.  Use this to get the mapping.


            # open the calculator file and isolate the 'Calculator' sheet.
            ws = grab_calc_worksheet_calc_sheet(calc_file_path, item)

            #find the coordinates of the 'Domain' header cell of the 'Calculator' worksheet
            domain_coord = find_calc_phrase_cell_coords(ws,'Domain')

            #store worksheet object as pandas dataframe
            calc_df = pd.DataFrame(ws.values)

            #rename column headers using the 'Domain' header row coordinates
            calc_df = calc_df.rename(columns=calc_df.iloc[domain_coord[0]-1])

            #use row coordinates of "Domain" header row to drop all rows before it.
            row_drop_list = [i for i in range(domain_coord[0])]
            calc_df = calc_df.drop(row_drop_list)

            #drop colum where all values are NA
            calc_df = calc_df.dropna(axis='columns', how="all")

            #drop rows where all values are NA
            calc_df = calc_df.dropna(how="all")

            #remove copyright row at the bottom.
            calc_df = calc_df[~calc_df['Domain'].str.contains('Copyright', na=False)]

            #only need 2 columns
            calc_df = calc_df[['Metric',"Measure"]]

            #clean things up to ensure best join
            calc_df['Metric'] = calc_df['Metric'].str.replace(' ','').str.upper()
            calc_df['Measure'] = calc_df['Measure'].str.replace(' ','').str.replace('-','_').str.upper()
            measure_df['measure_name'] = measure_df['measure_name'].str.replace(' ','').str.replace('-','_').str.upper()
            #join to measure_df dataframe to get db ids
            measure_names_keys_ids = pd.merge(measure_df, calc_df, how='inner', left_on=['measure_name'], right_on=['Measure'])
            
        
            # function that takes a calculator file path and name
            # and returns a pandas dataframe of just the domain name, measure name,
            # "What if" section, "Current ranking" section and "Target Performance Evaluation" section
            clean_df = parse_calculator_and_return_clean_dataframe(calc_file_path, item)
            
            #print('current hospital:',current_hospital_id)
            
            #UL003
            if current_hospital_id not in [472,471]:
                clean_df = clean_df[clean_df[' Domain Rank Result'] != 'No Rank']
            #else:
                #print(clean_df.head())
            #    clean_df = clean_df[~(clean_df[' Domain Rank Result'] == 'No Rank') & (clean_df['Domain'] == 'Equity')]

            # only need the measure name and metric value
            keep_cols = ['Measure', 'Metric Value']
            clean_df = clean_df[keep_cols]
            
            clean_df['Measure'] = clean_df['Measure'].str.replace(' ','').str.replace('-','_').str.upper()
            
            #get Percentile Distribution sheet info.  
            perc_dist_df = open_calc_perc_dist(calc_file_path, item)
            
            perc_dist_df = pd.merge(perc_dist_df, measure_names_keys_ids, how='left', left_on=['Keyword'], right_on=['Metric'])
            
            perc_dist_df = perc_dist_df[['Keyword','P10','P50','P90','measure_id']]

            # join calculator domain ranks dataframe with dataframe with measure names and ids from db
            clean_df = pd.merge(clean_df, measure_df, how='left', left_on=['Measure'], right_on=['measure_name'])

            # store the calc_id, hospital_id and period_id
            clean_df['calc_id'] = current_calc_id
            clean_df['hospital_id'] = current_hospital_id
            clean_df['period_id'] = period_id_result

            # join calculator measure value dataframe with dataframe with measure_value_types and ids from db
            clean_df = pd.merge(clean_df, hosp_hyperlinks, how='left', left_on=['Measure'], right_on=['Formal Name'])
            
            #add percentile distribution
            clean_df = pd.merge(clean_df, perc_dist_df, how='left', left_on=['measure_id'], right_on=['measure_id']) 
            
            print('current hospital:',current_hospital_id)
            print(clean_df.head())
            # if the calculator value is 'LV', store a flag for later reporting
            clean_df['baseline_lv_flg'] = clean_df.apply(lv_flg, axis=1)

            # fill nas in hyperlink column
            clean_df['Hyperlink'].fillna('None', inplace=True)

            # rename the hyperlink column to match the db
            clean_df = clean_df.rename(columns={'Hyperlink': 'report_hyperlink'})

            # reorder columns to match db
            cols = ['calc_id', 'hospital_id','Measure','measure_id', 'period_id', 'baseline_lv_flg', 'report_hyperlink','P10','P50','P90']
            clean_df = clean_df[cols]

            # append result dataframe to frames_list in order to union later.
            frames_list.append(clean_df)

    result_df = pd.concat(frames_list)
    result_df = result_df.drop_duplicates()
    
    #cannot insert 'NaN' into sql server.  Need to convert 'NaN' to None.
    result_df = result_df.where(pd.notnull(result_df), None)
    
    return (result_df)

##############################################################################################################################


# function takes the resulting df from parse_calculator_measures_and_links() function and
# inserts into calc_measure_values

def insert_into_calc_measure_values_tb(df):
    # connect to the NM_Analytics database
    conn = pyodbc.connect('Driver={SQL Server};'
                          'Server=edw00pd05wva.corp.nm.org\EDWIDS1;'
                          'Database=NM_Analytics_Prototype;'
                          'Trusted_Connection=yes;')

    cursor = conn.cursor()

    for index, row in df.iterrows():
        #print(row)
        cursor.execute(
            "INSERT INTO NM_Analytics_Prototype.vizient_qa.calc_measure_values([calc_id],[hospital_id],[measure_id],[period_id],[baseline_lv_flg],[report_hyperlink],[P10],[P50],[P90]) values (?,?,?,?,?,?,?,?,?)",
            row['calc_id'], row['hospital_id'], row['measure_id'], row['period_id'], row['baseline_lv_flg'],
            row['report_hyperlink'],row['P10'],row['P50'],row['P90'])  #UL003  update to insert percentile distribution columns
        conn.commit()

    cursor.close()
    conn.close()
    print('done inserting rows.')